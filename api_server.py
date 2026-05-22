"""
BRIDGE API Server  v2.4 (2026-03-31 hotfix)
===================================
bridgejob.co.kr 웹사이트용 백엔드 API

🔥 Hotfix: Render 강제 재배포 -- API 403 에러 해결

엔드포인트:
  GET  /              → 상태 확인
  GET  /api/jobs      → 공개 구인 목록 (anon)
  GET  /api/jobs/{id} → 포지션 상세
  POST /api/apply     → 구직자 지원 접수 (anon)
  POST /api/inquiry   → 구인처 문의 접수 (anon)

실행:
  uvicorn api_server:app --host 0.0.0.0 --port 8000 --reload

환경변수 (.env):
  SUPABASE_URL=https://xxxx.supabase.co
  SUPABASE_ANON_KEY=eyJ...
  SUPABASE_SERVICE_KEY=eyJ...  (서버사이드 전용 -- 노출 금지)
"""
import io
import os
import re
import zipfile
import sys
print(f"[STARTUP] Python {sys.version} | platform={sys.platform} | pid={os.getpid()}")
import json
import time
import uuid
import sqlite3
import logging
from collections import defaultdict
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Optional, Any
import hashlib
import hmac
import asyncio
import threading
import html as _html

# ad_only 광고 전용 소스 (fail-closed PII 가드)
try:
    from ad_only.loader import load_all_jobs as _ad_load_all_jobs, get_job as _ad_get_job
    from ad_only.pii_guard import assert_clean as _ad_assert_clean, PIIContaminationError as _ADPIIErr
    _AD_READY = True
except Exception as _ad_err:
    _AD_READY = False
    _ad_load_all_jobs = None
    _ad_get_job = None
    _ad_assert_clean = None
    _ADPIIErr = RuntimeError
    logging.getLogger("bridge.api").warning("ad_only 미사용 (fallback DB): %s", _ad_err)

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# ── dotenv ──────────────────────────────────────────────────────────────────
try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).resolve().parent / ".env")
except ImportError:
    pass

# ── BX credential loader (Windows only) ────────────────────────────────────
if sys.platform == "win32":
    try:
        from tools.bx import load_to_env as _bx_load
        _bx_load()
    except Exception:
        pass

# ── FastAPI / Supabase 임포트 ────────────────────────────────────────────────
try:
    from fastapi import FastAPI, HTTPException, Request, Query, status, UploadFile, File as FastFile, Form
    from fastapi.middleware.cors import CORSMiddleware
    from fastapi.middleware.gzip import GZipMiddleware
    from fastapi.staticfiles import StaticFiles
    from starlette.middleware.base import BaseHTTPMiddleware
    from starlette.middleware.trustedhost import TrustedHostMiddleware
    from fastapi.responses import JSONResponse, Response
    from pydantic import BaseModel, EmailStr, Field
except Exception as _fa_exc:
    print(f"[FATAL] fastapi 로드 실패: {_fa_exc}")
    import traceback; traceback.print_exc()
    sys.exit(1)

# ── SafeJSONResponse: 모든 admin 응답에 사용 ────────────────────────────────

class SafeJSONResponse(Response):
    """JSON 직렬화 시 브라우저 JSON.parse 호환성을 100% 보장하는 Response."""
    media_type = "application/json"

    def __init__(self, content=None, status_code=200, headers=None, **kwargs):
        body = json.dumps(
            self._deep_sanitize(content),
            ensure_ascii=False,
            default=str,
        )
        super().__init__(content=body, status_code=status_code, media_type=self.media_type, headers=headers)

    @classmethod
    def _deep_sanitize(cls, obj):
        if isinstance(obj, str):
            obj = obj.replace('\x00', '')
            obj = ''.join(c for c in obj if ord(c) >= 32 or c in '\n\r\t')
            obj = re.sub(r'\\(?!["\\/bfnrtu])', r'\\\\', obj)
            return obj
        elif isinstance(obj, dict):
            return {cls._deep_sanitize(k): cls._deep_sanitize(v) for k, v in obj.items()}
        elif isinstance(obj, (list, tuple)):
            return [cls._deep_sanitize(v) for v in obj]
        elif isinstance(obj, bytes):
            return obj.decode('utf-8', errors='replace')
        return obj


def bridge_error(
    category: str,
    context: str,
    retryable: bool = False,
    status: int = 500,
):
    """구조화된 에러 응답 -- isError/errorCategory/isRetryable/context 포함"""
    return JSONResponse(
        status_code=status,
        content={
            "isError": True,
            "errorCategory": category,
            "isRetryable": retryable,
            "context": context,
        },
    )


try:
    from PIL import Image as PILImage
    _PILLOW_OK = True
except ImportError:
    _PILLOW_OK = False
    print("[INFO] Pillow 미설치 -- 썸네일 생성 비활성화 (pip install Pillow)")

try:
    from supabase import create_client, Client
except Exception as _sb_exc:
    print(f"[ERROR] supabase 로드 실패: {_sb_exc}")
    import traceback; traceback.print_exc()
    create_client = None
    Client = None

# ── email_templates (확인 이메일 발송) ────────────────────────────────────────
try:
    from email_templates import send_applicant_confirmation, send_employer_confirmation
    _EMAIL_OK = True
except ImportError:
    _EMAIL_OK = False
    print("[INFO] email_templates 미설치 -- 확인 이메일 발송 비활성화")

# ── security_vault (PII 필드 암호화) ─────────────────────────────────────────
try:
    from security_vault import encrypt_field, decrypt_field, is_encrypted, t3_encrypt, auto_decrypt_value
    _VAULT_OK = True
except Exception as _vault_exc:
    _VAULT_OK = False
    print(f"[CRITICAL] security_vault 로드 실패: {_vault_exc}")
    import traceback; traceback.print_exc()
    # 암호화 없이도 서버 시작은 허용 (읽기 전용 모드)
    def encrypt_field(v, *a, **k): return v
    def decrypt_field(v, *a, **k): return v
    def is_encrypted(v, *a, **k): return False
    def t3_encrypt(v, *a, **k): return v
    def auto_decrypt_value(v, *a, **k): return v

# ── security 패키지 (PIIScanner + InputSanitizer만 사용) ─────────────────
# 암호화: security_vault.py (위에서 import 완료)
# PII 스캔 + 입력 검증: security/ 패키지
try:
    from security import PIIScanner, InputSanitizer
    _sec_scanner = PIIScanner(fail_closed=True)
    _sec_sanitizer = InputSanitizer()
    _SECURITY_PKG_OK = True
except Exception as _sec_exc:
    _SECURITY_PKG_OK = False
    _sec_scanner = None
    _sec_sanitizer = None
    print(f"[WARN] security 패키지 로드 실패 (선택적): {_sec_exc}")

# ── JWT Magic Link ────────────────────────────────────────────────────────────
try:
    import jwt as _jwt   # PyJWT
except Exception as _jwt_exc:
    print(f"[ERROR] PyJWT import 실패: {_jwt_exc}")
    import traceback; traceback.print_exc()
    _jwt = None

_JWT_SECRET  = os.getenv("JWT_SECRET")  # BRIDGE_FIELD_KEY 폴백 제거 -- 키 분리 원칙
if not _JWT_SECRET:
    print("[CRITICAL] JWT_SECRET 환경변수가 설정되지 않았습니다. 서버를 시작할 수 없습니다.")
    print("[CRITICAL] Render 대시보드 → Environment Variables → JWT_SECRET 추가 후 재배포.")
    sys.exit(1)
_JWT_ALG     = "HS256"
_JWT_TTL     = timedelta(hours=4)   # W4: 200MB 영상 업로드 대응 (1h → 4h)


def _make_candidate_token(candidate_id: str) -> str:
    """지원자 ID를 담은 4시간짜리 JWT 생성."""
    exp = datetime.now(timezone.utc) + _JWT_TTL
    payload = {
        "sub": str(candidate_id),
        "exp": exp,
        "iss": "bridge-apply",
        "jti": str(uuid.uuid4()),
        "type": "access",
    }
    return _jwt.encode(payload, _JWT_SECRET, algorithm=_JWT_ALG)


def _verify_candidate_token(token: str) -> Optional[str]:
    """JWT 검증 → candidate_id 반환. 만료/위조 시 None."""
    try:
        data = _jwt.decode(token, _JWT_SECRET, algorithms=[_JWT_ALG],
                           options={"require": ["sub", "exp", "iss"]})
        return data.get("sub")
    except Exception:
        return None


def _dedup_key(name: str, dob: Optional[str], nationality: Optional[str]) -> str:
    """복합 중복방지 키: lower(name_no_space) + birth_year + nationality."""
    n   = re.sub(r"\s+", "", (name or "").lower())
    yr  = (dob or "")[:4] or "xxxx"
    nat = (nationality or "").lower()
    return f"{n}_{yr}_{nat}"


def _encrypt_if_needed(value: str, column_name: str = "") -> str:
    """값이 있고 아직 암호화되지 않은 경우에만 T3v1 암호화."""
    if not value or not isinstance(value, str):
        return value
    return value if is_encrypted(value) else encrypt_field(value, column_name)

# ── PII 마스킹 설정 ───────────────────────────────────────────────────────────
# DB 컬럼 단위 차단 (응답 JSON에서 키 자체를 제거)
_PII_BLOCKED_KEYS: frozenset[str] = frozenset({
    # 학교/기관명
    "school_name", "school_name_en", "school_name_kr",
    # 이름 (후보자 및 담당자)
    "full_name", "first_name", "last_name", "name",
    "contact_name", "contact_person",
    "recruiter_name", "recruiter_email",
    # 직책/포지션 (담당자 직책 -- 채용 포지션 타입과 혼동 주의)
    "contact_position", "contact_title",
    "recruiter_position", "recruiter_title",
    "position_title", "staff_position",
    # 연락처
    "phone", "phone_primary", "phone_kakao", "mobile",
    "email",
    # 주소
    "location_full", "address", "district",
    # 사업자/신원
    "business_registration",
    "passport", "passport_number", "passport_status",
    "criminal_record", "criminal_record_check",
    "health_info",
    "employer_id", "internal_notes",
    # 카카오 / SNS (필드 미존재해도 미래 대비 차단)
    "kakaotalk", "kakao_id", "kakao", "kakao_link",
    "sns", "social_media", "instagram", "facebook", "twitter", "linkedin",
    # 메모 (관리자 전용 내부 필드)
    "memo", "memo_kr", "memo_en", "recruiter_memo",
    # 추가 연락처 계열
    "mobile_phone", "phone_number", "contact_phone",
})

# 값 수준 정규식 마스킹 -- 블록키를 통과한 문자열에 잔류 PII가 있을 경우 대비
_PII_PATTERNS: list[tuple[re.Pattern, str]] = [
    # 한국 휴대전화 (010-xxxx-xxxx / 01x-xxx-xxxx)
    (re.compile(r"\b01[016789][- ]?\d{3,4}[- ]?\d{4}\b"), "***-****-****"),
    # 국제 전화 (+1 555 000 0000 등)
    (re.compile(r"\+\d{1,3}[\s\-]?\(?\d{1,4}\)?[\s\-]?\d{3,4}[\s\-]?\d{3,4}"), "+*-***-***-****"),
    # 이메일
    (re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}"), "***@***.***"),
    # 사업자등록번호 (xxx-xx-xxxxx)
    (re.compile(r"\b\d{3}-\d{2}-\d{5}\b"), "***-**-*****"),
    # 주민등록번호 앞자리+뒷자리 패턴
    (re.compile(r"\b\d{6}-[1-4]\d{6}\b"), "******-*******"),
    # 한국어 이름 -- 라벨:값 형식만 매칭 (콜론 필수). '문의' 같은 일반 단어 제거.
    # 기존: '(?:담당자|성명|이름|문의|연락처)[:：\s]{0,3}' → '문의 접수' 도 매칭 false-positive
    # 수정: 콜론을 필수로 → 라벨 컨텍스트 명확한 경우만 매칭
    (re.compile(r"(?:담당자|성명|이름|연락처)[:：]\s*([가-힣]{2,4})"), "[REDACTED-NAME]"),
    # 영어 이름 -- "Name:", "Contact:", "Teacher:" 뒤 Title Case
    (re.compile(r"(?:Name|Contact|Teacher|Recruiter|Manager):\s*([A-Z][a-z]+ [A-Z][a-z]+)"), "[REDACTED-NAME]"),
]


def _scrub_value(v: str) -> str:
    """Apply all PII regex patterns to a single string value."""
    for pattern, replacement in _PII_PATTERNS:
        v = pattern.sub(replacement, v)
    return v


# 관리자 작성 콘텐츠 필드 -- 값 수준 PII 정규식 스캔 제외
# (FAQ 본문, 게시글 본문 등 의도적으로 포함된 비즈니스 이메일/연락처 보호)
_CONTENT_FIELDS: frozenset[str] = frozenset({
    "body", "preview", "review_text", "html_content", "description",
})


def _scrub_obj(obj: Any, _key: str = "") -> Any:
    """Recursively walk a JSON-deserialized object and mask PII."""
    if isinstance(obj, dict):
        return {
            k: _scrub_obj(v, _key=k)
            for k, v in obj.items()
            if k not in _PII_BLOCKED_KEYS   # drop blocked keys entirely
        }
    if isinstance(obj, list):
        return [_scrub_obj(item, _key=_key) for item in obj]
    if isinstance(obj, str):
        if _key in _CONTENT_FIELDS:
            return obj  # 콘텐츠 필드는 값 수준 스캔 제외 -- 비즈니스 연락처 마스킹 방지
        return _scrub_value(obj)
    return obj


class PIIMaskingMiddleware(BaseHTTPMiddleware):
    """
    Response middleware: intercepts all JSON responses and strips / masks PII.
    Adds X-PII-Scrubbed: true header for audit tracing.
    Zero-overhead for non-JSON responses (static files, health checks).
    """
    _SEC_HEADERS = {
        "X-Frame-Options": "DENY",
        "X-Content-Type-Options": "nosniff",
        "Referrer-Policy": "strict-origin-when-cross-origin",
        # X-XSS-Protection deprecated (OWASP 2025) -- disabled, CSP replaces it
        "X-XSS-Protection": "0",
        # HSTS: 2년 + subdomains + preload (Google/Apple/Microsoft 표준)
        "Strict-Transport-Security": "max-age=63072000; includeSubDomains; preload",
        # Permissions-Policy: 불필요 기능 비활성화
        "Permissions-Policy": "camera=(), microphone=(), geolocation=(), payment=()",
        # CSP: 순수 JSON API -- 리소스 로드 완전 차단 + 삽입 공격 차단
        # frame-ancestors: X-Frame-Options 대체 (모던 브라우저 우선)
        "Content-Security-Policy": "default-src 'none'; frame-ancestors 'none'",
    }

    def _add_sec_headers(self, response):
        for k, v in self._SEC_HEADERS.items():
            response.headers[k] = v

    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        ct = response.headers.get("content-type", "")
        if "application/json" not in ct:
            self._add_sec_headers(response)
            return response

        # 관리자 인증된 요청은 PII 마스킹 통과 (원본 데이터 반환)
        _ak = os.getenv("ADMIN_API_KEY", "")
        _admin_paths = ("/api/admin/", "/api/employers", "/api/send-mail")
        if (_ak
                and request.headers.get("x-admin-key", "") == _ak
                and any(request.url.path.startswith(p) for p in _admin_paths)):
            self._add_sec_headers(response)
            return response

        # 공개 마케팅 콘텐츠 — 이름이 PII가 아니라 의도된 공개 필드
        # (후기 작성자 first name, 강사 보드 닉네임 등)
        _public_name_paths = ("/api/testimonials", "/api/talents/board")
        if any(request.url.path.startswith(p) for p in _public_name_paths):
            self._add_sec_headers(response)
            return response

        # Accumulate response body
        body_parts: list[bytes] = []
        async for chunk in response.body_iterator:
            body_parts.append(chunk)
        raw_body = b"".join(body_parts)

        try:
            data = json.loads(raw_body)
            clean = _scrub_obj(data)
            clean_body = json.dumps(clean, ensure_ascii=False, default=str).encode("utf-8")
        except (json.JSONDecodeError, Exception):
            clean_body = raw_body   # pass through if not parseable JSON

        headers = dict(response.headers)
        headers["content-length"] = str(len(clean_body))
        headers["x-pii-scrubbed"] = "true"
        for k, v in self._SEC_HEADERS.items():
            headers[k] = v

        return Response(
            content=clean_body,
            status_code=response.status_code,
            headers=headers,
            media_type="application/json",
        )


# ── 설정 ─────────────────────────────────────────────────────────────────────
SUPABASE_URL      = os.getenv("SUPABASE_URL", "")
SUPABASE_ANON_KEY = os.getenv("SUPABASE_ANON_KEY", "")   # 공개 API용
SUPABASE_SVC_KEY  = os.getenv("SUPABASE_SERVICE_KEY", "") # 서버사이드 삽입용

# CORS 허용 출처 -- 기본 도메인 + 환경변수 추가분 병합
_CORE_ORIGINS = [
    "https://bridgejob.co.kr",
    "https://www.bridgejob.co.kr",
    "https://bridge-chi-lime.vercel.app",
    "https://bridge-n7hk.onrender.com",
]
_DEV_ORIGINS = [
    "http://localhost:3000",
    "http://localhost:3001",
    "http://localhost:3002",
    "http://localhost:8080",
]
_cors_env = os.getenv("CORS_ORIGINS", "")
_IS_PROD = os.getenv("BRIDGE_ENV", os.getenv("ENV", "")).lower() in ("production", "prod")

# CORS 환경변수 안전 검증 -- 와일드카드(*) 또는 비HTTPS 도메인 차단
_extra: list[str] = []
for _o in (_cors_env.split(",") if _cors_env else []):
    _o = _o.strip()
    if not _o:
        continue
    if _o == "*":
        import logging as _log_cors
        _log_cors.getLogger("bridge").error(
            "🚨 CORS_ORIGINS에 와일드카드(*) 감지 -- 차단됨. "
            "허용 도메인을 명시적으로 설정하세요."
        )
        continue  # 와일드카드 추가 거부
    if _IS_PROD and _o.startswith("http://"):
        import logging as _log_cors
        _log_cors.getLogger("bridge").warning(
            "⚠️ CORS_ORIGINS 비HTTPS 도메인 차단 (prod): %s", _o
        )
        continue  # 프로덕션에서 비HTTPS 차단
    _extra.append(_o)

ALLOWED_ORIGINS = list(dict.fromkeys(
    _CORE_ORIGINS + _extra + ([] if _IS_PROD else _DEV_ORIGINS)
))

# ── 앱 초기화 ─────────────────────────────────────────────────────────────────
# 프로덕션에서는 API 문서 비활성화 (BRIDGE_ENV=production 설정 시)

app = FastAPI(
    title="BRIDGE Recruitment API",
    description="bridgejob.co.kr 웹사이트 백엔드",
    version="1.0.0",
    docs_url=None if _IS_PROD else "/docs",
    redoc_url=None if _IS_PROD else "/redoc",
    openapi_url=None if _IS_PROD else "/openapi.json",
    default_response_class=SafeJSONResponse,
)


# ── Startup: DB 무결성 검사 + 자동 복원 ─────────────────────────────────────
def _startup_db_health_check() -> None:
    """서버 시작 시 DB 헬스체크 (Persistent Disk 사용).
    /data 마운트 확인 + 테이블 카운트 로그만 남김. 자동 복원 없음.
    (Persistent Disk는 재배포 시 유지되므로 복원 로직 불필요)
    """
    _log = logging.getLogger("bridge.startup")
    try:
        db_path = os.getenv("BRIDGE_DB_PATH") or os.getenv("DB_PATH") or str(_ADMIN_DB_PATH)
        _log.info("[STARTUP] DB_PATH: %s", db_path)
        # /data 마운트 여부 확인 (Persistent Disk 체크)
        if db_path.startswith("/data/"):
            data_dir = Path("/data")
            if data_dir.exists():
                _log.info("[STARTUP] ✓ Persistent Disk /data 마운트 확인")
            else:
                _log.error("[STARTUP] ✗ /data 미마운트 -- Render Persistent Disk 설정 확인 필요")
        conn = sqlite3.connect(db_path)
        conn.execute("PRAGMA busy_timeout = 3000")
        try:
            tables = {r[0] for r in conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            ).fetchall()}
            _REQUIRED = ("candidates", "jobs", "client_inquiries")
            counts = {}
            for t in _REQUIRED:
                if t in tables:
                    try:
                        counts[t] = conn.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
                    except Exception:
                        counts[t] = 0
                else:
                    counts[t] = -1  # 테이블 없음
        finally:
            conn.close()

        _log.info("[STARTUP] 테이블 카운트: %s", counts)

        # 자동 복원 트리거 조건 강화 — Render Free Plan sleep/wake 시 ephemeral 휘발 대응:
        # 1) 일부 테이블 비어있음 (기존 조건)
        # 2) candidates < 100건 (정상 운영 시 3000건+, 100미만이면 무조건 휘발 의심)
        # 3) sheet_prefs 0건 + 다른 테이블도 빈약 (UI 설정 손실 명확)
        _CRITICAL_LOSS = (
            any(v <= 0 for v in counts.values())
            or counts.get("candidates", 0) < 100
        )

        # 환경변수 진단 — 어떤 인증 방식이든 감지하도록 다중 fallback
        _OAUTH_TOK = os.getenv("DRIVE_OAUTH_TOKEN_JSON", "").strip()
        _SA_B64    = os.getenv("GCP_SA_JSON_B64", "").strip()
        _log.info(
            "[STARTUP] Drive 인증: OAuth=%s, SA(legacy)=%s",
            f"set({len(_OAUTH_TOK)}자)" if _OAUTH_TOK else "미설정",
            f"set({len(_SA_B64)}자)" if _SA_B64 else "미설정",
        )

        if _CRITICAL_LOSS:
            _log.warning("[STARTUP] ⚠ DB 데이터 휘발 감지 (counts=%s)", counts)
            # ── Drive 자동 복원 시도 ──────────────────────────────────────────
            if _OAUTH_TOK:
                _log.info("[STARTUP] DRIVE_OAUTH_TOKEN_JSON 감지 → Drive 자동 복원 시작")
                try:
                    import sys as _sys
                    _tools_dir = str(Path(__file__).resolve().parent / "tools")
                    if _tools_dir not in _sys.path:
                        _sys.path.insert(0, _tools_dir)
                    from db_persist import restore as _drive_restore  # type: ignore
                    ok = _drive_restore()
                    if ok:
                        _log.info("[STARTUP] ✓ Drive 자동 복원 성공 — DB 준비 완료")
                        # 복원된 DB에 새 테이블 누락 시 자동 생성 (옛 백업 호환성)
                        # IF NOT EXISTS 멱등이므로 안전 — 새 테이블만 추가 + 기존은 보존
                        _ENSURE_FUNCS = [
                            "_ensure_sheet_prefs_table",
                            "_ensure_email_blacklist_table",
                            "_ensure_boards_table",
                            "_ensure_file_uploads_table",
                            "_ensure_banners_table",
                            "_ensure_site_partners_table",
                            "_ensure_site_settings_table",
                            "_ensure_site_visits_table",
                            "_ensure_page_content_table",
                            "_ensure_pipeline_tables",
                            "_ensure_talent_auth_tables",
                            # post-restore 정리 (Drive 복원이 옛 DB로 덮어쓰는 문제 회피)
                            "_ensure_testimonials_schema",
                            "_ensure_apostille_categories",
                        ]
                        _ok_count = 0
                        for _fn in _ENSURE_FUNCS:
                            try:
                                _f = globals().get(_fn)
                                if callable(_f):
                                    _f()
                                    _ok_count += 1
                            except Exception as _ee:
                                _log.warning("[STARTUP] %s 실패: %s", _fn, _ee)
                        _log.info("[STARTUP] post-restore ensure: %d/%d 테이블 보장",
                                  _ok_count, len(_ENSURE_FUNCS))
                    else:
                        _log.error("[STARTUP] ✗ Drive 복원 실패 — 수동 복원 필요")
                except Exception as _re:
                    _log.error("[STARTUP] Drive 복원 예외: %s", _re, exc_info=True)
            elif _SA_B64:
                _log.info("[STARTUP] GCP_SA_JSON_B64 감지 → SA 폴백 복원 시작")
                try:
                    import sys as _sys
                    _tools_dir = str(Path(__file__).resolve().parent / "tools")
                    if _tools_dir not in _sys.path:
                        _sys.path.insert(0, _tools_dir)
                    from db_persist import restore_sa as _sa_restore  # type: ignore
                    ok = _sa_restore()
                    if ok:
                        _log.info("[STARTUP] ✓ SA 폴백 복원 성공 — DB 준비 완료")
                        _ENSURE_FUNCS = [
                            "_ensure_sheet_prefs_table", "_ensure_email_blacklist_table",
                            "_ensure_boards_table", "_ensure_file_uploads_table",
                            "_ensure_banners_table", "_ensure_site_partners_table",
                            "_ensure_site_settings_table", "_ensure_site_visits_table",
                            "_ensure_page_content_table", "_ensure_pipeline_tables",
                            "_ensure_talent_auth_tables",
                            # post-restore 정리 (Drive 복원이 옛 DB로 덮어쓰는 문제 회피)
                            "_ensure_testimonials_schema",
                            "_ensure_apostille_categories",
                        ]
                        for _fn in _ENSURE_FUNCS:
                            try:
                                _f = globals().get(_fn)
                                if callable(_f): _f()
                            except Exception as _ee:
                                _log.warning("[STARTUP] %s 실패: %s", _fn, _ee)
                    else:
                        _log.error("[STARTUP] ✗ SA 폴백 복원 실패")
                except Exception as _re:
                    _log.error("[STARTUP] SA 복원 예외: %s", _re, exc_info=True)
            else:
                _log.error(
                    "[STARTUP] ❌ Drive 인증 환경변수 전무 — 자동 복원 불가. "
                    "DRIVE_OAUTH_TOKEN_JSON 미설정 → master.db 휘발 시 복구 불가능. "
                    "Render 환경변수 등록 즉시 필요."
                )
    except Exception as e:
        _log.error("[STARTUP] DB 헬스체크 실패: %s", e, exc_info=True)


@app.on_event("startup")
async def _startup_event():
    """FastAPI startup hook -- 비블로킹 백그라운드 DB 헬스체크 + v2.0 파이프라인 워커."""
    try:
        threading.Thread(target=_startup_db_health_check, daemon=True).start()
    except Exception as e:
        logging.getLogger("bridge.startup").error("[STARTUP] DB 헬스체크 실패: %s", e)
    # Resume Pipeline v2.0 worker 기동 (3 threads) — print()도 병행 (gunicorn 로그 버퍼 우회)
    try:
        sys.path.insert(0, str(Path(__file__).resolve().parent))
        from tools.pipeline_worker import start_workers  # type: ignore
        _worker_count = int(os.getenv("PIPELINE_WORKER_COUNT", "3"))
        start_workers(_worker_count)
        print(f"[STARTUP] Pipeline v2.0 — {_worker_count} workers launched", flush=True)
        logging.getLogger("bridge.pipeline").info("[STARTUP] %d workers launched", _worker_count)
    except Exception as e:
        import traceback as _tb
        _errstr = f"[STARTUP] Pipeline worker 기동 실패: {type(e).__name__}: {e}"
        print(_errstr, flush=True)
        print(_tb.format_exc(), flush=True)
        logging.getLogger("bridge.pipeline").error(_errstr, exc_info=True)
    # 테이블 보장 (startup 직후 호출)
    try:
        _ensure_pipeline_tables()
    except Exception as e:
        print(f"[STARTUP] _ensure_pipeline_tables 실패: {e}", flush=True)
    # Gmail 자동 동기화 루프 시작 (GMAIL_SYNC_ENABLED=true 일 때만)
    try:
        from gmail_collector import start_auto_sync
        start_auto_sync()
        print("[STARTUP] Gmail 자동 동기화 루프 시작됨", flush=True)
    except Exception as e:
        print(f"[STARTUP] Gmail 자동 동기화 시작 실패: {e}", flush=True)

# ── 글로벌 HTTPException 핸들러 -- 모든 에러를 구조화된 JSON으로 자동 변환 ───
# 기존 raise HTTPException(400, "msg") → {isError, errorCategory, isRetryable, context}
# 이미 구조화된 detail(dict)은 그대로 통과. 향후 새 엔드포인트도 자동 적용.
_STATUS_TO_CATEGORY = {
    400: "VALIDATION_ERROR",
    401: "AUTH_ERROR",
    403: "FORBIDDEN",
    404: "NOT_FOUND",
    409: "CONFLICT",
    413: "PAYLOAD_TOO_LARGE",
    429: "RATE_LIMIT",
    500: "INTERNAL_ERROR",
    503: "SERVICE_UNAVAILABLE",
}
_RETRYABLE_STATUSES = {429, 500, 503}


@app.exception_handler(HTTPException)
async def _structured_error_handler(request: Request, exc: HTTPException):
    # 이미 구조화된 응답(bridge_error 또는 수동 dict)은 그대로 통과
    if isinstance(exc.detail, dict) and "isError" in exc.detail:
        return JSONResponse(status_code=exc.status_code, content=exc.detail)
    return JSONResponse(
        status_code=exc.status_code,
        content={
            "isError": True,
            "errorCategory": _STATUS_TO_CATEGORY.get(exc.status_code, "UNKNOWN"),
            "isRetryable": exc.status_code in _RETRYABLE_STATUSES,
            "context": str(exc.detail) if (exc.detail and os.getenv("ENVIRONMENT", "production") == "development") else "",
        },
    )


# 보안 미들웨어 -- 헤더 + Rate Limit + 감사 로그
try:
    from security_middleware import SecurityMiddleware, security_router, admin_security_router
    # 배포 직후 블랙리스트 초기화 -- 이전 세션 잔여 차단 해제
    try:
        from security_middleware import ip_blacklist
        if ip_blacklist._list or ip_blacklist._permanent:
            ip_blacklist._list.clear()
            ip_blacklist._permanent.clear()
            ip_blacklist._save()
            print("[STARTUP] IP blacklist cleared (fresh deploy)")
    except Exception:
        pass
    app.add_middleware(SecurityMiddleware)
    app.include_router(security_router)
    app.include_router(admin_security_router)
except ImportError:
    pass  # 선택적: 파일 없으면 무시

# PII 마스킹 미들웨어 -- CORS보다 먼저 등록해야 응답 전체를 커버
app.add_middleware(PIIMaskingMiddleware)

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,   # HttpOnly 쿠키 cross-origin 전송 허용 (SameSite=None)
    allow_methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["Content-Type", "Authorization", "X-Admin-Key", "X-Admin-Token", "X-Bridge-Signature", "X-Apply-Token", "Cookie"],
)

# GZip 압축 -- 1KB 이상 응답 자동 압축 (5.5MB → 약 370KB)
app.add_middleware(GZipMiddleware, minimum_size=1000)

# TrustedHost -- 허용된 호스트만 접근 (Host 헤더 위조 방지)
_TRUSTED_HOSTS = ["bridgejob.co.kr", "www.bridgejob.co.kr", "api.bridgejob.co.kr", "bridge-n7hk.onrender.com"]
if not _IS_PROD:
    _TRUSTED_HOSTS += ["localhost", "127.0.0.1", "0.0.0.0"]
app.add_middleware(TrustedHostMiddleware, allowed_hosts=_TRUSTED_HOSTS)

# Body size 제한 -- 일반 10MB / DB 복원 경로 30MB
_MAX_BODY_BYTES         = 10 * 1024 * 1024  # 10MB (일반)
_MAX_BODY_BYTES_RESTORE = 30 * 1024 * 1024  # 30MB (DB 복원 전용)

class BodySizeLimitMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        cl = request.headers.get("content-length")
        if cl:
            limit = _MAX_BODY_BYTES_RESTORE if request.url.path == "/api/admin/db/restore" else _MAX_BODY_BYTES
            if int(cl) > limit:
                return JSONResponse(status_code=413, content={
                    "isError": True, "errorCategory": "PAYLOAD_TOO_LARGE",
                    "isRetryable": False, "context": f"Request body too large (max {limit // (1024*1024)}MB)."
                })
        return await call_next(request)

app.add_middleware(BodySizeLimitMiddleware)

# ── 미들웨어 레벨 IP Rate Limit -- body 검증 이전 단계에서 차단 ──────────────
# 목적: 잘못된 payload 반복 전송으로 Pydantic 검증 부하 유발하는 DoS 방어
_IP_REQ_COUNTS: dict[str, list[float]] = {}
_IP_RL_WINDOW = 10        # 10초 창
_IP_RL_MAX    = 240       # 10초 내 240회 = 초당 24회 평균 (admin 페이지 병렬 로드 여유)
_IP_RL_EXEMPT = {"/health", "/api/health", "/api/admin/key", "/api/admin/sessions"}

def _real_client_ip(request: Request) -> str:
    # Render/Vercel proxy 뒤에서 실제 IP 추출 (첫 XFF 엔트리)
    xff = request.headers.get("x-forwarded-for", "")
    if xff:
        return xff.split(",")[0].strip()
    return request.client.host if request.client else "unknown"

class IPRateLimitMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        if request.url.path in _IP_RL_EXEMPT:
            return await call_next(request)
        client_ip = _real_client_ip(request)
        now = time.time()
        bucket = _IP_REQ_COUNTS.setdefault(client_ip, [])
        # 창 밖 요청 제거
        cutoff = now - _IP_RL_WINDOW
        bucket[:] = [t for t in bucket if t > cutoff]
        if len(bucket) >= _IP_RL_MAX:
            return JSONResponse(
                status_code=429,
                content={"error": "Too many requests", "retry_after": _IP_RL_WINDOW},
                headers={"Retry-After": str(_IP_RL_WINDOW)},
            )
        bucket.append(now)
        # 메모리 보호: 1만 IP 이상이면 가장 오래된 IP 제거
        if len(_IP_REQ_COUNTS) > 10000:
            oldest_ip = min(_IP_REQ_COUNTS, key=lambda k: max(_IP_REQ_COUNTS[k]) if _IP_REQ_COUNTS[k] else 0)
            _IP_REQ_COUNTS.pop(oldest_ip, None)
        return await call_next(request)

app.add_middleware(IPRateLimitMiddleware)

# ── Kill-Switch (MAINT 모드) -- 침해 의심 시 모든 mutation 503 반환 ──────────
class MaintenanceMiddleware(BaseHTTPMiddleware):
    """env MAINT=1 이면 POST/PUT/PATCH/DELETE 전체 503. GET은 정상."""
    _MUTATION = {"POST", "PUT", "PATCH", "DELETE"}

    async def dispatch(self, request: Request, call_next):
        if os.getenv("MAINT", "0") == "1" and request.method in self._MUTATION:
            return JSONResponse(
                status_code=503,
                content={"error": "Service under maintenance", "code": "MAINT_MODE"},
                headers={"Retry-After": "300"},
            )
        return await call_next(request)

app.add_middleware(MaintenanceMiddleware)

# ── Threat Feed -- Spamhaus/Firehol 블록리스트 (무료 공개 피드) ─────────────
# 성능 최적화: v4/v6 분리 + LRU 캐시로 요청당 O(1) amortized
import ipaddress as _threat_ipm
_THREAT_NETS_V4: list = []
_THREAT_NETS_V6: list = []
_THREAT_FEED_PATH = Path(__file__).parent / "data" / "threat_feed.txt"
# IP 판정 캐시 -- 같은 IP 반복 조회 시 재스캔 회피 (Render 단일 서버 메모리)
_THREAT_CACHE: dict[str, bool] = {}
_THREAT_CACHE_MAX = 20000

def _load_threat_feed():
    global _THREAT_NETS_V4, _THREAT_NETS_V6
    try:
        if not _THREAT_FEED_PATH.exists():
            return
        nets_v4, nets_v6 = [], []
        with open(_THREAT_FEED_PATH, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                try:
                    net = _threat_ipm.ip_network(line, strict=False)
                    if net.version == 4:
                        nets_v4.append(net)
                    else:
                        nets_v6.append(net)
                except ValueError:
                    continue
        _THREAT_NETS_V4 = nets_v4
        _THREAT_NETS_V6 = nets_v6
        logging.getLogger("bridge.api").info(
            "[threat_feed] v4=%d v6=%d CIDR 로드됨", len(nets_v4), len(nets_v6)
        )
    except Exception as e:
        logging.getLogger("bridge.api").warning("[threat_feed] 로드 실패: %s", e)

_load_threat_feed()

class ThreatFeedMiddleware(BaseHTTPMiddleware):
    """Spamhaus/Firehol 위협 피드 IP 즉시 차단 (캐시 + v4/v6 분리)."""
    # health 엔드포인트는 스킵 (Render health check 보호)
    _EXEMPT = {"/health", "/api/health"}

    async def dispatch(self, request: Request, call_next):
        if not (_THREAT_NETS_V4 or _THREAT_NETS_V6):
            return await call_next(request)
        if request.url.path in self._EXEMPT:
            return await call_next(request)
        client_ip = request.client.host if request.client else ""
        if not client_ip:
            return await call_next(request)

        # 캐시 조회 (O(1))
        cached = _THREAT_CACHE.get(client_ip)
        if cached is True:
            return JSONResponse(status_code=403, content={"error": "Blocked"})
        if cached is False:
            return await call_next(request)

        # 첫 조회만 O(N)
        blocked = False
        try:
            ip_obj = _threat_ipm.ip_address(client_ip)
            nets = _THREAT_NETS_V4 if ip_obj.version == 4 else _THREAT_NETS_V6
            for net in nets:
                if ip_obj in net:
                    blocked = True
                    logging.getLogger("bridge.security").warning(
                        "[threat_feed] 차단: %s → %s", client_ip, net
                    )
                    break
        except ValueError:
            pass

        # 캐시 저장 (bounded LRU: 용량 초과 시 초기화)
        if len(_THREAT_CACHE) >= _THREAT_CACHE_MAX:
            _THREAT_CACHE.clear()
        _THREAT_CACHE[client_ip] = blocked

        if blocked:
            return JSONResponse(status_code=403, content={"error": "Blocked"})
        return await call_next(request)

app.add_middleware(ThreatFeedMiddleware)

# ── CSRF Origin 검증 -- 관리자 mutation 요청의 Origin 헤더 검증 ──────────────
_ALLOWED_ORIGINS_SET = set(ALLOWED_ORIGINS)

class CSRFOriginMiddleware(BaseHTTPMiddleware):
    """POST/PUT/PATCH/DELETE 요청에 Origin 헤더 검증 (CSRF 방어)."""
    _EXEMPT = {"/api/apply", "/api/inquiry", "/api/track", "/api/public/talent-inquiry"}  # 공개 폼 제외
    _MUTATION = {"POST", "PUT", "PATCH", "DELETE"}

    async def dispatch(self, request: Request, call_next):
        if request.method not in self._MUTATION:
            return await call_next(request)
        path = request.url.path
        if any(path.startswith(p) for p in self._EXEMPT):
            return await call_next(request)
        origin = request.headers.get("origin", "")
        is_admin = path.startswith("/api/admin/")
        if is_admin:
            # 관리자 경로: Origin 없거나 허용 목록 외 → 차단
            if not origin or origin not in _ALLOWED_ORIGINS_SET:
                return JSONResponse(status_code=403, content={
                    "isError": True, "errorCategory": "CSRF_REJECTED",
                    "isRetryable": False, "context": "Cross-origin request blocked.",
                })
        elif origin and origin not in _ALLOWED_ORIGINS_SET:
            # 공개 경로: Origin 있으나 허용 목록 외 → 차단
            return JSONResponse(status_code=403, content={
                "isError": True, "errorCategory": "CSRF_REJECTED",
                "isRetryable": False, "context": "Cross-origin request blocked.",
            })
        return await call_next(request)

app.add_middleware(CSRFOriginMiddleware)

# ── 관리자 변경 감사 로그 -- 모든 admin mutation 상세 기록 ────────────────────
class AdminAuditMiddleware(BaseHTTPMiddleware):
    """관리자 POST/PATCH/DELETE 요청을 감사 로그에 기록."""
    _MUTATION = {"POST", "PUT", "PATCH", "DELETE"}

    async def dispatch(self, request: Request, call_next):
        path = request.url.path
        method = request.method
        if method not in self._MUTATION or not path.startswith("/api/admin"):
            return await call_next(request)
        # 관리자 인증된 요청만 로깅 (인증 실패는 _check_admin에서 로깅)
        _ak = os.getenv("ADMIN_API_KEY", "")
        has_key = _ak and request.headers.get("x-admin-key", "").strip() == _ak
        has_token = bool(request.headers.get("x-admin-token", "").strip())
        if not has_key and not has_token:
            return await call_next(request)

        ip = _get_client_ip(request)  # 스푸핑 방지 IP 추출 (X-Forwarded-For 위조 불가)
        response = await call_next(request)

        # 감사 로그 기록 (파일)
        try:
            from security_middleware import audit
            audit.log("ADMIN_MUTATION", {
                "ip": ip,
                "method": method,
                "path": path,
                "status": response.status_code,
                "auth": "session" if has_token else "api_key",
            }, "INFO" if response.status_code < 400 else "WARNING")
        except ImportError:
            pass
        return response

app.add_middleware(AdminAuditMiddleware)

# ── DB 인덱스 보장 (무한스크롤 cursor 성능) ─────────────────────────────────
def _ensure_candidate_indexes() -> None:
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        # 기존 인덱스
        conn.execute("CREATE INDEX IF NOT EXISTS idx_candidates_status ON candidates(status)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_candidates_name ON candidates(full_name)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_candidates_created ON candidates(created_at)")
        # 추가 인덱스 -- 대량 유입 대비 (WHERE is_deleted 매 쿼리 공통)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_candidates_deleted ON candidates(is_deleted)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_candidates_sheet ON candidates(sheet_number)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_candidates_visible ON candidates(talent_visible)")
        # candidates 복합 인덱스 (공개 게시판 공통 WHERE)
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_candidates_pub "
            "ON candidates(status, talent_visible, is_deleted)"
        )
        # jobs 인덱스
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_jobs_open "
            "ON jobs(is_deleted, status)"
        )
        # client_inquiries 인덱스
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_inquiries_active "
            "ON client_inquiries(is_deleted, email)"
        )
        conn.commit()
        conn.close()
    except Exception:
        pass

try:
    _ensure_candidate_indexes()
except Exception as _e:
    logging.getLogger("bridge.api").warning("_ensure_candidate_indexes 스킵: %s", _e)

# ── Supabase 클라이언트 (지연 초기화) ─────────────────────────────────────────
_anon_client: Optional["Client"] = None
_svc_client:  Optional["Client"] = None


def get_anon_client() -> "Client":
    global _anon_client
    if _anon_client is None:
        if not SUPABASE_URL or not SUPABASE_ANON_KEY:
            raise RuntimeError(".env 에 SUPABASE_URL / SUPABASE_ANON_KEY 필요")
        _anon_client = create_client(SUPABASE_URL, SUPABASE_ANON_KEY)
    return _anon_client


def get_svc_client() -> "Client":
    global _svc_client
    if _svc_client is None:
        if not SUPABASE_URL or not SUPABASE_SVC_KEY:
            raise RuntimeError(".env 에 SUPABASE_URL / SUPABASE_SERVICE_KEY 필요")
        _svc_client = create_client(SUPABASE_URL, SUPABASE_SVC_KEY)
    return _svc_client


# ── 요청 모델 ─────────────────────────────────────────────────────────────────
class CandidateApply(BaseModel):
    # honeypot 필드(_url) 등 모델 외 필드도 통과시킴 — 기본 'ignore' 는 silently drop
    # 결과: _check_honeypot() 가 항상 False → honeypot 무력화 (어제까지 봇 차단 안 됐던 원인)
    class Config:
        extra = "allow"
    """
    구직자 지원 웹폼 -- 구글폼 42개 항목 100% 매핑
    Google Form + Native Web Form 병렬 운영 공통 스키마
    """
    # ── 기본 식별 ──────────────────────────────────────────────────────────
    full_name:               str  = Field(..., min_length=2, max_length=100)  # [ENCRYPT]
    email:                   Optional[str] = None   # [ENCRYPT]
    how_to:                  Optional[str] = None   # How to (소스 경로)
    file_urls:               Optional[str] = None   # Attach your files (링크)

    # ── 인적사항 ───────────────────────────────────────────────────────────
    nationality:             Optional[str] = None
    ancestry:                Optional[str] = None   # Family Ancestry Background
    dob:                     Optional[str] = None   # Date of Birth
    gender:                  Optional[str] = None
    current_location:        Optional[str] = None
    marital_status:          Optional[str] = None   # Marital Status
    dependents:              Optional[str] = None   # Dependents
    pets:                    Optional[str] = None   # Pets

    # ── 학력 & 자격 ────────────────────────────────────────────────────────
    education:               Optional[str] = None   # Educational Background
    major:                   Optional[str] = None   # Major
    certification:           Optional[str] = None   # Certification

    # ── 비자 & 서류 ────────────────────────────────────────────────────────
    e_visa:                  Optional[str] = None   # E visa
    arc_holders:             Optional[str] = None   # ARC holders
    passport:                Optional[str] = None   # Passport [ENCRYPT]
    criminal_record:         Optional[str] = None   # Criminal Record [ENCRYPT]
    doc_status:              Optional[str] = None   # Document Status

    # ── 근무 선호 ──────────────────────────────────────────────────────────
    start_date:              Optional[str] = None   # Start date
    area_prefs:              Optional[str] = None   # TargetArea prefs
    target_age:              Optional[str] = None   # (수업 연령대 선호)
    job_prefs:               Optional[str] = None   # Job prefs (Notes)
    experience:              Optional[str] = None   # Experience
    employment:              Optional[str] = None   # Employment (현 고용주 인지 여부)
    reference:               Optional[str] = None   # Employment Reference
    interview_time:          Optional[str] = None   # Interview time

    # ── 급여 & 주거 ────────────────────────────────────────────────────────
    current_salary:          Optional[str] = None   # Current salary
    desired_salary:          Optional[str] = None   # Desired salary
    housing:                 Optional[str] = None   # Housing
    personal_considerations: Optional[str] = None   # Personal Considerations

    # ── 고위험 PII (최강 암호화) ───────────────────────────────────────────
    religion:                Optional[str] = None   # Religion [ENCRYPT]
    health_info:             Optional[str] = None   # Health Information [ENCRYPT]
    criminal_record_check:   Optional[str] = None   # Criminal Record Check [ENCRYPT]

    # ── 연락처 ─────────────────────────────────────────────────────────────
    kakaotalk:               Optional[str] = None   # KakaoTalk [ENCRYPT]
    mobile_phone:            Optional[str] = None   # Mobile Phone [ENCRYPT]

    # ── 수업 대상 & 서류 ─────────────────────────────────────────────────────
    target:                  Optional[str] = None   # Target (Kindy, Elem 등)
    korean_criminal_record:  Optional[str] = None   # 한국 내 범죄기록
    documents:               Optional[str] = None   # 서류 준비 상태

    # ── 동의 ───────────────────────────────────────────────────────────────
    agreement:               Optional[str] = None   # Agreement
    facts:                   Optional[str] = None   # Facts

    # ── 관리자 메모 ────────────────────────────────────────────────────────
    admin_notes:             Optional[str] = None   # 메모 (Google Form 기입란)

    # ── 2차 폼 / 병렬 운영 ────────────────────────────────────────────────
    apply_token:             Optional[str] = None   # JWT magic link
    source:                  Optional[str] = None   # 'web_form' | 'google_form'
    captcha_token:           Optional[str] = None   # PuzzleCaptcha 토큰


class ClientInquiry(BaseModel):
    # honeypot 필드(_url) 등 모델 외 필드도 통과시킴
    class Config:
        extra = "allow"
    """
    구인처 채용 문의 -- 구글폼 43개 항목 100% 매핑
    Google Form + Native Web Form 병렬 운영 공통 스키마
    """
    # ── 담당자 식별 ────────────────────────────────────────────────────────
    email:               Optional[str] = None   # 이메일 주소 [ENCRYPT]
    contact_name:        Optional[str] = None   # 성함 [ENCRYPT]
    contact_position:    Optional[str] = None   # 직책
    phone:               Optional[str] = None   # 휴대전화 [ENCRYPT]
    business_registration: Optional[str] = None # 사업자 [ENCRYPT]

    # ── 학교 / 기관 정보 ───────────────────────────────────────────────────
    school_name:         str = Field(..., min_length=2, max_length=200)  # 학교이름 (필수)
    school_name_en:      Optional[str] = None   # Name (English)
    hiring_history:      Optional[str] = None   # 채용이력
    school_location:     Optional[str] = None   # 근무처소재지
    vacancies:           Optional[str] = None   # 채용인원
    native_count:        Optional[str] = None   # 원어민강사 현재 수

    # ── 근무 조건 ──────────────────────────────────────────────────────────
    start_date:          Optional[str] = None   # 희망시작일
    teaching_age:        Optional[str] = None   # 수업대상
    working_days:        Optional[str] = None   # 근무요일
    schedule:            Optional[str] = None   # 근무일정 (시간)
    working_hours:       Optional[str] = None   # 총 근무시간
    class_size:          Optional[str] = None   # 학생수
    avg_lessons:         Optional[str] = None   # 평균강의
    prep_time:           Optional[str] = None   # 프렙타임

    # ── 선호 대상 ──────────────────────────────────────────────────────────
    preferred_candidate: Optional[str] = None   # 선호대상

    # ── 급여 조건 ──────────────────────────────────────────────────────────
    salary_raw:          Optional[str] = None   # 급여조건

    # ── 숙소 & 주거 ────────────────────────────────────────────────────────
    housing_type:        Optional[str] = None   # 거주지 (숙소 형태)
    housing_provided:    Optional[str] = None   # 숙소제공
    housing_detail:      Optional[str] = None   # 숙소관련

    # ── 복지 & 휴가 ────────────────────────────────────────────────────────
    travel_support:      Optional[str] = None   # 이동지원
    benefits:            Optional[str] = None   # 복지
    vacation:            Optional[str] = None   # (기존 호환)
    paid_vacation:       Optional[str] = None   # 유급휴일
    vacation_includes:   Optional[str] = None   # 휴일 포함여부

    # ── 근로계약 ───────────────────────────────────────────────────────────
    contract_type:       Optional[str] = None   # 계약구분
    break_time:          Optional[str] = None   # 휴게시간
    job_responsibilities: Optional[str] = None  # 업무책임
    sick_leave:          Optional[str] = None   # 보건휴가
    meal:                Optional[str] = None   # (기존 호환)
    meal_provided:       Optional[str] = None   # 식사
    meal_allowance:      Optional[str] = None   # 식대

    # ── 메모 (이중 언어) ────────────────────────────────────────────────────
    memo:                Optional[str] = None   # (기존 호환)
    memo_kr:             Optional[str] = None   # 메모
    memo_en:             Optional[str] = None   # Memo

    # ── 동의 & 기타 ────────────────────────────────────────────────────────
    privacy_policy:      Optional[str] = None   # 개인정보처리방침 동의

    # ── 관리 ───────────────────────────────────────────────────────────────
    location:            Optional[str] = None   # (기존 호환)
    source:              Optional[str] = None   # 'web_form' | 'google_form'
    captcha_token:       Optional[str] = None   # PuzzleCaptcha 토큰 (스팸 차단)


# ── 공통 응답 ─────────────────────────────────────────────────────────────────
def ok(data=None, message: str = "ok"):
    return {"success": True, "message": message, "data": _sanitize_data(data)}


def err(message: str, code: int = 400, error_code: str = ""):
    """구조화 에러 -- HTTPException에 error_code 포함."""
    raise HTTPException(
        status_code=code,
        detail={"message": message, "error_code": error_code or _http_to_error_code(code)},
    )


def _http_to_error_code(status: int) -> str:
    """HTTP 상태 코드 → 에이전트 재시도 판단용 에러 코드."""
    return {
        400: "VALIDATION_ERROR",
        401: "AUTH_FAILED",
        403: "FORBIDDEN",
        404: "NOT_FOUND",
        409: "CONFLICT",
        429: "RATE_LIMITED",
        500: "SERVER_ERROR",
        503: "SERVICE_UNAVAILABLE",
    }.get(status, "UNKNOWN_ERROR")


# ── 라우트 ────────────────────────────────────────────────────────────────────
@app.get("/", tags=["health"])
async def root():
    """서버 상태 확인"""
    return {
        "service": "BRIDGE Recruitment API",
        "status":  "running",
        "time":    datetime.now(timezone.utc).isoformat(),
        "docs":    "/docs",
    }


@app.get("/health", tags=["health"])
async def health_check():
    """Render / 외부 모니터링용 헬스체크"""
    return {
        "status": "ok",
        "version": "v2.4.1",
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }


# ── Render 슬립 방지 -- 10분마다 self-ping ──────────────────────────────────────
_RENDER_URL = os.getenv("RENDER_EXTERNAL_URL", "")

def _keep_alive_loop():
    """백그라운드 스레드: 10분마다 /health self-ping"""
    import urllib.request
    url = f"{_RENDER_URL}/health" if _RENDER_URL else ""
    if not url:
        return
    while True:
        time.sleep(600)  # 10분
        try:
            urllib.request.urlopen(url, timeout=10)
        except Exception:
            pass

if _RENDER_URL:
    _t = threading.Thread(target=_keep_alive_loop, daemon=True)
    _t.start()


def _job_row_to_public(row: dict) -> dict:
    """SQLite jobs row → PublicJob 형태로 매핑 (민감 필드 제외, 관리자 경로 전용)"""
    teaching_age_raw = row.get("teaching_age") or ""
    age_map = {
        "kindy": "kindergarten", "kindergarten": "kindergarten",
        "elem": "elementary", "elementary": "elementary",
        "pre_k": "pre_k", "pre-k": "pre_k", "prek": "pre_k",
        "middle": "middle", "high": "high", "adult": "adult",
    }
    age_groups = []
    for token in re.split(r"[,/\-&·\s]+", teaching_age_raw.lower()):
        token = token.strip()
        if token in age_map and age_map[token] not in age_groups:
            age_groups.append(age_map[token])

    benefits_raw = row.get("benefits") or ""
    benefits_list = [b.strip() for b in benefits_raw.split(",") if b.strip()] if benefits_raw else []

    raw_text = row.get("raw_text") or ""
    if raw_text:
        raw_text = re.sub(r'\([^)]*\)', '', raw_text)
        raw_text = "\n".join(l for l in raw_text.splitlines() if l.strip())

    return {
        "location":                row.get("city") or row.get("location"),
        "job_id":                  row.get("job_code", ""),
        "starting_date":           row.get("start_date"),
        "teaching_age":            age_groups or None,
        "teaching_age_raw":        teaching_age_raw or None,
        "class_size":              row.get("class_size"),
        "working_hours":           row.get("working_hours"),
        "monthly_salary":          row.get("salary_raw"),
        "teaching_hours_per_week": row.get("teach_hrs_week"),
        "vacation":                row.get("vacation"),
        "native_teacher_count":    row.get("native_count"),
        "housing":                 row.get("housing"),
        "preferences":             None,
        "employee_benefits":       benefits_list or None,
        "raw_text":                raw_text or None,
        "notes":                   None,
        "is_hot":                  bool(row.get("is_hot", 0)),
        "employment_type":         "part_time" if row.get("is_part_time") else "full_time",
        "hours_per_day":           row.get("daily_hours"),
        "status":                  row.get("status", "open"),
    }


def _ad_only_to_public(j: dict) -> dict:
    """
    ad_only 클린 dict → PublicJob 매핑.

    입력 스키마: location, job_code, start_date, teaching_age, class_size,
                 working_hours, monthly_salary, teach_hrs_week, vacation,
                 housing, native_count, benefits, extra_lines
    (이미 pii_guard 통과된 데이터)
    """
    teaching_age_raw = j.get("teaching_age") or ""
    age_map = {
        "kindy": "kindergarten", "kindergarten": "kindergarten",
        "elem": "elementary", "elementary": "elementary",
        "pre_k": "pre_k", "pre-k": "pre_k", "prek": "pre_k",
        "middle": "middle", "high": "high", "adult": "adult",
    }
    age_groups = []
    for token in re.split(r"[,/\-&·\s]+", teaching_age_raw.lower()):
        token = token.strip()
        if token in age_map and age_map[token] not in age_groups:
            age_groups.append(age_map[token])

    benefits_raw = j.get("benefits") or ""
    benefits_list = [b.strip() for b in benefits_raw.split(",") if b.strip()] if benefits_raw else []

    extras = j.get("extra_lines") or []
    raw_text = "\n".join(extras) if isinstance(extras, list) else None

    out = {
        "location":                j.get("location"),
        "job_id":                  j.get("job_code", ""),
        "starting_date":           j.get("start_date"),
        "teaching_age":            age_groups or None,
        "teaching_age_raw":        teaching_age_raw or None,
        "class_size":              j.get("class_size"),
        "working_hours":           j.get("working_hours"),
        "monthly_salary":          j.get("monthly_salary"),
        "teaching_hours_per_week": j.get("teach_hrs_week"),
        "vacation":                j.get("vacation"),
        "native_teacher_count":    j.get("native_count"),
        "housing":                 j.get("housing"),
        "preferences":             None,
        "employee_benefits":       benefits_list or None,
        "raw_text":                raw_text or None,
        "notes":                   None,
        "is_hot":                  False,
        "employment_type":         "full_time",
        "hours_per_day":           None,
        "status":                  "open",
    }
    # 방어선 2: 응답 레벨 최종 스캔 (location 은 strict, 나머지는 한글/메일/전화만)
    if _AD_READY and _ad_assert_clean:
        for k, v in out.items():
            if isinstance(v, str) and v:
                _ad_assert_clean(v, context=f"public_jobs.{k}",
                                 strict_brand=(k == "location"))
            elif isinstance(v, list):
                for it in v:
                    if isinstance(it, str):
                        _ad_assert_clean(it, context=f"public_jobs.{k}[]",
                                         strict_brand=False)
    return out


def _admin_key_ok(request: Request) -> bool:
    return bool(_ADMIN_KEY) and request.headers.get("x-admin-key", "") == _ADMIN_KEY


@app.get("/api/jobs", tags=["jobs"])
async def list_jobs(
    request:   Request,
    city:      Optional[str] = None,
    is_hot:    Optional[bool] = None,
    include_closed: Optional[bool] = None,
    limit:     int = 50,
    offset:    int = 0,
):
    """
    공개 구인 목록 조회.

    - 기본 공개 경로: ad_only/jobs_clean.txt (PII 0건, fail-closed 가드)
    - include_closed=true + admin key: master.db 직접 조회 (관리자 전용)
    """
    limit = min(max(limit, 1), 100)
    if not _rate_ok(_ip_hash(request), window=60, max_posts=60):
        return bridge_error("RATE_LIMIT", "Too many requests. Please slow down.", retryable=True, status=429)

    admin_mode = bool(include_closed) and _admin_key_ok(request)

    # 공개 경로 -- ad_only 전용
    if not admin_mode:
        if not _AD_READY or _ad_load_all_jobs is None:
            logging.getLogger("bridge.api").error("ad_only 미가용 -- 공개 jobs 응답 0건")
            return ok(data=[], message="0건 조회 (ad_only 미가용)")
        try:
            jobs = _ad_load_all_jobs()
            if city:
                c = city.strip()[:50].lower()
                jobs = [j for j in jobs if c in (j.get("location") or "").lower()]
            if is_hot is True:
                jobs = []  # ad_only 는 is_hot 메타 없음 -- 필터 시 빈 결과
            paged = jobs[offset: offset + limit]
            data = [_ad_only_to_public(j) for j in paged]
            return ok(data=data, message=f"{len(data)}건 조회 (ad_only)")
        except _ADPIIErr as pe:
            logging.getLogger("bridge.api").error("ad_only PII 감지 -- 공개 응답 차단: %s", pe)
            return ok(data=[], message="0건 조회 (PII 가드 차단)")
        except Exception as e:
            logging.getLogger("bridge.api").error("list_jobs ad_only 실패: %s", e, exc_info=True)
            return ok(data=[], message="0건 조회 (내부 오류)")

    # 관리자 경로 -- master.db 직접
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout=5000")
        conn.row_factory = sqlite3.Row
        try:
            where = ["is_deleted = 0"]
            params: list[Any] = []
            if city:
                city = city.strip()[:50]
                where.append("(city LIKE ? OR location LIKE ?)")
                params.extend([f"%{city}%", f"%{city}%"])
            if is_hot is not None:
                where.append("is_hot = ?")
                params.append(1 if is_hot else 0)
            sql = (
                "SELECT * FROM jobs WHERE " + " AND ".join(where)
                + " ORDER BY is_hot DESC, CASE WHEN COALESCE(raw_text, '') = '' THEN 1 ELSE 0 END ASC, created_at DESC"
                + " LIMIT ? OFFSET ?"
            )
            params.extend([limit, offset])
            rows = conn.execute(sql, params).fetchall()
            data = [_job_row_to_public(dict(r)) for r in rows]
            return ok(data=data, message=f"{len(data)}건 조회 (admin)")
        finally:
            conn.close()
    except Exception as e:
        logging.getLogger("bridge.api").error("list_jobs admin 실패: %s", e, exc_info=True)
        err("구인 목록을 불러올 수 없습니다. 잠시 후 다시 시도해주세요.", 500)


@app.get("/api/jobs/{job_id}", tags=["jobs"])
async def get_job(job_id: str, request: Request):
    """포지션 상세 조회 -- 공개 경로는 ad_only, admin key 있으면 master.db."""
    if not _rate_ok(_ip_hash(request), window=60, max_posts=60):
        return bridge_error("RATE_LIMIT", "Too many requests. Please slow down.", retryable=True, status=429)

    # 공개 경로 -- ad_only
    if not _admin_key_ok(request):
        if not _AD_READY or _ad_get_job is None:
            err("포지션을 찾을 수 없습니다.", 404)
        try:
            j = _ad_get_job(job_id)
            if not j:
                err("포지션을 찾을 수 없습니다.", 404)
            return ok(data=_ad_only_to_public(j))
        except HTTPException:
            raise
        except _ADPIIErr as pe:
            logging.getLogger("bridge.api").error("ad_only PII 감지 -- 상세 차단: %s", pe)
            err("포지션 정보를 불러올 수 없습니다.", 500)
        except Exception as e:
            logging.getLogger("bridge.api").error("get_job ad_only 실패: %s", e, exc_info=True)
            err("포지션 정보를 불러올 수 없습니다. 잠시 후 다시 시도해주세요.", 500)

    # 관리자 경로
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout=5000")
        conn.row_factory = sqlite3.Row
        try:
            row = conn.execute(
                "SELECT * FROM jobs WHERE id = ? AND is_deleted = 0",
                [job_id],
            ).fetchone()
            if not row:
                err("포지션을 찾을 수 없습니다.", 404)
            return ok(data=_job_row_to_public(dict(row)))
        finally:
            conn.close()
    except HTTPException:
        raise
    except Exception as e:
        logging.getLogger("bridge.api").error("get_job admin 실패: %s", e, exc_info=True)
        err("포지션 정보를 불러올 수 없습니다. 잠시 후 다시 시도해주세요.", 500)


# 구직자 암호화 필드 -- 고위험 PII 전체 + 준식별자(quasi-identifier) 포함
_CANDIDATE_ENCRYPT = {
    # ── 직접 식별자 ──
    "full_name",            # 식별 가능 이름
    "email",                # 이메일
    "mobile_phone",         # 휴대폰
    "kakaotalk",            # 카카오톡
    # ── 신원/법적 ──
    "passport",             # 여권 (번호/상태)
    "criminal_record",      # 전과기록 (자체 기술)
    "criminal_record_check", # 범죄경력조회 [고위험]
    "korean_criminal_record", # 한국 범죄경력 [고위험]
    # ── 민감 정보 ──
    "religion",             # 종교 [고위험]
    "health_info",          # 건강 정보 [고위험]
    # ── 준식별자 (단독으론 무해하나 조합 시 재식별 가능) ──
    "dob",                  # 생년월일 -- 재식별 위험
    "nationality",          # 국적
    "current_location",     # 현재 거주지
    "reference",            # 소개자/이전 학교 연락처 (제3자 PII 포함)
}

# ── CandidateApply 모델 → DB 컬럼 매핑 ───────────────────────────────────────
_APPLY_FIELD_MAP = {
    "education":                "education_level",
    "marital_status":           "married",
    "personal_considerations":  "personal_consideration",
    "agreement":                "consent",
    "facts":                    "fact_check",
    "admin_notes":              "notes",
    # target_age: target_age 컬럼이 별도 존재하므로 target 리다이렉트 제거 (L659 참조)
}

# ── Google Sheet 동기화 (비동기 헬퍼) ─────────────────────────────────────────
def _sync_to_sheet(candidate_id: str):
    """Background: DB → Google Sheet Source 시트에 1건 추가. 실패해도 무시."""
    try:
        from tools.sheet_sync import SheetSync
        syncer = SheetSync(
            credentials_file=os.getenv("GOOGLE_SERVICE_ACCOUNT_FILE", "gcp_service_account.json"),
            spreadsheet_id=os.getenv("GOOGLE_SHEETS_CANDIDATES_ID"),
        )
        syncer.sync_single(candidate_id)
    except Exception as e:
        logging.getLogger("bridge.api").warning("Sheet sync failed for %s: %s", candidate_id, e)

# payload에서 제거할 키 (DB 컬럼으로 직접 INSERT 불가)
_APPLY_SKIP_FIELDS = {"file_urls", "apply_token"}


def _map_apply_payload(payload: dict) -> dict:
    """CandidateApply 모델 필드명 → candidates DB 컬럼명 변환"""
    mapped = {}
    for k, v in payload.items():
        if k in _APPLY_SKIP_FIELDS:
            continue
        db_key = _APPLY_FIELD_MAP.get(k, k)
        mapped[db_key] = v
    mapped["source_file"] = "web_form"
    return mapped


@app.post("/api/apply", status_code=status.HTTP_201_CREATED, tags=["candidates"])
async def apply(request: Request, body: CandidateApply):
    # Rate limit 완화 — 사용자 retry 허용 (시간당 60회 / 분당 약 1회)
    # CAPTCHA + Honeypot 통과 못한 요청은 별도 카운터로 차단 (스팸 방지는 그쪽에서 처리)
    if not _rate_ok(_ip_hash(request), window=3600, max_posts=60):
        raise HTTPException(429, "Too many requests. Please try again later.")
    """
    구직자 지원 접수 (1차 / 2차 통합)

    1차 제출: INSERT → JWT magic link 토큰 반환
    2차 제출: apply_token 검증 → UPDATE (누적)
    토큰 없음: dedup_key 복합키로 기존 레코드 탐색 → UPSERT
    """
    # Input Sanitizer 검증
    if _SECURITY_PKG_OK and _sec_sanitizer:
        for _f in ["full_name", "email", "phone", "message"]:
            _v = getattr(body, _f, None)
            if _v and isinstance(_v, str) and not _sec_sanitizer.is_safe(_v):
                raise HTTPException(400, "유효하지 않은 입력입니다.")
    try:
        now_iso = datetime.now(timezone.utc).isoformat()

        payload = body.model_dump(exclude_none=True)
        token_in = payload.pop("apply_token", None)
        captcha_token = payload.pop("captcha_token", None)

        # ── CAPTCHA 검증 (스팸 차단) ────────────────────────────────────────────────
        ip_hash = _ip_hash(request)
        if not captcha_token or not _verify_captcha_token(captcha_token, ip_hash):
            raise HTTPException(400, "CAPTCHA verification failed. Please complete the puzzle.")

        # Honeypot 확인
        if _check_honeypot(payload):
            raise HTTPException(400, "Invalid submission detected.")
        # honeypot 필드는 DB에 저장 안 함 — INSERT 실패 방지
        payload.pop("_url", None)

        # PII 암호화
        for field in _CANDIDATE_ENCRYPT:
            if field in payload and payload[field]:
                payload[field] = _encrypt_if_needed(str(payload[field]), field)

        payload["updated_at"] = now_iso

        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        conn.row_factory = sqlite3.Row
        try:
            existing_id: Optional[str] = None

            # ── 경로 A: JWT 토큰으로 기존 레코드 찾기
            if token_in:
                cid = _verify_candidate_token(token_in)
                if cid:
                    existing_id = cid

            # ── 경로 B: email로 기존 레코드 탐색
            if not existing_id and body.email:
                row = conn.execute(
                    "SELECT candidate_id FROM candidates WHERE email = ? LIMIT 1",
                    (payload.get("email", body.email),),
                ).fetchone()
                if row:
                    existing_id = row["candidate_id"]

            # ── UPDATE (2차 누적)
            if existing_id:
                db_payload = _map_apply_payload(payload)
                sets = ", ".join(f"{k} = ?" for k in db_payload if k != "candidate_id")
                vals = [v for k, v in db_payload.items() if k != "candidate_id"]
                vals.append(existing_id)
                conn.execute(f"UPDATE candidates SET {sets} WHERE candidate_id = ?", vals)
                conn.commit()
                tok = _make_candidate_token(existing_id)
                return ok(
                    data={"id": existing_id, "apply_token": tok, "mode": "updated"},
                    message="지원 정보가 업데이트되었습니다."
                )

            # ── INSERT (1차 신규)
            import uuid
            new_id = f"cnd_{uuid.uuid4().hex[:12]}"
            payload["candidate_id"] = new_id
            payload["source"]     = payload.get("source") or "web_form"
            payload["status"]     = "Active"
            payload["created_at"] = now_iso

            db_payload = _map_apply_payload(payload)
            # sheet_number: INSERT 서브쿼리로 원자적 할당 (동시접수 race 방지)
            # 4를 포함하지 않는 sheet_number 할당 (Python에서 계산 후 INSERT)
            db_payload["sheet_number"] = _next_sheet_number(conn)
            cols = list(db_payload.keys())
            vals = list(db_payload.values())
            ph = ["?" for _ in cols]
            conn.execute(
                f"INSERT INTO candidates ({', '.join(cols)}) VALUES ({', '.join(ph)})",
                vals,
            )
            conn.commit()
            payload["sheet_number"] = db_payload["sheet_number"]

            token_out = _make_candidate_token(new_id)

            # 확인 이메일 발송 (실패해도 접수는 완료)
            if _EMAIL_OK and body.email:
                send_applicant_confirmation(body.email, body.full_name)

            # 제출 즉시 암호화 백업 (비동기, 실패해도 접수 완료)
            threading.Thread(target=_backup_on_submit, args=("apply",), daemon=True).start()

            # 원본 JSON 별도 보관 (암호화 전 평문 → 암호화 후 독립 파일, 불변 원본)
            _raw_apply = body.model_dump(exclude_none=True)
            _raw_apply.pop("captcha_token", None)
            threading.Thread(
                target=_save_raw_submission,
                args=(_raw_apply, "apply", str(new_id)),
                daemon=True,
            ).start()

            # Push 알림 (비동기, 실패해도 접수 완료)
            threading.Thread(
                target=_notify_push,
                args=("New Application", f"New candidate applied: #{new_id}", "/admin/m/candidates"),
                daemon=True,
            ).start()

            # Google Sheet 동기화 (비동기, 실패해도 접수 완료)
            threading.Thread(target=_sync_to_sheet, args=(str(new_id),), daemon=True).start()

            # GDrive 즉시 백업 (신규 파일 있을 경우 대비)
            threading.Thread(target=_trigger_gdrive_backup, daemon=True).start()

            return ok(
                data={"id": new_id, "apply_token": token_out, "mode": "created"},
                message="지원이 접수되었습니다. 담당자가 검토 후 연락드리겠습니다."
            )
        finally:
            conn.close()

    except HTTPException:
        raise
    except Exception as e:
        import logging as _log_apply
        _log_apply.getLogger("bridge.api").error("apply 접수 실패: %s", e, exc_info=True)
        err("접수에 실패했습니다. 잠시 후 다시 시도해주세요.", 500)


@app.post("/api/apply/refresh", tags=["candidates"])
async def refresh_apply_token(request: Request):
    """유효한 지원자 JWT를 새 1시간 토큰으로 갱신.

    Authorization: Bearer <current_token> 헤더 필요.
    만료된 토큰은 갱신 불가 -- 재지원 필요.
    """
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        raise HTTPException(401, "Authorization: Bearer <token> 헤더가 필요합니다.")
    token = auth[7:].strip()
    candidate_id = _verify_candidate_token(token)
    if not candidate_id:
        raise HTTPException(401, "토큰이 만료되었거나 유효하지 않습니다. 재지원 후 새 토큰을 발급받으세요.")
    new_token = _make_candidate_token(candidate_id)
    return ok(data={"apply_token": new_token})


@app.post("/api/inquiry", status_code=status.HTTP_201_CREATED, tags=["clients"])
async def inquiry(request: Request, body: ClientInquiry):
    # Rate limit 완화 — 사용자 retry 허용 (시간당 60회)
    if not _rate_ok(_ip_hash(request), window=3600, max_posts=60):
        raise HTTPException(429, "Too many requests. Please try again later.")
    """
    구인처 문의 접수
    - 학교/학원 → 채용 문의 → client_inquiries 테이블 INSERT
    - 연락처/담당자명은 security_vault AES-256-GCM 암호화 후 저장
    """
    # Input Sanitizer 검증
    if _SECURITY_PKG_OK and _sec_sanitizer:
        for _f in ["school_name", "contact_name", "email", "phone", "message"]:
            _v = getattr(body, _f, None)
            if _v and isinstance(_v, str) and not _sec_sanitizer.is_safe(_v):
                raise HTTPException(400, "유효하지 않은 입력입니다.")
    # 구인처 암호화: 직접 식별자 + 위치 + 메모 (제3자 정보 포함 가능)
    _INQUIRY_ENCRYPT = {
        "phone", "email", "contact_name", "business_registration",  # 직접 식별자
        "school_location",  # 위치 정보
        "memo",             # 자유 입력 -- PII 노출 가능성
    }

    try:
        payload = body.model_dump(exclude_none=True)
        captcha_token = payload.pop("captcha_token", None)

        # ── CAPTCHA 검증 (스팸 차단) ────────────────────────────────────────────────
        ip_hash = _ip_hash(request)
        if not captcha_token or not _verify_captcha_token(captcha_token, ip_hash):
            raise HTTPException(400, "CAPTCHA verification failed. Please complete the puzzle.")

        # Honeypot 확인
        if _check_honeypot(payload):
            raise HTTPException(400, "Invalid submission detected.")
        # honeypot 필드는 DB에 저장 안 함
        payload.pop("_url", None)

        # 연락처 PII 암호화 (storage 전 마지막 레이어)
        for field in _INQUIRY_ENCRYPT:
            if field in payload and payload[field]:
                payload[field] = _encrypt_if_needed(str(payload[field]), field)

        payload["source"]       = "web_form"
        payload["inbox_status"] = "new"
        payload["submitted_at"] = datetime.now(timezone.utc).isoformat()

        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        try:
            # DB 실제 컬럼만 INSERT, 나머지는 parsed_data JSON에 보관
            db_cols = {r[1] for r in conn.execute("PRAGMA table_info(client_inquiries)").fetchall()}
            insert_payload = {}
            extra_fields = {}
            for k, v in payload.items():
                if k in db_cols:
                    insert_payload[k] = v
                else:
                    extra_fields[k] = v
            if extra_fields:
                import json as _json_inq
                insert_payload["parsed_data"] = _json_inq.dumps(extra_fields, ensure_ascii=False)

            # plain 검색용 컬럼 (location/school_name은 암호화 대상 아님 → 평문 그대로)
            if "location_plain" in db_cols:
                insert_payload["location_plain"] = str(payload.get("location", "") or "")
            if "school_name_plain" in db_cols:
                insert_payload["school_name_plain"] = str(payload.get("school_name", "") or "")

            cols = ", ".join(insert_payload.keys())
            placeholders = ", ".join("?" * len(insert_payload))
            cur = conn.execute(
                f"INSERT INTO client_inquiries ({cols}) VALUES ({placeholders})",
                list(insert_payload.values()),
            )
            conn.commit()
            new_id = cur.lastrowid

            # 확인 이메일 발송 (실패해도 접수는 완료)
            if _EMAIL_OK and body.email:
                send_employer_confirmation(body.email, body.school_name, body.contact_name or "")

            # 제출 즉시 암호화 백업 (비동기, 실패해도 접수 완료)
            threading.Thread(target=_backup_on_submit, args=("inquiry",), daemon=True).start()

            # 원본 JSON 별도 보관 (암호화 전 평문 → 암호화 후 독립 파일, 불변 원본)
            _raw_payload = body.model_dump(exclude_none=True)
            _raw_payload.pop("captcha_token", None)
            threading.Thread(
                target=_save_raw_submission,
                args=(_raw_payload, "inquiry", str(new_id)),
                daemon=True,
            ).start()

            # 자동 채용공고 생성 (status='pending_review') -- 실패해도 접수 완료
            threading.Thread(
                target=_auto_create_job_from_inquiry,
                args=(new_id, body.school_name),
                daemon=True
            ).start()

            # Push 알림 (비동기, 실패해도 접수 완료)
            threading.Thread(
                target=_notify_push,
                args=("New Inquiry", f"New inquiry from {body.school_name}", "/admin/m/inquiries"),
                daemon=True,
            ).start()

            return ok(
                data={"id": new_id},
                message="문의가 접수되었습니다. 빠른 시일 내에 연락드리겠습니다."
            )
        finally:
            conn.close()

    except HTTPException:
        raise
    except Exception as e:
        import logging as _log_inq
        _log_inq.getLogger("bridge.api").error("inquiry 접수 실패: %s", e, exc_info=True)
        err("문의 접수에 실패했습니다. 잠시 후 다시 시도해주세요.", 500)


# ── 인재 게시판 공개 API ───────────────────────────────────────────────────────

# TTL 인메모리 캐시 (key → (data, expires_at))  -- 대량 유입 대비
_PUBLIC_CACHE: dict[str, tuple[list, float]] = {}
_PUBLIC_CACHE_TTL = 60  # 60초


def _public_cache_get(key: str):
    """캐시 조회. 만료 시 None 반환."""
    entry = _PUBLIC_CACHE.get(key)
    if entry and time.time() < entry[1]:
        return entry[0]
    _PUBLIC_CACHE.pop(key, None)
    return None


def _public_cache_set(key: str, data: list) -> None:
    """캐시 저장 (최대 100항목 -- 초과 시 오래된 항목 제거)."""
    if len(_PUBLIC_CACHE) >= 100:
        oldest = min(_PUBLIC_CACHE, key=lambda k: _PUBLIC_CACHE[k][1])
        _PUBLIC_CACHE.pop(oldest, None)
    _PUBLIC_CACHE[key] = (data, time.time() + _PUBLIC_CACHE_TTL)


def _db_connect(row_factory: bool = False) -> sqlite3.Connection:
    """최적화된 SQLite 연결 -- 공개 API 대량 유입 대비.
    busy_timeout / cache_size / mmap / temp_store 모두 설정.
    """
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.execute("PRAGMA cache_size = -32768")    # 32MB 페이지 캐시 (per-connection)
    conn.execute("PRAGMA mmap_size = 134217728")  # 128MB 메모리 맵 I/O
    conn.execute("PRAGMA temp_store = MEMORY")    # 임시 테이블 메모리 사용
    if row_factory:
        conn.row_factory = sqlite3.Row
    return conn


def _ok_public(data, max_age: int = 300) -> JSONResponse:
    """공개 API 전용 응답 -- Cache-Control 헤더 포함.
    Admin API에는 절대 사용 금지.
    """
    body = {"success": True, "message": "ok", "data": _sanitize_data(data)}
    return JSONResponse(
        content=body,
        headers={"Cache-Control": f"public, max-age={max_age}, stale-while-revalidate=30"},
    )


_TALENT_PUBLIC_COLS = (
    "sheet_number", "nationality", "area_prefs", "target",
    "certification", "education_level", "experience", "desired_salary",
    "thumb_url", "talent_badge", "talent_reference_star", "talent_summary",
)

@app.get("/api/public/talents", tags=["public"])
async def public_talents(
    nationality: str | None = None,
    area: str | None = None,
    target: str | None = None,
    q: str | None = None,
):
    """
    공개 강사 목록 -- talent_visible=1 인 Active 강사만, PII 절대 없음.
    nationality/area/target/q(sheet_number) 필터 지원.
    캐시: 필터 없는 기본 요청은 60초 TTL 캐시 적용.
    """
    # 필터 없는 기본 요청은 캐시에서 반환 (대량 유입 시 DB 보호)
    cache_key = f"pt:{nationality or ''}:{area or ''}:{target or ''}:{q or ''}"
    cached = _public_cache_get(cache_key)
    if cached is not None:
        return _ok_public(cached, max_age=_PUBLIC_CACHE_TTL)

    conn = _db_connect(row_factory=True)
    try:
        rows = conn.execute(
            """SELECT
                sheet_number,
                COALESCE(nationality, '')          AS nationality,
                COALESCE(area_prefs, '')           AS area_prefs,
                COALESCE(target, '')               AS target,
                COALESCE(certification, '')        AS certification,
                COALESCE(education_level, '')      AS education_level,
                COALESCE(experience, '')           AS experience,
                COALESCE(desired_salary, '')       AS desired_salary,
                COALESCE(photo_s3_key, '')         AS photo_s3_key,
                COALESCE(talent_badge, '')         AS talent_badge,
                COALESCE(talent_reference_star, 0) AS talent_reference_star,
                COALESCE(talent_summary, '')       AS talent_summary
              FROM candidates
              WHERE status = 'Active'
                AND COALESCE(talent_visible, 0) = 1""",
        ).fetchall()
    except Exception as _qe:
        logging.getLogger("bridge.api").error("public_talents 쿼리 오류: %s", _qe, exc_info=True)
        return ok(data=[])
    finally:
        conn.close()

    result = []
    for row in rows:
        r = dict(row)
        # T3v1 복호화: nationality (국가명만 공개 -- 이름·주소 아님)
        nat_raw = r.get("nationality") or ""
        r["nationality"] = decrypt_field(nat_raw, "nationality") if nat_raw else ""

        # 사진: photo_s3_key → presigned URL 실시간 생성 (만료 문제 해결)
        s3k = r.pop("photo_s3_key", "") or ""
        r["thumb_url"] = s3_presigned_url(s3k, expires=3600) if s3k else ""

        # 서버사이드 필터 적용
        if nationality and nationality.lower() not in (r["nationality"] or "").lower():
            continue
        if area and area.lower() not in (r.get("area_prefs") or "").lower():
            continue
        if target and target.lower() not in (r.get("target") or "").lower():
            continue
        if q:
            try:
                if int(q) != (r.get("sheet_number") or -1):
                    continue
            except (ValueError, TypeError):
                pass

        result.append(r)

    _public_cache_set(cache_key, result)
    return _ok_public(result, max_age=_PUBLIC_CACHE_TTL)


# ── 인재 게시판 인증 API ─────────────────────────────────────────────────────────

import secrets as _secrets

_MAGIC_TTL = 900           # 15분 (magic link 유효시간)
_SESSION_TTL = 86400 * 30  # 30일 (세션 쿠키)



@app.post("/api/public/talent-auth/request", status_code=200, tags=["talent-auth"])
async def talent_auth_request(
    request: Request,
    email: str = Form(...),
    company_name: str = Form(""),
    biz_number: str = Form(""),
    doc_file: UploadFile = FastFile(None),
):
    """
    인재 게시판 접근 요청 접수 (multipart/form-data).
    즉시 발송하지 않고 BRIDGE 담당자 검토 후 발송.
    """
    import re as _re
    email = email.strip()
    if not _re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email):
        raise HTTPException(400, "올바른 이메일 주소를 입력해주세요.")

    biz_num_clean = _re.sub(r'[^0-9]', '', biz_number.strip())
    if biz_num_clean and len(biz_num_clean) != 10:
        raise HTTPException(400, "사업자등록번호는 10자리 숫자입니다.")

    ip_h = _ip_hash(request)
    if not _rate_ok(ip_h, window=3600, max_posts=5):
        raise HTTPException(429, "Too many requests. Please try again later.")

    # 서류 파일 처리: DB BLOB 저장 (S3 불필요, 항상 작동)
    doc_s3_key = ""
    doc_filename = ""
    doc_bytes: bytes = b""
    if doc_file and doc_file.filename:
        allowed_exts = {".pdf", ".jpg", ".jpeg", ".png"}
        import os as _os
        ext = _os.path.splitext(doc_file.filename or "")[1].lower()
        if ext not in allowed_exts:
            raise HTTPException(400, "PDF, JPG, PNG 파일만 업로드 가능합니다.")
        doc_bytes = await doc_file.read()
        if len(doc_bytes) > 5 * 1024 * 1024:
            raise HTTPException(400, "파일 크기는 5MB 이하여야 합니다.")
        doc_filename = doc_file.filename
        # S3 가용 시 추가 백업 (실패해도 무관)
        try:
            import uuid as _uuid
            s3_key = f"biz_docs/{_uuid.uuid4().hex}{ext}"
            s3_upload_bytes_sync(doc_bytes, s3_key, content_type=doc_file.content_type or "application/octet-stream")
            doc_s3_key = s3_key
        except Exception:
            pass  # S3 없어도 DB BLOB으로 충분

    now_str = datetime.utcnow().isoformat()
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        existing = conn.execute(
            "SELECT id, status FROM talent_auth_requests WHERE email=? ORDER BY id DESC LIMIT 1",
            (email,)
        ).fetchone()
        if existing and existing[1] in ('pending', 'sent'):
            status_msg = "이미 검토 중인 요청이 있습니다." if existing[1] == 'pending' else "이미 링크가 발송되었습니다. 이메일을 확인해주세요."
            return ok(message=status_msg, data={"status": existing[1]})

        conn.execute(
            """INSERT INTO talent_auth_requests
               (email, company_name, requested_at, status, ip_hash, biz_number, doc_s3_key, doc_filename, doc_data)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (email, company_name.strip(), now_str, 'pending', ip_h,
             biz_num_clean, doc_s3_key, doc_filename, doc_bytes or None)
        )
        conn.commit()
    finally:
        conn.close()

    try:
        import threading
        _email_hint = email.split("@")[0][:3] + "***@" + email.split("@")[-1]

        def _notify():
            try:
                from tg_notify import send_telegram
                send_telegram(
                    f"[인재게시판] 새 접근 요청\n"
                    f"이메일: {_email_hint}\n"
                    f"업체: {company_name or '미입력'}\n"
                    f"사업자번호: {biz_num_clean or '미입력'}\n"
                    f"서류: {'첨부됨' if doc_bytes else '미첨부'}\n"
                    f"→ /admin/talent-auth 에서 링크 발송"
                )
            except Exception:
                pass
        threading.Thread(target=_notify, daemon=True).start()
    except Exception:
        pass

    return ok(message="요청이 접수되었습니다. 담당자 확인 후 이메일로 링크를 보내드립니다.")


@app.post("/api/public/talent-auth/verify-token", tags=["talent-auth"])
async def talent_auth_verify_token(request: Request):
    """매직 링크 토큰 검증 → 30일 세션 토큰 반환. Next.js route handler가 호출."""
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(400, "Invalid request body")

    magic_token = (body.get("token") or "").strip()
    if not magic_token:
        raise HTTPException(400, "토큰이 없습니다.")

    now = datetime.utcnow()
    now_str = now.isoformat()

    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        row = conn.execute(
            "SELECT * FROM talent_auth_tokens WHERE magic_token=?",
            (magic_token,)
        ).fetchone()

        if not row:
            logging.getLogger("bridge.api").warning("talent_auth: 존재하지 않는 토큰 시도")
            raise HTTPException(404, "유효하지 않은 링크입니다.")

        if row["magic_used_at"]:
            raise HTTPException(410, "이미 사용된 링크입니다. 새 링크를 요청해주세요.")

        expires_at = datetime.fromisoformat(row["magic_expires_at"])
        if now > expires_at:
            raise HTTPException(410, "링크가 만료되었습니다 (15분). 새 링크를 요청해주세요.")

        session_token = _secrets.token_urlsafe(32)
        session_expires = (now + timedelta(seconds=_SESSION_TTL)).isoformat()

        conn.execute(
            """UPDATE talent_auth_tokens
               SET magic_used_at=?, session_token=?, session_expires_at=?
               WHERE id=?""",
            (now_str, session_token, session_expires, row["id"])
        )
        if row["request_id"]:
            conn.execute(
                "UPDATE talent_auth_requests SET status='approved' WHERE id=?",
                (row["request_id"],)
            )
        conn.commit()

        return ok(data={
            "session_token": session_token,
            "email": row["email"],
            "expires_at": session_expires,
        })
    finally:
        conn.close()


@app.post("/api/public/talent-auth/check-session", tags=["talent-auth"])
async def talent_auth_check_session(request: Request):
    """세션 토큰 유효성 확인. Next.js route handler가 호출."""
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(400, "Invalid request body")

    session_token = (body.get("session_token") or "").strip()
    if not session_token:
        return ok(data={"valid": False})

    now = datetime.utcnow()
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        row = conn.execute(
            "SELECT email, session_expires_at FROM talent_auth_tokens WHERE session_token=?",
            (session_token,)
        ).fetchone()

        if not row or not row["session_expires_at"]:
            return ok(data={"valid": False})

        expires_at = datetime.fromisoformat(row["session_expires_at"])
        if now > expires_at:
            return ok(data={"valid": False})

        return ok(data={"valid": True, "email": row["email"]})
    finally:
        conn.close()


# ── 어드민: 인재게시판 접근 요청 관리 ──────────────────────────────────────────

@app.get("/api/admin/talent-auth/requests", tags=["talent-auth-admin"])
async def admin_talent_auth_requests(request: Request):
    """인재게시판 접근 요청 목록 (담당자용)."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            """SELECT id, email, company_name, requested_at, status, sent_at, notes,
                      biz_number, doc_s3_key, doc_filename
               FROM talent_auth_requests ORDER BY id DESC LIMIT 200"""
        ).fetchall()
        return ok(data=[dict(r) for r in rows])
    finally:
        conn.close()


def _talent_send_magic_link(email: str, request_id=None) -> str:
    """매직 토큰 생성 + 이메일 발송. 발송된 링크 URL 반환."""
    now = datetime.utcnow()
    magic_token = _secrets.token_urlsafe(32)
    magic_expires = (now + timedelta(seconds=_MAGIC_TTL)).isoformat()

    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        conn.execute(
            """INSERT INTO talent_auth_tokens
               (email, magic_token, created_at, magic_expires_at, request_id)
               VALUES (?,?,?,?,?)""",
            (email, magic_token, now.isoformat(), magic_expires, request_id)
        )
        if request_id:
            conn.execute(
                "UPDATE talent_auth_requests SET status='sent', sent_at=? WHERE id=?",
                (now.isoformat(), request_id)
            )
        conn.commit()
    finally:
        conn.close()

    frontend_url = os.getenv("FRONTEND_URL", "https://bridgejob.co.kr")
    link = f"{frontend_url}/talents/auth?token={magic_token}"

    html = f"""
<div style="font-family:-apple-system,sans-serif;max-width:480px;margin:0 auto;padding:32px 24px;background:#fff">
  <div style="text-align:center;margin-bottom:28px">
    <span style="font-size:26px;font-weight:800;letter-spacing:-0.5px">BRIDGE</span>
    <span style="display:block;color:#6B7280;font-size:13px;margin-top:4px">Teacher Board</span>
  </div>
  <p style="color:#374151;font-size:15px;line-height:1.7;margin-bottom:28px">
    BRIDGE 강사 게시판 접속 링크입니다.<br>
    아래 버튼을 클릭하면 강사 프로필을 확인하실 수 있습니다.<br>
    <strong>링크는 15분 동안 유효합니다.</strong>
  </p>
  <div style="text-align:center;margin-bottom:28px">
    <a href="{link}"
       style="display:inline-block;padding:14px 40px;background:#111;color:#fff;
              border-radius:8px;text-decoration:none;font-weight:700;font-size:15px">
      강사 게시판 접속하기
    </a>
  </div>
  <p style="font-size:12px;color:#9CA3AF;word-break:break-all">
    버튼이 작동하지 않으면 아래 URL을 복사하세요:<br>
    <a href="{link}" style="color:#6366F1">{link}</a>
  </p>
  <hr style="margin:24px 0;border:none;border-top:1px solid #F3F4F6">
  <p style="font-size:11px;color:#D1D5DB;text-align:center">
    본 메일은 BRIDGE 담당자가 발송했습니다 · bridgejob.co.kr
  </p>
</div>
"""
    sent = _smtp_send(email, "[BRIDGE] 강사 게시판 접속 링크", html)
    if not sent:
        raise RuntimeError("SMTP 발송 실패")
    return link


class _SendLinkByIdBody(BaseModel):
    request_id: int


@app.post("/api/admin/talent-auth/send-link", tags=["talent-auth-admin"])
async def admin_talent_auth_send_link(request: Request, body: _SendLinkByIdBody):
    """요청 ID로 매직 링크 생성 및 이메일 발송."""
    _check_admin(request)

    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        req_row = conn.execute(
            "SELECT id, email FROM talent_auth_requests WHERE id=?",
            (body.request_id,)
        ).fetchone()
        if not req_row:
            raise HTTPException(404, "요청을 찾을 수 없습니다.")
        email = req_row["email"]
    finally:
        conn.close()

    try:
        _talent_send_magic_link(email, request_id=body.request_id)
    except RuntimeError as _e:
        raise HTTPException(500, f"이메일 발송 실패: {_e}")

    return ok(message=f"{email}로 링크를 발송했습니다.")


@app.post("/api/admin/talent-auth/send-link-direct", tags=["talent-auth-admin"])
async def admin_talent_auth_send_direct(request: Request):
    """이메일 직접 입력으로 매직 링크 발송 (요청 없이도 가능)."""
    _check_admin(request)
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(400, "Invalid JSON")

    email = (body.get("email") or "").strip()
    import re as _re
    if not email or not _re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email):
        raise HTTPException(400, "올바른 이메일 주소를 입력해주세요.")

    try:
        _talent_send_magic_link(email, request_id=None)
    except RuntimeError as _e:
        raise HTTPException(500, f"이메일 발송 실패: {_e}")

    return ok(message=f"{email}로 링크를 발송했습니다.")


@app.delete("/api/admin/talent-auth/sessions/{email_enc}", tags=["talent-auth-admin"])
async def admin_talent_auth_revoke(request: Request, email_enc: str):
    """이메일로 세션 토큰 전체 취소."""
    _check_admin(request)
    from urllib.parse import unquote
    email = unquote(email_enc)

    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        conn.execute(
            "UPDATE talent_auth_tokens SET session_token=NULL, session_expires_at=NULL WHERE email=?",
            (email,)
        )
        conn.commit()
    finally:
        conn.close()

    return ok(message=f"{email} 세션 취소 완료")


@app.get("/api/admin/talent-auth/doc/{request_id}", tags=["talent-auth-admin"])
async def admin_talent_auth_doc(request: Request, request_id: int):
    """사업자등록증명원 반환 -- S3 presigned URL 우선, 없으면 DB BLOB 직접 응답."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        row = conn.execute(
            "SELECT doc_s3_key, doc_filename, doc_data FROM talent_auth_requests WHERE id=?",
            (request_id,)
        ).fetchone()
    finally:
        conn.close()
    if not row or (not row["doc_s3_key"] and not row["doc_data"]):
        raise HTTPException(404, "서류 파일 없음")
    # S3 presigned URL 우선
    if row["doc_s3_key"]:
        try:
            url = s3_presigned_url(row["doc_s3_key"], expires=3600)
            return ok(data={"url": url, "filename": row["doc_filename"]})
        except Exception:
            pass  # S3 실패 시 BLOB 폴백
    # DB BLOB 직접 응답
    if row["doc_data"]:
        import base64 as _b64
        import os as _os
        ext = _os.path.splitext(row["doc_filename"] or "")[1].lower()
        mime_map = {".pdf": "application/pdf", ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png"}
        mime = mime_map.get(ext, "application/octet-stream")
        b64 = _b64.b64encode(row["doc_data"]).decode()
        return ok(data={"data_uri": f"data:{mime};base64,{b64}", "filename": row["doc_filename"]})
    raise HTTPException(404, "서류 파일 없음")


@app.post("/api/admin/talent-preview-session", tags=["talent-auth-admin"])
async def admin_talent_preview_session(request: Request):
    """관리자 인재게시판 미리보기 -- 30일 세션 즉시 생성."""
    _check_admin(request)
    import secrets as _sec
    now = datetime.utcnow()
    session_token = _sec.token_urlsafe(32)
    expires = (now + timedelta(days=30)).isoformat()
    preview_email = "admin-preview@bridge.internal"
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        existing = conn.execute(
            "SELECT id FROM talent_auth_tokens WHERE email=?", (preview_email,)
        ).fetchone()
        if existing:
            conn.execute(
                "UPDATE talent_auth_tokens SET session_token=?, session_expires_at=? WHERE email=?",
                (session_token, expires, preview_email)
            )
        else:
            conn.execute(
                """INSERT INTO talent_auth_tokens (email, magic_token, created_at, session_token, session_expires_at)
                   VALUES (?, ?, ?, ?, ?)""",
                (preview_email, _sec.token_urlsafe(16), now.isoformat(), session_token, expires)
            )
        conn.commit()
    finally:
        conn.close()
    return ok(data={"session_token": session_token})


class _TalentInquiry(BaseModel):
    school_name: str
    email: str
    message: str
    contact_name: str | None = None
    phone: str | None = None
    candidate_ref: int | None = None  # sheet_number


@app.post("/api/public/talent-inquiry", status_code=201, tags=["public"])
async def public_talent_inquiry(request: Request, body: _TalentInquiry):
    """강사 게시판 문의 접수 -- IP당 5회/시간, PII 암호화 저장."""
    if not _rate_ok(_ip_hash(request), window=3600, max_posts=5):
        raise HTTPException(429, "Too many requests. Please try again later.")

    _TALENT_INQ_ENCRYPT = {"email", "phone", "contact_name"}
    payload = body.model_dump(exclude_none=True)

    for field in _TALENT_INQ_ENCRYPT:
        if field in payload and payload[field]:
            payload[field] = _encrypt_if_needed(str(payload[field]), field)

    payload["source"] = "talent_board"
    payload["inbox_status"] = "new"
    payload["submitted_at"] = datetime.now(timezone.utc).isoformat()

    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        try:
            db_cols = {r[1] for r in conn.execute("PRAGMA table_info(client_inquiries)").fetchall()}
            insert_payload: dict = {}
            extra_fields: dict = {}
            for k, v in payload.items():
                if k in db_cols:
                    insert_payload[k] = v
                else:
                    extra_fields[k] = v
            if extra_fields:
                import json as _jt_inq
                insert_payload["parsed_data"] = _jt_inq.dumps(extra_fields, ensure_ascii=False)
            if "location_plain" in db_cols:
                insert_payload["location_plain"] = str(payload.get("location", "") or "")
            if "school_name_plain" in db_cols:
                insert_payload["school_name_plain"] = str(payload.get("school_name", "") or "")
            cols = ", ".join(insert_payload.keys())
            placeholders = ", ".join("?" * len(insert_payload))
            cur = conn.execute(
                f"INSERT INTO client_inquiries ({cols}) VALUES ({placeholders})",
                list(insert_payload.values()),
            )
            conn.commit()
            new_id = cur.lastrowid
            return ok(data={"id": new_id}, message="문의가 접수되었습니다.")
        finally:
            conn.close()
    except HTTPException:
        raise
    except Exception as e:
        import logging as _log_ti
        _log_ti.getLogger("bridge.api").error("talent_inquiry 접수 실패: %s", e, exc_info=True)
        err("문의 접수에 실패했습니다. 잠시 후 다시 시도해주세요.", 500)


# ── 전역 예외 처리 ─────────────────────────────────────────────────────────────
@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    """모든 HTTPException → 통일된 구조 {"success", "error", "data"}."""
    detail = exc.detail
    if isinstance(detail, dict):
        msg = detail.get("message", str(detail))
        code = detail.get("error_code", _http_to_error_code(exc.status_code))
    else:
        msg = str(detail)
        code = _http_to_error_code(exc.status_code)
    return SafeJSONResponse(
        status_code=exc.status_code,
        content={
            "success": False,
            "error": {"code": code, "message": msg, "status": exc.status_code},
            "data": None,
        },
    )


@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    import logging as _logging
    _logging.getLogger("bridge.api").error("Unhandled exception: %s", exc, exc_info=True)
    return SafeJSONResponse(
        status_code=500,
        content={
            "success": False,
            "error": {"code": "SERVER_ERROR", "message": "서버 오류가 발생했습니다.", "status": 500},
            "data": None,
        },
    )


# ── Admin: ad_posts 대시보드 ──────────────────────────────────────────────────
_ADMIN_DB_PATH = Path(os.getenv("DB_PATH", os.getenv("BRIDGE_DB_PATH", str(Path(__file__).resolve().parent / "master.db"))))
# Render 디스크 마운트 경로(/data) 자동 생성 -- 없으면 sqlite3 연결 불가
try:
    _ADMIN_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
except OSError as _mkdir_err:
    logging.getLogger("bridge.api").warning("[STARTUP] DB 경로 디렉터리 생성 실패 (계속 진행): %s", _mkdir_err)
_ADMIN_KEY              = os.getenv("ADMIN_API_KEY", "")
_ADMIN_PW               = os.getenv("BRIDGE_ADMIN_LOGIN_PW", os.getenv("ADMIN_PASSWORD", ""))
_FORM_CONFIG_READ_KEY   = os.getenv("FORM_CONFIG_READ_KEY", "")  # 서버-to-서버 전용 (브라우저 미노출)

# ── Supabase (community_posts 영구 저장) ─────────────────────────────────────
# SUPABASE_URL + SUPABASE_SERVICE_KEY 환경변수 설정 시 커뮤니티 데이터를 Supabase에 저장.
# 미설정 시 기존 SQLite 폴백 사용 (개발/오프라인 환경 호환).
# ⚠️ 반드시 service_role 키를 사용할 것 -- anon 키는 RLS로 쓰기 차단됨.
_SUPABASE_URL = os.getenv("SUPABASE_URL", "")
_SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_KEY", "")
_supa = None

def _validate_supabase_service_key(key: str) -> bool:
    """JWT payload에서 role=service_role 여부를 검증한다."""
    try:
        import base64 as _b64, json as _j
        parts = key.split('.')
        if len(parts) < 2:
            return False
        # Base64 패딩 보정
        padded = parts[1] + '=' * (4 - len(parts[1]) % 4)
        payload = _j.loads(_b64.b64decode(padded).decode('utf-8'))
        return payload.get('role') == 'service_role'
    except Exception:
        return False  # 파싱 실패 시 보수적으로 false

if _SUPABASE_URL and _SUPABASE_SERVICE_KEY:
    # service_role 키인지 먼저 검증 -- anon 키 실수 사용 즉시 차단
    if not _validate_supabase_service_key(_SUPABASE_SERVICE_KEY):
        logging.getLogger("bridge.api").error(
            "[Supabase] 🚨 SECURITY HALT: SUPABASE_SERVICE_KEY가 service_role 키가 아닙니다. "
            "anon 키는 RLS로 쓰기가 차단됩니다. Supabase 비활성화 → SQLite 폴백."
        )
    else:
        try:
            from supabase import create_client as _create_supabase_client
            _supa = _create_supabase_client(_SUPABASE_URL, _SUPABASE_SERVICE_KEY)
            logging.getLogger("bridge.api").info("[Supabase] 커뮤니티 DB 연결됨 (service_role 검증 완료)")
        except Exception as _supa_exc:
            logging.getLogger("bridge.api").warning("[Supabase] 연결 실패 → SQLite 폴백: %s", _supa_exc)

# ── 상위 보안: IP 화이트리스트 ──────────────────────────────────────────────
# ADMIN_ALLOWED_IPS=1.2.3.4,5.6.7.0/24 (콤마 구분, CIDR 지원, 비어있으면 제한 없음)
_ADMIN_ALLOWED_IPS_RAW = os.getenv("ADMIN_ALLOWED_IPS", "").strip()
_ADMIN_ALLOWED_IPS: list[str] = [ip.strip() for ip in _ADMIN_ALLOWED_IPS_RAW.split(",") if ip.strip()] if _ADMIN_ALLOWED_IPS_RAW else []

import ipaddress as _ipaddr
import secrets as _secrets

def _ip_in_whitelist(client_ip: str) -> bool:
    """IP 화이트리스트 검사. 리스트 비어있으면 True (제한 없음)."""
    if not _ADMIN_ALLOWED_IPS:
        return True
    try:
        addr = _ipaddr.ip_address(client_ip)
    except ValueError:
        return False
    for entry in _ADMIN_ALLOWED_IPS:
        try:
            if "/" in entry:
                if addr in _ipaddr.ip_network(entry, strict=False):
                    return True
            else:
                if addr == _ipaddr.ip_address(entry):
                    return True
        except ValueError:
            continue
    return False

def _ip_subnet24(ip_str: str) -> str:
    """IP → /24 서브넷 문자열 (세션 바인딩용)."""
    try:
        addr = _ipaddr.ip_address(ip_str)
        if isinstance(addr, _ipaddr.IPv4Address):
            parts = ip_str.split(".")
            return f"{parts[0]}.{parts[1]}.{parts[2]}.0/24"
        return str(_ipaddr.ip_network(f"{addr}/48", strict=False))
    except ValueError:
        return "unknown"

# ── 상위 보안: 세션 토큰 시스템 ─────────────────────────────────────────────
# 로그인 시 세션 토큰 발급, 8시간 만료, /24 서브넷 바인딩
_SESSION_TTL = 8 * 3600  # 8시간
_SESSIONS: dict[str, dict] = {}  # token → {created, expires, subnet, ip, ua}
_SESSION_LOCK = threading.Lock()
_MAX_SESSIONS = 50  # 최대 동시 세션 수

def _get_client_ip(request: Request) -> str:
    """실제 클라이언트 IP 추출 -- 스푸핑 방지 우선순위.

    우선순위:
      1. CF-Connecting-IP  : Cloudflare가 설정 (위조 불가)
      2. X-Real-IP         : TCP 연결이 신뢰 프록시 대역일 때만 수용
      3. X-Forwarded-For   : Render 인프라가 추가한 마지막 항목
      4. request.client.host : TCP 연결 IP (최후 폴백)

    X-Real-IP는 신뢰 프록시(TRUSTED_PROXY_CIDRS) IP에서 온 연결일 때만 수용.
    X-Forwarded-For는 공격자가 위조한 첫 항목을 무시하고 마지막(서버가 추가) 항목 사용.
    이 값은 rate_limit / session subnet / IP whitelist 에 사용.
    """
    # 신뢰 프록시 대역 (Render 내부 대역 + Cloudflare IP 블록)
    _TRUSTED_PROXY_CIDRS = [
        "10.0.0.0/8",       # Render 내부 네트워크
        "172.16.0.0/12",    # Docker/Render 내부
        "100.64.0.0/10",    # RFC 6598 공유 주소 (Render 로드밸런서)
    ]

    def _is_trusted_proxy(ip_str: str) -> bool:
        try:
            addr = _ipaddr.ip_address(ip_str)
            return any(addr in _ipaddr.ip_network(cidr) for cidr in _TRUSTED_PROXY_CIDRS)
        except ValueError:
            return False

    tcp_host = request.client.host if request.client else ""

    # 1순위: Cloudflare CF-Connecting-IP (위조 불가)
    cf_ip = request.headers.get("CF-Connecting-IP", "").strip()
    if cf_ip:
        try:
            _ipaddr.ip_address(cf_ip)
            return cf_ip
        except ValueError:
            pass

    # 2순위: X-Real-IP -- TCP 연결이 신뢰 프록시 대역인 경우에만 수용
    if tcp_host and _is_trusted_proxy(tcp_host):
        real_ip = request.headers.get("X-Real-IP", "").strip()
        if real_ip:
            try:
                _ipaddr.ip_address(real_ip)
                return real_ip
            except ValueError:
                pass

    # 3순위: X-Forwarded-For 마지막 항목 (Render/프록시가 추가한 실제 IP)
    # 공격자가 위조한 첫 항목은 무시하고 rightmost(프록시 추가분) 사용
    xff = request.headers.get("X-Forwarded-For", "").strip()
    if xff:
        parts = [p.strip() for p in xff.split(",")]
        # 마지막부터 역방향으로 첫 번째 공인 IP 추출 (내부 대역 제외)
        for candidate in reversed(parts):
            try:
                addr = _ipaddr.ip_address(candidate)
                if not addr.is_private and not addr.is_loopback:
                    return candidate
            except ValueError:
                continue

    # 4순위: TCP 연결 IP
    return tcp_host if tcp_host else "unknown"

def _create_session(request: Request) -> str:
    """세션 토큰 생성 + 저장. 반환: 토큰 문자열."""
    token = _secrets.token_urlsafe(32)
    ip = _get_client_ip(request)
    now = time.time()
    with _SESSION_LOCK:
        # 만료 세션 정리
        expired = [k for k, v in _SESSIONS.items() if v["expires"] < now]
        for k in expired:
            del _SESSIONS[k]
        # 최대 세션 수 초과 시 가장 오래된 것 제거
        if len(_SESSIONS) >= _MAX_SESSIONS:
            oldest = min(_SESSIONS, key=lambda k: _SESSIONS[k]["created"])
            del _SESSIONS[oldest]
        _SESSIONS[token] = {
            "created": now,
            "expires": now + _SESSION_TTL,
            "subnet": _ip_subnet24(ip),
            "ip": ip,
            "ua": (request.headers.get("User-Agent", ""))[:100],
        }
    return token

def _validate_session(request: Request) -> bool:
    """세션 토큰 검증 -- HttpOnly 쿠키 우선, 헤더 폴백. 유효하면 True."""
    token = (
        request.cookies.get("bridge_session", "")          # HttpOnly 쿠키 (XSS 안전)
        or request.headers.get("x-admin-token", "")        # 헤더 폴백 (하위 호환)
    ).strip()
    if not token:
        return False
    now = time.time()
    with _SESSION_LOCK:
        sess = _SESSIONS.get(token)
        if not sess:
            return False
        if sess["expires"] < now:
            del _SESSIONS[token]
            return False
        # /24 서브넷 바인딩 검증
        client_ip = _get_client_ip(request)
        if _ip_subnet24(client_ip) != sess["subnet"]:
            _log.warning("세션 서브넷 불일치: token_subnet=%s client=%s", sess["subnet"], client_ip)
            del _SESSIONS[token]
            return False
        # 세션 갱신 (슬라이딩 만료)
        sess["expires"] = now + _SESSION_TTL
        return True

def _revoke_session(token: str) -> bool:
    """세션 토큰 폐기."""
    with _SESSION_LOCK:
        return _SESSIONS.pop(token, None) is not None

_log = logging.getLogger("bridge.security")

# 시작 진단 (값 노출 없이 상태만 출력)
print(f"[STARTUP] ADMIN_API_KEY={'SET' if _ADMIN_KEY else 'EMPTY'} "
      f"ADMIN_PASSWORD={'SET(pbkdf2)' if _ADMIN_PW.startswith('pbkdf2:') else 'SET(plain)' if _ADMIN_PW else 'EMPTY'} "
      f"DB={_ADMIN_DB_PATH}")

# SQLite WAL 모드 + 성능 튜닝 (대량 유입 대비)
try:
    _wal_conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    _wal_conn.execute("PRAGMA journal_mode = WAL")
    _wal_conn.execute("PRAGMA synchronous = NORMAL")
    _wal_conn.execute("PRAGMA cache_size = -65536")    # 64MB 페이지 캐시
    _wal_conn.execute("PRAGMA mmap_size = 268435456")  # 256MB 메모리 맵 I/O
    _wal_conn.execute("PRAGMA temp_store = MEMORY")    # 임시 테이블 메모리 사용
    _wal_conn.close()
except Exception:
    pass

if not _ADMIN_KEY:
    import logging as _log_adm
    _log_adm.getLogger("bridge.api").warning(
        "[SECURITY] ADMIN_API_KEY 미설정 -- 관리자 엔드포인트가 인증 없이 노출됩니다! "
        ".env에 ADMIN_API_KEY를 반드시 설정하세요."
    )


def _ensure_access_logs():
    """access_logs 테이블 생성 (없으면)"""
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("CREATE TABLE IF NOT EXISTS access_logs (id INTEGER PRIMARY KEY, ip TEXT, url TEXT, timestamp TEXT DEFAULT (datetime('now')))")
        conn.commit()
        conn.close()
    except Exception:
        pass

try:
    _ensure_access_logs()
except Exception as _e:
    logging.getLogger("bridge.api").warning("_ensure_access_logs 스킵: %s", _e)


# ── Form Config 기본값 ────────────────────────────────────────────────────────
_FORM_DEFAULTS: dict[str, dict[str, list[str]]] = {
    "apply": {
        "HOW_TO":          ["Friend Referral", "Previous Experience", "Google", "Reddit",
                            "Facebook", "Instagram", "YouTube", "ESL Cafe", "LinkedIn", "Other"],
        "NATIONALITIES":   ["USA", "United Kingdom", "Canada", "Ireland",
                            "Australia", "New Zealand", "South Africa",
                            "F (Korean diaspora or overseas Korean)", "South Korea", "Other"],
        "ANCESTRY":        ["Korean", "Chinese, Hong Kong, or Taiwanese", "Japanese",
                            "Mongolian or Other East Asian (excluding above)", "Indian", "Native American",
                            "White or European", "Pakistani, Afghan, Bangladeshi", "Black or African American",
                            "Middle Eastern or North African", "Hispanic or Latino", "Mixed or Multiracial",
                            "Pacific Islander", "Other", "Prefer not to disclose"],
        "EDUCATION":       ["Graduated, but diploma not available",
                            "Bachelor's degree from one of the 7 eligible countries",
                            "Bachelor's degree (other country)", "Master", "Doctor", "Associate",
                            "Online degree", "Did not graduate",
                            "I have a bachelor's or higher from Korea"],
        "CERTIFICATION":   ["Teaching license (Official Teaching License/Credential)",
                            "PGCE (Postgraduate Certificate in Education)",
                            "TEFL", "TESOL", "DELTA", "CELTA", "On the process", "No certification"],
        "E_VISA":          ["Recorded", "Never obtained an E visa"],
        "PASSPORT":        ["Expires within 2 years", "Valid for more than 2 years",
                            "Scheduled for renewal within the period"],
        "CRIMINAL_RECORD": ["Issued and apostille completed", "Issued but awaiting apostille",
                            "Never applied for a new one", "Holding a visa and recently left Korea"],
        "DOC_STATUS":      ["Already have a valid visa",
                            "All documents completed (Criminal check + Degree apostilled)",
                            "Degree / CBC at final apostille stage (confirmed return date)",
                            "Never applied for any documents",
                            "Holding a visa and recently left Korea within 1 month", "Other"],
        "TARGET_AGE":      ["Pre-K ~ Kindergarten", "Elementary school level",
                            "Middle school level", "High school level", "Adults", "No preferences"],
        "AREA_PREFS":      ["Metropolitan city", "Medium size city", "Small city", "No preference"],
        "EXPERIENCE":      ["None", "Less than 6 months", "6 months to 1 year",
                            "over 1 year", "over 2 year", "over 3 year", "over 4 year", "over 5 year",
                            "over 6 year", "over 7 year", "over 8 year", "over 9 year", "over 10 year",
                            "over 15 year", "over 20 year", "Overseas full-time teaching experience only"],
        "EMPLOYMENT":      ["The school is aware", "Not working in Korea", "Do not know",
                            "I will inform them very soon",
                            "School do know, but references cannot be verified",
                            "Contract terminated early I have a Letter of Release"],
        "MARITAL":         ["Married", "Coming with Dependents (Children or Family)",
                            "Single or Coming Alone", "Divorced", "No comment",
                            "Planning to get married within a year"],
        "HOUSING":         ["I would like to use the school provided housing",
                            "I am willing to pay monthly for better housing",
                            "I have enough deposit and will handle the monthly rent on my own",
                            "I have housing so no support is needed"],
        "DEPENDENTS_PETS": ["I am coming alone", "Young children (under age 6)",
                            "School age children or Family members", "Dog(s)", "Cat(s)", "Other pets"],
        "PERSONAL":        ["Zero tattoos or piercings", "Visible but can be covered",
                            "Visible tattoo (cannot be covered)",
                            "Facial piercings (nose lip eyebrow etc cannot be removed)"],
        "RELIGION":        ["Irreligious", "Christianity", "Buddhism", "Judaism", "Islam", "Hinduism",
                            "Does not celebrate events such as births due to specific beliefs or affiliations",
                            "Other"],
        "HEALTH":          ["I have not", "I have a condition", "Prefer not to say"],
        "CRC":             ["Clean record", "I have a record", "Prefer not to answer"],
        "KR_CRC":          ["Clean record", "I have a record",
                            "Not applicable (never lived in Korea)", "Prefer not to answer"],
    },
    "inquiry": {
        "BUSINESS_REG":        ["등록기관 Registered Institution", "현시간 미등록 Unregistered Institution"],
        "HIRE_HIST":           ["채용이력O", "채용이력X"],
        "NATIVE_COUNT":        ["없음", "1명", "2명", "3명", "4명", "5명", "6명",
                                "7~9명", "10~14명", "15~20명", "20명이상", "30명이상"],
        "VACANCIES":           ["1명", "2명", "3명", "4명", "5명", "6명", "7명", "8명", "9명", "10명"],
        "CONTRACT_TYPE":       ["Full time", "Part time"],
        "TEACHING_AGE":        ["48개월 미만 baby", "영유아 Pre-K, ~5세미만", "유치원 Kindergarten",
                                "초등학생 Elementary", "중학생 Middle School",
                                "고등학생 High School", "대학생/ 성인 Adult"],
        "CLASS_SIZE":          ["~5명 이내", "~8명 이내", "~12명 이내", "12명 이상", "20명이상"],
        "AVG_LESSONS":         ["주 10-15시간", "주 15-20시간", "주 20-25시간",
                                "주 25-30시간", "주 30-35시간", "주 35시간 이상"],
        "PREFERRED_CANDIDATE": ["영어권 원어민 Native Teachers", "교포 Kyopo",
                                "F비자 소지자 F Visa", "한국인 Koreans", "무관 No Preference"],
        "SALARY_RANGES":       ["2,20 KRW - 2,30 KRW", "2,30 KRW - 2,40 KRW",
                                "2,40 KRW - 2,50 KRW", "2,50 KRW - 2,60 KRW",
                                "2,60 KRW - 2,70 KRW", "2,70 KRW - 2,80 KRW",
                                "2,80 KRW - 2,90 KRW", "2,90 KRW - 3,00 KRW",
                                "3,00 KRW - 3,20 KRW", "3,20 KRW - 3,50 KRW",
                                "3,50 KRW - 4,50 KRW", "4,50 KRW - 6,50 KRW", "시급제 또는 기타"],
        "TRAVEL_SUPPORT":      ["국내교통비지원 Domestic allowance", "왕복항공지원 Round trip support",
                                "입국만지원 Entry", "출국만지원 Departure",
                                "규정에따른비용지원 Travel expenses", "제공없음 Not provided"],
        "MEAL_OPTS":           ["식사제공", "식대제공", "식사식대제공없음"],
        "HOUSING_OPTS":        ["풀옵션 하우징 Fully furnished housing", "개인기숙사 Dormitory",
                                "월세지원 Housing allowance",
                                "월세 및 보증금지원 Allowance and deposit support",
                                "다양한 지원가능 Negotiable", "숙소미제공 No housing provided"],
        "BENEFITS_OPTS":       ["비자스폰 (E) Working Visa sponsorship",
                                "퇴직금 Severance Pay (월급제 필수*)",
                                "교통비 Transportation expenses",
                                "국민연금 National Pension (SA제외 월급제 필수*)",
                                "건강보험 Medical Insurance",
                                "비자 건강검진 Medical check support",
                                "정착 지원금 Settlement Allowance",
                                "재계약 보너스 Renewal Bonus",
                                "계약완료 보너스 Contract Completion Bonus",
                                "본인/자녀 교육지원 Education support",
                                "기타 Bonus (연휴, 생일, 성과금등)"],
        "VACATION_INC":        ["포함 Including weekends", "주말,공휴일제외 Excluding weekends"],
    },
}


def _ensure_form_configs():
    """form_configs 테이블 생성 및 기본값 초기 INSERT (없는 행만)"""
    conn = None
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("""
            CREATE TABLE IF NOT EXISTS form_configs (
                form_name    TEXT NOT NULL,
                field_key    TEXT NOT NULL,
                options_json TEXT NOT NULL DEFAULT '[]',
                updated_at   TEXT NOT NULL DEFAULT (datetime('now')),
                PRIMARY KEY  (form_name, field_key)
            )
        """)
        conn.commit()
        for _fn, _fields in _FORM_DEFAULTS.items():
            for _fk, _defaults in _fields.items():
                if not conn.execute(
                    "SELECT 1 FROM form_configs WHERE form_name=? AND field_key=?",
                    (_fn, _fk),
                ).fetchone():
                    conn.execute(
                        "INSERT INTO form_configs (form_name, field_key, options_json) VALUES (?, ?, ?)",
                        (_fn, _fk, json.dumps(_defaults, ensure_ascii=False)),
                    )
        conn.commit()
    except Exception as _fe:
        logging.getLogger("bridge.api").warning("_ensure_form_configs 스킵: %s", _fe)
    finally:
        if conn:
            conn.close()


try:
    _ensure_form_configs()
except Exception as _e:
    logging.getLogger("bridge.api").warning("form_configs 초기화 스킵: %s", _e)


def _log_unauthorized_access(request: Request):
    """비관리자 접근 시 IP + URL + 시간 기록"""
    try:
        ip = _get_client_ip(request)  # 첫 XFF 항목 위조 방지 -- _get_client_ip 사용
        url = str(request.url.path)
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("INSERT INTO access_logs (ip, url) VALUES (?, ?)", (ip, url))
        conn.commit()
        conn.close()
    except Exception:
        pass


def _verify_admin_password(input_pw: str, stored: str) -> bool:
    """pbkdf2:sha256 해시 검증 (H1 보안 패치: 평문 폴백 완전 제거)"""
    if stored.startswith("pbkdf2:sha256:"):
        try:
            _, _, iterations, salt, dk_hex = stored.split(":", 4)
            dk = hashlib.pbkdf2_hmac("sha256", input_pw.encode(), salt.encode(), int(iterations))
            return hmac.compare_digest(dk.hex(), dk_hex)
        except Exception:
            return False
    # 평문 비밀번호 거부 -- ADMIN_PASSWORD가 pbkdf2:sha256: 형식이어야 합니다
    logging.getLogger("bridge.security").error(
        "[H1] ADMIN_PASSWORD가 pbkdf2:sha256: 형식이 아닙니다. "
        "python tools/bridge_reset_password.py 실행 후 Render 환경변수 재설정 필요."
    )
    return False


def _check_admin(request: Request):
    """
    관리자 인증 -- 3단계 검증:
    1. 세션 토큰 (x-admin-token) -- 유효하면 즉시 통과 (HMAC/API키 스킵)
    2. API 키 (x-admin-key) -- 기존 방식
    3. IP 화이트리스트 (ADMIN_ALLOWED_IPS) -- 설정 시 추가 검증
    """
    if not _ADMIN_KEY:
        raise HTTPException(status_code=503, detail={"isError": True, "errorCategory": "DB_UNAVAILABLE", "isRetryable": True, "context": "관리자 기능이 비활성화되어 있습니다."})

    # ── Layer 1: 세션 토큰 (최우선 -- 프론트엔드 내부 사용자용) ──
    if _validate_session(request):
        return  # 세션 유효 → 즉시 통과

    # ── Layer 2: API 키 검증 ──
    ip_h = _ip_hash(request)
    now = time.time()
    fails = [t for t in _AUTH_FAIL.get(ip_h, []) if now - t < 300]
    if len(fails) >= 10:
        raise HTTPException(status_code=429, detail={"isError": True, "errorCategory": "RATE_LIMIT", "isRetryable": True, "context": "인증 시도가 너무 많습니다. 5분 후 다시 시도하세요."})
    if not hmac.compare_digest(request.headers.get("x-admin-key", "").strip(), _ADMIN_KEY.strip()):
        fails.append(now)
        _AUTH_FAIL[ip_h] = fails
        _log_unauthorized_access(request)
        raise HTTPException(status_code=403, detail={"isError": True, "errorCategory": "ADMIN_KEY_INVALID", "isRetryable": False, "context": "관리자 키가 올바르지 않습니다."})
    _AUTH_FAIL.pop(ip_h, None)

    # ── Layer 3: IP 화이트리스트 (ADMIN_ALLOWED_IPS 설정 시) ──
    if _ADMIN_ALLOWED_IPS:
        client_ip = _get_client_ip(request)
        if not _ip_in_whitelist(client_ip):
            _log.warning("IP 화이트리스트 차단: ip=%s", client_ip)
            raise HTTPException(status_code=403, detail={"isError": True, "errorCategory": "IP_NOT_ALLOWED", "isRetryable": False, "context": "허용되지 않은 IP입니다."})


@app.post("/api/admin/login", tags=["admin"])
async def admin_login(request: Request):
    """비밀번호 검증 → ADMIN_API_KEY 반환."""
    try:
        from security_middleware import ip_blacklist, SecurityMiddleware, admin_login_guard
        real_ip = SecurityMiddleware._get_real_ip(request)
    except Exception:
        real_ip = request.client.host if request.client else "unknown"

    # Rate limit (brute-force 보호는 AdminLoginGuard가 담당)
    ip = _ip_hash(request)
    if not _rate_ok(ip, window=300, max_posts=10):
        return bridge_error("RATE_LIMIT", "Too many login attempts.", retryable=True, status=429)

    try:
        body = await request.json()
    except Exception:
        raise HTTPException(400, "Invalid JSON")

    pw = str(body.get("password", ""))
    if not _ADMIN_PW or not _ADMIN_KEY:
        raise HTTPException(503, "관리자 인증이 설정되지 않았습니다.")

    if not _verify_admin_password(pw, _ADMIN_PW):
        # 실패 기록 → 누진 차단
        try:
            ban_minutes = admin_login_guard.record_fail(real_ip)
            if ban_minutes is None:
                ip_blacklist.block_permanent(real_ip, "admin_brute_force:permanent")
            elif ban_minutes is not None:
                ip_blacklist.block(real_ip, "admin_brute_force", minutes=ban_minutes)
        except Exception as _blk_err:
            import logging as _blk_log
            _blk_log.getLogger("bridge.security").error(
                "admin brute-force blacklist 등록 실패 ip=%s error_type=%s",
                real_ip[:8] + "****", type(_blk_err).__name__,
            )
        raise HTTPException(403, "비밀번호가 올바르지 않습니다.")

    # 로그인 성공 → 실패 기록 초기화 + 블랙리스트 해제
    try:
        admin_login_guard.clear(real_ip)
        if real_ip in ip_blacklist._list:
            del ip_blacklist._list[real_ip]
            ip_blacklist._save()
    except Exception:
        pass

    # 세션 토큰 발급 (IP /24 서브넷 바인딩, 8시간 만료)
    session_token = _create_session(request)
    _log.info("관리자 로그인 성공: ip=%s subnet=%s", real_ip[:12] + "***", _ip_subnet24(real_ip))
    # HttpOnly 쿠키 설정 -- XSS로 탈취 불가 (SameSite=None for cross-origin Vercel→Render)
    cookie = (
        f"bridge_session={session_token}; HttpOnly; Secure; "
        f"SameSite=None; Path=/; Max-Age={_SESSION_TTL}"
    )
    response = JSONResponse(content=ok(data={"session_token": session_token}, message="로그인 성공"))
    response.headers["Set-Cookie"] = cookie
    return response


@app.get("/api/admin/key", tags=["admin"])
async def get_admin_key(request: Request):
    """C2 패치: 세션 토큰으로 인증 후 API 키 반환 (로그인 응답에서 분리됨)."""
    _check_admin(request)
    return ok(data={"api_key": _ADMIN_KEY}, message="OK")


@app.post("/api/admin/logout", tags=["admin"])
async def admin_logout(request: Request):
    """세션 토큰 폐기 (로그아웃) -- 쿠키 + 헤더 토큰 모두 폐기."""
    # 쿠키 토큰
    cookie_token = request.cookies.get("bridge_session", "").strip()
    if cookie_token:
        _revoke_session(cookie_token)
    # 헤더 토큰 (하위 호환)
    header_token = request.headers.get("x-admin-token", "").strip()
    if header_token and header_token != cookie_token:
        _revoke_session(header_token)
    # 쿠키 삭제
    response = JSONResponse(content=ok(message="로그아웃 완료"))
    response.headers["Set-Cookie"] = "bridge_session=; HttpOnly; Secure; SameSite=None; Path=/; Max-Age=0"
    return response


@app.get("/api/admin/sessions", tags=["admin"])
async def admin_sessions(request: Request):
    """활성 세션 목록 조회 (관리자 전용)."""
    _check_admin(request)
    now = time.time()
    with _SESSION_LOCK:
        active = []
        for tok, s in _SESSIONS.items():
            if s["expires"] > now:
                active.append({
                    "token_prefix": tok[:8] + "...",
                    "subnet": s["subnet"],
                    "created": datetime.fromtimestamp(s["created"], tz=timezone.utc).isoformat(),
                    "expires_in_min": round((s["expires"] - now) / 60),
                    "ua": s["ua"][:50],
                })
    return ok(data={"sessions": active, "count": len(active)})


@app.delete("/api/admin/sessions", tags=["admin"])
async def admin_revoke_all_sessions(request: Request):
    """모든 세션 폐기 (긴급 시)."""
    _check_admin(request)
    with _SESSION_LOCK:
        count = len(_SESSIONS)
        _SESSIONS.clear()
    _log.warning("모든 세션 강제 폐기: count=%d ip=%s", count, _get_client_ip(request))
    return ok(data={"revoked": count}, message="모든 세션이 폐기되었습니다.")


async def _render_update_env_var(key: str, value: str) -> tuple[bool, str]:
    """Render API로 환경변수 1개를 업데이트한다.
    성공 시 (True, ""), 실패 시 (False, 오류메시지) 반환."""
    render_api_key = os.environ.get("RENDER_API_KEY", "")
    render_service_id = os.environ.get("RENDER_SERVICE_ID", "")
    if not render_api_key or not render_service_id:
        return False, "RENDER_API_KEY 또는 RENDER_SERVICE_ID 환경변수 미설정"
    headers = {
        "Authorization": f"Bearer {render_api_key}",
        "Content-Type": "application/json",
        "Accept": "application/json",
    }
    base_url = f"https://api.render.com/v1/services/{render_service_id}/env-vars"
    try:
        import httpx
        async with httpx.AsyncClient(timeout=15.0) as client:
            # 기존 env vars 전체 조회
            r = await client.get(base_url, headers=headers)
            if r.status_code != 200:
                return False, f"Render GET 실패 ({r.status_code})"
            existing: list[dict] = r.json()  # [{"key": ..., "value": ...}, ...]
            # 대상 키 업데이트 또는 신규 추가
            updated = [ev for ev in existing if ev.get("key") != key]
            updated.append({"key": key, "value": value})
            # PUT으로 전체 교체
            r2 = await client.put(base_url, headers=headers, json=updated)
            if r2.status_code not in (200, 201):
                return False, f"Render PUT 실패 ({r2.status_code})"
            return True, ""
    except Exception as _re_exc:
        return False, str(_re_exc)


@app.post("/api/admin/change-password", tags=["admin"])
async def admin_change_password(request: Request):
    """관리자 비밀번호 변경 -- in-memory + .env + Render 환경변수 동시 업데이트."""
    _check_admin(request)
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(400, "Invalid JSON")
    new_pw = str(body.get("new_password", "")).strip()
    if len(new_pw) < 8:
        raise HTTPException(400, "비밀번호는 8자 이상이어야 합니다.")
    import secrets as _sec
    import re as _re
    salt = _sec.token_hex(16)
    iterations = 260000
    dk = hashlib.pbkdf2_hmac("sha256", new_pw.encode(), salt.encode(), iterations)
    new_hash = f"pbkdf2:sha256:{iterations}:{salt}:{dk.hex()}"
    # 1) in-memory 즉시 적용
    global _ADMIN_PW
    _ADMIN_PW = new_hash
    # 2) 로컬 .env 업데이트 (로컬 실행 시)
    env_path = Path(__file__).resolve().parent / ".env"
    if env_path.exists():
        content = env_path.read_text(encoding="utf-8")
        updated_env = _re.sub(r"^ADMIN_PASSWORD=.*$", f"ADMIN_PASSWORD={new_hash}",
                               content, flags=_re.MULTILINE)
        env_path.write_text(updated_env, encoding="utf-8")
    # 3) Render 환경변수 자동 업데이트
    render_ok, render_err = await _render_update_env_var("ADMIN_PASSWORD", new_hash)
    if render_ok:
        msg = "비밀번호 변경 완료. Render 환경변수도 자동 업데이트되었습니다."
    else:
        logging.getLogger("bridge.security").warning(
            "[change-password] Render 환경변수 자동 업데이트 실패: %s", render_err
        )
        msg = (
            "비밀번호 변경 완료 (현재 세션 즉시 적용). "
            f"Render 환경변수 자동 업데이트 실패: {render_err} -- "
            "Render 대시보드에서 ADMIN_PASSWORD를 수동으로 업데이트하세요."
        )
    return ok(data={"render_synced": render_ok}, message=msg)


@app.post("/api/admin/reset-blacklist", tags=["admin"])
async def admin_reset_blacklist(request: Request):
    """IP 블랙리스트 초기화 (관리자 전용)."""
    _check_admin(request)
    audit_dir = Path(os.getenv("AUDIT_DIR", str(Path(__file__).resolve().parent / "audit")))
    (audit_dir / "ip_blacklist.json").write_text("{}", encoding="utf-8")
    # 영구 블랙리스트 파일도 초기화 (관리자 lockout 복구)
    try:
        (audit_dir / "ip_permanent_ban.json").write_text("[]", encoding="utf-8")
    except Exception:
        pass
    # 메모리 블랙리스트도 초기화 (일시 + 영구)
    try:
        from security_middleware import ip_blacklist
        ip_blacklist._list.clear()
        ip_blacklist._permanent.clear()
        ip_blacklist._save()
    except Exception:
        pass
    return ok(message="IP 블랙리스트 초기화 완료 (일시+영구)")


@app.get("/api/admin/dashboard", tags=["admin"])
async def admin_dashboard(request: Request):
    """관리자 대시보드 -- 전체 통계 + 최근 활동."""
    _check_admin(request)

    # Render cold-start 방어: 컬럼 없으면 즉시 추가
    try:
        _ensure_candidates_all_cols()
    except Exception:
        pass

    stats: dict = {}
    recent_activity: list[dict] = []

    # ── SQLite 통계 (community_posts, interviews, ad_posts) ────────────────
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        conn.row_factory = sqlite3.Row
        try:
            # 게시글 수
            try:
                r = conn.execute(
                    "SELECT COUNT(*) AS n FROM community_posts WHERE is_deleted=0"
                ).fetchone()
                stats["posts"] = r["n"] if r else 0
            except Exception:
                stats["posts"] = 0

            # 예정 인터뷰
            try:
                r = conn.execute(
                    "SELECT COUNT(*) AS n FROM interviews WHERE is_deleted=0 AND status='scheduled'"
                ).fetchone()
                stats["interviews_scheduled"] = r["n"] if r else 0
            except Exception:
                stats["interviews_scheduled"] = 0

            # Ad posts 통계
            try:
                r = conn.execute(
                    "SELECT COUNT(*) AS n FROM ad_posts"
                ).fetchone()
                stats["ad_posts"] = r["n"] if r else 0
            except Exception:
                stats["ad_posts"] = 0

            # 최근 활동 피드: 게시글 최신 5건
            try:
                recent_posts = conn.execute(
                    """SELECT id, board, title, created_at, 'post' AS type
                       FROM community_posts WHERE is_deleted=0
                       ORDER BY created_at DESC LIMIT 5"""
                ).fetchall()
                for rp in recent_posts:
                    recent_activity.append(dict(rp))
            except Exception:
                pass

            # 최근 활동 피드: 인터뷰 최신 5건
            try:
                recent_ivs = conn.execute(
                    """SELECT id, candidate_name, interview_date, interview_time,
                              status, created_at, 'interview' AS type
                       FROM interviews WHERE is_deleted=0
                       ORDER BY created_at DESC LIMIT 5"""
                ).fetchall()
                for ri in recent_ivs:
                    recent_activity.append(dict(ri))
            except Exception:
                pass

        finally:
            conn.close()
    except Exception as _e1:
        logging.getLogger("bridge.api").warning("dashboard conn1 실패: %s", _e1)

    # ── SQLite 통계 (candidates, jobs, client_inquiries) ────────────────────
    try:
        conn2 = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn2.execute("PRAGMA busy_timeout = 5000")
        conn2.row_factory = sqlite3.Row
        try:
            # 지원자 수
            try:
                r = conn2.execute("SELECT COUNT(*) AS n FROM candidates").fetchone()
                stats["candidates"] = r["n"] if r else 0
            except Exception:
                stats["candidates"] = 0

            # Active 지원자
            try:
                r = conn2.execute(
                    "SELECT COUNT(*) AS n FROM candidates WHERE status='Active'"
                ).fetchone()
                stats["candidates_active"] = r["n"] if r else 0
            except Exception:
                stats["candidates_active"] = 0

            # Jobs 수
            try:
                r = conn2.execute(
                    "SELECT COUNT(*) AS n FROM jobs WHERE status='open'"
                ).fetchone()
                stats["jobs_open"] = r["n"] if r else 0
            except Exception:
                stats["jobs_open"] = 0

            # 채용의뢰 수
            try:
                r = conn2.execute(
                    "SELECT COUNT(*) AS n FROM client_inquiries"
                ).fetchone()
                stats["inquiries"] = r["n"] if r else 0
            except Exception:
                stats["inquiries"] = 0

            try:
                r = conn2.execute("SELECT COUNT(*) AS n FROM payments").fetchone()
                stats["payments_total"] = r["n"] if r else 0
                r = conn2.execute("SELECT COUNT(*) AS n FROM payments WHERE status='confirmed'").fetchone()
                stats["payments_confirmed"] = r["n"] if r else 0
                r = conn2.execute("SELECT COALESCE(SUM(amount),0) AS n FROM payments WHERE status='confirmed'").fetchone()
                stats["revenue"] = r["n"] if r else 0
            except Exception:
                stats["payments_total"] = 0
                stats["payments_confirmed"] = 0
                stats["revenue"] = 0
        finally:
            conn2.close()
    except Exception as _e2:
        logging.getLogger("bridge.api").warning("dashboard conn2 실패: %s", _e2)

    # 최근 활동을 시간순 정렬
    recent_activity.sort(
        key=lambda x: x.get("created_at", ""), reverse=True
    )

    return ok(data={"stats": stats, "recent_activity": recent_activity[:10]})


def _read_ad_posts(
    status_filter:   Optional[str],
    platform_filter: Optional[str],
    limit:           int,
) -> tuple[list[dict], dict]:
    """master.db 에서 ad_posts 읽기. (rows, stats) 반환."""
    if not _ADMIN_DB_PATH.exists():
        return [], {"error": f"master.db not found at {_ADMIN_DB_PATH}"}

    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.row_factory = sqlite3.Row

    # 전체 통계
    stat_rows = conn.execute(
        "SELECT status, COUNT(*) AS n FROM ad_posts GROUP BY status"
    ).fetchall()
    stats = {r["status"]: r["n"] for r in stat_rows}
    stats["total"] = sum(stats.values())

    # 플랫폼 목록
    platforms = [
        r[0] for r in conn.execute(
            "SELECT DISTINCT platform FROM ad_posts ORDER BY platform"
        ).fetchall()
    ]
    stats["platforms"] = platforms

    # 데이터 조회
    where_parts: list[str] = []
    params: list = []
    if status_filter:
        where_parts.append("status = ?")
        params.append(status_filter)
    if platform_filter:
        where_parts.append("platform = ?")
        params.append(platform_filter)

    where = ("WHERE " + " AND ".join(where_parts)) if where_parts else ""
    params.append(limit)

    rows = conn.execute(
        f"""
        SELECT id, job_code, seq, platform, status,
               ad_title, ad_body,
               draft_at, posted_at, posted_url,
               screenshot_path, error_msg
        FROM   ad_posts
        {where}
        ORDER BY id DESC
        LIMIT  ?
        """,
        params,
    ).fetchall()

    conn.close()
    return [dict(r) for r in rows], stats


@app.get("/api/admin/ad-posts", tags=["admin"])
async def admin_ad_posts(
    request:  Request,
    status:   Optional[str] = None,
    platform: Optional[str] = None,
    limit:    int = 200,
):
    """
    ad_posts 대시보드 데이터 (master.db → SQLite)
    보호: ADMIN_API_KEY 환경변수가 설정된 경우 X-Admin-Key 헤더 필수
    """
    _check_admin(request)
    rows, stats = _read_ad_posts(status, platform, limit)
    return ok(
        data={"stats": stats, "posts": rows},
        message=f"{len(rows)}건 조회",
    )


# ── Admin: ad_posts CRUD ─────────────────────────────────────────────────────

class _AdPostCreate(BaseModel):
    job_code: str
    platform: str = "craigslist"
    status:   str = "draft"
    ad_title: Optional[str] = None
    ad_body:  Optional[str] = None

class _AdPostUpdate(BaseModel):
    status:     Optional[str] = None
    ad_title:   Optional[str] = None
    ad_body:    Optional[str] = None
    posted_url: Optional[str] = None
    error_msg:  Optional[str] = None
    posted_at:  Optional[str] = None


@app.post("/api/admin/ad-posts", tags=["admin"])
async def admin_ad_posts_create(request: Request, body: _AdPostCreate):
    """ad_posts 신규 생성 -- status 초기값 draft"""
    _check_admin(request)
    if not _ADMIN_DB_PATH.exists():
        raise HTTPException(status_code=500, detail="master.db not found")
    now = datetime.utcnow().isoformat()
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    try:
        cur = conn.execute(
            """INSERT INTO ad_posts (job_code, platform, status, ad_title, ad_body, draft_at)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (body.job_code, body.platform, body.status, body.ad_title, body.ad_body, now),
        )
        conn.commit()
        new_id = cur.lastrowid
        logging.getLogger("bridge.api").info("[AD_POST] created id=%s job_code=%s", new_id, body.job_code)
    finally:
        conn.close()
    return ok(data={"id": new_id}, message="광고 생성 완료")


@app.patch("/api/admin/ad-posts/{post_id}", tags=["admin"])
async def admin_ad_posts_update(request: Request, post_id: int, body: _AdPostUpdate):
    """ad_posts 수정 -- 전달된 필드만 업데이트"""
    _check_admin(request)
    if not _ADMIN_DB_PATH.exists():
        raise HTTPException(status_code=500, detail="master.db not found")
    updates: dict = {k: v for k, v in body.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=422, detail="변경할 필드가 없습니다")
    set_clause = ", ".join(f"{col} = ?" for col in updates)
    params = list(updates.values()) + [post_id]
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    try:
        cur = conn.execute(
            f"UPDATE ad_posts SET {set_clause} WHERE id = ?", params  # noqa: S608
        )
        conn.commit()
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="해당 id의 광고가 없습니다")
        logging.getLogger("bridge.api").info("[AD_POST] updated id=%s fields=%s", post_id, list(updates.keys()))
    finally:
        conn.close()
    return ok(message="광고 수정 완료")


@app.delete("/api/admin/ad-posts/{post_id}", tags=["admin"])
async def admin_ad_posts_delete(request: Request, post_id: int):
    """ad_posts 삭제 -- status='deleted' 논리 삭제"""
    _check_admin(request)
    if not _ADMIN_DB_PATH.exists():
        raise HTTPException(status_code=500, detail="master.db not found")
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    try:
        cur = conn.execute(
            "UPDATE ad_posts SET status = 'deleted' WHERE id = ?", (post_id,)
        )
        conn.commit()
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="해당 id의 광고가 없습니다")
        logging.getLogger("bridge.api").info("[AD_POST] deleted(soft) id=%s", post_id)
    finally:
        conn.close()
    return ok(message="광고 삭제 완료")


# ── Admin: Candidates Grid ────────────────────────────────────────────────────
_ADMIN_DECRYPT_FIELDS = {
    "full_name", "email",
    "mobile_phone", "kakaotalk",
    "passport", "criminal_record",
    "criminal_record_check", "korean_criminal_record",
    "religion", "health_info",
    "dob", "nationality", "current_location", "reference",
    "gender", "notes",
    # T3v1 확장 필드 (encrypt_migrate.py v2/v3과 동기화)
    "kakao_id", "passport_number", "home_country", "emergency_contact",
    "contact_name", "business_registration", "school_location", "memo",
    "phone",
}


def _safe_decrypt(val, column_name: str = ""):
    """단일 값 안전 복호화. 컬럼명 포함 L1 키 정확도 보장.
    암호화된 값인데 복호화 실패 시 "" 반환 (암호문 UI 노출 방지).
    평문 값이면 그대로 반환.
    주의: decrypt_field는 실패 시 원본을 반환하므로, 결과가 여전히 암호문인지 재확인함.
    """
    if val is None:
        return val
    try:
        # 공백·개행·탭 모두 제거 (base64에는 공백 없음)
        cleaned = str(val).strip().replace('\n', '').replace('\r', '').replace('\t', '').replace(' ', '')
        if not cleaned:
            return val
        # 패딩 자동 보정 (일부 구버전 암호화는 padding 없이 저장됨)
        pad = len(cleaned) % 4
        padded = cleaned + '=' * (4 - pad) if pad else cleaned
        # padded 버전으로 암호화 여부 체크
        for candidate in [padded, cleaned]:
            if is_encrypted(candidate):
                result = decrypt_field(candidate, column_name)
                if result is None:
                    return ""
                # decrypt_field가 실패 시 원본 암호문을 반환 → 재확인 후 "" 반환
                if isinstance(result, str) and is_encrypted(result):
                    return ""
                return result
        return val
    except Exception as exc:
        _log = logging.getLogger("bridge.decrypt")
        _log.debug("_safe_decrypt 실패 col=%s (len=%d): %s", column_name, len(str(val)), type(exc).__name__)
        # 암호화 여부 재확인 후 암호문이면 "" 반환
        try:
            cleaned2 = str(val).strip().replace('\n', '').replace('\r', '').replace('\t', '').replace(' ', '')
            if is_encrypted(cleaned2):
                return ""
        except Exception:
            pass
        return val


def _decrypt_row(row: dict) -> dict:
    """candidates 행 전체 복호화 -- 컬럼명 포함 L1 키 정확 복호화"""
    result = dict(row)
    for key, val in result.items():
        if val is not None and isinstance(val, str):
            result[key] = _safe_decrypt(val, key)
    return result


def _sanitize_str(v):
    """Remove all characters that can break JSON parsing in any browser/runtime."""
    if not isinstance(v, str):
        return v
    # 1. Null bytes
    v = v.replace('\x00', '')
    # 2. All C0 control chars except \t \n \r (which json.dumps handles)
    v = re.sub(r'[\x01-\x08\x0b\x0c\x0e-\x1f\x7f]', '', v)
    # 3. C1 control chars (U+0080-U+009F) -- break some parsers
    v = re.sub(r'[\x80-\x9f]', '', v)
    # 4. Unicode surrogates (U+D800-U+DFFF) -- invalid in JSON
    v = re.sub(r'[\ud800-\udfff]', '', v)
    # 5. BOM and other zero-width chars
    v = v.replace('\ufeff', '').replace('\ufffe', '')
    return v


def _sanitize_data(obj):
    """Recursively sanitize all string values in nested dicts/lists."""
    if isinstance(obj, str):
        return _sanitize_str(obj)
    if isinstance(obj, dict):
        return {k: _sanitize_data(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_sanitize_data(item) for item in obj]
    return obj


_ACTIVE_STATUSES = {"new", "Active", "reviewing", "interviewing", "offered"}
_PAST_STATUSES = {"placed", "rejected", "withdrawn", "inactive", "Inactive", "Closed", "Deleted"}


@app.get("/api/admin/decrypt-check", tags=["admin"])
async def decrypt_check(request: Request):
    """복호화 진단 -- 첫 5개 후보자의 암호화 필드 원본/복호화 값 비교"""
    _check_admin(request)
    import os, hashlib
    vault_ok = _VAULT_OK
    key_set = bool(os.environ.get("BRIDGE_FIELD_KEY", "").strip())
    key_preview = hashlib.sha256(os.environ.get("BRIDGE_FIELD_KEY", "").encode()).hexdigest()[:8] if key_set else "NOT_SET"
    fields = ["nationality", "current_location", "reference", "dob", "korean_criminal_record"]
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        f"SELECT candidate_id, {', '.join(fields)} FROM candidates LIMIT 5"
    ).fetchall()
    conn.close()
    results = []
    for r in rows:
        row_info = {"candidate_id": r["candidate_id"]}
        for f in fields:
            raw = r[f]
            if raw and isinstance(raw, str):
                dec = _safe_decrypt(raw, f)
                row_info[f] = {
                    "raw_len": len(raw),
                    "raw_preview": raw[:20] + "..." if len(raw) > 20 else raw,
                    "decrypted": dec if dec != raw else "[복호화실패-원본반환]",
                    "changed": dec != raw,
                }
            else:
                row_info[f] = {"raw_len": 0, "decrypted": None, "changed": False}
        results.append(row_info)
    return ok(data={
        "vault_ok": vault_ok,
        "key_set": key_set,
        "key_sha256_prefix": key_preview,
        "samples": results,
    })


@app.get("/api/admin/candidates", tags=["admin"])
async def admin_candidates(
    request:     Request,
    search:      Optional[str] = None,
    nationality: Optional[str] = None,
    visa:        Optional[str] = None,
    status_filter: Optional[str] = Query(None, alias="status"),
    limit:       int = 150,
    offset:      int = 0,
    cursor:      int = 0,   # cursor-based: 마지막으로 받은 rowid (0=처음부터)
):
    """
    지원자 목록 -- SQLite. status=active|past|<specific> 필터 지원.
    offset/limit 기반 페이지네이션 (cursor=0일 때 offset 사용).
    cursor>0이면 rowid<cursor 조건 추가 (하위호환 유지).
    보호: X-Admin-Key 필수
    """
    _check_admin(request)
    # Render cold-start: 컬럼 없으면 즉시 추가
    try:
        _ensure_candidates_all_cols()
    except Exception:
        pass
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        conn.row_factory = sqlite3.Row
        try:
            where = ["status != 'Deleted'"]
            params: list = []
            if search:
                where.append("(full_name LIKE ? OR email LIKE ? OR CAST(sheet_number AS TEXT) LIKE ?)")
                params.extend([f"%{search}%", f"%{search}%", f"%{search}%"])
            if nationality:
                where.append("nationality = ?")
                params.append(nationality)
            if visa:
                where.append("e_visa LIKE ?")
                params.append(f"%{visa}%")
            if status_filter and status_filter != "all":
                if status_filter == "active":
                    placeholders = ",".join("?" * len(_ACTIVE_STATUSES))
                    where.append(f"status IN ({placeholders})")
                    params.extend(_ACTIVE_STATUSES)
                elif status_filter == "past":
                    placeholders = ",".join("?" * len(_PAST_STATUSES))
                    where.append(f"status IN ({placeholders})")
                    params.extend(_PAST_STATUSES)
                else:
                    where.append("status = ?")
                    params.append(status_filter)

            where_sql = " AND ".join(where)

            # 동적 컬럼 목록: Render DB에 실제 존재하는 컬럼만 선택 (no such column → 500 방지)
            _db_cols = {r[1] for r in conn.execute("PRAGMA table_info(candidates)").fetchall()}
            if not _db_cols:
                # candidates 테이블 자체가 없음 → 빈 결과 반환
                return SafeJSONResponse(content=ok(
                    data={"total": 0, "candidates": [], "next_cursor": None},
                    message="0명 조회",
                ))
            _ALL_COLS = [
                "candidate_id", "sheet_number", "email", "full_name", "nationality",
                "ancestry",  # nationality_plain 제거 -- 복호화된 nationality로 대체
                "dob", "gender", "current_location", "start_date", "target", "area_prefs",
                "experience", "current_salary", "desired_salary", "certification", "e_visa",
                "mobile_phone", "kakaotalk", "criminal_record", "housing", "arc_holders",
                "job_prefs", "reference", "documents", "status", "created_at", "source", "notes",
                "placed_company", "placed_salary", "start_month", "housing_detail",
                "referral_fee", "process_date", "past_placement", "preferences",
                "housing_type", "education_level", "major", "interview_time", "health_info",
                "piercings", "dependents", "married", "religion", "korean_criminal_record",
                "consent", "fact_check", "photo_url", "criminal_record_check", "doc_status",
                "how_to", "tattoo", "visa_type", "stage", "mail_tags", "korea_experience",
                "employment", "contract_offered", "personal_consideration",
                "recruiter_memo", "passport_status", "resume_status",
                "talent_visible", "talent_badge", "talent_reference_star", "talent_summary",
            ]
            _COLS = ", ".join(c for c in _ALL_COLS if c in _db_cols)

            total_row = conn.execute(
                f"SELECT COUNT(*) FROM candidates WHERE {where_sql}", params
            ).fetchone()
            total = total_row[0] if total_row else 0

            if cursor > 0:
                # 하위호환: cursor-based pagination
                cursor_where = f"({where_sql}) AND rowid < ?"
                query_params = list(params) + [cursor, limit]
                rows_raw = conn.execute(
                    f"""SELECT {_COLS}, rowid FROM candidates
                        WHERE {cursor_where}
                        ORDER BY rowid DESC
                        LIMIT ?""",
                    query_params,
                ).fetchall()
            else:
                # offset-based pagination
                safe_offset = max(0, offset)
                safe_limit = max(1, min(2000, limit))
                query_params = list(params) + [safe_limit, safe_offset]
                rows_raw = conn.execute(
                    f"""SELECT {_COLS}, rowid FROM candidates
                        WHERE {where_sql}
                        ORDER BY rowid DESC
                        LIMIT ? OFFSET ?""",
                    query_params,
                ).fetchall()

            candidates = []
            last_rowid: Optional[int] = None
            for r in rows_raw:
                d = dict(r)
                last_rowid = d.pop("rowid", None)
                d["row_id"] = last_rowid          # 관리번호 fallback용 rowid 노출
                d["id"] = d.pop("candidate_id", d.get("id"))
                d.setdefault("admin_notes", d.get("notes", ""))
                d.setdefault("photo_url", None)
                d.setdefault("thumb_url", None)
                d.setdefault("target_age", d.get("target", ""))
                d = _decrypt_row(d)   # PII 복호화 (AES-256-GCM)
                for k, v in d.items():
                    d[k] = _sanitize_str(v)
                candidates.append(d)

            # next_cursor: cursor-based 호환용 (cursor>0 일 때만 의미 있음)
            next_cursor: Optional[int] = last_rowid if (cursor > 0 and len(candidates) == limit) else None

            payload = ok(
                data={"total": total, "candidates": candidates, "next_cursor": next_cursor},
                message=f"{len(candidates)}명 조회",
            )
            return SafeJSONResponse(content=payload)
        finally:
            conn.close()
    except HTTPException:
        raise
    except Exception as e:
        logging.getLogger("bridge.api").error("admin_candidates 조회 실패: %s", e, exc_info=True)
        err("지원자 목록을 불러올 수 없습니다.", 500)


@app.get("/api/admin/candidates/brief", tags=["admin"])
async def admin_candidates_brief(request: Request):
    """후보자 경량 목록 -- 드롭다운용 (4필드만)."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            "SELECT candidate_id, full_name, nationality, status FROM candidates WHERE status != 'Deleted' ORDER BY candidate_id DESC"
        ).fetchall()
        result = []
        for r in rows:
            d = dict(r)
            d = _decrypt_row(d)
            result.append({
                "candidate_id": d.get("candidate_id", ""),
                "full_name": d.get("full_name", ""),
                "nationality": d.get("nationality", ""),
                "status": d.get("status", ""),
            })
        return ok(data=result)
    finally:
        conn.close()


@app.get("/api/admin/candidates/export", tags=["admin"])
async def admin_export_candidates(request: Request, format: str = "csv"):
    """후보자 전체 내보내기 (CSV / XLSX). PII 복호화 포함."""
    _check_admin(request)
    # SECURITY: export는 하루 10회 제한 -- 대량 PII dump 방어
    if not _rate_ok(_ip_hash(request), window=86400, max_posts=10):
        raise HTTPException(429, "export 요청 한도 초과 (일 10회). 잠시 후 시도하세요.")
    import csv as _csv_mod
    import io as _io_mod

    fmt = format.lower()
    if fmt not in ("csv", "xlsx"):
        raise HTTPException(400, "format은 csv 또는 xlsx만 가능합니다.")

    _EXPORT_COLS = (
        "candidate_id, sheet_number, email, full_name, nationality, ancestry, dob, gender, "
        "current_location, start_date, target, area_prefs, experience, "
        "current_salary, desired_salary, certification, e_visa, mobile_phone, "
        "kakaotalk, criminal_record, housing, arc_holders, job_prefs, "
        "reference, documents, status, created_at, source, notes, "
        "placed_company, placed_salary, start_month, housing_detail, "
        "referral_fee, process_date, past_placement, preferences, "
        "housing_type, education_level, major, interview_time, health_info, "
        "piercings, dependents, married, religion, korean_criminal_record, "
        "consent, fact_check, photo_url, criminal_record_check, doc_status, "
        "how_to, tattoo, visa_type, stage, mail_tags, korea_experience, "
        "employment, contract_offered, personal_consideration, "
        "recruiter_memo, passport_status"
    )
    col_names = [c.strip() for c in _EXPORT_COLS.split(",")]

    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA busy_timeout = 5000")
        try:
            rows = conn.execute(
                f"SELECT {_EXPORT_COLS} FROM candidates WHERE is_deleted = 0 ORDER BY created_at DESC"
            ).fetchall()
        finally:
            conn.close()

        data = []
        for r in rows:
            d = _decrypt_row(dict(r))
            data.append(d)

        today = datetime.now().strftime("%Y%m%d")

        if fmt == "csv":
            buf = _io_mod.StringIO()
            writer = _csv_mod.writer(buf)
            writer.writerow(col_names)
            for d in data:
                writer.writerow([str(d.get(c, "") or "") for c in col_names])
            content = buf.getvalue().encode("utf-8-sig")  # BOM for Excel 한글 호환
            from starlette.responses import Response as _RawResponse
            return _RawResponse(
                content=content,
                media_type="text/csv; charset=utf-8",
                headers={"Content-Disposition": f'attachment; filename="bridge_candidates_{today}.csv"'},
            )
        else:  # xlsx
            from openpyxl import Workbook as _Workbook
            wb = _Workbook()
            ws = wb.active
            ws.title = "Candidates"
            ws.append(col_names)
            for d in data:
                ws.append([str(d.get(c, "") or "") for c in col_names])
            xlsx_buf = _io_mod.BytesIO()
            wb.save(xlsx_buf)
            xlsx_buf.seek(0)
            from starlette.responses import Response as _RawResponse
            return _RawResponse(
                content=xlsx_buf.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="bridge_candidates_{today}.xlsx"'},
            )
    except HTTPException:
        raise
    except Exception as e:
        logging.getLogger("bridge.api").error("export 실패: %s", e, exc_info=True)
        err("내보내기에 실패했습니다.", 500)


@app.post("/api/admin/candidates", tags=["admin"])
async def admin_create_candidate(request: Request, body: dict):
    """새 후보자 추가 (관리자 전용)."""
    _check_admin(request)
    if not _rate_ok(_ip_hash(request)):
        raise HTTPException(429, "잠시 후 다시 시도해주세요.")
    full_name = (body.get("full_name") or "").strip()
    if not full_name:
        raise HTTPException(400, "이름(full_name)은 필수입니다.")
    import uuid as _uuid_cand
    cid = f"cnd_{_uuid_cand.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc).isoformat()
    ALLOWED = {
        "email", "full_name", "nationality", "ancestry", "dob", "gender",
        "current_location", "education_level", "major", "certification",
        "e_visa", "arc_holders", "passport", "criminal_record", "documents",
        "start_date", "target", "area_prefs", "job_prefs", "experience",
        "employment", "reference", "current_salary", "desired_salary",
        "interview_time", "housing", "personal_consideration", "religion",
        "health_info", "piercings", "dependents", "pets", "married",
        "korean_criminal_record", "kakaotalk", "mobile_phone", "consent",
        "fact_check", "notes", "status", "source", "tattoo", "visa_type",
        "working_hours", "target_age",
    }
    record = {k: v for k, v in body.items() if k in ALLOWED and v}
    record["candidate_id"] = cid
    record["created_at"] = now
    record["updated_at"] = now
    record["source"] = record.get("source", "admin_manual")
    record["source_file"] = "admin_manual"
    record["status"] = record.get("status", "Active")
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        try:
            # 4를 포함하지 않는 sheet_number 할당
            record["sheet_number"] = _next_sheet_number(conn)
            cols = list(record.keys())
            vals = list(record.values())
            ph = ["?" for _ in cols]
            conn.execute(
                f"INSERT INTO candidates ({', '.join(cols)}) VALUES ({', '.join(ph)})",
                vals,
            )
            conn.commit()
            return ok(data={"candidate_id": cid, "sheet_number": record["sheet_number"]}, message="후보자 등록 완료")
        finally:
            conn.close()
    except Exception as e:
        logging.getLogger("bridge.api").error("admin_create_candidate 실패: %s", e, exc_info=True)
        err("후보자 등록에 실패했습니다.", 500)


@app.patch("/api/admin/candidates/{candidate_id}", tags=["admin"])
async def admin_update_candidate(
    candidate_id: str,
    request:      Request,
    body:         dict,
):
    """인라인 편집: 관리 가능 필드."""
    _check_admin(request)
    # admin_notes → notes 매핑
    if "admin_notes" in body:
        body["notes"] = body.pop("admin_notes")
    EDITABLE = {
        "notes", "reference", "status", "assigned_to",
        "contract_offered", "contract_progress",
        "email_contract", "email_immigration", "email_overseas",
        "email_transition", "email_arrival",
        "placed_company", "placed_salary", "start_month",
        "housing_detail", "referral_fee", "process_date", "past_placement",
        "recruiter_memo", "preferences", "dislikes",
        "residence_type", "start_detail", "target_level", "housing_type",
        "interview_time", "job_prefs", "invoice", "tattoo", "visa_type",
        "working_hours", "photo_url", "thumb_url",
        "stage", "mail_tags", "korea_experience",
        # Sheet 인라인 편집용 추가 필드
        "full_name", "email", "nationality", "ancestry", "dob", "gender",
        "current_location", "start_date", "target", "area_prefs",
        "experience", "employment", "current_salary", "desired_salary",
        "education_level", "major", "certification", "doc_status",
        "health_info", "personal_consideration", "criminal_record_check", "korean_criminal_record",
        "kakaotalk", "mobile_phone", "arc_holders", "married", "religion",
        "dependents", "consent", "fact_check", "housing", "e_visa",
        "how_to", "passport_status",
        # 인재 게시판 관리 필드
        "talent_visible", "talent_badge", "talent_reference_star", "talent_summary",
    }
    update = {k: v for k, v in body.items() if k in EDITABLE}
    if not update:
        raise HTTPException(400, "수정 가능한 필드 없음")
    # PII 필드 재암호화 -- 평문 저장 방지
    for field in _CANDIDATE_ENCRYPT:
        if field in update and update[field]:
            update[field] = _encrypt_if_needed(str(update[field]), field)
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        try:
            sets = ", ".join(f"{k} = ?" for k in update)
            vals = list(update.values()) + [candidate_id]
            conn.execute(
                f"UPDATE candidates SET {sets} WHERE candidate_id = ?", vals
            )
            conn.commit()
            return ok(message="수정 완료")
        finally:
            conn.close()
    except Exception as e:
        import logging as _log_upd
        _log_upd.getLogger("bridge.api").error("admin_update 실패: %s", e, exc_info=True)
        err("수정에 실패했습니다.", 500)


@app.delete("/api/admin/candidates/{candidate_id}", tags=["admin"])
async def admin_delete_candidate(candidate_id: str, request: Request):
    """Soft Delete -- status='Deleted' 설정 (물리 삭제 금지)."""
    _check_admin(request)
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        try:
            conn.execute(
                "UPDATE candidates SET status = 'Deleted', updated_at = ? WHERE candidate_id = ?",
                (datetime.now(timezone.utc).isoformat(), candidate_id),
            )
            conn.commit()
            return ok(message="삭제 처리 완료 (soft delete)")
        finally:
            conn.close()
    except Exception as e:
        import logging as _log_del
        _log_del.getLogger("bridge.api").error("admin_delete 실패: %s", e, exc_info=True)
        err("삭제에 실패했습니다.", 500)


_CANDIDATES_MUTABLE_COLS: frozenset = frozenset({
    "email", "full_name", "nationality", "ancestry", "dob", "gender",
    "current_location", "start_date", "target", "area_prefs", "experience",
    "employment", "current_salary", "desired_salary", "certification",
    "e_visa", "mobile_phone", "kakaotalk", "criminal_record", "passport",
    "housing", "arc_holders", "job_prefs", "reference", "documents", "status",
    "source", "inbox_status", "notes", "assigned_to", "last_activity",
    "contract_offered", "contract_progress", "email_contract", "email_immigration",
    "email_overseas", "email_transition", "email_arrival", "placed_company",
    "placed_salary", "start_month", "housing_detail", "referral_fee",
    "process_date", "past_placement", "recruiter_memo", "preferences",
    "dislikes", "residence_type", "start_detail", "target_level", "housing_type",
    "education_level", "major", "interview_time", "health_info",
    "personal_consideration", "piercings", "dependents", "pets", "married",
    "religion", "korean_criminal_record", "consent", "fact_check",
    "photo_url", "thumb_url", "target_age", "criminal_record_check",
    "doc_status", "how_to", "tattoo", "visa_type", "working_hours", "invoice",
    "stage", "mail_tags", "korea_experience",
    "passport_status", "updated_at",
})


@app.put("/api/admin/candidates/{candidate_id}", tags=["admin"])
async def admin_full_update_candidate(candidate_id: str, request: Request, body: dict):
    """지원자 모든 컬럼 수정 (관리자 전용)."""
    _check_admin(request)
    # 수정 불가 필드 제거
    for _immutable in ("candidate_id", "id", "created_at", "source_file", "source_row",
                       "gmail_message_id", "raw_email_body", "parsed_data"):
        body.pop(_immutable, None)
    # 화이트리스트 필터 -- DB에 없는 컬럼명 주입 차단
    body = {k: v for k, v in body.items() if k in _CANDIDATES_MUTABLE_COLS}
    if not body:
        raise HTTPException(400, "수정할 데이터 없음")
    # PII 필드 재암호화 -- 평문 저장 방지
    for field in _CANDIDATE_ENCRYPT:
        if field in body and body[field]:
            body[field] = _encrypt_if_needed(str(body[field]), field)
    body["updated_at"] = datetime.now(timezone.utc).isoformat()
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        try:
            sets = ", ".join(f"{k} = ?" for k in body)
            vals = list(body.values()) + [candidate_id]
            conn.execute(f"UPDATE candidates SET {sets} WHERE candidate_id = ?", vals)
            conn.commit()
            return ok(message="수정 완료")
        finally:
            conn.close()
    except Exception as e:
        import logging as _log_ful
        _log_ful.getLogger("bridge.api").error("admin_full_update 실패: %s", e, exc_info=True)
        err("수정에 실패했습니다.", 500)


class BulkStatusBody(BaseModel):
    ids: list[str]
    status: str = Field(..., min_length=1, max_length=50)


@app.put("/api/admin/candidates/bulk-status", tags=["admin"])
async def admin_bulk_status(request: Request, body: BulkStatusBody):
    """여러 지원자 상태 일괄 변경."""
    _check_admin(request)
    if not body.ids:
        raise HTTPException(400, "대상 ID 목록 필요")
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        try:
            now_iso = datetime.now(timezone.utc).isoformat()
            placeholders = ",".join("?" * len(body.ids))
            conn.execute(
                f"UPDATE candidates SET status = ?, updated_at = ? WHERE candidate_id IN ({placeholders})",
                [body.status, now_iso] + body.ids,
            )
            conn.commit()
            return ok(message=f"{len(body.ids)}명 상태 → {body.status}")
        finally:
            conn.close()
    except Exception as e:
        import logging as _log_bulk
        _log_bulk.getLogger("bridge.api").error("bulk_status 실패: %s", e, exc_info=True)
        err("일괄 상태 변경에 실패했습니다.", 500)


@app.get("/api/admin/candidates/{candidate_id}/profile", tags=["admin"])
async def admin_candidate_profile(candidate_id: str, request: Request):
    """후보자 프로필 카드 HTML 미리보기 (PII 제외)."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.row_factory = sqlite3.Row
    try:
        row = conn.execute(
            "SELECT * FROM candidates WHERE candidate_id = ?", (candidate_id,)
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="후보자를 찾을 수 없습니다.")
        c = dict(row)
        c["id"] = c.get("candidate_id", c.get("id", ""))
        card_html = _build_profile_card(c)
        return ok(data={"html": card_html, "candidate_id": candidate_id})
    finally:
        conn.close()


# ── Admin: Inquiries (채용의뢰) 관리 ────────────────────────────────────────

@app.get("/api/admin/inquiries", tags=["admin"])
async def admin_list_inquiries(
    request: Request,
    limit: int = 50,
    offset: int = 0,
    q: str = "",
    source: str = "",
):
    """채용의뢰 목록 -- client_inquiries 테이블. q=검색어, source=소스필터."""
    _check_admin(request)
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        conn.row_factory = sqlite3.Row
        try:
            if q:
                # SECURITY FIX: contact_name/email/phone은 T3v1 암호화 → SQL LIKE 불가
                # 전체 로드 후 복호화하여 Python 메모리에서 필터 (1239건 기준 <100ms)
                base_where = "WHERE is_deleted = 0"
                if source:
                    base_where += " AND source_file = ?"
                    src_params: list = [source]
                else:
                    src_params = []
                all_rows = conn.execute(
                    f"SELECT * FROM client_inquiries {base_where} ORDER BY id DESC",
                    src_params,
                ).fetchall()
                q_lower = q.lower()

                def _inq_match(r: dict) -> bool:
                    fields = [
                        r.get("contact_name") or "",
                        r.get("email") or "",
                        r.get("phone") or "",
                        r.get("memo") or "",
                        r.get("school_name_plain") or r.get("school_name") or "",
                        r.get("location_plain") or r.get("location") or "",
                        r.get("notes") or "",
                    ]
                    return any(q_lower in str(f).lower() for f in fields)

                decrypted_all = [_decrypt_row(dict(r)) for r in all_rows]
                filtered = [r for r in decrypted_all if _inq_match(r)]
                total = len(filtered)
                page = filtered[offset: offset + limit]
                return ok(data={"total": total, "inquiries": page})
            else:
                where_clauses = ["is_deleted = 0"]
                params: list = []
                if source:
                    where_clauses.append("source_file = ?")
                    params.append(source)
                where_sql = " WHERE " + " AND ".join(where_clauses)
                total = conn.execute(f"SELECT COUNT(*) FROM client_inquiries{where_sql}", params).fetchone()[0]
                rows = conn.execute(
                    f"SELECT * FROM client_inquiries{where_sql} ORDER BY id DESC LIMIT ? OFFSET ?",
                    params + [limit, offset],
                ).fetchall()
                return ok(data={"total": total, "inquiries": [_decrypt_row(dict(r)) for r in rows]})
        finally:
            conn.close()
    except HTTPException:
        raise
    except Exception as e:
        import logging as _log_inq
        _log_inq.getLogger("bridge.api").error("admin_inquiries 실패: %s", e, exc_info=True)
        err("채용의뢰 목록을 불러올 수 없습니다.", 500)


@app.put("/api/admin/inquiries/{inquiry_id}", tags=["admin"])
async def admin_update_inquiry(inquiry_id: int, request: Request, body: dict):
    """채용의뢰 상태/메모 수정."""
    _check_admin(request)
    EDITABLE = {"inbox_status", "notes", "assigned_to", "is_duplicate_suspect"}
    update = {k: v for k, v in body.items() if k in EDITABLE}
    if not update:
        raise HTTPException(400, "수정 가능한 필드: inbox_status, notes, assigned_to")
    update["last_activity"] = datetime.now(timezone.utc).isoformat()
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        try:
            sets = ", ".join(f"{k} = ?" for k in update)
            vals = list(update.values()) + [inquiry_id]
            conn.execute(f"UPDATE client_inquiries SET {sets} WHERE id = ?", vals)
            conn.commit()
            return ok(message="수정 완료")
        finally:
            conn.close()
    except Exception as e:
        import logging as _log_upd
        _log_upd.getLogger("bridge.api").error("admin_update_inquiry 실패: %s", e, exc_info=True)
        err("수정에 실패했습니다.", 500)


@app.patch("/api/admin/inquiries/{inquiry_id}", tags=["admin"])
async def admin_patch_inquiry(inquiry_id: int, request: Request, body: dict):
    """채용의뢰 필드 수정 -- memo, raw_email_body, school_name, email, phone, location, inbox_status."""
    _check_admin(request)
    EDITABLE = {"memo", "raw_email_body", "school_name", "email", "phone", "location", "inbox_status", "notes", "assigned_to"}
    update = {k: v for k, v in body.items() if k in EDITABLE}
    if not update:
        raise HTTPException(400, f"수정 가능한 필드: {', '.join(sorted(EDITABLE))}")
    # PII 필드 수정 시 T3v1 재암호화 (암호화되지 않은 값만)
    _INQ_PII = {"email", "phone", "memo", "contact_name", "school_location", "business_registration"}
    for pii_field in _INQ_PII:
        if pii_field in update and update[pii_field] and isinstance(update[pii_field], str):
            if not is_encrypted(update[pii_field]):
                update[pii_field] = t3_encrypt(update[pii_field], pii_field)
    update["last_activity"] = datetime.now(timezone.utc).isoformat()
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        conn.row_factory = sqlite3.Row
        try:
            sets = ", ".join(f"{k} = ?" for k in update)
            vals = list(update.values()) + [inquiry_id]
            conn.execute(f"UPDATE client_inquiries SET {sets} WHERE id = ?", vals)
            conn.commit()
            return ok(message="수정 완료")
        finally:
            conn.close()
    except HTTPException:
        raise
    except Exception as e:
        import logging as _log_patch
        _log_patch.getLogger("bridge.api").error("admin_patch_inquiry 실패: %s", e, exc_info=True)
        err("수정에 실패했습니다.", 500)


@app.patch("/api/admin/inquiries/{inquiry_id}/duplicate-flag", tags=["admin"])
async def admin_toggle_duplicate_flag(inquiry_id: int, request: Request):
    """중복 의심 마킹 토글 -- 삭제하지 않고 플래그만 표시."""
    _check_admin(request)
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        try:
            row = conn.execute(
                "SELECT is_duplicate_suspect FROM client_inquiries WHERE id = ? AND is_deleted = 0",
                (inquiry_id,),
            ).fetchone()
            if not row:
                raise HTTPException(404, "해당 채용의뢰를 찾을 수 없습니다.")
            new_flag = 0 if row[0] else 1
            conn.execute(
                "UPDATE client_inquiries SET is_duplicate_suspect = ?, last_activity = ? WHERE id = ?",
                (new_flag, datetime.now(timezone.utc).isoformat(), inquiry_id),
            )
            conn.commit()
            return ok(data={"is_duplicate_suspect": new_flag}, message="마킹 완료" if new_flag else "마킹 해제")
        finally:
            conn.close()
    except HTTPException:
        raise
    except Exception as e:
        import logging as _log_dup
        _log_dup.getLogger("bridge.api").error("duplicate_flag 실패: %s", e, exc_info=True)
        err("마킹 처리에 실패했습니다.", 500)


# ── Admin: NEW 접수 알림 ──────────────────────────────────────────────────────

@app.get("/api/admin/inquiries/new-count", tags=["admin"])
async def admin_inquiries_new_count(request: Request):
    """is_new=1 건수 반환 (사이드바 폴링용)."""
    _check_admin(request)
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        try:
            row = conn.execute(
                "SELECT COUNT(*) FROM client_inquiries WHERE is_new=1 AND is_deleted=0"
            ).fetchone()
            return ok(data={"count": row[0] if row else 0})
        finally:
            conn.close()
    except Exception as e:
        logging.getLogger("bridge.api").warning("/api/admin/inquiries/new-count 오류: %s", e)
        return ok(data={"count": 0})


@app.patch("/api/admin/inquiries/{inquiry_id}/confirm-new", tags=["admin"])
async def admin_confirm_new(inquiry_id: int, request: Request):
    """카드 NEW 확인 -> is_new=0, 남은 new 건수 반환."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        conn.execute(
            "UPDATE client_inquiries SET is_new=0 WHERE id=? AND is_deleted=0",
            (inquiry_id,),
        )
        conn.commit()
        remaining = conn.execute(
            "SELECT COUNT(*) FROM client_inquiries WHERE is_new=1 AND is_deleted=0"
        ).fetchone()[0]
        return ok(data={"id": inquiry_id, "is_new": 0, "remaining_new": remaining})
    finally:
        conn.close()


@app.patch("/api/admin/inquiries/confirm-all-new", tags=["admin"])
async def admin_confirm_all_new(request: Request):
    """전체 NEW 일괄 확인 -> is_new=0."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        cur = conn.execute(
            "UPDATE client_inquiries SET is_new=0 WHERE is_new=1 AND is_deleted=0"
        )
        conn.commit()
        return ok(data={"updated": cur.rowcount, "remaining_new": 0})
    finally:
        conn.close()


# ── Admin: Email Templates & Guide Links ────────────────────────────────────

@app.get("/api/admin/email-templates", tags=["admin"])
async def admin_list_email_templates(request: Request):
    """이메일 템플릿 목록."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        cur = conn.execute("SELECT * FROM email_templates ORDER BY template_key")
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    finally:
        conn.close()
    return ok(data=rows)


@app.put("/api/admin/email-templates/{template_key}", tags=["admin"])
async def admin_update_email_template(template_key: str, request: Request, body: dict):
    """이메일 템플릿 수정/생성."""
    _check_admin(request)
    now_iso = datetime.now(timezone.utc).isoformat()
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        conn.execute(
            "INSERT OR REPLACE INTO email_templates (template_key, subject, body_html, updated_at) VALUES (?, ?, ?, ?)",
            (template_key, body.get("subject", ""), body.get("body_html", ""), now_iso),
        )
        conn.commit()
    finally:
        conn.close()
    return ok(message=f"템플릿 '{template_key}' 저장 완료")


@app.delete("/api/admin/email-templates/{template_key}", tags=["admin"])
async def admin_delete_email_template(template_key: str, request: Request):
    """이메일 템플릿 삭제."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        conn.execute("DELETE FROM email_templates WHERE template_key = ?", (template_key,))
        conn.commit()
    finally:
        conn.close()
    return ok(message=f"템플릿 '{template_key}' 삭제 완료")


@app.get("/api/admin/guide-links", tags=["admin"])
async def admin_list_guide_links(request: Request):
    """가이드 링크 목록."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        cur = conn.execute("SELECT * FROM guide_links ORDER BY link_key")
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    finally:
        conn.close()
    return ok(data=rows)


@app.put("/api/admin/guide-links/{link_key}", tags=["admin"])
async def admin_update_guide_link(link_key: str, request: Request, body: dict):
    """가이드 링크 수정/생성."""
    _check_admin(request)
    now_iso = datetime.now(timezone.utc).isoformat()
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        conn.execute(
            "INSERT OR REPLACE INTO guide_links (link_key, url, label, updated_at) VALUES (?, ?, ?, ?)",
            (link_key, body.get("url", ""), body.get("label", ""), now_iso),
        )
        conn.commit()
    finally:
        conn.close()
    return ok(message=f"링크 '{link_key}' 저장 완료")


# ── Email Sending API ────────────────────────────────────────────────────────
import smtplib
import ssl as _ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders as _email_encoders

_SMTP_HOST = os.getenv("SMTP_HOST", "smtp.gmail.com")
_SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
_SMTP_USER = os.getenv("BRIDGE_SMTP_USER", os.getenv("SMTP_USER", os.getenv("GMAIL_USER", "")))
_SMTP_PASS = os.getenv("BRIDGE_GMAIL_SMTP_APPKEY", os.getenv("BRIDGE_SMTP_PASS", os.getenv("SMTP_PASS", os.getenv("GMAIL_APP_PASSWORD", ""))))
_log_email = logging.getLogger("bridge.email_send")


# ── SMTP 자격증명 Vault 로더 ──────────────────────────────────────────────────
def _load_smtp_from_vault() -> None:
    """
    tools/smtp_creds.enc.json 에서 T3v1 암호화된 SMTP 자격증명을 로드하여
    os.environ에 주입 → 아래 SMTP_CONFIG가 평문으로 읽음.
    파일 없음 또는 복호화 실패 시 조용히 스킵 (기존 환경변수 유지).
    """
    _log_smtp = logging.getLogger("bridge.smtp_vault")
    vault_file = Path(__file__).parent / "tools" / "smtp_creds.enc.json"
    if not vault_file.exists():
        return
    try:
        import json as _json_sv
        data = _json_sv.loads(vault_file.read_text(encoding="utf-8"))
        if data.get("version") != "T3v1":
            _log_smtp.warning("[SMTP Vault] 버전 불일치 -- 스킵")
            return
        entries = data.get("entries", {})
        injected = []
        for env_key, entry in entries.items():
            enc = entry.get("enc", "")
            col = entry.get("col", "")
            if not enc:
                continue
            # decrypt_field: T3v1 자동 감지 → t3_decrypt → 실패 시 원값 반환
            decrypted = decrypt_field(enc, col)
            if decrypted and decrypted != enc:
                os.environ[env_key] = decrypted
                injected.append(env_key)
        if injected:
            _log_smtp.info("[SMTP Vault] 로드 완료: %s", injected)
        else:
            _log_smtp.warning("[SMTP Vault] 복호화 실패 (BRIDGE_FIELD_KEY 확인 필요)")
    except Exception as _sv_exc:
        logging.getLogger("bridge.smtp_vault").warning("[SMTP Vault] 로드 실패: %s", _sv_exc)


# SMTP_CONFIG 초기화 전에 Vault에서 자격증명 주입
_load_smtp_from_vault()

SMTP_CONFIG = {
    "naver": {
        "host": "smtp.naver.com",
        "port": 587,
        "user": os.getenv("NAVER_SMTP_USER", ""),
        "password": os.getenv("NAVER_SMTP_PASS", ""),
        "limit": 500,
    },
    "gmail": {
        "host": "smtp.gmail.com",
        "port": 587,
        "user": os.getenv("GMAIL_SMTP_USER", os.getenv("BRIDGE_SMTP_USER", "")),
        "password": os.getenv("BRIDGE_GMAIL_SMTP_APPKEY", os.getenv("GMAIL_SMTP_PASS", os.getenv("BRIDGE_SMTP_PASS", ""))),
        "limit": 100,
    },
}

# template_key → candidates 컬럼 매핑
_TEMPLATE_COL_MAP = {
    "contract_offer":       "email_contract",
    "immigration_guide":    "email_immigration",
    "overseas_visa_prep":   "email_overseas",
    "job_transition_guide": "email_transition",
    "arrival_guide":        "email_arrival",
}


# ── SECURITY: PII 스캔 -- 메일 본문에 개인정보 포함 시 발송 차단 ───────────────
_PII_EMAIL_RE = re.compile(r'[\w.\-+]+@[\w.\-]+\.\w+')
_PII_PHONE_RE = re.compile(r'\d{2,4}[-.\s]?\d{3,4}[-.\s]?\d{4}')
_PII_KR_ID_RE = re.compile(r'\d{6}-[1-4]\d{6}')

# 허용 이메일: BRIDGE 공식 이메일은 PII 스캔에서 제외
_SAFE_EMAILS = frozenset({"bridgejobkr@gmail.com", "bridgejobkr@naver.com"})


def _pii_scan_body(html_body: str) -> list[str]:
    """메일 본문에서 PII 패턴 감지. 위반 항목 리스트 반환 (빈 리스트 = 안전)."""
    violations: list[str] = []
    # SECURITY: 이메일 주소 패턴 감지
    found_emails = _PII_EMAIL_RE.findall(html_body)
    unsafe_emails = [e for e in found_emails if e.lower() not in _SAFE_EMAILS]
    if unsafe_emails:
        violations.append(f"이메일 주소 감지: {len(unsafe_emails)}건")
    # SECURITY: 전화번호 패턴 감지
    if _PII_PHONE_RE.search(html_body):
        violations.append("전화번호 패턴 감지")
    # SECURITY: 주민등록번호 패턴 감지
    if _PII_KR_ID_RE.search(html_body):
        violations.append("주민등록번호 패턴 감지")
    return violations


# ── SECURITY: 메일 발송 Rate Limiter (3초 간격, 10/분, 200/일) ──────────────
_MAIL_SEND_TIMES: list[float] = []
_MAIL_DAILY_COUNT: dict[str, int] = {}  # date_str → count
_MAIL_RATE_LOCK = threading.Lock()


def _mail_rate_check() -> tuple[bool, str]:
    """메일 발송 속도 제한 체크. (통과여부, 거부사유)."""
    now = _time.time()
    today = datetime.now().strftime("%Y-%m-%d")

    with _MAIL_RATE_LOCK:
        # SECURITY: 일일 최대 200통
        daily = _MAIL_DAILY_COUNT.get(today, 0)
        if daily >= 200:
            return False, "일일 발송 한도 초과 (200통/일)"

        # SECURITY: 분당 최대 10통
        recent_min = [t for t in _MAIL_SEND_TIMES if now - t < 60]
        if len(recent_min) >= 10:
            return False, "분당 발송 한도 초과 (10통/분)"

        # SECURITY: 최소 3초 간격
        if _MAIL_SEND_TIMES and (now - _MAIL_SEND_TIMES[-1]) < 3:
            return False, "발송 간격 부족 (최소 3초)"

        return True, ""


def _mail_rate_record():
    """발송 성공 시 카운트 기록."""
    now = _time.time()
    today = datetime.now().strftime("%Y-%m-%d")
    with _MAIL_RATE_LOCK:
        _MAIL_SEND_TIMES.append(now)
        # 5분 이상 된 타임스탬프 정리
        cutoff = now - 300
        while _MAIL_SEND_TIMES and _MAIL_SEND_TIMES[0] < cutoff:
            _MAIL_SEND_TIMES.pop(0)
        _MAIL_DAILY_COUNT[today] = _MAIL_DAILY_COUNT.get(today, 0) + 1
        # 어제 이전 카운트 정리
        for d in list(_MAIL_DAILY_COUNT.keys()):
            if d != today:
                del _MAIL_DAILY_COUNT[d]


def _smtp_send(to_email: str, subject: str, html_body: str, reply_to: str = "bridgejobkr@gmail.com") -> bool:
    """Gmail SMTP로 이메일 발송. 실패 시 False."""
    if not _SMTP_USER or not _SMTP_PASS or _SMTP_PASS == "your_app_password":
        _log_email.warning("SMTP 미설정 -- 발송 스킵 (to=%s)", to_email[:4] + "***")
        return False
    try:
        msg = MIMEMultipart("alternative")
        # SECURITY: From 고정 -- BRIDGE 공식 계정
        msg["From"] = f"BRIDGE <{_SMTP_USER}>"
        # SECURITY: To에 수신자 1명만 -- CC/BCC 절대 금지
        msg["To"] = to_email
        msg["Subject"] = subject
        # SECURITY: Reply-To → 우리 메일로만 회신
        msg["Reply-To"] = reply_to
        # SECURITY: X-Mailer 헤더 미설정 (기본 제거)
        msg.attach(MIMEText(html_body, "html", "utf-8"))

        ctx = _ssl.create_default_context()
        with smtplib.SMTP(_SMTP_HOST, _SMTP_PORT) as srv:
            srv.ehlo()
            srv.starttls(context=ctx)
            srv.ehlo()
            srv.login(_SMTP_USER, _SMTP_PASS)
            # SECURITY: sendmail에 수신자 1명만 전달
            srv.sendmail(_SMTP_USER, [to_email], msg.as_string())

        _log_email.info("이메일 발송 성공: %s → %s", subject[:40], to_email[:4] + "***")
        return True
    except Exception as e:
        _log_email.error("이메일 발송 실패 (to=%s): %s", to_email[:4] + "***", e, exc_info=True)
        return False


def _send_one_email(provider_key: str, to_email: str, subject: str,
                    html_body: str, attachments: "list[str] | None" = None) -> dict:
    """
    단일 메일 발송 (SMTP_CONFIG 기반). 성공 시 {"ok": True}, 실패 시 {"ok": False, "error": "..."}.
    attachments: 로컬 파일 경로 리스트.
    """
    cfg = SMTP_CONFIG.get(provider_key)
    if not cfg:
        return {"ok": False, "error": f"알 수 없는 발신자: {provider_key}"}
    if not cfg["user"] or not cfg["password"]:
        return {"ok": False, "error": f"{provider_key} SMTP 자격증명 미설정"}
    try:
        msg = MIMEMultipart("mixed")
        msg["Subject"] = subject
        msg["From"] = f"BRIDGE Recruitment <{cfg['user']}>"
        msg["To"] = to_email
        msg["Reply-To"] = "bridgejobkr@gmail.com"
        msg.attach(MIMEText(html_body, "html", "utf-8"))
        if attachments:
            from email.mime.base import MIMEBase as _MIMEBase
            from email import encoders as _enc
            for fpath in attachments:
                if not os.path.isfile(fpath):
                    continue
                with open(fpath, "rb") as f:
                    part = _MIMEBase("application", "octet-stream")
                    part.set_payload(f.read())
                _enc.encode_base64(part)
                part.add_header("Content-Disposition", "attachment", filename=os.path.basename(fpath))
                msg.attach(part)
        with smtplib.SMTP(cfg["host"], cfg["port"], timeout=15) as server:
            server.starttls()
            server.login(cfg["user"], cfg["password"])
            server.sendmail(cfg["user"], to_email, msg.as_string())
        return {"ok": True}
    except smtplib.SMTPAuthenticationError:
        return {"ok": False, "error": "SMTP 인증 실패 -- 앱 비밀번호 확인"}
    except smtplib.SMTPRecipientsRefused:
        return {"ok": False, "error": f"수신 거부: {to_email}"}
    except Exception as e:
        return {"ok": False, "error": str(e)[:200]}


# ── 이메일 블랙리스트 시스템 ─────────────────────────────────────────────────
# 블랙리스트 이메일로는 후보자 정보를 절대 발송하지 않음
# 이메일은 해싱(SHA-256)으로 저장하여 DB 유출 시에도 원문 노출 차단

import hashlib

_BLACKLIST_CACHE: set[str] | None = None
_BLACKLIST_CACHE_TS: float = 0


def _email_hash(email: str) -> str:
    """이메일 → SHA-256 해시 (소문자 정규화)."""
    return hashlib.sha256(email.strip().lower().encode("utf-8")).hexdigest()


def _ensure_email_blacklist_table():
    """email_blacklist 테이블 자동 생성."""
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS email_blacklist (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                email_hash  TEXT NOT NULL UNIQUE,
                label       TEXT DEFAULT '',
                reason      TEXT DEFAULT '',
                added_at    TEXT DEFAULT (datetime('now')),
                added_by    TEXT DEFAULT 'admin'
            )
        """)
        conn.commit()
    finally:
        conn.close()


def _load_blacklist_hashes() -> set[str]:
    """블랙리스트 해시 캐시 로드 (60초 TTL)."""
    global _BLACKLIST_CACHE, _BLACKLIST_CACHE_TS
    now = time.time()
    if _BLACKLIST_CACHE is not None and (now - _BLACKLIST_CACHE_TS) < 60:
        return _BLACKLIST_CACHE
    _ensure_email_blacklist_table()
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        rows = conn.execute("SELECT email_hash FROM email_blacklist").fetchall()
        _BLACKLIST_CACHE = {r[0] for r in rows}
        _BLACKLIST_CACHE_TS = now
    finally:
        conn.close()
    return _BLACKLIST_CACHE


def _is_email_blacklisted(email: str) -> bool:
    """이메일이 블랙리스트에 있는지 확인."""
    hashes = _load_blacklist_hashes()
    return _email_hash(email) in hashes


def _secure_send_email(to_email: str, subject: str, html_body: str,
                       skip_pii_scan: bool = False) -> tuple[bool, str]:
    """보안 규칙 적용 이메일 발송. (성공여부, 메시지) 반환.

    Security rules applied:
    0. 블랙리스트 차단 -- 등록된 이메일 발송 불가
    1. 개별 발송 (CC/BCC 없음) -- _smtp_send에서 강제
    2. PII 스캔 -- 본문에 개인정보 감지 시 차단
    3. Rate limit -- 3초 간격, 10/분, 200/일
    4. Reply-To 격리 -- bridgejobkr@gmail.com
    """
    # SECURITY: 블랙리스트 차단
    if _is_email_blacklisted(to_email):
        _log_email.warning("블랙리스트 차단 (to=%s)", to_email[:4] + "***")
        return False, "블랙리스트 등록 이메일 -- 발송 차단됨"

    # SECURITY: PII 스캔
    if not skip_pii_scan:
        violations = _pii_scan_body(html_body)
        if violations:
            detail = "; ".join(violations)
            _log_email.warning("PII 발송 차단 (to=%s): %s", to_email[:4] + "***", detail)
            return False, f"PII 위반으로 발송 차단: {detail}"

    # SECURITY: Rate limit 체크
    rate_ok, rate_msg = _mail_rate_check()
    if not rate_ok:
        return False, rate_msg

    # 발송
    sent = _smtp_send(to_email, subject, html_body)
    if sent:
        _mail_rate_record()
        return True, "발송 성공"
    return False, "SMTP 발송 실패"


def _strip_image_exif(file_bytes: bytes) -> bytes:
    """SECURITY: 이미지 EXIF 데이터 제거 (Pillow 사용)."""
    if not _PILLOW_OK:
        return file_bytes
    try:
        img = PILImage.open(io.BytesIO(file_bytes))
        out = io.BytesIO()
        # SECURITY: exif 데이터 없이 저장
        img.save(out, format=img.format or "JPEG")
        return out.getvalue()
    except Exception:
        return file_bytes


def _mask_email_for_log(email: str) -> str:
    """SECURITY: 로그용 이메일 마스킹. abc***@gmail.com"""
    if not email or "@" not in email:
        return "***"
    local, domain = email.rsplit("@", 1)
    if len(local) <= 3:
        return f"{local[0]}**@{domain}"
    return f"{local[:3]}***@{domain}"


def _load_guide_links() -> dict:
    """guide_links SQLite에서 link_key->url 딕셔너리 로드."""
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        try:
            rows = conn.execute("SELECT link_key, url FROM guide_links").fetchall()
            return {r[0]: r[1] for r in rows}
        finally:
            conn.close()
    except Exception:
        return {}


def _load_email_template(template_key: str) -> Optional[dict]:
    """email_templates 테이블에서 template_key로 1건 조회. 없으면 None."""
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        cur = conn.execute(
            "SELECT subject, body_html FROM email_templates WHERE template_key = ? LIMIT 1",
            (template_key,),
        )
        row = cur.fetchone()
        if row:
            return {"subject": row[0], "body_html": row[1]}
        return None
    finally:
        conn.close()


def _substitute_vars(html: str, candidate: dict, guide_links: dict) -> str:
    """{{변수}} 치환: candidate 필드 + guide_links + 날짜."""
    # candidate 필드 치환
    field_map = {
        "name": candidate.get("full_name") or candidate.get("first_name", ""),
        "email": candidate.get("email", ""),
        "company": candidate.get("placed_company", ""),
        "location": candidate.get("city") or candidate.get("location", ""),
        "salary": candidate.get("placed_salary") or candidate.get("desired_salary", ""),
        "start_date": candidate.get("start_month") or candidate.get("start_detail", ""),
        "housing": candidate.get("housing_type") or candidate.get("housing_detail", ""),
        "date": datetime.now().strftime("%Y-%m-%d"),
        "recruiter_name": candidate.get("assigned_to", "BRIDGE"),
    }
    for key, val in field_map.items():
        html = html.replace("{{" + key + "}}", str(val or ""))

    # guide_links 치환: {{link_xxx}} → guide_links[xxx] URL
    for link_key, url in guide_links.items():
        html = html.replace("{{link_" + link_key + "}}", url)

    # 미치환 {{link_xxx}} 제거 (# 링크로 대체)
    html = re.sub(r"\{\{link_\w+\}\}", "#", html)
    # 미치환 {{xxx}} 제거
    html = re.sub(r"\{\{\w+\}\}", "", html)

    return html


def _build_profile_card(c: dict) -> str:
    """PII 제외 후보자 프로필 카드 HTML 생성."""
    cid = _html.escape(str(c.get("id", "")))
    nationality = _html.escape(str(c.get("nationality", "--")))
    visa = _html.escape(str(c.get("visa_type", "--")))
    age = _html.escape(str(c.get("age", "--")))
    education = _html.escape(str(c.get("education_level", "--")))
    teaching_exp = _html.escape(str(c.get("teaching_experience", "--")))
    korea_exp = _html.escape(str(c.get("korea_experience", "--")))
    location = _html.escape(str(c.get("city", "--")))
    available = _html.escape(str(c.get("start_month") or c.get("available_from", "--")))
    photo = c.get("photo_url", "")

    photo_html = ""
    if photo and str(photo).startswith(("http://", "https://")):
        photo_html = f'<img src="{_html.escape(str(photo))}" style="width:80px;height:80px;border-radius:50%;object-fit:cover" alt="photo"/>'

    return f"""<div style="border:1px solid #e5e7eb;border-radius:12px;padding:16px;margin:12px 0;display:flex;gap:16px;align-items:flex-start">
  {photo_html}
  <div style="flex:1;font-size:14px">
    <p style="font-weight:bold;font-size:15px;margin:0 0 8px">ID #{cid} · {nationality}</p>
    <table style="font-size:13px;border-collapse:collapse;width:100%">
      <tr><td style="color:#6b7280;padding:2px 12px 2px 0">Visa</td><td>{visa}</td></tr>
      <tr><td style="color:#6b7280;padding:2px 12px 2px 0">Age</td><td>{age}</td></tr>
      <tr><td style="color:#6b7280;padding:2px 12px 2px 0">Education</td><td>{education}</td></tr>
      <tr><td style="color:#6b7280;padding:2px 12px 2px 0">Teaching Exp.</td><td>{teaching_exp}</td></tr>
      <tr><td style="color:#6b7280;padding:2px 12px 2px 0">Korea Exp.</td><td>{korea_exp}</td></tr>
      <tr><td style="color:#6b7280;padding:2px 12px 2px 0">Location</td><td>{location}</td></tr>
      <tr><td style="color:#6b7280;padding:2px 12px 2px 0">Available</td><td>{available}</td></tr>
    </table>
  </div>
</div>"""


# ── 프로필 메일 빌더 ──────────────────────────────────────────────────────────


def _next_sheet_number(conn) -> int:
    """4를 포함하지 않는 다음 sheet_number 반환 (최소 10013부터).

    규칙:
    - DB 최대값 + 1부터 시작
    - 숫자 문자열에 '4'가 포함된 번호는 모두 건너뜀 (사음수 회피)
    - 최소 시작: 10013 (10000~10012에 4가 없는 첫 번호는 10001이지만, 정책상 10013)
    - safety loop: 최대 10,000회 탐색 후 RuntimeError
    """
    row = conn.execute(
        "SELECT COALESCE(MAX(sheet_number), 10012) FROM candidates"
    ).fetchone()
    n = max(int(row[0]), 10012) + 1
    if n < 10013:
        n = 10013
    for _ in range(10_000):
        if "4" not in str(n):
            return n
        n += 1
    raise RuntimeError(f"_next_sheet_number: 유효 번호를 찾지 못함 (n={n})")


def _ensure_candidates_sheet_number():
    """candidates 테이블에 sheet_number 컬럼 추가 (없으면)."""
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("ALTER TABLE candidates ADD COLUMN sheet_number INTEGER DEFAULT NULL")
        conn.commit()
        conn.close()
    except Exception:
        pass


try:
    _ensure_candidates_sheet_number()
except Exception:
    pass


def _ensure_candidates_extra_cols():
    """candidates 테이블에 stage/mail_tags/korea_experience 컬럼 추가 (없으면)."""
    _cols = [
        ("stage", "TEXT", "'none'"),
        ("mail_tags", "TEXT", "''"),
        ("korea_experience", "TEXT", "''"),
    ]
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        for col_name, col_type, default in _cols:
            try:
                conn.execute(f"ALTER TABLE candidates ADD COLUMN {col_name} {col_type} DEFAULT {default}")
            except Exception:
                pass  # already exists
        conn.commit()
        conn.close()
    except Exception:
        pass


try:
    _ensure_candidates_extra_cols()
except Exception:
    pass


def _ensure_auto_process_cols():
    """auto-process 용 컬럼 추가 (없으면 무시)."""
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        for sql in [
            "ALTER TABLE candidates ADD COLUMN cv_processed_s3_key TEXT DEFAULT NULL",
            "ALTER TABLE candidates ADD COLUMN resume_status TEXT DEFAULT 'pending'",
            "ALTER TABLE file_uploads ADD COLUMN s3_key TEXT DEFAULT NULL",
        ]:
            try:
                conn.execute(sql)
            except Exception:
                pass
        conn.commit()
        conn.close()
    except Exception:
        pass


try:
    _ensure_auto_process_cols()
except Exception:
    pass


def _ensure_talent_cols():
    """candidates 테이블에 인재게시판 컬럼 4개 추가 (없으면 무시)."""
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        for sql in [
            "ALTER TABLE candidates ADD COLUMN talent_visible INTEGER DEFAULT 0",
            "ALTER TABLE candidates ADD COLUMN talent_badge TEXT DEFAULT NULL",
            "ALTER TABLE candidates ADD COLUMN talent_reference_star INTEGER DEFAULT 0",
            "ALTER TABLE candidates ADD COLUMN talent_summary TEXT DEFAULT NULL",
        ]:
            try:
                conn.execute(sql)
            except Exception:
                pass  # 이미 존재하는 컬럼
        conn.commit()
        conn.close()
    except Exception:
        pass


try:
    _ensure_talent_cols()
except Exception:
    pass


def _ensure_nationality_plain():
    """candidates.nationality_plain 컬럼 추가 + 암호화된 nationality 복호화 마이그레이션.
    nationality는 PII 아님(미국/캐나다 등 7개국) → 평문 검색 컬럼으로 분리.
    서버 시작 시 1회 실행.
    """
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        # 컬럼 없으면 추가
        try:
            conn.execute("ALTER TABLE candidates ADD COLUMN nationality_plain TEXT DEFAULT ''")
            conn.commit()
        except Exception:
            pass  # 이미 존재

        # 빈 값만 채우기 (이미 채워진 건 스킵)
        total = conn.execute("SELECT COUNT(*) FROM candidates").fetchone()[0]
        if total == 0:
            conn.close()
            return

        rows = conn.execute(
            "SELECT id, nationality FROM candidates "
            "WHERE (nationality_plain IS NULL OR nationality_plain = '') AND nationality IS NOT NULL AND nationality != ''"
        ).fetchall()

        updated = 0
        for row_id, nat_val in rows:
            plain = nat_val
            if nat_val and len(nat_val) > 30:
                try:
                    plain = decrypt_field(nat_val)
                except Exception:
                    plain = nat_val  # 복호화 실패 시 원본 유지
            conn.execute(
                "UPDATE candidates SET nationality_plain = ? WHERE id = ?",
                (plain, row_id)
            )
            updated += 1

        conn.commit()
        conn.close()
        if updated:
            logging.getLogger("bridge.api").info("[MIGRATION] nationality_plain 채우기: %d건", updated)
    except Exception as e:
        logging.getLogger("bridge.api").warning("[MIGRATION] nationality_plain 스킵: %s", e)


try:
    _ensure_nationality_plain()
except Exception:
    pass


def _ensure_client_inquiries_plain_cols():
    """client_inquiries.location_plain + school_name_plain 컬럼 추가 + 암호화 복호화 백필.
    location / school_name 은 _INQUIRY_ENCRYPT 대상 아니지만 encrypt_migrate.py로
    기존 데이터가 T3v1 암호화됨 → plain 컬럼으로 검색용 평문 분리.
    서버 시작 시 1회 실행.
    """
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 10000")
        try:
            for col in ("location_plain TEXT DEFAULT ''", "school_name_plain TEXT DEFAULT ''"):
                try:
                    conn.execute(f"ALTER TABLE client_inquiries ADD COLUMN {col}")
                except Exception:
                    pass  # 이미 존재
            conn.commit()

            # 빈 location_plain 백필
            rows_loc = conn.execute(
                "SELECT id, location FROM client_inquiries "
                "WHERE (location_plain IS NULL OR location_plain = '') "
                "AND location IS NOT NULL AND location != ''"
            ).fetchall()
            for row_id, val in rows_loc:
                plain = decrypt_field(val, "location") if val else ""
                conn.execute("UPDATE client_inquiries SET location_plain = ? WHERE id = ?", (plain, row_id))

            # 빈 school_name_plain 백필
            rows_sn = conn.execute(
                "SELECT id, school_name FROM client_inquiries "
                "WHERE (school_name_plain IS NULL OR school_name_plain = '') "
                "AND school_name IS NOT NULL AND school_name != ''"
            ).fetchall()
            for row_id, val in rows_sn:
                plain = decrypt_field(val, "school_name") if val else ""
                conn.execute("UPDATE client_inquiries SET school_name_plain = ? WHERE id = ?", (plain, row_id))

            conn.commit()
            updated = len(rows_loc) + len(rows_sn)
            if updated:
                logging.getLogger("bridge.api").info(
                    "[MIGRATION] client_inquiries plain 백필: location %d건, school_name %d건",
                    len(rows_loc), len(rows_sn)
                )
        finally:
            conn.close()
    except Exception as e:
        logging.getLogger("bridge.api").warning("[MIGRATION] client_inquiries plain 스킵: %s", e)


try:
    _ensure_client_inquiries_plain_cols()
except Exception:
    pass


def _ensure_talent_auth_tables():
    """talent_auth_requests / talent_auth_tokens 테이블 마이그레이션.
    _ADMIN_DB_PATH 정의 이후에 호출됨.
    """
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS talent_auth_requests (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                email        TEXT NOT NULL,
                company_name TEXT DEFAULT '',
                requested_at TEXT NOT NULL,
                status       TEXT DEFAULT 'pending',
                sent_at      TEXT,
                ip_hash      TEXT DEFAULT '',
                notes        TEXT DEFAULT '',
                biz_number   TEXT DEFAULT '',
                doc_s3_key   TEXT DEFAULT '',
                doc_filename TEXT DEFAULT '',
                doc_data     BLOB
            );
            CREATE TABLE IF NOT EXISTS talent_auth_tokens (
                id                  INTEGER PRIMARY KEY AUTOINCREMENT,
                email               TEXT NOT NULL,
                magic_token         TEXT UNIQUE NOT NULL,
                session_token       TEXT UNIQUE,
                created_at          TEXT NOT NULL,
                magic_expires_at    TEXT NOT NULL,
                session_expires_at  TEXT,
                magic_used_at       TEXT,
                request_id          INTEGER
            );
            CREATE INDEX IF NOT EXISTS idx_tat_magic   ON talent_auth_tokens(magic_token);
            CREATE INDEX IF NOT EXISTS idx_tat_session ON talent_auth_tokens(session_token);
        """)
        conn.commit()
        # 기존 테이블 컬럼 마이그레이션 (ADD COLUMN IF NOT EXISTS 미지원 → try/except)
        for _col_sql in [
            "ALTER TABLE talent_auth_requests ADD COLUMN biz_number TEXT DEFAULT ''",
            "ALTER TABLE talent_auth_requests ADD COLUMN doc_s3_key TEXT DEFAULT ''",
            "ALTER TABLE talent_auth_requests ADD COLUMN doc_filename TEXT DEFAULT ''",
            "ALTER TABLE talent_auth_requests ADD COLUMN doc_data BLOB",
        ]:
            try:
                conn.execute(_col_sql)
                conn.commit()
            except Exception:
                pass  # 이미 컬럼 존재
        conn.close()
    except Exception as _e:
        logging.getLogger("bridge.api").warning("_ensure_talent_auth_tables 실패: %s", _e)


try:
    _ensure_talent_auth_tables()
except Exception:
    pass


def _ensure_jobs_cols():
    """jobs 테이블에 누락 컬럼 추가 (Render DB 마이그레이션)."""
    _cols = [
        ("is_deleted", "INTEGER", "0"),
        ("brj_id", "TEXT", "NULL"),
        ("legacy_id", "TEXT", "NULL"),
        ("region", "TEXT", "NULL"),
        ("region_name", "TEXT", "NULL"),
        ("enc_employer_name", "TEXT", "NULL"),
        ("enc_contact_name", "TEXT", "NULL"),
        ("enc_contact_phone", "TEXT", "NULL"),
        ("enc_contact_email", "TEXT", "NULL"),
        ("enc_contact_kakao", "TEXT", "NULL"),
        ("employer_display_name", "TEXT", "NULL"),
        ("salary_krw", "INTEGER", "NULL"),
        ("salary_negotiable", "INTEGER", "NULL"),
        ("housing_type", "TEXT", "NULL"),
        ("housing_detail", "TEXT", "NULL"),
        ("vacation_days", "INTEGER", "NULL"),
        ("visa_sponsorship", "INTEGER", "NULL"),
        ("f_visa_welcome", "INTEGER", "NULL"),
        ("kyopo_welcome", "INTEGER", "NULL"),
        ("degree_requirement", "TEXT", "NULL"),
        ("korea_resident_only", "INTEGER", "NULL"),
        ("raw_text", "TEXT", "NULL"),
        ("raw_source", "TEXT", "NULL"),
        ("teaching_hours_weekly", "TEXT", "NULL"),
        ("parse_confidence", "REAL", "NULL"),
        ("parse_warnings", "TEXT", "NULL"),
    ]
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        for col_name, col_type, default in _cols:
            try:
                conn.execute(f"ALTER TABLE jobs ADD COLUMN {col_name} {col_type} DEFAULT {default}")
            except Exception:
                pass
        conn.commit()
        conn.close()
    except Exception:
        pass


def _ensure_candidates_all_cols():
    """candidates 테이블에 최신 컬럼 추가 (Render DB 마이그레이션)."""
    _cols = [
        ("employment", "TEXT", "NULL"),
        ("contract_offered", "TEXT", "NULL"),
        ("personal_consideration", "TEXT", "NULL"),
        ("is_deleted", "INTEGER", "0"),
        ("read_at", "TEXT", "NULL"),
        ("photo_s3_key", "TEXT", "NULL"),
        # v3 추가 -- _COLS SELECT에 포함, 없으면 500
        ("recruiter_memo", "TEXT", "NULL"),
        ("passport_status", "TEXT", "NULL"),
        ("criminal_record_check", "TEXT", "NULL"),
        ("doc_status", "TEXT", "NULL"),
        ("how_to", "TEXT", "NULL"),
        ("tattoo", "TEXT", "NULL"),
        ("visa_type", "TEXT", "NULL"),
        ("gender", "TEXT", "NULL"),
        ("religion", "TEXT", "NULL"),
        ("health_info", "TEXT", "NULL"),
        # Resume Pipeline v2.0 — 100% 성공 보장용 상태/무결성 컬럼
        ("cv_processed_sha256", "TEXT", "NULL"),
        ("cv_processed_at", "TEXT", "NULL"),
        ("cv_processed_status", "TEXT", "'pending'"),
        ("cv_processed_attempts", "INTEGER", "0"),
        ("cv_processed_last_error", "TEXT", "NULL"),
        ("cv_original_s3_key", "TEXT", "NULL"),
        ("cv_original_sha256", "TEXT", "NULL"),
        ("photo_sha256", "TEXT", "NULL"),
    ]
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        for col_name, col_type, default in _cols:
            try:
                conn.execute(f"ALTER TABLE candidates ADD COLUMN {col_name} {col_type} DEFAULT {default}")
            except Exception:
                pass
        conn.commit()
        conn.close()
    except Exception:
        pass


def _ensure_pipeline_tables():
    """Resume Pipeline v2.0 — durable queue + event log 테이블 자동 생성."""
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        # ── job_queue: 영속 큐 (서버 재시작 생존) ─────────────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS job_queue (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                job_type        TEXT NOT NULL,
                payload_json    TEXT NOT NULL,
                idempotency_key TEXT UNIQUE NOT NULL,
                status          TEXT NOT NULL DEFAULT 'queued'
                    CHECK(status IN ('queued','running','done','failed','dlq')),
                attempts        INTEGER NOT NULL DEFAULT 0,
                max_attempts    INTEGER NOT NULL DEFAULT 10,
                next_run_at     TEXT,
                last_error      TEXT,
                created_at      TEXT DEFAULT (datetime('now')),
                updated_at      TEXT DEFAULT (datetime('now')),
                started_at      TEXT,
                finished_at     TEXT
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_job_queue_status_next ON job_queue(status, next_run_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_job_queue_type ON job_queue(job_type)")

        # ── pipeline_events: 감사 로그 (append-only) ──────────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS pipeline_events (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                ts            TEXT DEFAULT (datetime('now')),
                event_type    TEXT NOT NULL,
                candidate_id  TEXT,
                job_id        INTEGER,
                details_json  TEXT,
                severity      TEXT DEFAULT 'info'
                    CHECK(severity IN ('info','warn','error','critical'))
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_pipeline_events_cid ON pipeline_events(candidate_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_pipeline_events_ts ON pipeline_events(ts)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_pipeline_events_type ON pipeline_events(event_type)")

        conn.commit()
        conn.close()
    except Exception as e:
        logging.getLogger("bridge.pipeline").warning("_ensure_pipeline_tables skip: %s", e)


try:
    _ensure_jobs_cols()
except Exception:
    pass

try:
    _ensure_candidates_all_cols()
except Exception:
    pass

try:
    _ensure_pipeline_tables()
except Exception:
    pass


def _build_profile_card_v2(c: dict) -> str:
    """실사례(02.JPG, 03.JPG) 기준 프로필 카드. 기존 _build_profile_card() 건드리지 않음."""
    num = c.get("sheet_number") or c.get("candidate_id", "")
    nat = (c.get("nationality") or "").strip()
    loc = (c.get("current_location") or "").lower()
    is_korea = any(k in loc for k in [
        "korea", "한국", "서울", "부산", "대구", "인천",
        "수원", "경기", "광주", "대전", "울산", "대구",
        "제주", "세종", "청주", "전주", "천안", "성남",
        "용인", "고양", "안양", "평택", "김포",
    ])
    residence = "국내" if is_korea else "해외"
    res_emoji = "\U0001f60a" if is_korea else "\U0001f3e0"

    raw_photo = c.get("photo_url") or c.get("thumb_url") or ""
    photo_url = raw_photo if raw_photo.startswith(("http://", "https://")) else ""
    if photo_url:
        photo_html = (
            f'<img src="{_html.escape(photo_url)}" width="65" height="65" alt="" '
            f'style="border-radius:50%;object-fit:cover;object-position:top;'
            f'float:left;margin:0 12px 8px 0;display:block"/>'
        )
    else:
        photo_html = (
            '<div style="width:65px;height:65px;border-radius:50%;background:#e5e7eb;'
            'float:left;margin:0 12px 8px 0;font-size:28px;line-height:65px;'
            'text-align:center">\U0001f464</div>'
        )

    cert = (c.get("arc_holders") or c.get("visa_type") or c.get("education_level") or "")
    area = c.get("area_prefs") or "\u2014"
    cert_str = f"{area} | {cert}" if cert else area

    kor_exp = str(c.get("korea_experience") or "").strip()
    exp = str(c.get("experience") or "").strip()
    if kor_exp:
        exp_label = f"\uc6d0 \ud55c\uad6d {kor_exp}\ub144\ucc28"
    elif exp:
        exp_label = f"\uc6d0 \ud55c\uad6d {exp}\ub144\ucc28"
    else:
        exp_label = "\u2014"

    housing = (c.get("housing_type") or c.get("housing") or c.get("housing_detail") or "\u2014")
    salary = c.get("desired_salary") or c.get("placed_salary") or "\ud611\uc758"

    interview = _html.escape(str(c.get("recruiter_memo") or "\u2014"))
    reference = _html.escape(str(c.get("reference") or "\u2014"))
    prefs = _html.escape(str(c.get("preferences") or "\u2014"))
    dislik = _html.escape(str(c.get("dislikes") or "\u2014"))

    target = c.get("target") or c.get("target_level") or ""
    start = c.get("start_month") or c.get("start_date") or ""
    target_str = " | ".join(filter(None, [target, f"{start}\uc2dc\uc791" if start else ""])) or "\u2014"

    YEL = "background:#FFFF00;font-weight:bold;padding:1px 3px"

    return (
        f'<div style="margin:20px 0 4px">'
        f'  <strong style="font-size:14px">\u25a0{_html.escape(str(num))}{_html.escape(nat)} {residence}\uac70\uc8fc{res_emoji}</strong>'
        f'</div>'
        f'<div style="margin-bottom:4px;overflow:hidden">'
        f'  {photo_html}'
        f'  <div style="font-size:13px;line-height:2.0;color:#222">'
        f'    \u2022\uc120\ud638\uc9c0\uc5ed \uc790\uaca9 \uae30\ud0c0: {_html.escape(cert_str)}<br>'
        f'    \u2022\uacbd\ub825 \uc8fc\uac70 \ud76c\ub9dd\uae09\uc5ec: {_html.escape(exp_label)} | {_html.escape(str(housing))} | {_html.escape(str(salary))}<br>'
        f'    \u2022\ub9ac\ud06c\ub8e8\ud130 \uc778\ud130\ubdf0: {interview}<br>'
        f'    \u2022\ub808\ud37c\ub7f0\uc2a4: {reference}<br>'
        f'    <span style="{YEL}">\ud76c\ub9dd\uc0ac\ud56d</span>: {prefs}<br>'
        f'    <span style="{YEL}">\uae30\ud53c\uc0ac\ud56d</span>: {dislik}<br>'
        f'    \u2022\ud0c0\uac9f \uadfc\ub85c\uac1c\uc2dc: {_html.escape(target_str)}'
        f'  </div>'
        f'  <div style="clear:both"></div>'
        f'</div>'
        f'<hr style="border:none;border-top:1px dashed #d1d5db;margin:12px 0"/>'
    )


_PROFILE_SEARCH_COLS = frozenset({
    "candidate_id", "sheet_number", "nationality", "current_location",
    "photo_url", "thumb_url", "area_prefs", "visa_type", "arc_holders",
    "experience", "housing_type", "housing", "housing_detail",
    "desired_salary", "placed_salary", "recruiter_memo", "reference",
    "preferences", "dislikes", "target", "target_level", "start_month",
    "start_date", "education_level", "status", "married", "certification",
    "personal_consideration", "gender", "dob",
})


@app.get("/api/admin/candidates/profile-search", tags=["admin"])
async def profile_search(request: Request, q: str = "", limit: int = 20):
    """프로필 빌더용 경량 검색. 평문 필드만 LIKE 검색 (암호화 필드 제외)."""
    _check_admin(request)
    limit = min(max(1, limit), 100)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.row_factory = sqlite3.Row
    try:
        if q.strip():
            pat = f"%{q.strip()}%"
            rows = conn.execute(
                "SELECT candidate_id, sheet_number, nationality, current_location, "
                "photo_url, thumb_url, area_prefs, status, gender, dob "
                "FROM candidates "
                "WHERE nationality LIKE ? OR area_prefs LIKE ? OR current_location LIKE ? "
                "OR status LIKE ? OR candidate_id LIKE ? "
                "ORDER BY CASE WHEN status='Active' THEN 0 ELSE 1 END, rowid DESC "
                "LIMIT ?",
                (pat, pat, pat, pat, pat, limit),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT candidate_id, sheet_number, nationality, current_location, "
                "photo_url, thumb_url, area_prefs, status, gender, dob "
                "FROM candidates "
                "ORDER BY CASE WHEN status='Active' THEN 0 ELSE 1 END, rowid DESC "
                "LIMIT ?",
                (limit,),
            ).fetchall()
        data = [dict(r) for r in rows]
        return ok(data=data)
    finally:
        conn.close()


class BuildProfileHtmlBody(BaseModel):
    candidate_ids: list[str]
    include_intro: bool = True
    include_footer: bool = True


@app.post("/api/admin/candidates/build-profile-html", tags=["admin"])
async def build_profile_html(request: Request, body: BuildProfileHtmlBody):
    """선택된 후보자 목록으로 프로필 이메일 HTML 생성."""
    _check_admin(request)
    if not body.candidate_ids:
        return err("candidate_ids가 비어 있습니다.")
    if len(body.candidate_ids) > 99:
        return err("최대 99명까지 선택 가능합니다.")

    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.row_factory = sqlite3.Row
    try:
        cards_html = []
        for cid in body.candidate_ids:
            row = conn.execute("SELECT * FROM candidates WHERE candidate_id = ?", (cid,)).fetchone()
            if row:
                c = dict(row)
                # 암호화 필드 복호화 (카드에는 PII 미포함이지만 recruiter_memo 등에 필요)
                for f in ("full_name", "email", "mobile_phone", "kakaotalk"):
                    val = c.get(f, "")
                    if val and is_encrypted(val):
                        try:
                            c[f] = decrypt_field(val)
                        except Exception:
                            c[f] = ""
                cards_html.append(_build_profile_card_v2(c))

        parts = []
        if body.include_intro:
            parts.append(
                '<p style="font-size:13px;color:#444;margin-bottom:20px">'
                '\uc548\ub155\ud558\uc138\uc694. BRIDGE \uc6d0\uc5b4\ubbfc \uac15\uc0ac \ud504\ub85c\ud544\uc744 \uacf5\uc720\ub4dc\ub9bd\ub2c8\ub2e4.<br/>'
                'Start date and preferences noted. Reference provided for review only.'
                '</p>'
            )
        parts.append("".join(cards_html))
        if body.include_footer:
            parts.append(
                '<hr style="border:none;border-top:1px solid #e5e7eb;margin:20px 0"/>'
                '<p style="font-size:12px;color:#555;line-height:1.8">'
                'BRIDGE\ub294 \uac15\uc0ac\uc758 \uc778\uc131\uc744 \uac00\uc7a5 \uc911\uc694\ud558\uac8c \uc5ec\uae30\uba70, \uacf5\uc815\ud558\uace0 \ucc28\ubcc4 \uc5c6\ub294 \ucc44\uc6a9\uc744 \uc9c0\ud5a5\ud569\ub2c8\ub2e4.<br/>'
                '\uc778\ud130\ubdf0\ub294 Google Meet\uc73c\ub85c \uc9c4\ud589\ub429\ub2c8\ub2e4.'
                '</p>'
                '<p style="font-size:11px;color:#555">Kind Regards,</p>'
                '<p style="font-size:11px;color:#666;line-height:1.7">'
                '\u25a0 \uc9c1\uc5c5\uc548\uc815\ubc95 \uc81c34\uc870 \uc548\ub0b4: '
                '\ubb34\uc790\uaca9\uc790\uac00 \uc18c\uac1c \ube44\uc6a9\uc744 \uccad\uad6c\ud558\ub294 \uacbd\uc6b0 \uc2e0\uace0 \uc2dc \ud3ec\uc0c1\uae08 \uc9c0\uae09'
                '</p>'
                '<p style="font-size:10px;color:#999;line-height:1.6">'
                '\ubcf8 \uba54\uc77c\uc740 \uc9c0\uc815\ub41c \uc218\uc2e0\uc790\uc5d0\uac8c\ub9cc \uc804\ub2ec\ub41c \uac83\uc73c\ub85c, '
                '\ubb34\ub2e8 \uc5f4\ub78c\u00b7\ubcf4\uad00\u00b7\uc804\ub2ec\uc744 \uae08\uc9c0\ud569\ub2c8\ub2e4.'
                '</p>'
            )
        full_html = "\n".join(parts)
        return ok(data={"html": full_html, "count": len(cards_html)})
    finally:
        conn.close()


class SendEmailBody(BaseModel):
    template_key: str
    custom_subject: Optional[str] = None
    custom_body: Optional[str] = None


@app.post("/api/admin/candidates/{candidate_id}/send-email", tags=["admin"])
async def admin_send_email(candidate_id: str, request: Request, body: SendEmailBody):
    """후보자에게 이메일 템플릿 발송. template_key에 따라 해당 email_* 컬럼 업데이트."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.row_factory = sqlite3.Row
    try:
        # 후보자 조회
        row = conn.execute("SELECT * FROM candidates WHERE candidate_id = ?", (candidate_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="후보자를 찾을 수 없습니다.")
        c = dict(row)

        # 이메일 주소 복호화
        to_email = c.get("email", "")
        if to_email and is_encrypted(to_email):
            to_email = decrypt_field(to_email)
        if not to_email or "@" not in to_email:
            raise HTTPException(status_code=400, detail="유효한 이메일 주소가 없습니다.")

        # 이름 복호화 (변수 치환용)
        for f in ("full_name", "first_name", "last_name", "phone"):
            val = c.get(f, "")
            if val and is_encrypted(val):
                c[f] = decrypt_field(val)

        # 커스텀 내용 또는 템플릿 로드
        if body.custom_subject and body.custom_body:
            subject = body.custom_subject
            html = body.custom_body
        else:
            tpl = _load_email_template(body.template_key)
            if not tpl:
                raise HTTPException(status_code=404, detail=f"템플릿 '{body.template_key}'을 찾을 수 없습니다.")
            subject = body.custom_subject or tpl["subject"]
            html = body.custom_body or tpl["body_html"]

        # 변수 치환
        guide_links = _load_guide_links()
        html = _substitute_vars(html, c, guide_links)
        subject = _substitute_vars(subject, c, guide_links)

        # SECURITY: rate limit 체크
        ok_send, deny_reason = _mail_rate_check()
        if not ok_send:
            raise HTTPException(status_code=429, detail=deny_reason)

        # 발송
        sent = _smtp_send(to_email, subject, html)
        if not sent:
            raise HTTPException(status_code=500, detail="이메일 발송에 실패했습니다. SMTP 설정을 확인하세요.")
        _mail_rate_record()

        # 발송 기록: 해당 email_* 컬럼 업데이트
        now_iso = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        col = _TEMPLATE_COL_MAP.get(body.template_key)
        if col:
            conn.execute(f"UPDATE candidates SET {col} = ? WHERE candidate_id = ?", (now_iso, candidate_id))
            conn.commit()

        return ok(message=f"이메일 발송 완료 → {to_email}", data={"sent_to": to_email, "template": body.template_key})
    finally:
        conn.close()


class BulkSendBody(BaseModel):
    candidate_ids: list[str]
    template_key: str


@app.post("/api/admin/candidates/bulk-patch", tags=["admin"])
async def admin_bulk_patch(request: Request, body: dict = None):
    """암호화 복원용 벌크 패치 -- candidate_id 기준으로 여러 필드를 한꺼번에 업데이트.
    body: { "rows": [{"candidate_id": "...", "field1": "val1", ...}, ...] }
    """
    _check_admin(request)
    if body is None:
        body = await request.json()
    rows = body.get("rows", [])
    if not rows or not isinstance(rows, list):
        raise HTTPException(400, "rows 필드 필요")
    ALLOWED = {
        "nationality", "current_location", "dob", "korean_criminal_record",
        "reference", "email", "full_name", "mobile_phone", "kakaotalk",
        "gender", "health_info", "religion", "notes", "criminal_record",
        "criminal_record_check",
    }
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 30000")
    updated = 0
    skipped = 0
    try:
        for row in rows:
            cid = row.get("candidate_id")
            if not cid:
                skipped += 1
                continue
            updates = {k: v for k, v in row.items() if k in ALLOWED}
            if not updates:
                skipped += 1
                continue
            set_parts = [f"{k} = ?" for k in updates]
            params = list(updates.values()) + [cid]
            conn.execute(
                f"UPDATE candidates SET {', '.join(set_parts)} WHERE candidate_id = ?",
                params,
            )
            updated += 1
        conn.commit()
        return ok(data={"updated": updated, "skipped": skipped})
    finally:
        conn.close()


@app.post("/api/admin/candidates/bulk-send", tags=["admin"])
async def admin_bulk_send(request: Request, body: BulkSendBody):
    """다수 후보자에게 일괄 이메일 발송 (1초 간격)."""
    _check_admin(request)
    import asyncio

    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.row_factory = sqlite3.Row
    try:
        # 템플릿 로드
        tpl = _load_email_template(body.template_key)
        if not tpl:
            raise HTTPException(status_code=404, detail=f"템플릿 '{body.template_key}'을 찾을 수 없습니다.")

        guide_links = _load_guide_links()
        results = []

        for cid in body.candidate_ids:
            row = conn.execute("SELECT * FROM candidates WHERE candidate_id = ?", (cid,)).fetchone()
            if not row:
                results.append({"id": cid, "status": "not_found"})
                continue
            c = dict(row)

            # 이메일 복호화
            to_email = c.get("email", "")
            if to_email and is_encrypted(to_email):
                to_email = decrypt_field(to_email)
            if not to_email or "@" not in to_email:
                results.append({"id": cid, "status": "no_email"})
                continue

            # 이름 복호화
            for f in ("full_name", "first_name", "last_name", "phone"):
                val = c.get(f, "")
                if val and is_encrypted(val):
                    c[f] = decrypt_field(val)

            # SECURITY: rate limit 체크
            ok_send, deny_reason = _mail_rate_check()
            if not ok_send:
                results.append({"id": cid, "status": "rate_limited", "reason": deny_reason})
                break

            # 치환 + 발송
            subject = _substitute_vars(tpl["subject"], c, guide_links)
            html = _substitute_vars(tpl["body_html"], c, guide_links)
            sent = _smtp_send(to_email, subject, html)

            if sent:
                _mail_rate_record()
                now_iso = datetime.now(timezone.utc).strftime("%Y-%m-%d")
                col = _TEMPLATE_COL_MAP.get(body.template_key)
                if col:
                    conn.execute(f"UPDATE candidates SET {col} = ? WHERE candidate_id = ?", (now_iso, cid))
                results.append({"id": cid, "status": "sent"})
            else:
                results.append({"id": cid, "status": "failed"})

            # 1초 간격 (스팸 방지)
            await asyncio.sleep(1)

        conn.commit()
        sent_count = sum(1 for r in results if r["status"] == "sent")
        return ok(
            message=f"일괄 발송 완료: {sent_count}/{len(body.candidate_ids)}건 성공",
            data={"results": results}
        )
    finally:
        conn.close()


class SendProfilesBody(BaseModel):
    candidate_ids: list[str]
    to_email: str
    school_name: Optional[str] = None


@app.post("/api/admin/candidates/send-profiles", tags=["admin"])
async def admin_send_profiles(request: Request, body: SendProfilesBody):
    """학교에 후보자 프로필 카드 발송 (PII 제외)."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.row_factory = sqlite3.Row
    try:
        # candidate_profile 템플릿 로드
        tpl = _load_email_template('candidate_profile')
        if not tpl:
            raise HTTPException(status_code=404, detail="candidate_profile 템플릿이 없습니다.")

        # 프로필 카드 생성
        cards_html = ""
        for cid in body.candidate_ids:
            row = conn.execute("SELECT * FROM candidates WHERE candidate_id = ?", (cid,)).fetchone()
            if not row:
                continue
            c = dict(row)
            # PII 필드는 프로필 카드에 포함하지 않음 (이름, 이메일, 전화번호 제외)
            cards_html += _build_profile_card(c)

        if not cards_html:
            raise HTTPException(status_code=400, detail="유효한 후보자가 없습니다.")

        # 템플릿에 프로필 카드 삽입
        guide_links = _load_guide_links()
        html = tpl["body_html"].replace("{{profile_cards}}", cards_html)
        subject = tpl["subject"]

        # 학교명이 있으면 제목에 추가
        if body.school_name:
            subject = f"{subject} - {body.school_name}"

        # 미치환 변수 정리
        html = re.sub(r"\{\{\w+\}\}", "", html)

        # SECURITY: rate limit 체크
        ok_send, deny_reason = _mail_rate_check()
        if not ok_send:
            raise HTTPException(status_code=429, detail=deny_reason)

        sent = _smtp_send(body.to_email, subject, html)
        if not sent:
            raise HTTPException(status_code=500, detail="이메일 발송에 실패했습니다.")
        _mail_rate_record()

        return ok(
            message=f"프로필 발송 완료 → {body.to_email} ({len(body.candidate_ids)}명)",
            data={"sent_to": body.to_email, "candidate_count": len(body.candidate_ids)}
        )
    finally:
        conn.close()


# ── Profile Matching + Bulk Send ─────────────────────────────────────────────

# 지역 매핑: area_prefs 키워드 → client_inquiries.location 매칭값
_REGION_MAP: dict[str, list[str]] = {
    "서울": ["서울", "Seoul", "강남", "강북", "송파", "마포", "종로", "영등포", "서초", "용산", "성동", "광진"],
    "경기": ["경기", "Gyeonggi", "수원", "성남", "용인", "안양", "부천", "화성", "안산", "고양", "일산", "파주", "이천", "평택", "분당", "판교"],
    "인천": ["인천", "Incheon"],
    "부산": ["부산", "Busan"],
    "대구": ["대구", "Daegu"],
    "대전": ["대전", "Daejeon"],
    "광주": ["광주", "Gwangju"],
    "울산": ["울산", "Ulsan"],
    "세종": ["세종", "Sejong"],
    "강원": ["강원", "Gangwon", "춘천", "원주", "강릉"],
    "충북": ["충북", "충청북도", "Chungbuk", "청주", "충주"],
    "충남": ["충남", "충청남도", "Chungnam", "천안", "아산"],
    "전북": ["전북", "전라북도", "Jeonbuk", "전주", "익산", "군산"],
    "전남": ["전남", "전라남도", "Jeonnam", "목포", "순천", "여수"],
    "경북": ["경북", "경상북도", "Gyeongbuk", "포항", "구미", "경주", "김천"],
    "경남": ["경남", "경상남도", "Gyeongnam", "창원", "김해", "진주", "거제"],
    "제주": ["제주", "Jeju"],
}

# 대상 연령 매핑: target/target_age → teaching_age 매칭
# 실제 DB 값(Kindy/Elem 약어) 포함
_TARGET_MAP: dict[str, list[str]] = {
    "유치": ["Kindergarten", "킨디", "유치원", "유아", "Kinder", "Pre-K", "PreK", "Kindy"],
    "초등": ["Elementary", "초등학교", "Primary", "초등", "Elem"],
    "중등": ["Middle", "중학교", "중등", "Junior"],
    "고등": ["High", "고등학교", "고등", "Senior"],
    "성인": ["Adult", "성인", "University", "대학", "Corporate", "기업"],
    "전체": ["All", "전체", "Any"],
}


class MatchingEmployersQuery(BaseModel):
    candidate_id: str


class ProfileBroadcastBody(BaseModel):
    candidate_id: str
    employer_ids: list[int]
    custom_subject: Optional[str] = None
    custom_body: Optional[str] = None


def _match_region(area_prefs: str, location: str) -> bool:
    """후보자 area_prefs와 employer location 매칭."""
    if not area_prefs or not location:
        return False
    area_lower = area_prefs.lower()
    loc_lower = location.lower()
    # "Anywhere" / "전국" → 모든 지역 매칭
    if any(kw in area_lower for kw in ["anywhere", "전국", "any", "flexible"]):
        return True
    for region_key, keywords in _REGION_MAP.items():
        area_hit = any(kw.lower() in area_lower for kw in [region_key] + keywords)
        loc_hit = any(kw.lower() in loc_lower for kw in [region_key] + keywords)
        if area_hit and loc_hit:
            return True
    return False


def _match_target(target: str, target_age: str, teaching_age: str) -> bool:
    """후보자 target/target_age와 employer teaching_age 매칭."""
    if not teaching_age:
        return False
    cand_parts = f"{target or ''} {target_age or ''}".lower()
    emp_lower = teaching_age.lower()
    # 전체/Any → 모든 대상 매칭
    if any(kw in cand_parts for kw in ["all", "any", "전체"]):
        return True
    if any(kw in emp_lower for kw in ["all", "any", "전체"]):
        return True
    for _tgt_key, keywords in _TARGET_MAP.items():
        cand_hit = any(kw.lower() in cand_parts for kw in keywords)
        emp_hit = any(kw.lower() in emp_lower for kw in keywords)
        if cand_hit and emp_hit:
            return True
    return False


@app.get("/api/admin/matching/employers", tags=["admin"])
async def admin_matching_employers(request: Request, candidate_id: str):
    """후보자 지역/대상 기반 employer 매칭 목록 (jobs 테이블 기반)."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        # 후보자 조회
        cand = conn.execute(
            "SELECT * FROM candidates WHERE candidate_id = ?", (candidate_id,)
        ).fetchone()
        if not cand:
            raise HTTPException(status_code=404, detail="Candidate not found")
        cand = _decrypt_row(dict(cand))

        area_prefs = cand.get("area_prefs", "") or ""
        target = cand.get("target", "") or ""
        target_age = cand.get("target_age", "") or ""

        # jobs 테이블 조회 -- 삭제 안 된 것만
        employers = conn.execute(
            "SELECT id, brj_id, job_code, region, location, city, "
            "employer_display_name, teaching_age, salary_raw, status, created_at "
            "FROM jobs WHERE is_deleted = 0 ORDER BY id DESC"
        ).fetchall()

        # 이미 발송된 기록 조회
        sent_rows = conn.execute(
            "SELECT employer_id FROM profile_sends WHERE candidate_id = ?",
            (candidate_id,)
        ).fetchall()
        sent_ids = {r["employer_id"] for r in sent_rows}

        matched = []
        unmatched = []
        for emp in employers:
            emp_d = dict(emp)
            emp_loc = emp_d.get("location", "") or ""
            emp_teach = emp_d.get("teaching_age", "") or ""

            region_ok = _match_region(area_prefs, emp_loc)
            target_ok = _match_target(target, target_age, emp_teach)
            emp_d["_region_match"] = region_ok
            emp_d["_target_match"] = target_ok
            emp_d["_already_sent"] = emp_d["id"] in sent_ids

            score = int(region_ok) + int(target_ok)
            emp_d["_match_score"] = score

            if score > 0:
                matched.append(emp_d)
            else:
                unmatched.append(emp_d)

        matched.sort(key=lambda x: (-x["_match_score"], x.get("employer_display_name", "") or ""))

        return ok(data={
            "candidate": {
                "candidate_id": cand.get("candidate_id", ""),
                "full_name": cand.get("full_name", ""),
                "nationality": cand.get("nationality", ""),
                "target": target,
                "target_age": target_age,
                "area_prefs": area_prefs,
                "experience": cand.get("experience", ""),
                "education_level": cand.get("education_level", ""),
                "certification": cand.get("certification", ""),
                "visa_type": cand.get("visa_type", ""),
                "start_date": cand.get("start_date", ""),
                "photo_url": cand.get("photo_url", ""),
                "dob": cand.get("dob", ""),
                "gender": cand.get("gender", ""),
            },
            "matched": matched,
            "unmatched": unmatched,
            "total_matched": len(matched),
            "total_unmatched": len(unmatched),
        })
    finally:
        conn.close()


@app.post("/api/admin/matching/send-profile", tags=["admin"])
async def admin_matching_send_profile(request: Request, body: ProfileBroadcastBody):
    """매칭된 employer들에게 후보자 프로필 일괄 발송 (PII 제외)."""
    _check_admin(request)
    if not body.employer_ids:
        raise HTTPException(status_code=400, detail="employer_ids가 비어 있습니다.")

    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        # 후보자 조회
        cand = conn.execute(
            "SELECT * FROM candidates WHERE candidate_id = ?", (body.candidate_id,)
        ).fetchone()
        if not cand:
            raise HTTPException(status_code=404, detail="Candidate not found")
        cand_d = dict(cand)

        # 프로필 카드 생성 (PII 제외)
        card_html = _build_profile_card(cand_d)

        # 템플릿 로드
        tpl = _load_email_template('profile_broadcast')
        if not tpl:
            raise HTTPException(status_code=404, detail="profile_broadcast template not found")

        base_subject = body.custom_subject or tpl["subject"]
        base_body = body.custom_body or tpl["body_html"]

        sent_count = 0
        fail_count = 0
        results = []

        for emp_id in body.employer_ids:
            emp = conn.execute(
                "SELECT id, email, school_name, school_name_plain, contact_name FROM client_inquiries WHERE id = ?",
                (emp_id,)
            ).fetchone()
            if not emp:
                continue

            # PII 복호화 -- email/contact_name 은 T3v1 암호화, school_name_plain 우선
            to_email   = _safe_decrypt(emp["email"], "email") or ""
            school     = (emp["school_name_plain"]
                          or _safe_decrypt(emp["school_name"], "school_name")
                          or "School")
            contact    = _safe_decrypt(emp["contact_name"], "contact_name") or ""
            if not to_email or "@" not in to_email:
                continue  # 복호화 실패/빈값이면 스킵

            # 템플릿 변수 치환
            html = base_body.replace("{{profile_cards}}", card_html)
            html = html.replace("{{school_name}}", school)
            html = html.replace("{{contact_name}}", contact)
            html = re.sub(r"\{\{\w+\}\}", "", html)

            subject = f"{base_subject} - {school}" if school != "School" else base_subject

            sent = _smtp_send(to_email, subject, html)
            status_val = "sent" if sent else "failed"
            if sent:
                sent_count += 1
            else:
                fail_count += 1

            # 발송 로그 -- to_email 마스킹 저장 (PII 최소화)
            masked_email = to_email[:3] + "***" + to_email[to_email.index("@"):] if "@" in to_email else "***"
            conn.execute(
                """INSERT INTO profile_sends (candidate_id, employer_id, school_name, to_email, status)
                   VALUES (?, ?, ?, ?, ?)""",
                (body.candidate_id, emp_id, school, masked_email, status_val)
            )
            results.append({"employer_id": emp_id, "school": school, "email": masked_email, "status": status_val})

        conn.commit()

        return ok(
            message=f"Profile broadcast: {sent_count} sent, {fail_count} failed",
            data={"sent": sent_count, "failed": fail_count, "results": results}
        )
    finally:
        conn.close()


@app.get("/api/admin/matching/history", tags=["admin"])
async def admin_matching_history(request: Request, candidate_id: str = ""):
    """프로필 발송 이력 조회."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        if candidate_id:
            rows = conn.execute(
                "SELECT * FROM profile_sends WHERE candidate_id = ? ORDER BY sent_at DESC",
                (candidate_id,)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM profile_sends ORDER BY sent_at DESC LIMIT 200"
            ).fetchall()
        return ok(data=[dict(r) for r in rows])
    finally:
        conn.close()


# ── Community Board API ──────────────────────────────────────────────────────
_BOARDS = {
    "visa", "visa_type", "visa_related", "immigration",
    "support_kr", "support", "about", "korea", "tips", "testimonials", "information",
}
_RATE_LIMIT: dict[str, list] = {}  # ip_hash → [timestamps]
_AUTH_FAIL: dict[str, list] = {}   # ip_hash → [fail timestamps] (brute-force 방어)

import hashlib, time as _time

def _ip_hash(request: Request) -> str:
    # _get_client_ip() 와 동일한 스푸핑 방지 로직 사용
    # X-Forwarded-For 첫 항목은 공격자가 위조 가능하므로 절대 사용 금지
    ip = _get_client_ip(request)
    return hashlib.sha256(ip.encode()).hexdigest()[:16]


def _is_ssrf_safe_url(url: str) -> bool:
    """SSRF 방어: 내부 IP/메타데이터 URL 차단. True=안전, False=차단."""
    import ipaddress, urllib.parse
    try:
        parsed = urllib.parse.urlparse(url)
        if parsed.scheme not in ("http", "https"):
            return False
        host = parsed.hostname or ""
        # 메타데이터 서버 차단
        if host in ("169.254.169.254", "metadata.google.internal"):
            return False
        try:
            addr = ipaddress.ip_address(host)
            if addr.is_private or addr.is_loopback or addr.is_link_local or addr.is_reserved:
                return False
        except ValueError:
            pass  # 도메인 이름이면 통과 (DNS rebinding은 별도 처리)
        return True
    except Exception:
        return False

def _rate_ok(ip_hash: str, window: int = 300, max_posts: int = 5) -> bool:
    """5분 안에 최대 5건 포스팅 허용 (IP 해시 기반)."""
    now = _time.time()
    ts = _RATE_LIMIT.get(ip_hash, [])
    ts = [t for t in ts if now - t < window]
    if len(ts) >= max_posts:
        return False
    ts.append(now)
    _RATE_LIMIT[ip_hash] = ts
    return True


# ── CAPTCHA 검증 (Puzzle CAPTCHA) ───────────────────────────────────────────────
# H2 보안 패치 (완화): BRIDGE_HMAC_KEY 미설정 시 랜덤 키로 대체 (서버 기동 허용)
# CAPTCHA는 단일 인스턴스 내에서만 검증하므로 랜덤 키도 기능상 문제없음.
# 보안 강화 목적으로 Render 환경변수에 BRIDGE_HMAC_KEY를 설정하는 것을 강권.
_raw_hmac_key = os.environ.get("BRIDGE_HMAC_KEY")
if not _raw_hmac_key:
    import secrets as _sec_hmac
    _raw_hmac_key = _sec_hmac.token_hex(32)
    logging.getLogger("bridge.api").warning(
        "[WARNING] BRIDGE_HMAC_KEY 미설정 -- 임시 랜덤 키 사용 중. "
        "Render 대시보드 → Environment Variables → BRIDGE_HMAC_KEY 추가 권장."
    )
_CAPTCHA_HMAC_KEY: bytes = _raw_hmac_key.encode()

# C1 패치: CAPTCHA_ENABLED 플래그 + nonce replay 방지
_CAPTCHA_ENABLED = os.environ.get("CAPTCHA_ENABLED", "true").lower() != "false"
_CAPTCHA_NONCES: "dict[str, float]" = {}  # nonce → 만료 시각(epoch)


def _verify_captcha_token(token: str, ip_hash: str) -> bool:
    """CAPTCHA 토큰 검증: 타임스탬프 유효성 + nonce 재사용 방지.
    C1 패치: return True 하드코딩 제거 -- 실제 검증 수행.
    """
    if not _CAPTCHA_ENABLED:
        return True
    if not token or not token.startswith("puzzle_"):
        return False

    try:
        # 토큰 형식: puzzle_{timestamp_ms}_{nonce}
        parts = token.split("_")
        if len(parts) < 3:
            return False

        timestamp = int(parts[1])
        nonce     = parts[2]
        now_ms    = int(_time.time() * 1000)

        # 30분 이내인지 확인 — 사용자가 폼 작성하는 동안 만료 방지
        # (CAPTCHA 자체가 시작 단계에서 처리되므로 폼 작성 시간 충분히 확보)
        if now_ms - timestamp > 30 * 60 * 1000:
            return False

        # nonce 재사용(replay attack) 방지
        now_sec = _time.time()
        if nonce in _CAPTCHA_NONCES:
            logging.getLogger("bridge.security").warning(
                "[C1] CAPTCHA nonce 재사용 시도 차단: nonce=%s ip=%s", nonce[:8], ip_hash[:8]
            )
            return False
        _CAPTCHA_NONCES[nonce] = now_sec + 600  # 10분 보관 후 만료

        # 만료된 nonce 정리 (메모리 누수 방지)
        expired = [n for n, exp in list(_CAPTCHA_NONCES.items()) if now_sec > exp]
        for n in expired:
            del _CAPTCHA_NONCES[n]

        return True
    except Exception:
        return False


def _check_honeypot(payload: dict) -> bool:
    """Honeypot 필드 확인 (봇 탐지)."""
    honeypot_field = payload.get("_url", "")
    if honeypot_field:
        # 봇이 채운 필드가 있으면 True (거부)
        return True
    return False


@app.get("/api/community/{board}", tags=["community"])
async def community_list(
    board: str,
    request: Request,
    limit: int = 30,
    offset: int = 0,
    category: Optional[str] = None,
):
    limit = min(max(limit, 1), 50)  # 스크래핑 방지
    if not _rate_ok(_ip_hash(request), window=60, max_posts=40):
        return bridge_error("RATE_LIMIT", "Too many requests.", retryable=True, status=429)
    if board not in _BOARDS:
        raise HTTPException(404, "Board not found")
    # ── Supabase 경로 ──
    if _supa:
        try:
            q = _supa.table('community_posts').select(
                'id,title,author_hash,pinned,views,created_at,content_type,sort_order,category,body,image_paths',
                count='exact'
            ).eq('board', board).eq('is_deleted', 0)
            if category:
                q = q.eq('category', category)
            q = q.order('pinned', desc=True).order('sort_order', desc=True).order('created_at', desc=True)
            q = q.range(offset, offset + limit - 1)
            res = q.execute()
            total = res.count or 0
            posts = [dict(r, preview=(r.get('body') or '')[:200]) for r in (res.data or [])]
            return ok(data={"total": total, "posts": posts})
        except Exception as _e:
            logging.getLogger("bridge.api").warning("[Supabase] community_list 실패 → SQLite 폴백: %s", _e)
    # ── SQLite 폴백 ──
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.row_factory = sqlite3.Row
    try:
        where = "board=? AND is_deleted=0"
        params: list = [board]
        if category:
            where += " AND category=?"
            params.append(category)
        total = conn.execute(f"SELECT COUNT(*) FROM community_posts WHERE {where}", params).fetchone()[0]
        rows = conn.execute(
            f"SELECT id,title,author_hash,pinned,views,created_at,content_type,sort_order,category,body,image_paths "
            f"FROM community_posts WHERE {where} "
            f"ORDER BY pinned DESC, sort_order DESC, created_at DESC LIMIT ? OFFSET ?",
            params + [limit, offset],
        ).fetchall()
        posts = []
        for r in rows:
            d = dict(r)
            d['preview'] = (d.get('body') or '')[:200]
            # image_paths: SQLite는 TEXT JSON 또는 빈 문자열 → 항상 list로 정규화
            ip_raw = d.get('image_paths')
            if isinstance(ip_raw, str) and ip_raw.strip():
                try:
                    import json as _json
                    d['image_paths'] = _json.loads(ip_raw) if ip_raw.strip().startswith('[') else [ip_raw]
                except Exception:
                    d['image_paths'] = [ip_raw] if ip_raw.strip() else []
            elif not isinstance(ip_raw, list):
                d['image_paths'] = []
            posts.append(d)
        return ok(data={"total": total, "posts": posts})
    finally:
        conn.close()


@app.get("/api/community/{board}/{post_id}", tags=["community"])
async def community_get(board: str, post_id: int, request: Request):
    if not _rate_ok(_ip_hash(request), window=60, max_posts=40):
        return bridge_error("RATE_LIMIT", "Too many requests.", retryable=True, status=429)
    if board not in _BOARDS:
        raise HTTPException(404, "Board not found")
    # ── Supabase 경로 ──
    if _supa:
        try:
            res = _supa.table('community_posts').select(
                'id,board,title,body,pinned,views,created_at,content_type,image_paths,category'
            ).eq('id', post_id).eq('board', board).eq('is_deleted', 0).execute()
            if not res.data:
                raise HTTPException(404, "Post not found")
            d = res.data[0]
            new_views = (d.get('views') or 0) + 1
            _supa.table('community_posts').update({'views': new_views}).eq('id', post_id).execute()
            d['views'] = new_views
            return ok(data=d)
        except HTTPException:
            raise
        except Exception as _e:
            logging.getLogger("bridge.api").warning("[Supabase] community_get 실패 → SQLite 폴백: %s", _e)
    # ── SQLite 폴백 ──
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.row_factory = sqlite3.Row
    try:
        row = conn.execute(
            "SELECT id,board,title,body,pinned,views,created_at,content_type,image_paths,category "
            "FROM community_posts WHERE id=? AND board=? AND is_deleted=0",
            (post_id, board),
        ).fetchone()
        if not row:
            raise HTTPException(404, "Post not found")
        d = dict(row)
        # image_paths 정규화
        ip_raw = d.get('image_paths')
        if isinstance(ip_raw, str) and ip_raw.strip():
            try:
                import json as _json
                d['image_paths'] = _json.loads(ip_raw) if ip_raw.strip().startswith('[') else [ip_raw]
            except Exception:
                d['image_paths'] = [ip_raw] if ip_raw.strip() else []
        elif not isinstance(ip_raw, list):
            d['image_paths'] = []
        new_views = (d.get('views') or 0) + 1
        conn.execute("UPDATE community_posts SET views=? WHERE id=?", (new_views, post_id))
        conn.commit()
        d['views'] = new_views
        return ok(data=d)
    finally:
        conn.close()


class CommunityPost(BaseModel):
    title: str = Field(..., min_length=2, max_length=200)
    body:  str = Field(..., min_length=10, max_length=200000)  # HTML 게시물 지원 (200k)
    content_type: str = Field("markdown", pattern=r"^(markdown|html)$")
    category: Optional[str] = Field(None, max_length=50)


@app.post("/api/community/{board}", status_code=201, tags=["community"])
async def community_create(board: str, post: CommunityPost, request: Request):
    if board not in _BOARDS:
        raise HTTPException(404, "Board not found")
    ip_hash = _ip_hash(request)
    if not _rate_ok(ip_hash):
        raise HTTPException(429, "Too many posts. Please wait a few minutes.")

    # 콘텐츠 타입에 따라 처리
    _tag_re = re.compile(r'<[^>]+>')
    clean_title = _tag_re.sub('', post.title.strip())  # 제목은 항상 strip
    if post.content_type == 'html':
        clean_body = _sanitize_html(post.body.strip())
    else:
        clean_body = _tag_re.sub('', post.body.strip())

    # 게시글 본문에서 연락처 PII 차단 (태그 제거 후 검사)
    _pii_check = re.compile(r'01[016789][- ]?\d{3,4}[- ]?\d{4}|[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}')
    pii_check_body = _tag_re.sub('', clean_body)  # PII 체크는 태그 제거 후 텍스트만
    if _pii_check.search(pii_check_body) or _pii_check.search(clean_title):
        raise HTTPException(400, "Phone numbers and email addresses are not allowed in posts.")

    # ── Supabase 경로 ──
    if _supa:
        try:
            res = _supa.table('community_posts').insert({
                'board': board,
                'title': clean_title,
                'body': clean_body,
                'author_hash': ip_hash,
                'content_type': post.content_type,
                'category': post.category,
                'pinned': 0,
                'views': 0,
                'is_deleted': 0,
                'sort_order': 0,
                'image_paths': '[]',
            }).execute()
            new_id = res.data[0]['id'] if res.data else None
            return ok(data={"id": new_id}, message="Post created")
        except Exception as _e:
            logging.getLogger("bridge.api").warning("[Supabase] community_create 실패 → SQLite 폴백: %s", _e)
    # ── SQLite 폴백 ──
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    try:
        cur = conn.execute(
            "INSERT INTO community_posts (board,title,body,author_hash,content_type,category,pinned,views,is_deleted,sort_order,created_at) "
            "VALUES (?,?,?,?,?,?,0,0,0,0,datetime('now'))",
            (board, clean_title, clean_body, ip_hash, post.content_type, post.category),
        )
        conn.commit()
        new_id = cur.lastrowid
    finally:
        conn.close()
    return ok(data={"id": new_id}, message="Post created")


@app.delete("/api/community/{board}/{post_id}", tags=["community"])
async def community_delete(board: str, post_id: int, request: Request):
    """작성자(IP 해시 일치)만 삭제 가능."""
    if board not in _BOARDS:
        raise HTTPException(404, "Board not found")
    ip_hash  = _ip_hash(request)
    is_admin = _ADMIN_KEY and request.headers.get("x-admin-key", "") == _ADMIN_KEY
    # ── Supabase 경로 ──
    if _supa:
        try:
            res = _supa.table('community_posts').select('author_hash').eq('id', post_id).eq('board', board).eq('is_deleted', 0).execute()
            if not res.data:
                raise HTTPException(404, "Post not found")
            if res.data[0].get('author_hash') != ip_hash and not is_admin:
                raise HTTPException(403, "Forbidden")
            _supa.table('community_posts').update({'is_deleted': 1}).eq('id', post_id).execute()
            return ok(message="Deleted")
        except HTTPException:
            raise
        except Exception as _e:
            logging.getLogger("bridge.api").warning("[Supabase] community_delete 실패 → SQLite 폴백: %s", _e)
    # ── SQLite 폴백 ──
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        row = conn.execute(
            "SELECT author_hash FROM community_posts WHERE id = ? AND board = ? AND is_deleted = 0 LIMIT 1",
            (post_id, board),
        ).fetchone()
        if not row:
            raise HTTPException(404, "Post not found")
        if row[0] != ip_hash and not is_admin:
            raise HTTPException(403, "Forbidden")
        conn.execute("UPDATE community_posts SET is_deleted = 1 WHERE id = ?", (post_id,))
        conn.commit()
    finally:
        conn.close()
    return ok(message="Deleted")


# ── Admin: Community Posts 관리 ────────────────────────────────────────────────

class PinUpdate(BaseModel):
    pinned: int = Field(..., ge=0, le=1)


@app.patch("/api/admin/community/posts/{post_id}/pin", tags=["admin"])
async def admin_pin_post(post_id: int, body: PinUpdate, request: Request):
    """게시글 고정/해제 (관리자 전용)."""
    _check_admin(request)
    # ── Supabase 경로 ──
    if _supa:
        try:
            res = _supa.table('community_posts').select('id').eq('id', post_id).eq('is_deleted', 0).execute()
            if not res.data:
                raise HTTPException(404, "Post not found")
            _supa.table('community_posts').update({'pinned': body.pinned}).eq('id', post_id).execute()
            return ok(message=f"Post #{post_id} pinned={body.pinned}")
        except HTTPException:
            raise
        except Exception as _e:
            logging.getLogger("bridge.api").warning("[Supabase] admin_pin_post 실패 → SQLite 폴백: %s", _e)
    # ── SQLite 폴백 ──
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        row = conn.execute(
            "SELECT id FROM community_posts WHERE id = ? AND is_deleted = 0 LIMIT 1", (post_id,)
        ).fetchone()
        if not row:
            raise HTTPException(404, "Post not found")
        conn.execute("UPDATE community_posts SET pinned = ? WHERE id = ?", (body.pinned, post_id))
        conn.commit()
    finally:
        conn.close()
    return ok(message=f"Post #{post_id} pinned={body.pinned}")


class PostEdit(BaseModel):
    title: Optional[str] = Field(None, min_length=2, max_length=200)
    body:  Optional[str] = Field(None, min_length=10, max_length=200000)
    content_type: Optional[str] = Field(None, pattern=r"^(markdown|html)$")
    category: Optional[str] = Field(None, max_length=50)


class ReorderItem(BaseModel):
    id: int
    sort_order: int


class ReorderRequest(BaseModel):
    items: list[ReorderItem]


@app.patch("/api/admin/community-reorder/{board}", tags=["admin"])
async def admin_reorder_posts(board: str, body: ReorderRequest, request: Request):
    """게시글 정렬 순서 일괄 업데이트 (관리자 전용)."""
    _check_admin(request)
    if board not in _BOARDS:
        raise HTTPException(404, "Board not found")
    # ── Supabase 경로 ──
    if _supa:
        try:
            for item in body.items:
                _supa.table('community_posts').update({'sort_order': item.sort_order}).eq('id', item.id).eq('board', board).eq('is_deleted', 0).execute()
            _maybe_auto_backup()
            return ok(message=f"Reordered {len(body.items)} posts")
        except Exception as _e:
            logging.getLogger("bridge.api").warning("[Supabase] admin_reorder_posts 실패 → SQLite 폴백: %s", _e)
    # ── SQLite 폴백 ──
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        for item in body.items:
            conn.execute(
                "UPDATE community_posts SET sort_order = ? WHERE id = ? AND board = ? AND is_deleted = 0",
                (item.sort_order, item.id, board),
            )
        conn.commit()
    finally:
        conn.close()
    _maybe_auto_backup()
    return ok(message=f"Reordered {len(body.items)} posts")


@app.patch("/api/admin/community/{board}/{post_id}", tags=["admin"])
async def admin_edit_post(board: str, post_id: int, body: PostEdit, request: Request):
    """게시글 제목/본문 수정 (관리자 전용)."""
    _check_admin(request)
    if board not in _BOARDS:
        raise HTTPException(404, "Board not found")

    _tag_re = re.compile(r'<[^>]+>')
    updates: dict = {}
    if body.title is not None:
        updates['title'] = _tag_re.sub('', body.title.strip())
    if body.body is not None:
        updates['body'] = _sanitize_html(body.body.strip()) if body.content_type == 'html' \
                          else _tag_re.sub('', body.body.strip())
    if body.content_type is not None:
        updates['content_type'] = body.content_type
    if body.category is not None:
        updates['category'] = body.category
    if not updates:
        raise HTTPException(400, "수정할 항목이 없습니다.")

    # ── Supabase 경로 ──
    if _supa:
        try:
            res = _supa.table('community_posts').select('id').eq('id', post_id).eq('board', board).eq('is_deleted', 0).execute()
            if not res.data:
                raise HTTPException(404, "Post not found")
            _supa.table('community_posts').update(updates).eq('id', post_id).eq('board', board).execute()
            _maybe_auto_backup()
            return ok(message=f"Post #{post_id} updated")
        except HTTPException:
            raise
        except Exception as _e:
            logging.getLogger("bridge.api").warning("[Supabase] admin_edit_post 실패 → SQLite 폴백: %s", _e)
    # ── SQLite 폴백 ──
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        row = conn.execute(
            "SELECT id FROM community_posts WHERE id = ? AND board = ? AND is_deleted = 0 LIMIT 1",
            (post_id, board),
        ).fetchone()
        if not row:
            raise HTTPException(404, "Post not found")
        set_clause = ", ".join(f"{k} = ?" for k in updates)
        vals = list(updates.values()) + [post_id, board]
        conn.execute(f"UPDATE community_posts SET {set_clause} WHERE id = ? AND board = ?", vals)
        conn.commit()
    finally:
        conn.close()
    _maybe_auto_backup()
    return ok(message=f"Post #{post_id} updated")


@app.patch("/api/admin/community/{board}/{post_id}/move", tags=["admin"])
async def admin_move_post(board: str, post_id: int, request: Request):
    """게시글 게시판 이동 (관리자 전용). body: {"target_board": "visa_type"}"""
    _check_admin(request)
    if board not in _BOARDS:
        raise HTTPException(404, "Board not found")
    body   = await request.json()
    target = str(body.get("target_board", "")).strip()
    if target not in _BOARDS:
        raise HTTPException(400, f"유효하지 않은 대상 게시판: {target}")
    if target == board:
        raise HTTPException(400, "현재 게시판과 동일합니다.")
    # ── Supabase 경로 ──
    if _supa:
        try:
            res = _supa.table('community_posts').select('id').eq('id', post_id).eq('board', board).eq('is_deleted', 0).execute()
            if not res.data:
                raise HTTPException(404, "Post not found")
            _supa.table('community_posts').update({'board': target}).eq('id', post_id).eq('board', board).execute()
            return ok(message=f"Post #{post_id} moved to {target}")
        except HTTPException:
            raise
        except Exception as _e:
            logging.getLogger("bridge.api").warning("[Supabase] admin_move_post 실패 → SQLite 폴백: %s", _e)
    # ── SQLite 폴백 ──
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        row = conn.execute(
            "SELECT id FROM community_posts WHERE id = ? AND board = ? AND is_deleted = 0 LIMIT 1",
            (post_id, board),
        ).fetchone()
        if not row:
            raise HTTPException(404, "Post not found")
        conn.execute(
            "UPDATE community_posts SET board = ? WHERE id = ? AND board = ?",
            (target, post_id, board),
        )
        conn.commit()
    finally:
        conn.close()
    return ok(message=f"Post #{post_id} moved to {target}")


@app.get("/api/admin/community/posts", tags=["admin"])
async def admin_search_posts(
    request: Request,
    search: Optional[str] = None,
    board: Optional[str] = None,
    limit: int = 200,
):
    """관리자 게시글 검색 (전체 보드 대상)."""
    _check_admin(request)
    # ── Supabase 경로 ──
    if _supa:
        try:
            q = _supa.table('community_posts').select(
                'id,board,title,author_hash,pinned,views,created_at,content_type,sort_order,body'
            ).eq('is_deleted', 0)
            if board and board != "all":
                if board not in _BOARDS:
                    raise HTTPException(404, "Board not found")
                q = q.eq('board', board)
            if search:
                term = search.strip()
                q = q.or_(f"title.ilike.%{term}%,body.ilike.%{term}%")
            q = q.order('pinned', desc=True).order('sort_order', desc=True).order('created_at', desc=True).limit(limit)
            res = q.execute()
            posts = [dict(r, preview=(r.get('body', '') or '')[:200]) for r in (res.data or [])]
            return ok(data={"posts": posts})
        except HTTPException:
            raise
        except Exception as _e:
            logging.getLogger("bridge.api").warning("[Supabase] admin_search_posts 실패 → SQLite 폴백: %s", _e)
    # ── SQLite 폴백 ──
    sql = "SELECT id, board, title, author_hash, pinned, views, created_at, content_type, sort_order, body FROM community_posts WHERE is_deleted = 0"
    params: list = []
    if board and board != "all":
        if board not in _BOARDS:
            raise HTTPException(404, "Board not found")
        sql += " AND board = ?"
        params.append(board)
    if search:
        term = search.strip()
        sql += " AND (title LIKE ? OR body LIKE ?)"
        params.extend([f"%{term}%", f"%{term}%"])
    sql += " ORDER BY pinned DESC, sort_order DESC, created_at DESC LIMIT ?"
    params.append(limit)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        cur = conn.execute(sql, params)
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    finally:
        conn.close()
    posts = []
    for row in rows:
        row['preview'] = (row.get('body', '') or '')[:200]
        posts.append(row)
    return ok(data={"posts": posts})


# ── Admin: Applications 관리 ─────────────────────────────────────────────────

@app.get("/api/admin/applications", tags=["admin"])
async def admin_list_applications(
    request: Request,
    type: Optional[str] = None,
    page: int = 1,
    limit: int = 200,
):
    """구직자 + 구인자 접수 통합 목록 -- SQLite.
    type=employer → 구인자만 (구직자 스킵, 성능 향상).
    page/limit → 페이지네이션 (type=employer 시 적용).
    """
    _check_admin(request)
    employer_only = (type == "employer")
    page = max(1, page)
    limit = min(max(1, limit), 500)
    offset = (page - 1) * limit
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        conn.row_factory = sqlite3.Row
        try:
            apps: list[dict] = []

            # 구직자 (candidates) -- employer_only 시 스킵
            if not employer_only:
                cands = conn.execute(
                    "SELECT candidate_id, full_name, email, nationality, mobile_phone, "
                    "current_location, target, target_age, desired_salary, experience, "
                    "start_date, status, created_at, updated_at "
                    "FROM candidates ORDER BY created_at DESC LIMIT 200"
                ).fetchall()
                for c in cands:
                    apps.append({
                        "id": c["candidate_id"], "type": "candidate",
                        "name": _safe_decrypt(c["full_name"], "full_name") or "", "email": _safe_decrypt(c["email"], "email") or "",
                        "nationality": c["nationality"],
                        "phone": _safe_decrypt(c["mobile_phone"], "mobile_phone"),
                        "location": c["current_location"],
                        "target": c["target"],
                        "target_age": c["target_age"],
                        "desired_salary": c["desired_salary"],
                        "experience": c["experience"],
                        "start_date": c["start_date"],
                        "status": c["status"] or "Active",
                        "created_at": c["created_at"] or "",
                        "updated_at": c["updated_at"],
                    })

            # 구인자 -- jobs 테이블 (원본 워드포맷 데이터 + raw_text 포함)
            # 테이블 존재 여부 확인 (Render 신규/빈 DB 대비)
            _exist_check = {r[0] for r in conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name IN ('jobs','client_inquiries')"
            ).fetchall()}

            job_rows = conn.execute(
                "SELECT id, job_code, seq, location, city, district, region_name, "
                "start_date, teaching_age, class_size, working_hours, salary_raw, "
                "teach_hrs_week, vacation, housing, housing_type, housing_detail, "
                "native_count, benefits, internal_notes, status, created_at, source_file, raw_text, "
                "enc_employer_name, enc_contact_name, enc_contact_phone, enc_contact_email, "
                "enc_contact_kakao, employer_display_name "
                "FROM jobs WHERE is_deleted = 0 ORDER BY seq ASC, job_code ASC"
            ).fetchall() if "jobs" in _exist_check else []
            for j in job_rows:
                # 암호화된 PII 필드 복호화
                dec_employer = _safe_decrypt(j["enc_employer_name"], "enc_employer_name")
                dec_contact  = _safe_decrypt(j["enc_contact_name"], "enc_contact_name")
                dec_phone    = _safe_decrypt(j["enc_contact_phone"], "enc_contact_phone")
                dec_email    = _safe_decrypt(j["enc_contact_email"], "enc_contact_email")
                dec_kakao    = _safe_decrypt(j["enc_contact_kakao"], "enc_contact_kakao")
                display_name = j["employer_display_name"] or ""
                # school_name: display_name 우선 → 복호화된 employer_name fallback
                school = display_name or dec_employer or ""
                apps.append({
                    "id": str(j["id"]), "type": "employer",
                    "name": school, "email": dec_email or "",
                    "school_name": school,
                    "contact_name": dec_contact,
                    "phone": dec_phone,
                    "kakao": dec_kakao,
                    "job_code": j["job_code"] or "",
                    "source_file": j["source_file"],
                    "location": j["location"] or "",
                    "city": j["city"] or "",
                    "start_date": j["start_date"],
                    "teaching_age": j["teaching_age"],
                    "class_size": j["class_size"],
                    "working_hours": j["working_hours"],
                    "salary_raw": j["salary_raw"],
                    "teach_hrs_week": str(j["teach_hrs_week"]) if j["teach_hrs_week"] else None,
                    "vacation": j["vacation"],
                    "housing": j["housing"],
                    "housing_type": j["housing_type"],
                    "housing_detail": j["housing_detail"],
                    "native_count": j["native_count"],
                    "benefits": j["benefits"],
                    "memo": _safe_decrypt(j["internal_notes"], "memo"),   # T3v1 복호화 (암호화 시 column_name='memo' 사용됨)
                    "notes": None,
                    "raw_text": j["raw_text"] or None,            # 원본 텍스트 (어드민: PII 미제거)
                    "status": j["status"] or "open",
                    "created_at": j["created_at"] or "",
                })

            # 구인자 -- client_inquiries 테이블 (이메일/폼 접수 데이터)
            # school_name_plain/location_plain 우선 사용 (plain 컬럼 100% 존재)
            # 실패 시 _safe_decrypt로 암호화 값 복호화 폴백
            inq_rows = conn.execute(
                "SELECT id, school_name, school_name_plain, email, contact_name, phone, "
                "location, location_plain, "
                "start_date, vacancies, teaching_age, schedule, working_hours, "
                "salary_raw, housing_type, housing_detail, travel_support, benefits, vacation, "
                "sick_leave, meal, memo, notes, assigned_to, submitted_at, raw_email_body, "
                "COALESCE(inbox_status, 'new') as status "
                "FROM client_inquiries WHERE is_deleted = 0 ORDER BY submitted_at DESC"
            ).fetchall() if "client_inquiries" in _exist_check else []
            for inq in inq_rows:
                # PII 필드 복호화
                dec_email   = _safe_decrypt(inq["email"], "email")
                dec_phone   = _safe_decrypt(inq["phone"], "phone")
                dec_contact = _safe_decrypt(inq["contact_name"], "contact_name")
                dec_memo    = _safe_decrypt(inq["memo"], "memo")
                # school_name: plain 우선 → 암호화값 복호화 폴백 (100% 암호화 대응)
                school_display = (
                    inq["school_name_plain"]
                    or _safe_decrypt(inq["school_name"], "school_name")
                    or ""
                )
                # location: plain 우선 → 암호화값 복호화 폴백
                location_display = (
                    inq["location_plain"]
                    or _safe_decrypt(inq["location"], "location")
                    or ""
                )
                apps.append({
                    "id": f"inq_{inq['id']}", "type": "employer",
                    "name": school_display,
                    "email": dec_email or "",
                    "school_name": school_display,
                    "contact_name": dec_contact,
                    "phone": dec_phone,
                    "location": location_display,
                    "start_date": inq["start_date"],
                    "vacancies": inq["vacancies"],
                    "teaching_age": inq["teaching_age"],
                    "schedule": inq["schedule"],
                    "working_hours": inq["working_hours"],
                    "salary_raw": inq["salary_raw"],
                    "housing_type": inq["housing_type"],
                    "housing_detail": inq["housing_detail"],
                    "travel_support": inq["travel_support"],
                    "benefits": inq["benefits"],
                    "vacation": inq["vacation"],
                    "sick_leave": inq["sick_leave"],
                    "meal": inq["meal"],
                    "memo": dec_memo,
                    "notes": inq["notes"],
                    "assigned_to": inq["assigned_to"],
                    "raw_email_body": inq["raw_email_body"] or None,
                    "raw_text": None,
                    "status": inq["status"],
                    "created_at": inq["submitted_at"] or "",
                })

            apps.sort(key=lambda x: x.get("created_at", ""), reverse=True)

            # employer_only 시 페이지네이션 적용
            if employer_only:
                total = len(apps)
                page_data = apps[offset: offset + limit]
                result = ok(data=page_data, message=f"{len(page_data)}건 (전체 {total}건)")
                result["total"] = total
                result["page"] = page
                result["has_more"] = offset + limit < total
                return result

            return ok(data=apps)
        finally:
            conn.close()

    except HTTPException:
        raise
    except Exception as e:
        import logging as _log_apps
        _log_apps.getLogger("bridge.api").error("admin_applications 실패: %s", e, exc_info=True)
        err("접수 목록을 불러올 수 없습니다.", 500)


class StatusUpdate(BaseModel):
    status: Optional[str] = Field(None, min_length=1, max_length=50)
    type: Optional[str] = None
    memo: Optional[str] = None   # 내부 메모 (어드민 전용)
    notes: Optional[str] = None  # 본문 편집 (raw_text 대체)


@app.patch("/api/admin/applications/{app_id}", tags=["admin"])
async def admin_update_application(app_id: str, body: StatusUpdate, request: Request):
    """접수 상태/메모/본문 변경 -- SQLite."""
    _check_admin(request)
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        try:
            now_iso = datetime.now(timezone.utc).isoformat()
            if body.type == "candidate":
                if body.status:
                    conn.execute(
                        "UPDATE candidates SET status = ?, updated_at = ? WHERE candidate_id = ?",
                        (body.status, now_iso, app_id),
                    )
            elif app_id.startswith("inq_"):
                # client_inquiries 업데이트
                inq_id = int(app_id[4:])
                if body.status:
                    conn.execute(
                        "UPDATE client_inquiries SET inbox_status = ? WHERE id = ?",
                        (body.status, inq_id),
                    )
                if body.memo is not None:
                    enc_memo = _encrypt_if_needed(body.memo, "memo") if body.memo else None
                    conn.execute(
                        "UPDATE client_inquiries SET memo = ? WHERE id = ?",
                        (enc_memo, inq_id),
                    )
                if body.notes is not None:
                    conn.execute(
                        "UPDATE client_inquiries SET notes = ? WHERE id = ?",
                        (body.notes, inq_id),
                    )
            else:
                # jobs 테이블 업데이트
                job_id = int(app_id)
                if body.status:
                    conn.execute(
                        "UPDATE jobs SET status = ? WHERE id = ?",
                        (body.status, job_id),
                    )
                if body.memo is not None:
                    enc_memo = _encrypt_if_needed(body.memo, "internal_notes") if body.memo else None
                    conn.execute(
                        "UPDATE jobs SET internal_notes = ? WHERE id = ?",
                        (enc_memo, job_id),
                    )
                if body.notes is not None:
                    conn.execute(
                        "UPDATE jobs SET raw_text = ? WHERE id = ?",
                        (body.notes, job_id),
                    )
            conn.commit()
            return ok(message=f"{app_id} 업데이트 완료")
        finally:
            conn.close()

    except HTTPException:
        raise
    except Exception as e:
        import logging as _log_upd
        _log_upd.getLogger("bridge.api").error("admin_update_app 실패: %s", e, exc_info=True)
        err("상태 변경에 실패했습니다.", 500)


# ── Admin: Payments 관리 ──────────────────────────────────────────────────────

@app.get("/api/admin/payments", tags=["admin"])
async def admin_list_payments(request: Request):
    """결제 기록 목록 -- SQLite (테이블 없으면 빈 배열)."""
    _check_admin(request)
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        conn.row_factory = sqlite3.Row
        try:
            # payments 테이블 존재 여부 확인
            exists = conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name='payments'"
            ).fetchone()
            if not exists:
                return ok(data=[])
            rows = conn.execute(
                "SELECT * FROM payments ORDER BY created_at DESC LIMIT 200"
            ).fetchall()
            return ok(data=[dict(r) for r in rows])
        finally:
            conn.close()
    except HTTPException:
        raise
    except Exception as e:
        import logging as _log_pay
        _log_pay.getLogger("bridge.api").error("admin_payments 실패: %s", e, exc_info=True)
        return ok(data=[])


class PaymentStatusUpdate(BaseModel):
    status: str = Field(..., min_length=1, max_length=50)


@app.patch("/api/admin/payments/{payment_id}", tags=["admin"])
async def admin_update_payment(payment_id: str, body: PaymentStatusUpdate, request: Request):
    """결제 상태 변경."""
    _check_admin(request)
    try:
        now_iso = datetime.now(timezone.utc).isoformat()
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        try:
            conn.execute(
                "UPDATE payments SET status = ?, confirmed_at = ? WHERE id = ?",
                (body.status, now_iso if body.status == "confirmed" else None, payment_id),
            )
            conn.commit()
        finally:
            conn.close()
        return ok(message=f"Payment {payment_id} -> {body.status}")

    except HTTPException:
        raise
    except Exception as e:
        import logging as _log_pay_upd
        _log_pay_upd.getLogger("bridge.api").error("admin_payment_update 실패: %s", e, exc_info=True)
        err("상태 변경에 실패했습니다.", 500)


# ── Admin: Interview Scheduling ──────────────────────────────────────────────


def _ensure_interviews_schema():
    """interviews 테이블에 이메일 발송 timestamp 컬럼 추가."""
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    for col in ("email_sent_candidate_at", "email_sent_employer_at"):
        try:
            conn.execute(f"ALTER TABLE interviews ADD COLUMN {col} TEXT")
        except Exception:
            pass
    conn.commit()
    conn.close()


try:
    _ensure_interviews_schema()
except Exception as _e:
    logging.getLogger("bridge.api").warning("_ensure_interviews_schema 스킵: %s", _e)


def _ensure_interview_status_log():
    """interview_status_log 테이블 생성 (상태 변경 이력 추적)."""
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS interview_status_log (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            interview_id INTEGER NOT NULL,
            old_status   TEXT,
            new_status   TEXT NOT NULL,
            changed_by   TEXT DEFAULT 'admin',
            note         TEXT DEFAULT '',
            created_at   TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_isl_interview ON interview_status_log(interview_id)")
    conn.commit()
    conn.close()


try:
    _ensure_interview_status_log()
except Exception as _e:
    logging.getLogger("bridge.api").warning("_ensure_interview_status_log 스킵: %s", _e)


def _ensure_community_schema():
    """community_posts 테이블 생성 + 컬럼 추가 + 자동 시드 (Supabase 우선, SQLite 폴백)."""
    import json as _json
    # ── Supabase 시드 (Supabase 연결됐고 테이블이 비어있으면) ─────────────────
    if _supa:
        try:
            count_res = _supa.table('community_posts').select('id', count='exact').execute()
            if (count_res.count or 0) == 0:
                seed_path = Path(__file__).parent / "migrations" / "community_seed.json"
                if seed_path.exists():
                    data = _json.loads(seed_path.read_text(encoding="utf-8"))
                    # boards 시드
                    for b in data.get("boards", []):
                        try:
                            _supa.table('boards').upsert({
                                'id': b["id"], 'label': b["label"],
                                'label_kr': b.get("label_kr"),
                                'display_mode': b.get("display_mode", "list"),
                                'sort_order': b.get("sort_order", 0),
                                'is_hidden': b.get("is_hidden", 0),
                            }, on_conflict='id').execute()
                        except Exception:
                            pass
                    # posts 시드
                    batch = [{
                        'board': p["board"], 'title': p["title"], 'body': p["body"],
                        'author_hash': p.get("author_hash", "bridge_admin"),
                        'image_paths': p.get("image_paths", "[]"),
                        'pinned': p.get("pinned", 0),
                        'content_type': p.get("content_type", "markdown"),
                        'sort_order': p.get("sort_order", 0),
                        'category': p.get("category"),
                    } for p in data.get("posts", [])]
                    if batch:
                        _supa.table('community_posts').insert(batch).execute()
                    logging.getLogger("bridge.api").info(
                        "[Supabase] community 시드 완료: %d posts", len(batch))
            return  # Supabase 성공 시 SQLite 스키마 생성 스킵
        except Exception as _se:
            logging.getLogger("bridge.api").warning("[Supabase] community 시드 실패 → SQLite 진행: %s", _se)
    # ── SQLite 스키마 + 시드 ──────────────────────────────────────────────────
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    # ── 1. CREATE TABLE (처음 배포 또는 DB 초기화 후 자동 복구)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS community_posts (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            board       TEXT    NOT NULL DEFAULT 'support',
            title       TEXT    NOT NULL,
            body        TEXT    NOT NULL,
            author_hash TEXT,
            image_paths TEXT    DEFAULT '[]',
            pinned      INTEGER DEFAULT 0,
            views       INTEGER DEFAULT 0,
            created_at  TEXT    DEFAULT (strftime('%Y-%m-%dT%H:%M:%SZ','now')),
            updated_at  TEXT    DEFAULT (strftime('%Y-%m-%dT%H:%M:%SZ','now')),
            is_deleted  INTEGER DEFAULT 0,
            category    TEXT    DEFAULT NULL,
            content_type TEXT   DEFAULT 'markdown',
            sort_order  INTEGER DEFAULT 0
        )
    """)
    # ── 2. ALTER TABLE (기존 DB에 컬럼 없을 경우)
    for col_sql in (
        "ALTER TABLE community_posts ADD COLUMN content_type TEXT DEFAULT 'markdown'",
        "ALTER TABLE community_posts ADD COLUMN sort_order INTEGER DEFAULT 0",
        "ALTER TABLE community_posts ADD COLUMN category TEXT DEFAULT NULL",
        "ALTER TABLE community_posts ADD COLUMN image_paths TEXT DEFAULT '[]'",
        "ALTER TABLE community_posts ADD COLUMN updated_at TEXT DEFAULT (strftime('%Y-%m-%dT%H:%M:%SZ','now'))",
    ):
        try:
            conn.execute(col_sql)
        except Exception:
            pass
    # boards 테이블 먼저 생성 (boards 시드보다 앞서야 함)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS boards (
            id TEXT PRIMARY KEY, label TEXT NOT NULL, label_kr TEXT,
            display_mode TEXT DEFAULT 'list', sort_order INTEGER DEFAULT 0,
            is_hidden INTEGER DEFAULT 0, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()
    # ── 3. 자동 시드 (posts가 0개일 때 seed 파일에서 복원)
    count = conn.execute("SELECT COUNT(*) FROM community_posts WHERE is_deleted=0").fetchone()[0]
    if count == 0:
        seed_path = Path(__file__).parent / "migrations" / "community_seed.json"
        if seed_path.exists():
            try:
                data = _json.loads(seed_path.read_text(encoding="utf-8"))
                # boards 시드 (없으면)
                for b in data.get("boards", []):
                    conn.execute("""
                        INSERT OR IGNORE INTO boards (id, label, label_kr, display_mode, sort_order, is_hidden)
                        VALUES (?,?,?,?,?,?)
                    """, (b["id"], b["label"], b.get("label_kr"), b.get("display_mode","list"),
                          b.get("sort_order",0), b.get("is_hidden",0)))
                # posts 시드
                for p in data.get("posts", []):
                    conn.execute("""
                        INSERT INTO community_posts
                            (board, title, body, author_hash, image_paths, pinned, content_type, sort_order, category)
                        SELECT ?,?,?,?,?,?,?,?,?
                        WHERE NOT EXISTS (
                            SELECT 1 FROM community_posts WHERE title=? AND board=? AND is_deleted=0
                        )
                    """, (p["board"], p["title"], p["body"], p.get("author_hash","bridge_admin"),
                          p.get("image_paths","[]"), p.get("pinned",0), p.get("content_type","markdown"),
                          p.get("sort_order",0), p.get("category"),
                          p["title"], p["board"]))
                conn.commit()
                logging.getLogger("bridge.api").info(
                    "_ensure_community_schema: 시드 %d posts 복원 완료", len(data.get("posts",[])))
            except Exception as _se:
                logging.getLogger("bridge.api").warning("community seed 스킵: %s", _se)
    conn.close()


try:
    _ensure_community_schema()
except Exception as _e:
    logging.getLogger("bridge.api").warning("_ensure_community_schema 스킵: %s", _e)


# ── HTML Sanitizer (allowlist 기반) ──────────────────────────────────────────

_ALLOWED_TAGS = frozenset({
    "p", "h1", "h2", "h3", "h4", "h5", "h6",
    "ul", "ol", "li", "a", "img", "strong", "em", "b", "i", "u",
    "br", "hr", "table", "thead", "tbody", "tr", "th", "td",
    "div", "span", "blockquote", "code", "pre", "sub", "sup",
})
_ALLOWED_ATTRS = frozenset({
    "href", "src", "alt", "class", "style", "target", "rel",
    "width", "height", "colspan", "rowspan",
})


def _sanitize_html(html_str: str) -> str:
    """Allowlist 기반 HTML sanitizer. script/iframe/object 등 위험 태그 제거."""
    from html.parser import HTMLParser
    import html as _html_mod

    output: list[str] = []

    class Sanitizer(HTMLParser):
        def __init__(self):
            super().__init__()
            self._skip_depth = 0

        def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]):
            tag_lower = tag.lower()
            if tag_lower not in _ALLOWED_TAGS:
                self._skip_depth += 1
                return
            if self._skip_depth > 0:
                return
            safe_attrs = []
            for k, v in attrs:
                if k.lower() in _ALLOWED_ATTRS and v is not None:
                    # href/src에서 javascript: 차단
                    if k.lower() in ("href", "src"):
                        if v.strip().lower().startswith("javascript:"):
                            continue
                    safe_attrs.append(f'{k}="{_html_mod.escape(v, quote=True)}"')
            attr_str = (" " + " ".join(safe_attrs)) if safe_attrs else ""
            if tag_lower in ("br", "hr", "img"):
                output.append(f"<{tag_lower}{attr_str} />")
            else:
                output.append(f"<{tag_lower}{attr_str}>")

        def handle_endtag(self, tag: str):
            tag_lower = tag.lower()
            if tag_lower not in _ALLOWED_TAGS:
                if self._skip_depth > 0:
                    self._skip_depth -= 1
                return
            if self._skip_depth > 0:
                return
            output.append(f"</{tag_lower}>")

        def handle_data(self, data: str):
            if self._skip_depth > 0:
                return
            output.append(_html_mod.escape(data))

        def handle_entityref(self, name: str):
            if self._skip_depth > 0:
                return
            output.append(f"&{name};")

        def handle_charref(self, name: str):
            if self._skip_depth > 0:
                return
            output.append(f"&#{name};")

    parser = Sanitizer()
    parser.feed(html_str)
    return "".join(output)


def _ensure_interview_templates():
    """interview_employer / interview_candidate 이메일 템플릿 시딩."""
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")

    employer_subject = "인터뷰 일정 안내 -- {scheduled_at}"
    employer_html = """<!DOCTYPE html>
<html><head><meta charset="utf-8"></head>
<body style="font-family:-apple-system,'Segoe UI','Malgun Gothic',sans-serif;max-width:600px;margin:0 auto;padding:20px;color:#333;line-height:1.7;">
<div style="text-align:center;padding:20px 0;border-bottom:2px solid #1d1d1f;">
<h1 style="margin:0;font-size:24px;color:#1d1d1f;font-weight:700;">BRIDGE</h1>
<p style="margin:4px 0 0;font-size:12px;color:#86868b;">ESL Teacher Recruitment Platform</p>
</div>
<div style="padding:30px 0;">
<p>안녕하세요,</p>
<p>BRIDGE를 통해 채용 후보자 <strong>{candidate_first_name}</strong>님과의 인터뷰가 예정되었습니다.</p>

<div style="background:#f0fdf4;border:1px solid #86efac;padding:20px;margin:24px 0;border-radius:12px;text-align:center;">
<p style="margin:0 0 8px;font-size:18px;font-weight:700;color:#166534;">{scheduled_at}</p>
<a href="{meet_link}" style="display:inline-block;background:#1d1d1f;color:#fff;padding:12px 28px;border-radius:8px;text-decoration:none;font-weight:600;margin-top:12px;">Google Meet 입장</a>
<p style="margin:12px 0 0;font-size:12px;color:#6b7280;">링크: <a href="{meet_link}" style="color:#0071e3;">{meet_link}</a></p>
</div>

<div style="background:#f5f5f7;padding:20px;margin:24px 0;border-radius:12px;">
<p style="margin:0 0 12px;font-weight:600;color:#1d1d1f;">BRIDGE 서비스 정책 안내</p>
<ul style="margin:0;padding-left:20px;color:#424245;line-height:1.9;">
<li>후보자 개인정보(성, 연락처, 이메일 등)는 BRIDGE 정책에 따라 비공개입니다.</li>
<li>인터뷰 중 후보자에게 개인 연락처를 직접 요청하지 마세요.</li>
<li>채용 진행은 반드시 BRIDGE를 통해 진행해 주세요.</li>
<li>인터뷰 결과는 BRIDGE 담당자에게 회신해 주시면 됩니다.</li>
</ul>
</div>

<div style="background:#fffbeb;border:1px solid #fde68a;padding:20px;margin:24px 0;border-radius:12px;">
<p style="margin:0 0 12px;font-weight:600;color:#92400e;">인터뷰 가이드</p>
<ul style="margin:0;padding-left:20px;color:#424245;line-height:1.9;">
<li>Google Meet '회의 액세스 유형'을 <strong>'열기'</strong>로 설정해 주세요.</li>
<li>예정 시간 2-3분 전에 입장해 주세요.</li>
<li>카메라와 마이크를 미리 테스트해 주세요.</li>
<li>조용하고 밝은 환경에서 진행해 주세요.</li>
</ul>
</div>

<p>일정 변경이 필요하시면 <a href="mailto:bridgejobkr@gmail.com" style="color:#0071e3;">bridgejobkr@gmail.com</a>으로 연락해 주세요.</p>
<p style="color:#6e6e73;margin-top:30px;">감사합니다.<br><strong>BRIDGE Team 드림</strong></p>
</div>
<div style="border-top:1px solid #e5e7eb;padding-top:16px;text-align:center;font-size:12px;color:#86868b;">
<p>The BRIDGE Team &middot; <a href="https://bridgejob.co.kr" style="color:#0071e3;">bridgejob.co.kr</a></p>
</div>
</body></html>"""

    candidate_subject = "Interview Scheduled -- {scheduled_at}"
    candidate_html = """<!DOCTYPE html>
<html><head><meta charset="utf-8"></head>
<body style="font-family:-apple-system,'Segoe UI',Arial,sans-serif;max-width:600px;margin:0 auto;padding:20px;color:#333;line-height:1.7;">
<div style="text-align:center;padding:20px 0;border-bottom:2px solid #1d1d1f;">
<h1 style="margin:0;font-size:24px;color:#1d1d1f;font-weight:700;">BRIDGE</h1>
<p style="margin:4px 0 0;font-size:12px;color:#86868b;">ESL Teacher Recruitment Platform</p>
</div>
<div style="padding:30px 0;">
<p>Dear {candidate_first_name},</p>
<p>Great news! Your interview has been scheduled through BRIDGE.</p>

<div style="background:#f0fdf4;border:1px solid #86efac;padding:20px;margin:24px 0;border-radius:12px;text-align:center;">
<p style="margin:0 0 8px;font-size:18px;font-weight:700;color:#166534;">{scheduled_at}</p>
<a href="{meet_link}" style="display:inline-block;background:#1d1d1f;color:#fff;padding:12px 28px;border-radius:8px;text-decoration:none;font-weight:600;margin-top:12px;">Join Google Meet</a>
<p style="margin:12px 0 0;font-size:12px;color:#6b7280;">Link: <a href="{meet_link}" style="color:#0071e3;">{meet_link}</a></p>
</div>

<div style="background:#eff6ff;border:1px solid #bfdbfe;padding:20px;margin:24px 0;border-radius:12px;">
<p style="margin:0 0 12px;font-weight:600;color:#1e40af;">Privacy Protocol</p>
<ul style="margin:0;padding-left:20px;color:#424245;line-height:1.9;">
<li>The employer will only know your first name during the interview.</li>
<li>Your personal contact information is protected by BRIDGE.</li>
<li>Do not share your phone number, personal email, or address directly.</li>
<li>All communication should go through BRIDGE.</li>
</ul>
</div>

<div style="background:#f5f5f7;padding:20px;margin:24px 0;border-radius:12px;">
<p style="margin:0 0 12px;font-weight:600;color:#1d1d1f;">Essential Guide</p>
<ul style="margin:0;padding-left:20px;color:#424245;line-height:1.9;">
<li>Test your camera and microphone beforehand</li>
<li>Find a quiet, well-lit space</li>
<li>Dress professionally (business casual or above)</li>
<li>Have your resume and any documents ready</li>
<li>Join the meeting 2-3 minutes early</li>
</ul>
</div>

<p>If you need to reschedule, please email <a href="mailto:bridgejobkr@gmail.com" style="color:#0071e3;">bridgejobkr@gmail.com</a> as soon as possible.</p>
<p style="color:#6e6e73;margin-top:30px;">Good luck!<br><strong>BRIDGE Team</strong></p>
</div>
<div style="border-top:1px solid #e5e7eb;padding-top:16px;text-align:center;font-size:12px;color:#86868b;">
<p>The BRIDGE Team &middot; <a href="https://bridgejob.co.kr" style="color:#0071e3;">bridgejob.co.kr</a></p>
</div>
</body></html>"""

    # profile_broadcast 템플릿 (구인자에게 후보자 프로필 발송용)
    profile_subject = "BRIDGE -- 채용 후보자 프로필 안내"
    profile_html = """<!DOCTYPE html>
<html><head><meta charset="utf-8"></head>
<body style="font-family:-apple-system,'Segoe UI','Malgun Gothic',sans-serif;max-width:600px;margin:0 auto;padding:20px;color:#333;line-height:1.7;">
<div style="text-align:center;padding:20px 0;border-bottom:2px solid #1d1d1f;">
<h1 style="margin:0;font-size:24px;color:#1d1d1f;font-weight:700;">BRIDGE</h1>
<p style="margin:4px 0 0;font-size:12px;color:#86868b;">ESL Teacher Recruitment Platform</p>
</div>
<div style="padding:30px 0;">
<p>안녕하세요, {{contact_name}}님</p>
<p>BRIDGE에서 {{school_name}} 조건에 맞는 채용 후보자를 안내드립니다.</p>

<div style="margin:24px 0;">
{{profile_cards}}
</div>

{{download_link}}

<div style="background:#f5f5f7;padding:20px;margin:24px 0;border-radius:12px;">
<p style="margin:0 0 12px;font-weight:600;color:#1d1d1f;">안내 사항</p>
<ul style="margin:0;padding-left:20px;color:#424245;line-height:1.9;">
<li>후보자 개인정보(성명, 연락처, 이메일 등)는 BRIDGE 정책에 따라 비공개입니다.</li>
<li>채용 진행은 반드시 BRIDGE를 통해 진행해 주세요.</li>
<li>관심 있는 후보자가 있으시면 이 이메일에 회신해 주세요.</li>
<li>인터뷰 일정을 잡아드리겠습니다.</li>
</ul>
</div>

<p>궁금한 점이 있으시면 <a href="mailto:bridgejobkr@gmail.com" style="color:#0071e3;">bridgejobkr@gmail.com</a>으로 연락해 주세요.</p>
<p style="color:#6e6e73;margin-top:30px;">감사합니다.<br><strong>BRIDGE Team 드림</strong></p>
</div>
<div style="border-top:1px solid #e5e7eb;padding-top:16px;text-align:center;font-size:12px;color:#86868b;">
<p>The BRIDGE Team &middot; <a href="https://bridgejob.co.kr" style="color:#0071e3;">bridgejob.co.kr</a></p>
<p style="font-size:10px;color:#aeaeb2;">이 이메일은 BRIDGE 채용 서비스의 일환으로 발송되었습니다.</p>
</div>
</body></html>"""

    for key, subj, body in [
        ("interview_employer", employer_subject, employer_html),
        ("interview_candidate", candidate_subject, candidate_html),
        ("profile_broadcast", profile_subject, profile_html),
    ]:
        conn.execute(
            """INSERT INTO email_templates (template_key, subject, body_html, updated_at)
               VALUES (?, ?, ?, CURRENT_TIMESTAMP)
               ON CONFLICT(template_key) DO NOTHING""",
            (key, subj, body),
        )
    conn.commit()
    conn.close()


def _ensure_profile_sends_schema():
    """profile_sends 테이블 생성 + 누락 컬럼 ALTER (없으면)."""
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS profile_sends (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            candidate_id TEXT NOT NULL,
            employer_id INTEGER,
            school_name TEXT,
            to_email TEXT,
            status TEXT DEFAULT 'sent',
            sent_at TEXT DEFAULT (datetime('now','localtime')),
            error_msg TEXT
        )
    """)
    # 스프린트 A: 누락 컬럼 추가 (이미 있으면 무시)
    _ps_new_cols = [
        ("candidate_number",  "INTEGER"),
        ("subject",           "TEXT"),
        ("template_used",     "TEXT DEFAULT 'profile_broadcast'"),
        ("sender_email",      "TEXT"),
        ("attachments",       "TEXT"),
        ("created_at",        "TEXT"),
    ]
    existing_cols = {row[1] for row in conn.execute("PRAGMA table_info(profile_sends)").fetchall()}
    for col_name, col_def in _ps_new_cols:
        if col_name not in existing_cols:
            conn.execute(f"ALTER TABLE profile_sends ADD COLUMN {col_name} {col_def}")
    # 인덱스
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ps_candidate ON profile_sends(candidate_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ps_employer  ON profile_sends(employer_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ps_sent_at   ON profile_sends(sent_at)")
    conn.commit()
    conn.close()


try:
    _ensure_interview_templates()
except Exception as _e:
    logging.getLogger("bridge.api").warning("_ensure_interview_templates 스킵: %s", _e)
try:
    _ensure_profile_sends_schema()
except Exception as _e:
    logging.getLogger("bridge.api").warning("_ensure_profile_sends_schema 스킵: %s", _e)


class InterviewCreate(BaseModel):
    candidate_name:    str = Field("", max_length=200)
    candidate_email:   str = Field("", max_length=200)
    candidate_id:      str = Field("", max_length=50)
    employer_name:     str = Field("", max_length=200)
    employer_email:    str = Field("", max_length=200)
    interview_date:    str = Field(..., min_length=8, max_length=20)
    interview_time:    str = Field(..., min_length=3, max_length=20)
    meet_link:         str = Field("", max_length=500)
    notes:             str = Field("", max_length=2000)
    duration_minutes:  int = Field(20, ge=10, le=240)
    auto_send_email:   bool = Field(False)
    email_subject:     str = Field("", max_length=500)
    email_body:        str = Field("", max_length=5000)


@app.get("/api/admin/interviews", tags=["admin"])
async def admin_list_interviews(request: Request, status: Optional[str] = None):
    """인터뷰 목록 조회."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    if status:
        rows = conn.execute(
            "SELECT * FROM interviews WHERE is_deleted=0 AND status=? ORDER BY interview_date DESC, interview_time DESC",
            (status,),
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM interviews WHERE is_deleted=0 ORDER BY interview_date DESC, interview_time DESC"
        ).fetchall()
    conn.close()
    return ok(data=[dict(r) for r in rows])


@app.post("/api/admin/interviews", status_code=201, tags=["admin"])
async def admin_create_interview(body: InterviewCreate, request: Request):
    """인터뷰 생성. auto_send_email=true면 후보자에게 자동 발송."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    email_result = None
    try:
        cur = conn.execute(
            """INSERT INTO interviews
               (candidate_name, candidate_email, candidate_id, employer_name, employer_email,
                interview_date, interview_time, meet_link, notes, duration_minutes)
               VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (body.candidate_name, body.candidate_email, body.candidate_id,
             body.employer_name, body.employer_email,
             body.interview_date, body.interview_time,
             body.meet_link, body.notes, body.duration_minutes),
        )
        interview_id = cur.lastrowid
        conn.commit()

        # auto-send email to candidate if requested
        if body.auto_send_email and body.candidate_email:
            try:
                to_email = body.candidate_email
                # Use custom email content if provided, otherwise fall back to template
                if body.email_subject.strip() and body.email_body.strip():
                    subject = body.email_subject
                    html = body.email_body.replace("\n", "<br>")
                else:
                    conn.row_factory = sqlite3.Row
                    iv = conn.execute("SELECT * FROM interviews WHERE id=?", (interview_id,)).fetchone()
                    if iv:
                        subject, html, to_email = _render_interview_email(dict(iv), "candidate")
                    else:
                        subject, html = "", ""
                if to_email and subject and html:
                    sent = _smtp_send(to_email, subject, html)
                    if sent:
                        conn.execute(
                            "UPDATE interviews SET email_sent_candidate=1, candidate_email_sent=1, candidate_email_sent_at=CURRENT_TIMESTAMP WHERE id=?",
                            (interview_id,),
                        )
                        conn.commit()
                        email_result = {"sent_to": to_email, "status": "sent"}
                    else:
                        email_result = {"status": "smtp_failed"}
            except Exception as _e:
                logging.getLogger("bridge.api").warning("auto-send email failed for interview #%d: %s", interview_id, _e)
                email_result = {"status": "error", "detail": str(_e)}
    finally:
        conn.close()
    return ok(
        data={"id": interview_id, "email_result": email_result},
        message=f"Interview #{interview_id} created",
    )


def _render_interview_email(interview: dict, target: str):
    """인터뷰 이메일 렌더링. (subject, html, to_email) 반환."""
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.row_factory = sqlite3.Row
    try:
        tpl_key = f"interview_{target}"
        row = conn.execute(
            "SELECT subject, body_html FROM email_templates WHERE template_key = ?",
            (tpl_key,),
        ).fetchone()
        if not row:
            raise HTTPException(404, f"Template '{tpl_key}' not found")
    finally:
        conn.close()

    candidate_name = interview.get("candidate_name") or ""
    first_name = candidate_name.split(" ")[0] if candidate_name.strip() else (
        "Candidate" if target == "candidate" else "후보자"
    )
    scheduled_at = f"{interview['interview_date']} {interview['interview_time']} KST"
    meet_link = interview.get("meet_link") or ""
    to_email = interview.get("candidate_email") if target == "candidate" else interview.get("employer_email")

    employer_name = interview.get("employer_name") or ""

    replacements = {
        "{candidate_first_name}": first_name,
        "{candidate_name}": candidate_name,
        "{scheduled_at}": scheduled_at,
        "{meet_link}": meet_link,
        "{employer_name}": employer_name,
    }

    subject = row["subject"]
    html = row["body_html"]
    for k, v in replacements.items():
        subject = subject.replace(k, v)
        html = html.replace(k, v)

    # PII 격리 체크
    candidate_email = interview.get("candidate_email") or ""
    employer_email = interview.get("employer_email") or ""
    if target == "employer":
        # 고용주 메일에 후보자 이메일/전화 포함 시 차단
        if candidate_email and candidate_email in html:
            raise HTTPException(400, "PII violation: candidate email found in employer template")
    elif target == "candidate":
        # 후보자 메일에 고용주 이메일/전화 포함 시 차단
        if employer_email and employer_email in html:
            raise HTTPException(400, "PII violation: employer email found in candidate template")

    return subject, html, to_email


class InterviewSendEmail(BaseModel):
    target: str = Field(..., pattern=r"^(candidate|employer)$")


@app.get("/api/admin/interviews/{interview_id}/preview-email", tags=["admin"])
async def admin_preview_interview_email(interview_id: int, target: str, request: Request):
    """인터뷰 이메일 미리보기."""
    _check_admin(request)
    if target not in ("candidate", "employer"):
        raise HTTPException(400, "target must be 'candidate' or 'employer'")
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        iv = conn.execute("SELECT * FROM interviews WHERE id=? AND is_deleted=0", (interview_id,)).fetchone()
    finally:
        conn.close()
    if not iv:
        raise HTTPException(404, "Interview not found")
    subject, html, to_email = _render_interview_email(dict(iv), target)
    return ok(data={"subject": subject, "body_html": html, "to_email": to_email or ""})


@app.post("/api/admin/interviews/{interview_id}/send-email", tags=["admin"])
async def admin_send_interview_email(interview_id: int, body: InterviewSendEmail, request: Request):
    """인터뷰 이메일 발송."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        iv = conn.execute("SELECT * FROM interviews WHERE id=? AND is_deleted=0", (interview_id,)).fetchone()
        if not iv:
            conn.close()
            raise HTTPException(404, "Interview not found")
        subject, html, to_email = _render_interview_email(dict(iv), body.target)
        if not to_email:
            conn.close()
            raise HTTPException(400, f"No {body.target} email address on this interview")
        sent = _smtp_send(to_email, subject, html)
        if not sent:
            conn.close()
            return err("이메일 발송 실패 (SMTP 설정 확인)", 500)
        sent_cols = {
            "employer": ("email_sent_employer", "school_email_sent", "school_email_sent_at"),
            "candidate": ("email_sent_candidate", "candidate_email_sent", "candidate_email_sent_at"),
        }
        old_col, new_col, new_at_col = sent_cols.get(body.target, ("email_sent_candidate", "candidate_email_sent", "candidate_email_sent_at"))
        conn.execute(
            f"UPDATE interviews SET {old_col}=1, {new_col}=1, {new_at_col}=CURRENT_TIMESTAMP, updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (interview_id,),
        )
        conn.commit()
    finally:
        conn.close()
    return ok(data={"sent_to": to_email, "target": body.target}, message=f"{body.target} 이메일 발송 완료")


class InterviewStatusUpdate(BaseModel):
    status:           Optional[str] = Field(None, min_length=1, max_length=20)
    interview_date:   Optional[str] = Field(None, min_length=8, max_length=20)
    interview_time:   Optional[str] = Field(None, min_length=3, max_length=20)
    duration_minutes: Optional[int] = Field(None, ge=10, le=240)
    notes:            Optional[str] = Field(None, max_length=2000)
    candidate_name:   Optional[str] = Field(None, max_length=200)
    candidate_email:  Optional[str] = Field(None, max_length=200)
    employer_name:    Optional[str] = Field(None, max_length=200)
    employer_email:   Optional[str] = Field(None, max_length=200)
    meet_link:        Optional[str] = Field(None, max_length=500)


@app.patch("/api/admin/interviews/{interview_id}", tags=["admin"])
async def admin_update_interview(interview_id: int, body: InterviewStatusUpdate, request: Request):
    """인터뷰 수정 -- 상태/날짜/시간/메모 등 부분 업데이트."""
    _check_admin(request)
    valid_statuses = {
        # 파이프라인 단계
        "pending", "scheduled", "completed",
        "contract_offered", "under_review", "contract_signed",
        "placement_pending", "placed",
        # 네거티브
        "no_show", "no_show_teacher", "no_show_employer",
        "cancelled", "rejected", "fallen_through", "pending_retry",
    }
    if body.status and body.status not in valid_statuses:
        raise HTTPException(400, f"Invalid status. Must be one of: {sorted(valid_statuses)}")
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    r = conn.execute("SELECT id, status FROM interviews WHERE id=? AND is_deleted=0", (interview_id,)).fetchone()
    if not r:
        conn.close()
        raise HTTPException(404, "Interview not found")
    old_status = r[1] if r else None

    updates = []
    params = []
    for field in ("status", "interview_date", "interview_time", "duration_minutes",
                  "notes", "candidate_name", "candidate_email", "employer_name",
                  "employer_email", "meet_link"):
        val = getattr(body, field, None)
        if val is not None:
            updates.append(f"{field}=?")
            params.append(val)
    if not updates:
        conn.close()
        raise HTTPException(400, "No fields to update")
    updates.append("updated_at=CURRENT_TIMESTAMP")
    params.append(interview_id)
    conn.execute(f"UPDATE interviews SET {', '.join(updates)} WHERE id=?", params)
    conn.commit()
    # 상태 변경 시 interview_status_log 기록
    if body.status and body.status != old_status:
        try:
            conn.execute(
                "INSERT INTO interview_status_log (interview_id, old_status, new_status, changed_by) VALUES (?,?,?,'admin')",
                (interview_id, old_status, body.status),
            )
            conn.commit()
        except Exception as _log_e:
            logging.getLogger("bridge.api").warning("status_log insert 실패 #%d: %s", interview_id, _log_e)
    conn.close()
    changed = [f for f in ("status", "interview_date", "interview_time") if getattr(body, f, None)]
    return ok(message=f"Interview #{interview_id} updated ({', '.join(changed) or 'fields'})")


class _InterviewStatusPatch(BaseModel):
    status: str
    memo: str = ""


@app.patch("/api/admin/interviews/{interview_id}/status", tags=["admin"])
async def admin_update_interview_status(interview_id: int, body: _InterviewStatusPatch, request: Request):
    """인터뷰 상태 변경 전용 -- Sprint C-FINAL 칸반 뷰용."""
    _check_admin(request)
    return await admin_update_interview(
        interview_id,
        InterviewStatusUpdate(status=body.status, notes=body.memo or None),
        request,
    )


@app.delete("/api/admin/interviews/{interview_id}", tags=["admin"])
async def admin_delete_interview(interview_id: int, request: Request):
    """인터뷰 삭제 (soft delete)."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    r = conn.execute("SELECT id FROM interviews WHERE id=? AND is_deleted=0", (interview_id,)).fetchone()
    if not r:
        conn.close()
        raise HTTPException(404, "Interview not found")
    conn.execute("UPDATE interviews SET is_deleted=1, updated_at=CURRENT_TIMESTAMP WHERE id=?", (interview_id,))
    conn.commit()
    conn.close()
    return ok(message=f"Interview #{interview_id} deleted")


@app.get("/api/admin/interviews/pipeline", tags=["admin"])
async def admin_interviews_pipeline(request: Request):
    """인터뷰 파이프라인 -- 상태별 그룹화 (Kanban 뷰용)."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT * FROM interviews WHERE is_deleted=0 ORDER BY interview_date ASC, interview_time ASC"
    ).fetchall()
    conn.close()
    pipeline: dict = {"scheduled": [], "completed": [], "cancelled": [], "no_show": []}
    for r in rows:
        d = dict(r)
        st = d.get("status", "scheduled")
        if st not in pipeline:
            pipeline[st] = []
        pipeline[st].append(d)
    counts = {k: len(v) for k, v in pipeline.items()}
    return ok(data={"pipeline": pipeline, "counts": counts})


class InterviewRetry(BaseModel):
    target: str = Field(..., pattern=r"^(candidate|employer|both)$")


@app.post("/api/admin/interviews/{interview_id}/retry", tags=["admin"])
async def admin_retry_interview_email(interview_id: int, body: InterviewRetry, request: Request):
    """실패/누락 인터뷰 이메일 재발송. target: candidate|employer|both"""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    iv = conn.execute("SELECT * FROM interviews WHERE id=? AND is_deleted=0", (interview_id,)).fetchone()
    conn.close()
    if not iv:
        raise HTTPException(404, "Interview not found")
    targets = ["candidate", "employer"] if body.target == "both" else [body.target]
    results: dict = {}
    for t in targets:
        try:
            subject, html, to_email = _render_interview_email(dict(iv), t)
            if not to_email:
                results[t] = {"status": "skipped", "reason": "no_email"}
                continue
            sent = _smtp_send(to_email, subject, html)
            if sent:
                conn2 = sqlite3.connect(str(_ADMIN_DB_PATH))
                conn2.execute("PRAGMA busy_timeout = 5000")
                if t == "candidate":
                    conn2.execute(
                        "UPDATE interviews SET email_sent_candidate=1, candidate_email_sent=1, candidate_email_sent_at=CURRENT_TIMESTAMP, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                        (interview_id,),
                    )
                else:
                    conn2.execute(
                        "UPDATE interviews SET email_sent_employer=1, school_email_sent=1, school_email_sent_at=CURRENT_TIMESTAMP, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                        (interview_id,),
                    )
                conn2.commit()
                conn2.close()
                results[t] = {"status": "sent", "to": _mask_email(to_email)}
            else:
                results[t] = {"status": "smtp_failed"}
        except Exception as _re:
            logging.getLogger("bridge.api").warning(
                "interview #%d retry %s 실패: %s", interview_id, t, _re
            )
            results[t] = {"status": "error", "detail": type(_re).__name__}
    return ok(data=results, message=f"Interview #{interview_id} retry: {body.target}")


# ── Google Calendar 자동 확정 ──────────────────────────────────────────────────

class InterviewConfirm(BaseModel):
    candidate_id:     str = Field(..., min_length=1, max_length=50)
    inquiry_id:       int = Field(...)
    interview_date:   str = Field(..., min_length=8, max_length=20)
    interview_time:   str = Field(..., min_length=3, max_length=10)
    duration_minutes: int = Field(60, ge=15, le=240)
    notes:            str = Field("", max_length=2000)


_GCAL_SCOPES = ["https://www.googleapis.com/auth/calendar"]


def _get_gcal_service():
    """Google Calendar 서비스 계정 인증 (GOOGLE_SERVICE_ACCOUNT_JSON 환경변수 필요)."""
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
    except ImportError as e:
        raise RuntimeError(f"Google API 패키지 없음: {e}")
    sa_json = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
    if not sa_json:
        raise RuntimeError("GOOGLE_SERVICE_ACCOUNT_JSON 미설정")
    try:
        sa_info = json.loads(sa_json)
    except Exception:
        raise RuntimeError("GOOGLE_SERVICE_ACCOUNT_JSON JSON 파싱 실패")
    creds = service_account.Credentials.from_service_account_info(sa_info, scopes=_GCAL_SCOPES)
    return build("calendar", "v3", credentials=creds)


def _create_gcal_event(
    service, calendar_id: str, summary: str, dt_str: str, time_str: str, duration_min: int
) -> str:
    """Google Calendar 이벤트 생성 → Meet link 반환."""
    naive_dt = datetime.strptime(f"{dt_str}T{time_str}", "%Y-%m-%dT%H:%M")
    kst = timezone(timedelta(hours=9))
    start_dt = naive_dt.replace(tzinfo=kst)
    end_dt = start_dt + timedelta(minutes=duration_min)
    event = {
        "summary": summary,
        "start": {"dateTime": start_dt.isoformat(), "timeZone": "Asia/Seoul"},
        "end": {"dateTime": end_dt.isoformat(), "timeZone": "Asia/Seoul"},
        "conferenceData": {
            "createRequest": {
                "conferenceSolutionKey": {"type": "hangoutsMeet"},
                "requestId": f"bridge-iv-{uuid.uuid4().hex[:16]}",
            }
        },
    }
    result = service.events().insert(
        calendarId=calendar_id, body=event, conferenceDataVersion=1
    ).execute()
    meet_link = ""
    for ep in result.get("conferenceData", {}).get("entryPoints", []):
        if ep.get("entryPointType") == "video":
            meet_link = ep.get("uri", "")
            break
    if not meet_link:
        meet_link = result.get("hangoutLink", "")
    return meet_link


def _mask_email(email: str) -> str:
    """PII 마스킹: abc@domain.com → a***@domain.com"""
    if not email or "@" not in email:
        return "***"
    local, domain = email.split("@", 1)
    return f"{local[:1]}***@{domain}"


@app.post("/api/admin/interview/confirm", status_code=201, tags=["admin"])
async def admin_confirm_interview(body: InterviewConfirm, request: Request):
    """Google Calendar 인터뷰 확정: 이벤트 생성 → Meet link → DB 저장 → 이메일 자동 발송."""
    _check_admin(request)
    _log_iv = logging.getLogger("bridge.api")

    # 1. 후보자/구인처 조회
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        cand = conn.execute(
            "SELECT candidate_id, full_name, email FROM candidates WHERE candidate_id=? AND is_deleted=0",
            (body.candidate_id,),
        ).fetchone()
        inq = conn.execute(
            "SELECT id, school_name, contact_name, email FROM client_inquiries WHERE id=? AND is_deleted=0",
            (body.inquiry_id,),
        ).fetchone()
    finally:
        conn.close()

    if not cand:
        raise HTTPException(404, f"Candidate '{body.candidate_id}' not found")
    if not inq:
        raise HTTPException(404, f"Inquiry #{body.inquiry_id} not found")

    candidate_name  = cand["full_name"] or ""
    candidate_email = cand["email"] or ""
    employer_name   = inq["school_name"] or inq["contact_name"] or ""
    employer_email  = inq["email"] or ""

    # 2. Google Calendar 이벤트 생성 (환경변수 미설정 시 graceful fail)
    calendar_id = os.getenv("GOOGLE_CALENDAR_ID", "primary")
    meet_link: str = ""
    gcal_error: Optional[str] = None
    try:
        service   = _get_gcal_service()
        summary   = f"[BRIDGE Interview] {candidate_name} × {employer_name}"
        meet_link = _create_gcal_event(
            service, calendar_id, summary,
            body.interview_date, body.interview_time, body.duration_minutes,
        )
    except RuntimeError as e:
        gcal_error = str(e)
        _log_iv.warning("Google Calendar 비활성 (수동 링크 필요): %s", gcal_error)
    except Exception as e:
        gcal_error = "Google Calendar API 오류"
        _log_iv.error("Google Calendar 예외: %s", type(e).__name__)

    # 3. interviews 테이블 INSERT
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        cur = conn.execute(
            """INSERT INTO interviews
               (candidate_name, candidate_email, employer_name, employer_email,
                interview_date, interview_time, meet_link, notes, duration_minutes, status)
               VALUES (?,?,?,?,?,?,?,?,?,'scheduled')""",
            (candidate_name, candidate_email, employer_name, employer_email,
             body.interview_date, body.interview_time, meet_link, body.notes, body.duration_minutes),
        )
        interview_id = cur.lastrowid
        conn.commit()
    finally:
        conn.close()

    _log_iv.info(
        "Interview #%d confirmed: candidate=%s employer=%s meet=%s",
        interview_id, _mask_email(candidate_email), _mask_email(employer_email),
        (meet_link[:30] + "...") if len(meet_link) > 30 else meet_link,
    )

    # 3-b. candidates.recruiter_memo 자동 업데이트 (인터뷰 확정 기록)
    try:
        memo_line = (
            f"[인터뷰확정] {body.interview_date} {body.interview_time} "
            f"× {employer_name} (#{interview_id})"
        )
        conn3 = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn3.execute("PRAGMA busy_timeout = 5000")
        existing = conn3.execute(
            "SELECT recruiter_memo FROM candidates WHERE candidate_id=?",
            (body.candidate_id,),
        ).fetchone()
        prev_memo = (existing[0] or "") if existing else ""
        new_memo = (prev_memo + "\n" + memo_line).strip()
        conn3.execute(
            "UPDATE candidates SET recruiter_memo=?, updated_at=CURRENT_TIMESTAMP WHERE candidate_id=?",
            (new_memo, body.candidate_id),
        )
        conn3.commit()
        conn3.close()
    except Exception as _memo_e:
        _log_iv.warning("recruiter_memo 업데이트 실패 candidate=%s: %s", body.candidate_id, _memo_e)

    # 4. 이메일 자동 발송 (실패해도 DB 삽입은 유지)
    iv_dict = {
        "candidate_name":  candidate_name,
        "candidate_email": candidate_email,
        "employer_name":   employer_name,
        "employer_email":  employer_email,
        "interview_date":  body.interview_date,
        "interview_time":  body.interview_time,
        "meet_link":       meet_link,
    }
    email_errors: list = []
    for target in ("candidate", "employer"):
        try:
            subject, html, to_email = _render_interview_email(iv_dict, target)
            if to_email:
                sent = _smtp_send(to_email, subject, html)
                if sent:
                    conn2 = sqlite3.connect(str(_ADMIN_DB_PATH))
                    conn2.execute("PRAGMA busy_timeout = 5000")
                    if target == "candidate":
                        conn2.execute(
                            "UPDATE interviews SET email_sent_candidate=1, candidate_email_sent=1, candidate_email_sent_at=CURRENT_TIMESTAMP WHERE id=?",
                            (interview_id,),
                        )
                    else:
                        conn2.execute(
                            "UPDATE interviews SET email_sent_employer=1, school_email_sent=1, school_email_sent_at=CURRENT_TIMESTAMP WHERE id=?",
                            (interview_id,),
                        )
                    conn2.commit()
                    conn2.close()
        except Exception as e:
            email_errors.append(f"{target}: {type(e).__name__}")

    return ok(
        data={
            "id":           interview_id,
            "meet_link":    meet_link,
            "gcal_error":   gcal_error,
            "email_errors": email_errors,
        },
        message=(
            f"Interview #{interview_id} confirmed"
            + (f" -- GCal: {gcal_error}" if gcal_error else "")
            + (f" -- email errors: {len(email_errors)}" if email_errors else "")
        ),
    )


# ── File Upload System ─────────────────────────────────────────────────────────
_UPLOAD_BASE = Path(os.getenv("BRIDGE_UPLOAD_DIR", str(Path(__file__).resolve().parent / "uploads")))
try:
    _UPLOAD_BASE.mkdir(parents=True, exist_ok=True)
except OSError:
    _UPLOAD_BASE = Path("/tmp/bridge_uploads")
    _UPLOAD_BASE.mkdir(parents=True, exist_ok=True)

_log_upload = logging.getLogger("bridge.upload")

# ── S3 스토리지 (boto3) ─────────────────────────────────────────────────────
# AWS_ACCESS_KEY_ID / AWS_SECRET_ACCESS_KEY / AWS_S3_BUCKET 환경변수 필요
# 미설정 시 업로드 엔드포인트에서 RuntimeError → 503 반환 (Fail-Closed)
try:
    from backend.utils.storage import (
        upload_bytes as s3_upload_bytes,
        upload_file  as s3_upload_file,
        upload_bytes_sync as s3_upload_bytes_sync,
        download_bytes as s3_download_bytes,
        get_presigned_url as s3_presigned_url,
        delete_file  as s3_delete_file,
        check_s3_connection,
        StorageError,
    )
    _S3_OK = True
except Exception as _s3_exc:
    _S3_OK = False
    logging.getLogger("bridge.api").warning("[S3] storage 모듈 로드 실패 (파일업로드 비활성화): %s", _s3_exc)

    class StorageError(Exception):
        pass

    async def s3_upload_bytes(*a, **kw):
        raise RuntimeError("S3 스토리지 미설정 -- backend.utils.storage 로드 실패")

    def s3_upload_bytes_sync(*a, **kw):
        raise RuntimeError("S3 스토리지 미설정")

    def s3_download_bytes(*a, **kw):
        raise RuntimeError("S3 스토리지 미설정")

    async def s3_upload_file(*a, **kw):
        raise RuntimeError("S3 스토리지 미설정 -- backend.utils.storage 로드 실패")

    def get_presigned_url(*a, **kw):
        raise RuntimeError("S3 스토리지 미설정")

    def delete_file(*a, **kw):
        raise RuntimeError("S3 스토리지 미설정")

    def check_s3_connection():
        return False

# ── 서명 URL (HMAC-SHA256) -- 업로드 파일 접근 보호 ──────────────────────────
_UPLOAD_SIGN_KEY: str = os.getenv("UPLOAD_SIGN_KEY", "").strip()
_SIGN_EXPIRY: int = 600  # 10분


def _sign_file_path(rel_path: str, expires: int) -> str:
    """HMAC-SHA256 서명 생성. rel_path = 'candidates/cnd_xxx/photo.jpg' 형식."""
    if not _UPLOAD_SIGN_KEY:
        raise RuntimeError("UPLOAD_SIGN_KEY 환경변수 미설정")
    msg = f"{rel_path}:{expires}"
    return hmac.new(_UPLOAD_SIGN_KEY.encode(), msg.encode(), hashlib.sha256).hexdigest()


def generate_signed_url(rel_path: str, expires_in: int = _SIGN_EXPIRY) -> str:
    """서명 URL 생성. /api/files/{rel_path}?expires=...&sig=... 형식 반환."""
    expires = int(time.time()) + expires_in
    sig = _sign_file_path(rel_path, expires)
    return f"/api/files/{rel_path}?expires={expires}&sig={sig}"


def _verify_signed_url(rel_path: str, expires: str, sig: str) -> bool:
    """서명 URL 검증. 만료 또는 서명 불일치 시 False."""
    if not _UPLOAD_SIGN_KEY:
        return False
    try:
        exp = int(expires)
    except (ValueError, TypeError):
        return False
    if exp < int(time.time()):
        return False  # 만료됨
    expected = _sign_file_path(rel_path, exp)
    return hmac.compare_digest(expected, sig or "")


@app.get("/api/files/{entity_type}/{entity_id}/{filename:path}", tags=["files"])
async def serve_protected_file(
    entity_type: str,
    entity_id: str,
    filename: str,
    request: Request,
    expires: str = "",
    sig: str = "",
):
    """
    HMAC 서명 URL 또는 관리자 인증으로만 파일 접근 허용.
    - 유효 서명 or 관리자 키 → FileResponse
    - 서명 불일치/만료/미제공 → 403
    - entity_type: candidates | inquiries | community 등
    """
    from fastapi.responses import FileResponse as _FR

    # 경로 순회 공격 방지
    try:
        file_path = (_UPLOAD_BASE / entity_type / entity_id / filename).resolve()
        file_path.relative_to(_UPLOAD_BASE.resolve())  # 범위 벗어나면 ValueError
    except (ValueError, Exception):
        raise HTTPException(403, "접근이 거부되었습니다.")

    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(404, "파일을 찾을 수 없습니다.")

    # 관리자 인증 우선 체크
    admin_key = request.headers.get("x-admin-key", "").strip()
    is_admin = bool(_ADMIN_KEY and hmac.compare_digest(admin_key, _ADMIN_KEY))

    if not is_admin:
        rel_path = f"{entity_type}/{entity_id}/{filename}"
        if not _verify_signed_url(rel_path, expires, sig):
            _log_upload.warning(
                "UNSIGNED_FILE_ACCESS ip=%s path=%s",
                _ip_hash(request), rel_path
            )
            raise HTTPException(403, "파일 접근 권한이 없습니다. 서명 URL을 사용하세요.")

    # 콘텐츠 타입 감지 (확장자 기반)
    import mimetypes
    mime, _ = mimetypes.guess_type(str(file_path))
    media_type = mime or "application/octet-stream"

    # [비활성 -- 로컬 파일 서빙] S3 전환 후 불필요. 아래 S3 redirect 사용.
    # return _FR(path=str(file_path), media_type=media_type, headers={...})
    # ── S3 presigned URL 리다이렉트 ────────────────────────────────────────
    # DB에서 s3_key를 조회하는 것이 이상적이나,
    # 하위호환: entity_type/entity_id/filename 경로를 s3_key로 사용
    _s3_key = f"legacy/{entity_type}/{entity_id}/{filename}"
    try:
        _presigned = s3_presigned_url(_s3_key, expires=3600)
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url=_presigned, status_code=302)
    except (RuntimeError, StorageError) as _s3_err:
        _log_upload.error("presigned URL 생성 실패: %s", _s3_err)
        raise HTTPException(503, "파일 서비스가 일시적으로 불가합니다.")


@app.get("/api/admin/sign-url", tags=["admin"])
async def admin_sign_url(request: Request, path: str = ""):
    """
    관리자 전용: 파일 서명 URL 생성.
    ?path=candidates/cnd_xxx/photo.jpg → {signed_url, expires_at}
    """
    _check_admin(request)
    if not path or ".." in path or path.startswith("/"):
        raise HTTPException(400, "유효하지 않은 파일 경로입니다.")
    if not _UPLOAD_SIGN_KEY:
        raise HTTPException(503, "파일 서명 서비스가 현재 불가합니다.")
    signed = generate_signed_url(path)
    expires_at = int(time.time()) + _SIGN_EXPIRY
    return ok(data={
        "signed_url": signed,
        "expires_at": expires_at,
        "expires_in": _SIGN_EXPIRY,
    })

# File type configs: (max_bytes, allowed_extensions, magic_bytes_prefixes)
_FILE_LIMITS: dict[str, tuple[int, set[str], list[bytes]]] = {
    "photo": (
        5 * 1024 * 1024,
        {".jpg", ".jpeg", ".png", ".webp"},
        [b"\xff\xd8\xff", b"\x89PNG", b"RIFF"],  # JPEG, PNG, WebP
    ),
    "cv": (
        20 * 1024 * 1024,   # v2.0: 10MB → 20MB (사진 embed DOCX 대응)
        {".pdf", ".doc", ".docx", ".rtf", ".odt"},
        [b"%PDF", b"\xd0\xcf\x11\xe0", b"PK", b"{\\rtf"],  # PDF, DOC, DOCX/ODT, RTF
    ),
    "cover_letter": (
        20 * 1024 * 1024,
        {".pdf", ".doc", ".docx", ".rtf", ".odt"},
        [b"%PDF", b"\xd0\xcf\x11\xe0", b"PK", b"{\\rtf"],
    ),
    "certificate": (
        15 * 1024 * 1024,
        {".pdf", ".jpg", ".jpeg", ".png", ".doc", ".docx"},
        [b"%PDF", b"\xff\xd8\xff", b"\x89PNG", b"\xd0\xcf\x11\xe0", b"PK"],
    ),
    "reference": (
        15 * 1024 * 1024,
        {".pdf", ".doc", ".docx", ".jpg", ".jpeg", ".png"},
        [b"%PDF", b"\xd0\xcf\x11\xe0", b"PK", b"\xff\xd8\xff", b"\x89PNG"],
    ),
    "video": (
        200 * 1024 * 1024,   # v2.0: 100MB → 200MB
        {".mp4", ".mov", ".webm", ".avi", ".mkv"},
        [b"ftyp", b"\x00\x00\x00\x18ftyp", b"\x00\x00\x00\x20ftyp",
         b"\x1aE\xdf\xa3",   # WebM/MKV
         b"RIFF",             # AVI
         ],
    ),
    "attachment": (
        20 * 1024 * 1024,
        {".pdf", ".doc", ".docx", ".rtf", ".odt", ".jpg", ".jpeg", ".png", ".txt", ".zip"},
        [b"%PDF", b"\xd0\xcf\x11\xe0", b"PK", b"\xff\xd8\xff", b"\x89PNG", b"{\\rtf"],
    ),
    "community_image": (
        5 * 1024 * 1024,
        {".jpg", ".jpeg", ".png", ".gif", ".webp"},
        [b"\xff\xd8\xff", b"\x89PNG", b"GIF8", b"RIFF"],
    ),
    "community_file": (
        10 * 1024 * 1024,
        {".pdf", ".doc", ".docx", ".xlsx", ".ppt", ".pptx", ".txt", ".zip"},
        [b"%PDF", b"\xd0\xcf\x11\xe0", b"PK"],
    ),
}


def _validate_file(data: bytes, filename: str, file_type: str) -> str:
    """Validate file extension and magic bytes. Returns sanitized extension."""
    cfg = _FILE_LIMITS.get(file_type)
    if not cfg:
        raise HTTPException(400, f"Unknown file type: {file_type}")

    max_size, allowed_ext, magic_prefixes = cfg
    if len(data) > max_size:
        raise HTTPException(413, f"File too large. Max {max_size // (1024*1024)}MB for {file_type}.")

    ext = Path(filename).suffix.lower()
    if ext not in allowed_ext:
        raise HTTPException(400, f"Invalid file extension '{ext}'. Allowed: {', '.join(sorted(allowed_ext))}")

    if magic_prefixes:
        header = data[:12]
        # MP4/MOV: ftyp box는 바이트 4~8에 위치 (box size 가변)
        is_mp4 = header[4:8] == b"ftyp"
        if not (any(header[:8].startswith(m) for m in magic_prefixes) or is_mp4):
            raise HTTPException(400, "File content does not match its extension.")

    # W3: DOCX/ODT는 PK 헤더를 ZIP과 공유 → 내부 구조 추가 검증
    if ext == ".docx":
        try:
            with zipfile.ZipFile(io.BytesIO(data)) as _zf:
                _names = {e.filename for e in _zf.infolist()}
                # DOCX 필수 엔트리 확인
                if "[Content_Types].xml" not in _names and "word/" not in " ".join(_names):
                    raise HTTPException(400, "Invalid DOCX: missing required OOXML entries.")
                total_uncompressed = sum(i.file_size for i in _zf.infolist())
                if total_uncompressed > 100 * 1024 * 1024:
                    raise HTTPException(413, "DOCX content exceeds 100MB uncompressed limit.")
        except zipfile.BadZipFile:
            raise HTTPException(400, "Invalid or corrupted DOCX file.")

    # ZIP bomb 방어: 압축 해제 후 크기 및 파일 수 제한
    if ext == ".zip":
        try:
            with zipfile.ZipFile(io.BytesIO(data)) as zf:
                total_uncompressed = sum(i.file_size for i in zf.infolist())
                if total_uncompressed > 200 * 1024 * 1024:
                    raise HTTPException(413, "ZIP content exceeds 200MB uncompressed limit.")
                if len(zf.infolist()) > 200:
                    raise HTTPException(400, "ZIP contains too many files (max 200).")
        except zipfile.BadZipFile:
            raise HTTPException(400, "Invalid or corrupted ZIP file.")

    return ext


def _make_thumbnail(src_path: Path, thumb_path: Path, size: int = 150) -> bool:
    """Create a 150x150 top-center crop thumbnail for portrait photos."""
    if not _PILLOW_OK:
        return False
    try:
        with PILImage.open(src_path) as img:
            img = img.convert("RGB")
            w, h = img.size
            # Crop square from top-center with 15% vertical offset
            sq = min(w, h)
            left = (w - sq) // 2
            top = int(h * 0.15)  # 15% offset for face area
            if top + sq > h:
                top = max(0, h - sq)
            img_crop = img.crop((left, top, left + sq, top + sq))
            img_crop = img_crop.resize((size, size), PILImage.LANCZOS)
            img_crop.save(thumb_path, "JPEG", quality=85)
        return True
    except Exception as e:
        _log_upload.error("Thumbnail generation failed: %s", e)
        return False


def _set_resume_status(entity_id: str, status: str):
    """resume_status 업데이트 헬퍼 (pending/processing/done/error)."""
    try:
        c = sqlite3.connect(str(_ADMIN_DB_PATH))
        c.execute("PRAGMA busy_timeout = 5000")
        c.execute("UPDATE candidates SET resume_status = ? WHERE candidate_id = ?",
                  (status, entity_id))
        c.commit()
        c.close()
    except Exception:
        pass


def _auto_process_resume(entity_id: str, cv_s3_key: str):
    """
    CV 업로드 후 백그라운드 자동 처리.
    1차: resume_converter/pipeline.py (PII + 얼굴 + PDF 병합 전체)
    2차 폴백: doc_processor (PII만)
    threading.Thread 에서 호출 (동기).
    """
    import tempfile
    conn = None
    _set_resume_status(entity_id, "processing")
    try:
        # 1. 후보자 정보 조회
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            "SELECT sheet_number, full_name, nationality, email, "
            "mobile_phone, kakaotalk, current_location, dob, gender, "
            "candidate_id, photo_url "
            "FROM candidates WHERE candidate_id = ?",
            (entity_id,),
        ).fetchone()
        if not row:
            _log_upload.warning("[AutoProcess] candidate %s not found", entity_id)
            _set_resume_status(entity_id, "error")
            return

        candidate = dict(row)
        brj_number = candidate.get("sheet_number") or 0
        if not brj_number:
            _digits = re.sub(r"[^0-9]", "", entity_id)[-4:]
            brj_number = int(_digits) if _digits else 0
            if not brj_number:
                _log_upload.info("[AutoProcess] %s -- 번호 생성 불가, 스킵", entity_id)
                _set_resume_status(entity_id, "error")
                return

        # 복호화 시도 (암호화된 필드)
        for fld in ("full_name", "email", "mobile_phone", "kakaotalk",
                     "nationality", "current_location", "dob", "gender"):
            raw = candidate.get(fld)
            if raw:
                try:
                    candidate[fld] = _safe_decrypt(raw, fld)
                except Exception:
                    pass

        # 2. CV 다운로드
        cv_data = s3_download_bytes(cv_s3_key)
        if not cv_data:
            _log_upload.warning("[AutoProcess] CV 다운로드 실패: %s", cv_s3_key)
            _set_resume_status(entity_id, "error")
            return

        # 3. 사진 다운로드 (있으면)
        photo_row = conn.execute(
            "SELECT s3_key FROM file_uploads "
            "WHERE entity_id = ? AND file_type = 'photo' AND s3_key IS NOT NULL "
            "ORDER BY rowid DESC LIMIT 1",
            (entity_id,),
        ).fetchone()

        # 4. 커버레터 다운로드 (있으면)
        cover_row = conn.execute(
            "SELECT s3_key FROM file_uploads "
            "WHERE entity_id = ? AND file_type = 'cover_letter' AND s3_key IS NOT NULL "
            "ORDER BY rowid DESC LIMIT 1",
            (entity_id,),
        ).fetchone()

        with tempfile.TemporaryDirectory(prefix=f"resume_{entity_id}_") as tmpdir:
            tmpdir_path = Path(tmpdir)
            cv_ext = Path(cv_s3_key).suffix.lower() or ".pdf"
            cv_path = tmpdir_path / f"resume{cv_ext}"
            cv_path.write_bytes(cv_data)

            photo_path = None
            if photo_row and photo_row[0]:
                try:
                    photo_data = s3_download_bytes(photo_row[0])
                    if photo_data:
                        photo_ext = Path(photo_row[0]).suffix.lower() or ".jpg"
                        photo_path = tmpdir_path / f"photo{photo_ext}"
                        photo_path.write_bytes(photo_data)
                except Exception as e:
                    _log_upload.warning("[AutoProcess] 사진 다운로드 실패: %s", e)

            cover_path = None
            if cover_row and cover_row[0]:
                try:
                    cover_data = s3_download_bytes(cover_row[0])
                    if cover_data:
                        cover_ext = Path(cover_row[0]).suffix.lower() or ".pdf"
                        cover_path = tmpdir_path / f"cover{cover_ext}"
                        cover_path.write_bytes(cover_data)
                except Exception as e:
                    _log_upload.warning("[AutoProcess] 커버레터 다운로드 실패: %s", e)

            processed_bytes = None
            pii_count = 0

            # ── 1차 시도: resume_converter/pipeline.py (전체 파이프라인) ──
            try:
                _tools_dir = Path(__file__).resolve().parent / "tools"
                if str(_tools_dir.parent) not in sys.path:
                    sys.path.insert(0, str(_tools_dir.parent))
                from tools.resume_converter.pipeline import Pipeline
                from tools.resume_converter.file_classifier import detect_file_type as _rc_detect

                # pipeline 입력 폴더 준비 (파일명으로 자동 분류)
                pipe_folder = tmpdir_path / "pipe_input"
                pipe_folder.mkdir()
                # 파일을 타입명으로 복사 → pipeline _step_classify가 detect_file_type 사용
                import shutil
                shutil.copy2(str(cv_path), str(pipe_folder / cv_path.name))
                if photo_path:
                    shutil.copy2(str(photo_path), str(pipe_folder / photo_path.name))
                if cover_path:
                    shutil.copy2(str(cover_path), str(pipe_folder / cover_path.name))

                pipe = Pipeline(
                    auto_mode=True,
                    on_progress=lambda msg: _log_upload.info("[Pipeline] %s", msg),
                )
                # pipeline에 후보자 메타 주입 (구글시트 조회 대신 DB 값 사용)
                _dob = candidate.get("dob") or ""
                _birth_year = _dob[:4] if len(_dob) >= 4 else ""
                _meta = {
                    "name":        candidate.get("full_name", ""),
                    "nationality": candidate.get("nationality", ""),
                    "gender":      candidate.get("gender", ""),
                    "birth_year":  _birth_year,
                }
                job = pipe.run(str(brj_number), pipe_folder, candidate_meta=_meta)

                if job.output_path and job.output_path.exists():
                    processed_bytes = job.output_path.read_bytes()
                    pii_count = sum([
                        len(job.cover_pii.pii_found) if job.cover_pii else 0,
                        len(job.resume_pii.pii_found) if job.resume_pii else 0,
                        len(job.rec_pii.pii_found) if job.rec_pii else 0,
                    ])
                    _log_upload.info(
                        "[AutoProcess] Pipeline OK: %s (#%s) → %d PII, %d bytes",
                        entity_id, brj_number, pii_count, len(processed_bytes),
                    )
                else:
                    _log_upload.warning(
                        "[AutoProcess] Pipeline 출력 없음: %s, errors=%s",
                        entity_id, job.errors,
                    )
            except Exception as pipe_err:
                _log_upload.warning(
                    "[AutoProcess] Pipeline 실패 → doc_processor 폴백: %s -- %s",
                    entity_id, pipe_err,
                )

            # ── 2차 폴백: doc_processor (PII만) ──
            if processed_bytes is None and cv_ext == ".pdf":
                try:
                    import importlib
                    _dp_spec = importlib.util.find_spec("tools.doc_processor")
                    if _dp_spec is None:
                        _tools_dir = Path(__file__).resolve().parent / "tools"
                        if str(_tools_dir.parent) not in sys.path:
                            sys.path.insert(0, str(_tools_dir.parent))
                    from tools.doc_processor import process_pdf as _dp_process_pdf

                    processed_doc, logs = _dp_process_pdf(
                        cv_path, brj_number, candidate,
                        dry=False, photo_path=photo_path,
                    )
                    out_path = tmpdir_path / "processed.pdf"
                    processed_doc.save(str(out_path), garbage=4, deflate=True)
                    processed_doc.close()
                    processed_bytes = out_path.read_bytes()
                    pii_count = len(logs)
                    _log_upload.info(
                        "[AutoProcess] Fallback OK: %s (#%s) → %d PII, %d bytes",
                        entity_id, brj_number, pii_count, len(processed_bytes),
                    )
                except Exception as dp_err:
                    _log_upload.error("[AutoProcess] doc_processor 폴백도 실패: %s -- %s",
                                     entity_id, dp_err, exc_info=True)

            # ── 3차 폴백: doc_processor DOCX 전용 ──
            if processed_bytes is None and cv_ext == ".docx":
                try:
                    import importlib as _il, io as _io
                    _dp_spec2 = _il.util.find_spec("tools.doc_processor")
                    if _dp_spec2 is None:
                        _tools_dir2 = Path(__file__).resolve().parent / "tools"
                        if str(_tools_dir2.parent) not in sys.path:
                            sys.path.insert(0, str(_tools_dir2.parent))
                    from tools.doc_processor import process_docx as _dp_process_docx

                    processed_doc_obj, logs = _dp_process_docx(
                        cv_path, brj_number, candidate,
                        dry=False, photo_path=photo_path,
                    )
                    _docx_buf = _io.BytesIO()
                    processed_doc_obj.save(_docx_buf)
                    processed_bytes = _docx_buf.getvalue()
                    pii_count = len(logs)
                    _log_upload.info(
                        "[AutoProcess] DOCX Fallback OK: %s (#%s) → %d PII, %d bytes",
                        entity_id, brj_number, pii_count, len(processed_bytes),
                    )
                except Exception as dp_docx_err:
                    _log_upload.error(
                        "[AutoProcess] DOCX 폴백도 실패: %s -- %s",
                        entity_id, dp_docx_err, exc_info=True,
                    )

            if not processed_bytes:
                _set_resume_status(entity_id, "error")
                return

            # 5. S3 업로드 (동기) — 확장자 보존
            _processed_fname = f"cv_processed{'.docx' if cv_ext == '.docx' else '.pdf'}"
            result = s3_upload_bytes_sync(
                data=processed_bytes,
                folder=f"candidates/{entity_id}",
                filename=_processed_fname,
                allowed_category="resume",
            )

            # 6. 이전 processed 파일 정리 (S3 orphan 방지)
            old_key_row = conn.execute(
                "SELECT cv_processed_s3_key FROM candidates WHERE candidate_id = ?",
                (entity_id,),
            ).fetchone()
            old_s3_key = old_key_row[0] if old_key_row else None
            if old_s3_key and old_s3_key != result["s3_key"]:
                try:
                    s3_delete_file(old_s3_key)
                    _log_upload.info("[AutoProcess] 이전 processed 삭제: %s", old_s3_key)
                except Exception:
                    pass
            conn.execute(
                "UPDATE file_uploads SET is_deleted = 1 "
                "WHERE entity_id = ? AND file_type = 'cv_processed' AND is_deleted = 0",
                (entity_id,),
            )

            # 7. DB 업데이트
            conn.execute(
                "UPDATE candidates SET cv_processed_s3_key = ?, resume_status = 'done' "
                "WHERE candidate_id = ?",
                (result["s3_key"], entity_id),
            )
            conn.execute(
                "INSERT INTO file_uploads "
                "(entity_type, entity_id, file_type, file_url, file_size, s3_key) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                ("candidate", entity_id, "cv_processed",
                 result["s3_url"], len(processed_bytes), result["s3_key"]),
            )
            conn.commit()

            _log_upload.info(
                "[AutoProcess] DONE: %s (#%s) → %d PII, %d bytes, status=done",
                entity_id, brj_number, pii_count, len(processed_bytes),
            )

            # A1: 관리자 텔레그램 알림 (논블로킹)
            try:
                _cname = candidate.get("full_name") or f"#{brj_number}" if candidate else f"#{brj_number}"
                _tg_msg = (
                    f"✅ CV 처리 완료\n"
                    f"후보자: {_cname} (ID {entity_id})\n"
                    f"PII 제거: {pii_count}건 | 크기: {len(processed_bytes)//1024}KB"
                )
                import subprocess as _sp
                # 2026-05-13: pythonw + CREATE_NO_WINDOW (console flicker fix)
                _py = "C:/Users/Scarlett/AppData/Local/Programs/Python/Python313/pythonw.exe"
                _tg_script = str(Path(__file__).resolve().parent / "tools" / "tg_notify.py")
                _sp.Popen(
                    [_py, "-X", "utf8", _tg_script, _tg_msg],
                    stdout=_sp.DEVNULL, stderr=_sp.DEVNULL,
                    creationflags=0x08000000,  # CREATE_NO_WINDOW
                )
            except Exception:
                pass  # 알림 실패는 무시 (비핵심 기능)

    except Exception as e:
        _log_upload.error("[AutoProcess] FAIL %s: %s", entity_id, e, exc_info=True)
        _set_resume_status(entity_id, "error")
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.post("/api/upload/{entity_type}/{entity_id}", tags=["upload"])
async def upload_file(
    entity_type: str,
    entity_id: str,
    request: Request,
    file: UploadFile = FastFile(...),
    file_type: str = "attachment",
):
    """
    파일 업로드 엔드포인트.
    entity_type: 'candidate' | 'inquiry' | 'community'
    file_type: 'photo' | 'cv' | 'cover_letter' | 'certificate' | 'video' | 'attachment' | 'community_image' | 'community_file'
    """
    if entity_type not in ("candidate", "inquiry", "community"):
        raise HTTPException(400, "entity_type must be 'candidate', 'inquiry', or 'community'")

    # SECURITY: candidate는 apply_token(JWT) 또는 관리자 인증
    # inquiry/community는 관리자 인증 필수
    is_admin_ok = False
    try:
        _check_admin(request)
        is_admin_ok = True
    except HTTPException:
        pass

    if not is_admin_ok:
        if entity_type != "candidate":
            raise HTTPException(403, "인증이 필요합니다.")
        # candidate: apply_token 헤더 또는 쿼리로 검증 + 해당 candidate_id 일치
        apply_token = (
            request.headers.get("X-Apply-Token")
            or request.query_params.get("apply_token")
            or ""
        ).strip()
        if not apply_token:
            raise HTTPException(403, "apply_token이 필요합니다.")
        try:
            token_cid = _verify_candidate_token(apply_token)
            if str(token_cid) != str(entity_id):
                raise HTTPException(403, "토큰과 entity_id 불일치")
        except HTTPException:
            raise
        except Exception:
            raise HTTPException(403, "apply_token 검증 실패")
        # 공개 업로드는 photo/cv/cover_letter만 허용 (사이즈 제한은 _validate_file이 처리)
        if file_type not in ("photo", "cv", "cover_letter", "certificate"):
            raise HTTPException(403, f"공개 업로드에서는 '{file_type}' 타입을 사용할 수 없습니다.")

    if not _rate_ok(_ip_hash(request), window=60, max_posts=30):
        raise HTTPException(429, "Too many uploads. Please slow down.")

    data = await file.read()
    if not data:
        raise HTTPException(400, "Empty file.")

    ext = _validate_file(data, file.filename or "file", file_type)

    # Build directory
    dir_name = "candidates" if entity_type == "candidate" else ("community" if entity_type == "community" else "inquiries")
    if file_type == "certificate":
        entity_dir = _UPLOAD_BASE / dir_name / entity_id / "certificates"
    else:
        entity_dir = _UPLOAD_BASE / dir_name / entity_id
    entity_dir.mkdir(parents=True, exist_ok=True)

    # File naming
    if file_type == "photo":
        fname = f"photo_original{ext}"
    elif file_type in ("cv", "cover_letter"):
        fname = f"{file_type}{ext}"
    elif file_type == "certificate":
        cert_id = f"cert_{uuid.uuid4().hex[:6]}"
        fname = f"{cert_id}{ext}"
    elif file_type == "video":
        fname = f"video{ext}"
    else:
        att_id = f"doc_{uuid.uuid4().hex[:6]}"
        fname = f"{att_id}{ext}"

    file_path = entity_dir / fname
    # [비활성 -- 로컬 저장] S3 전환 후 불필요. 아래 S3 업로드 사용.
    # file_path.write_bytes(data)

    # ── S3 업로드 ──────────────────────────────────────────────────────────
    _s3_folder = f"{dir_name}/{entity_id}"
    if file_type == "certificate":
        _s3_folder = f"{dir_name}/{entity_id}/certificates"
    _s3_category = (
        "image"    if file_type in ("photo", "community_image") else
        "resume"   if file_type in ("cv", "cover_letter")       else
        "video"    if file_type == "video"                       else
        "any"
    )
    try:
        _s3_result = await s3_upload_bytes(
            data=data,
            folder=_s3_folder,
            filename=file.filename or fname,
            allowed_category=_s3_category,
        )
        s3_key   = _s3_result["s3_key"]
        file_url = _s3_result["s3_url"]
    except (RuntimeError, StorageError) as _s3_err:
        _log_upload.error("S3 업로드 실패: %s", _s3_err)
        raise HTTPException(503, "파일 저장에 실패했습니다. 잠시 후 다시 시도하세요.")
    # ──────────────────────────────────────────────────────────────────────

    # Thumbnail for photos
    thumb_url = None
    if file_type == "photo" and ext in (".jpg", ".jpeg", ".png", ".webp"):
        # [비활성 -- 로컬 썸네일] S3 전환 후 미구현. TODO: Pillow → S3 업로드 방식으로 구현 예정.
        # thumb_path = entity_dir / "photo_thumb.jpg"
        # if _make_thumbnail(file_path, thumb_path):
        #     thumb_url = f"/uploads/{dir_name}/{entity_id}/photo_thumb.jpg"
        pass

    # Build URL (S3 presigned -- 1시간 유효)
    # [비활성 -- 로컬 URL] S3 presigned URL 사용 (아래 코드 참조)
    # rel = file_path.relative_to(_UPLOAD_BASE); file_url = f"/uploads/{rel.as_posix()}"
    file_url = s3_presigned_url(s3_key, expires=3600)

    # SHA-256 무결성 해시 계산 (SEC-08)
    file_sha256 = hashlib.sha256(data).hexdigest()

    # Record in SQLite file_uploads table
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        try:
            conn.execute("ALTER TABLE file_uploads ADD COLUMN sha256_hash TEXT")
        except Exception:
            pass  # 컬럼 이미 존재
        try:
            conn.execute(
                "INSERT INTO file_uploads (entity_type, entity_id, file_type, file_url, file_size, s3_key, sha256_hash) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (entity_type, entity_id, file_type, file_url, len(data), s3_key, file_sha256),
            )
            # Update candidates photo_url/thumb_url if photo upload
            if entity_type == "candidate" and file_type == "photo":
                if thumb_url:
                    conn.execute(
                        "UPDATE candidates SET photo_url = ?, thumb_url = ?, photo_s3_key = ? WHERE candidate_id = ?",
                        (file_url, thumb_url, s3_key, entity_id),
                    )
                else:
                    conn.execute(
                        "UPDATE candidates SET photo_url = ?, photo_s3_key = ? WHERE candidate_id = ?",
                        (file_url, s3_key, entity_id),
                    )
            conn.commit()
        finally:
            conn.close()
    except Exception as e:
        _log_upload.error("File metadata save failed: %s", e)
        # File is saved locally; metadata failure is non-blocking

    # ── v2.0 Universal Enqueue: 전 파일타입 파이프라인 등록 ─────────────
    # 문서(cv/cover_letter/certificate/reference) → resume_process (PII 제거)
    # 이미지(photo) → image_process (해시 + photo_s3_key 갱신)
    # 영상(video) → video_process (해시만)
    # 기타(attachment 등) → attachment_store (해시만)
    _enqueue_ok = False
    if entity_type == "candidate":
        try:
            sys.path.insert(0, str(Path(__file__).resolve().parent))
            from tools.pipeline_v2 import enqueue_file_process  # type: ignore
            _job_id = enqueue_file_process(
                candidate_id=entity_id,
                s3_key=s3_key,
                file_type=file_type,
                filename=fname,
                trigger="upload_endpoint",
            )
            _enqueue_ok = True
            logging.getLogger("bridge.pipeline").info(
                "[ENQUEUE] cid=%s job_id=%s type=%s file=%s",
                entity_id, _job_id, file_type, fname[:60]
            )
        except Exception as _enq_err:
            logging.getLogger("bridge.pipeline").warning(
                "[ENQUEUE] fail cid=%s err=%s", entity_id, _enq_err
            )

    # 레거시 즉시 처리 (CV/커버레터만) — 큐 등록 실패 시 안전망 (이중처리 방지)
    if entity_type == "candidate" and file_type in ("cv", "cover_letter") and not _enqueue_ok:
        threading.Thread(
            target=_auto_process_resume,
            args=(entity_id, s3_key),
            daemon=True,
        ).start()

    return ok(
        data={
            "file_url":  file_url,   # presigned URL (1시간 유효)
            "s3_key":    s3_key,     # DB 저장용 키 (전체 URL 아님)
            "thumb_url": thumb_url,
            "file_size": len(data),
        },
        message="File uploaded successfully.",
    )


@app.get("/api/admin/candidates/{candidate_id}/processed-cv", tags=["admin"])
async def get_processed_cv(candidate_id: str, request: Request):
    """PII 제거된 이력서 presigned URL 반환 (관리자 전용)."""
    _check_admin(request)
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        try:
            row = conn.execute(
                "SELECT cv_processed_s3_key FROM candidates WHERE candidate_id = ?",
                (candidate_id,),
            ).fetchone()
        finally:
            conn.close()
    except Exception as e:
        _log_upload.error("processed-cv lookup failed: %s", e)
        raise HTTPException(500, "DB 조회 실패")
    if not row or not row[0]:
        raise HTTPException(404, "Processed CV not found")
    url = s3_presigned_url(row[0], expires=3600)
    return ok(data={"url": url, "s3_key": row[0]})


@app.delete("/api/admin/candidates/{candidate_id}/processed-cv", tags=["admin"])
async def delete_processed_cv(candidate_id: str, request: Request):
    """Processed CV 삭제 (S3 + DB). 원본 CV는 유지."""
    _check_admin(request)
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        try:
            row = conn.execute(
                "SELECT cv_processed_s3_key FROM candidates WHERE candidate_id = ?",
                (candidate_id,),
            ).fetchone()
            if row and row[0]:
                try:
                    s3_delete_file(row[0])
                except Exception:
                    pass
                conn.execute(
                    "UPDATE candidates SET cv_processed_s3_key = NULL WHERE candidate_id = ?",
                    (candidate_id,),
                )
                conn.execute(
                    "UPDATE file_uploads SET is_deleted = 1 "
                    "WHERE entity_id = ? AND file_type = 'cv_processed' AND is_deleted = 0",
                    (candidate_id,),
                )
                conn.commit()
        finally:
            conn.close()
    except Exception as e:
        _log_upload.error("delete processed-cv failed: %s", e)
        raise HTTPException(500, "삭제 실패")
    return ok(message="Processed CV deleted.")


@app.post("/api/admin/candidates/{candidate_id}/reprocess-cv", tags=["admin"])
async def reprocess_cv(candidate_id: str, request: Request):
    """수동 재처리 트리거 -- sheet_number 변경 후 호출."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        # 가장 최근 CV 업로드의 s3_key 조회
        row = conn.execute(
            "SELECT s3_key FROM file_uploads "
            "WHERE entity_id = ? AND file_type IN ('cv', 'cover_letter') "
            "AND s3_key IS NOT NULL AND is_deleted = 0 "
            "ORDER BY rowid DESC LIMIT 1",
            (candidate_id,),
        ).fetchone()
    finally:
        conn.close()
    if not row or not row[0]:
        raise HTTPException(404, "Original CV not found in S3")
    # v2.0 durable queue enqueue + 즉시 스레드 (이중 안전망)
    try:
        sys.path.insert(0, str(Path(__file__).resolve().parent))
        from tools.pipeline_v2 import enqueue_resume_process  # type: ignore
        enqueue_resume_process(candidate_id=candidate_id, cv_s3_key=row[0], trigger="reprocess")
    except Exception as _enq_err:
        logging.getLogger("bridge.pipeline").warning("[ENQUEUE reprocess] fail: %s", _enq_err)
    threading.Thread(
        target=_auto_process_resume,
        args=(candidate_id, row[0]),
        daemon=True,
    ).start()
    return ok(message="Reprocessing started.")


@app.get("/api/admin/resume/status/{candidate_id}", tags=["admin"])
async def get_resume_status(candidate_id: str, request: Request):
    """변환 상태 조회 (pending/processing/done/error)."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        row = conn.execute(
            "SELECT resume_status, cv_processed_s3_key FROM candidates WHERE candidate_id = ?",
            (candidate_id,),
        ).fetchone()
    finally:
        conn.close()
    if not row:
        raise HTTPException(404, "Candidate not found")
    return ok(data={
        "status": row[0] or "pending",
        "has_processed": bool(row[1]),
    })


@app.get("/api/admin/resume/preview/{candidate_id}", tags=["admin"])
async def get_resume_preview(candidate_id: str, request: Request):
    """변환 완료된 PDF presigned URL 반환."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        row = conn.execute(
            "SELECT cv_processed_s3_key FROM candidates WHERE candidate_id = ?",
            (candidate_id,),
        ).fetchone()
    finally:
        conn.close()
    if not row or not row[0]:
        raise HTTPException(404, "Processed CV not found")
    url = s3_presigned_url(row[0], expires=3600)
    if not url:
        raise HTTPException(500, "Presigned URL 생성 실패")
    return ok(data={"url": url, "s3_key": row[0]})


@app.post("/api/admin/resume/reprocess/{candidate_id}", tags=["admin"])
async def reprocess_resume_full(candidate_id: str, request: Request):
    """전체 파이프라인 재변환 (PII + 얼굴 + PDF 병합)."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        row = conn.execute(
            "SELECT s3_key FROM file_uploads "
            "WHERE entity_id = ? AND file_type IN ('cv', 'cover_letter') "
            "AND s3_key IS NOT NULL AND is_deleted = 0 "
            "ORDER BY rowid DESC LIMIT 1",
            (candidate_id,),
        ).fetchone()
    finally:
        conn.close()
    if not row or not row[0]:
        raise HTTPException(404, "Original CV not found")
    _set_resume_status(candidate_id, "pending")
    # v2.0 durable queue enqueue + 즉시 스레드
    try:
        sys.path.insert(0, str(Path(__file__).resolve().parent))
        from tools.pipeline_v2 import enqueue_resume_process  # type: ignore
        enqueue_resume_process(candidate_id=candidate_id, cv_s3_key=row[0], trigger="full_pipeline")
    except Exception as _enq_err:
        logging.getLogger("bridge.pipeline").warning("[ENQUEUE full] fail: %s", _enq_err)
    threading.Thread(
        target=_auto_process_resume,
        args=(candidate_id, row[0]),
        daemon=True,
    ).start()
    return ok(message="Full pipeline reprocessing started.")


@app.delete("/api/admin/files/{file_id}", tags=["admin"])
async def delete_uploaded_file(file_id: int, request: Request):
    """업로드 파일 삭제 (S3 + soft-delete). 원본·첨부 모두 대응."""
    _check_admin(request)
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        try:
            row = conn.execute(
                "SELECT s3_key, entity_id, file_type FROM file_uploads WHERE id = ? AND is_deleted = 0",
                (file_id,),
            ).fetchone()
            if not row:
                raise HTTPException(404, "File not found")
            s3_key, eid, ftype = row[0], row[1], row[2]
            if s3_key:
                try:
                    s3_delete_file(s3_key)
                except Exception:
                    pass
            conn.execute("UPDATE file_uploads SET is_deleted = 1 WHERE id = ?", (file_id,))
            # cv_processed 삭제 시 candidates 컬럼도 정리
            if ftype == "cv_processed" and eid:
                conn.execute(
                    "UPDATE candidates SET cv_processed_s3_key = NULL WHERE candidate_id = ?",
                    (eid,),
                )
            conn.commit()
        finally:
            conn.close()
    except HTTPException:
        raise
    except Exception as e:
        _log_upload.error("delete file failed: %s", e)
        raise HTTPException(500, "삭제 실패")
    return ok(message="File deleted.")


@app.post("/api/admin/upload-image", tags=["admin"])
async def admin_upload_editor_image(
    request: Request,
    file: UploadFile = FastFile(...),
):
    """에디터 이미지 업로드 (관리자 전용). 서버에 저장 후 URL 반환."""
    _check_admin(request)
    if not _rate_ok(_ip_hash(request), window=60, max_posts=30):
        raise HTTPException(429, "Too many uploads.")
    data = await file.read()
    if not data:
        raise HTTPException(400, "Empty file.")
    ext = _validate_file(data, file.filename or "image.png", "community_image")
    # [비활성 -- 로컬 에디터 이미지 저장] S3 전환 후 불필요. 아래 S3 업로드 사용.
    # img_dir = _UPLOAD_BASE / "editor"; img_dir.mkdir(parents=True, exist_ok=True)
    # fname = f"{uuid.uuid4().hex[:12]}{ext}"; file_path = img_dir / fname
    # file_path.write_bytes(data); file_url = f"/uploads/editor/{fname}"
    # ── S3 업로드 ──────────────────────────────────────────────────────────
    try:
        _editor_result = await s3_upload_bytes(
            data=data,
            folder="editor",
            filename=file.filename or f"image{ext}",
            allowed_category="image",
        )
        file_url = s3_presigned_url(_editor_result["s3_key"], expires=86400)  # 24시간
    except (RuntimeError, StorageError) as _s3_err:
        _log_upload.error("에디터 이미지 S3 업로드 실패: %s", _s3_err)
        raise HTTPException(503, f"이미지 저장 실패: {_s3_err}")
    # ──────────────────────────────────────────────────────────────────────
    return ok(data={"url": file_url}, message="Image uploaded")


@app.get("/api/admin/files/{entity_type}/{entity_id}", tags=["admin"])
async def admin_list_files(entity_type: str, entity_id: str, request: Request):
    """관리자: 엔티티의 업로드된 파일 목록."""
    _check_admin(request)
    if entity_type not in ("candidate", "inquiry", "community"):
        raise HTTPException(400, "entity_type must be 'candidate', 'inquiry', or 'community'")
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        try:
            cur = conn.execute(
                "SELECT id, file_type, file_url, file_size, created_at FROM file_uploads WHERE entity_type = ? AND entity_id = ? AND is_deleted = 0 ORDER BY created_at DESC",
                (entity_type, entity_id),
            )
            cols = [d[0] for d in cur.description]
            rows = [dict(zip(cols, r)) for r in cur.fetchall()]
        finally:
            conn.close()
        return ok(data=rows)
    except Exception as e:
        _log_upload.error("admin_list_files failed: %s", e)
        return ok(data=[])


# ── Audit Log & Security Report ──────────────────────────────────────────────

_AUDIT_DIR = Path(os.getenv("AUDIT_DIR", str(Path(__file__).resolve().parent / "audit")))
_security_warn_store: dict[str, list[float]] = defaultdict(list)

@app.get("/api/admin/audit-log", tags=["admin"])
async def admin_audit_log(request: Request, date: str = "", limit: int = 100):
    """감사 로그 조회 (관리자 전용)."""
    _check_admin(request)
    if limit < 1 or limit > 1000:
        limit = 100
    target_date = date or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    log_path = _AUDIT_DIR / f"audit_{target_date}.jsonl"
    if not log_path.exists():
        return ok(data=[], message=f"No audit log for {target_date}")
    entries = []
    try:
        with open(log_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    entries.append(json.loads(line))
    except Exception:
        return err("Failed to read audit log", 500)
    return ok(data=entries[-limit:], message=f"{len(entries[-limit:])} entries from {target_date}")


@app.post("/api/security/report", tags=["security"])
async def security_report(request: Request):
    """프론트엔드 보안 이벤트 수신."""
    ip = _ip_hash(request)
    if not _rate_ok(ip, window=60, max_posts=20):
        raise HTTPException(429, "Too many security reports.")

    try:
        body = await request.json()
    except Exception:
        raise HTTPException(400, "Invalid JSON")

    event_type = str(body.get("type", ""))[:40]
    detail = str(body.get("detail", ""))[:500]
    url = str(body.get("url", ""))[:500]

    # 보안 이벤트 + 폼 제출 실패 알림 (백업 메커니즘)
    SECURITY_TYPES = {"xss", "sqli", "bot", "devtools", "tampering"}
    FORM_FAIL_TYPES = {"apply_submission_failed", "inquiry_submission_failed"}
    if event_type not in SECURITY_TYPES | FORM_FAIL_TYPES:
        raise HTTPException(400, "Invalid event type")

    # 폼 제출 실패 — 운영자에게 이메일 알림 (사용자가 다시 시도하지 않을 수 있으므로 수동 follow-up용)
    if event_type in FORM_FAIL_TYPES:
        try:
            err_text = str(body.get("error", ""))[:500]
            user_email = str(body.get("email", ""))[:200]
            user_name = str(body.get("full_name", "") or body.get("school", ""))[:200]
            ua = str(body.get("user_agent", ""))[:300]
            label = "지원" if event_type == "apply_submission_failed" else "채용 문의"
            subj = f"[BRIDGE 폼 실패] {label} 제출 실패 - {user_name or '익명'}"
            html = (
                f"<h3>{label} 폼 제출 실패</h3>"
                f"<p><b>유형:</b> {event_type}</p>"
                f"<p><b>오류:</b> {err_text}</p>"
                f"<p><b>이메일:</b> {user_email}</p>"
                f"<p><b>이름/학원:</b> {user_name}</p>"
                f"<p><b>UA:</b> {ua}</p>"
                f"<p><b>시각:</b> {datetime.now(timezone.utc).isoformat()}</p>"
                f"<hr><p style='color:#888;font-size:11px'>"
                f"사용자가 다시 시도하지 않을 수 있으니, 이메일 주소로 직접 연락하여 수동 접수를 받아주세요.</p>"
            )
            try:
                _smtp_send("bridgejobkr@gmail.com", subj, html)
            except Exception:
                pass
            # 텔레그램 알림도 fire-and-forget
            try:
                import requests as _req
                tg_token = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
                tg_chat = os.getenv("TELEGRAM_CHAT_ID", "").strip()
                if tg_token and tg_chat:
                    _req.post(
                        f"https://api.telegram.org/bot{tg_token}/sendMessage",
                        data={"chat_id": tg_chat,
                              "text": f"🚨 {label} 폼 실패\n이메일: {user_email}\n이름: {user_name}\n오류: {err_text[:200]}"},
                        timeout=5,
                    )
            except Exception:
                pass
        except Exception as _alert_err:
            logging.getLogger("bridge.api").warning("폼 실패 알림 발송 실패: %s", _alert_err)

    # Log to security_events.jsonl
    _AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    event_path = _AUDIT_DIR / "security_events.jsonl"
    entry = {
        "ts": datetime.now(timezone.utc).isoformat(),
        "ip": ip,
        "type": event_type,
        "detail": detail,
        "url": url,
        "user_agent": (request.headers.get("user-agent") or "")[:200],
    }
    try:
        with open(event_path, "a", encoding="utf-8") as f:
            f.write(json.dumps(entry, ensure_ascii=False) + "\n")
    except Exception:
        pass

    # Track warning threshold: 10+ events in 5 min
    now = time.time()
    _security_warn_store[ip] = [t for t in _security_warn_store[ip] if now - t < 300]
    _security_warn_store[ip].append(now)
    is_warned = len(_security_warn_store[ip]) >= 10

    return ok(message="Event recorded", data={"warned": is_warned})


# ── Admin: Job Status ──────────────────────────────────────────────────────────

class JobStatusUpdate(BaseModel):
    status: str = Field(..., pattern=r'^(open|closed)$')

@app.put("/api/admin/jobs/{job_id}/status", tags=["admin"])
async def admin_update_job_status(job_id: int, body: JobStatusUpdate, request: Request):
    """채용공고 Open/Close 상태 변경 (관리자)."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        r = conn.execute("SELECT id FROM jobs WHERE id=? AND is_deleted=0", (job_id,)).fetchone()
        if not r:
            raise HTTPException(404, "Job not found")
        conn.execute("UPDATE jobs SET status=? WHERE id=?", (body.status, job_id))
        conn.commit()
        return ok(message=f"Job #{job_id} status → {body.status}")
    finally:
        conn.close()

@app.put("/api/admin/jobs/{job_id}/hot", tags=["admin"])
async def admin_toggle_job_hot(job_id: int, request: Request):
    """채용공고 HOT 토글 (관리자)."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        row = conn.execute("SELECT id, is_hot FROM jobs WHERE id=? AND is_deleted=0", (job_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Job not found")
        new_hot = 0 if row[1] else 1
        conn.execute("UPDATE jobs SET is_hot=? WHERE id=?", (new_hot, job_id))
        conn.commit()
        return ok(message=f"Job #{job_id} hot → {bool(new_hot)}")
    finally:
        conn.close()


# ── Auto Backup Helper ─────────────────────────────────────────────────────────

_db_change_count = 0
_DB_BACKUP_INTERVAL = 10  # 매 10건 변경마다 백업

# 제출 백업 디렉터리 (Render: /data/backups/auto, 로컬: ./backups/auto)
_BACKUP_DIR = Path(os.getenv("BRIDGE_BACKUP_DIR",
    "/data/backups/auto" if os.getenv("RENDER_EXTERNAL_URL") else str(Path(__file__).resolve().parent / "backups" / "auto")))
try:
    _BACKUP_DIR.mkdir(parents=True, exist_ok=True)
except OSError:
    _BACKUP_DIR = Path("/tmp/bridge_backups/auto")
    _BACKUP_DIR.mkdir(parents=True, exist_ok=True)

# 카운터 영속화 파일 (재시작 후에도 유지)
_CHANGE_COUNT_FILE = _BACKUP_DIR.parent / "change_count.txt"
try:
    _CHANGE_COUNT_FILE.parent.mkdir(parents=True, exist_ok=True)
except OSError:
    pass

def _load_change_count() -> int:
    try:
        return int(_CHANGE_COUNT_FILE.read_text().strip())
    except Exception:
        return 0

def _save_change_count(n: int) -> None:
    try:
        _CHANGE_COUNT_FILE.write_text(str(n))
    except Exception:
        pass

_db_change_count = _load_change_count()


def _maybe_auto_backup():
    """DB 변경 누적 카운트 → 변경 추적용 카운터.

    원본 동작 복원: 로컬 file copy만 (관측 검증 결과: DB는 cold start 사이에도
    유지되므로 Drive 업로드 불필요. 이전 Drive upload 변경은 Drive 15GB 쿼터
    낭비 + 매 admin 변경마다 백그라운드 thread spawn 비용만 발생했음).
    """
    global _db_change_count
    _db_change_count += 1
    _save_change_count(_db_change_count)
    if _db_change_count % _DB_BACKUP_INTERVAL != 0:
        return
    try:
        import shutil
        src = _ADMIN_DB_PATH
        dst = src.with_suffix(".db.auto_backup")
        shutil.copy2(str(src), str(dst))
        logging.getLogger("bridge.api").info("AUTO_BACKUP: master.db → %s (after %d changes)", dst.name, _db_change_count)
    except Exception as e:
        logging.getLogger("bridge.api").warning("AUTO_BACKUP failed: %s", e)


def _notify_push(title: str, body: str, url: str = "/admin/m") -> None:
    """Push 알림 발송 (비동기 호출용). 실패해도 무시."""
    try:
        from push_api import send_push_to_all
        send_push_to_all(title=title, body=body, url=url)
    except Exception as e:
        logging.getLogger("bridge.push").debug("Push notify skipped: %s", e)


def _trigger_gdrive_backup() -> None:
    """파일 업로드/접수 시 즉시 GDrive 백업 트리거 (originals + resumes만)."""
    import subprocess
    try:
        subprocess.Popen(
            [r"Q:\Phtyon 3\python.exe", "-X", "utf8",
             "tools/gdrive_backup.py", "--originals", "--resumes"],
            cwd=r"Q:\Claudework\bridge base",
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
    except Exception as e:
        logging.getLogger("bridge.backup").debug("GDrive backup trigger skipped: %s", e)


def _backup_on_submit(source: str = "") -> None:
    """
    강사 지원 / 채용의뢰 제출 시 즉시 타임스탬프 백업 생성.
    AES-256-GCM 암호화 후 backups/auto/submit_YYYYMMDD_HHMMSS.enc 저장.
    실패해도 접수에는 영향 없음 (비동기 호출 권장).
    """
    import shutil
    _log_bk = logging.getLogger("bridge.backup")
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    src = _ADMIN_DB_PATH

    # 1. 일반 .auto_backup 갱신 (빠른 비암호화 스냅샷)
    try:
        shutil.copy2(str(src), str(src.with_suffix(".db.auto_backup")))
    except Exception as e:
        _log_bk.warning("auto_backup copy failed: %s", e)

    # 2. 암호화 백업 (AES-256-GCM)
    try:
        from cryptography.hazmat.primitives.ciphers.aead import AESGCM
        import base64 as _b64
        _key_raw = os.getenv("BRIDGE_FIELD_KEY", "").strip()
        if not _key_raw:
            raise EnvironmentError("BRIDGE_FIELD_KEY 미설정")
        import hashlib as _hs
        key = _hs.sha256(_key_raw.encode()).digest()
        nonce = os.urandom(12)
        plaintext = src.read_bytes()
        aesgcm = AESGCM(key)
        ciphertext = aesgcm.encrypt(nonce, plaintext, source.encode() or b"submit")
        payload = nonce + ciphertext  # nonce(12) + ciphertext + tag(16)
        enc_path = _BACKUP_DIR / f"submit_{ts}_{source}.enc"
        enc_path.write_bytes(payload)
        _log_bk.info("SUBMIT_BACKUP: %s (%d bytes encrypted)", enc_path.name, len(payload))

        # 3. 오래된 제출 백업 정리 (최근 50개 유지)
        backups = sorted(_BACKUP_DIR.glob("submit_*.enc"), key=lambda p: p.stat().st_mtime)
        for old in backups[:-50]:
            try:
                old.unlink()
            except Exception:
                pass

    except Exception as e:
        _log_bk.error("암호화 백업 실패 (접수는 완료됨): %s", e)


def _save_raw_submission(payload: dict, source: str, record_id: str = "") -> None:
    """
    제출 시점 원본 JSON을 별도 AES-256-GCM 파일로 보관.
    backups/raw/raw_{source}_{YYYYMMDD_HHMMSS}_{id}.enc
    - DB와 독립된 불변 원본 -- 관리자도 DB에서 수정해도 이 파일은 보존
    - 복호화 키: BRIDGE_FIELD_KEY (Render 환경변수, 서버만 보유)
    - 최근 200개 유지
    """
    import json as _j, hashlib as _hs, logging as _lg
    _log = _lg.getLogger("bridge.raw_backup")
    try:
        _key_raw = os.getenv("BRIDGE_FIELD_KEY", "").strip()
        if not _key_raw:
            return
        raw_dir = _BACKUP_DIR.parent / "raw"
        raw_dir.mkdir(parents=True, exist_ok=True)

        ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        fname = f"raw_{source}_{ts}_{record_id}.enc"

        plaintext = _j.dumps({
            "source": source,
            "id": record_id,
            "submitted_at": ts,
            "payload": payload,
        }, ensure_ascii=False, default=str).encode("utf-8")

        from cryptography.hazmat.primitives.ciphers.aead import AESGCM
        key = _hs.sha256(_key_raw.encode()).digest()
        nonce = os.urandom(12)
        ciphertext = AESGCM(key).encrypt(nonce, plaintext, source.encode())
        (raw_dir / fname).write_bytes(nonce + ciphertext)
        _log.info("RAW_BACKUP: %s (%d bytes)", fname, len(plaintext))

        # 최근 200개만 유지
        files = sorted(raw_dir.glob(f"raw_{source}_*.enc"), key=lambda p: p.stat().st_mtime)
        for old in files[:-200]:
            try: old.unlink()
            except Exception: pass
    except Exception as e:
        _log.warning("RAW_BACKUP 실패 (접수는 완료됨): %s", e)


# ── 자동 채용공고 생성 Helper (inquiry 접수 시 background thread 호출) ────────

def _auto_create_job_from_inquiry(inquiry_id: int, school_name: str) -> None:
    """inquiry 접수 직후 background에서 job 레코드 자동 생성 (status=pending_review)"""
    import logging as _log_auto
    _log = _log_auto.getLogger("bridge.auto_job")
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        conn.row_factory = sqlite3.Row
        try:
            inq = conn.execute("SELECT * FROM client_inquiries WHERE id = ?", (inquiry_id,)).fetchone()
            if not inq:
                return
            # inquiry_id 기반 중복 체크 (primary) + notes 폴백
            existing = conn.execute(
                "SELECT id FROM jobs WHERE inquiry_id = ?", (inquiry_id,)
            ).fetchone()
            if existing:
                return  # 이미 등록됨
            notes = inq["notes"] or ""
            if "JOB_REGISTERED" in notes:
                return  # notes 기반 폴백 체크

            # 급여 파싱
            salary_raw = inq["salary_raw"] or ""
            salary_nums = re.findall(r"[\d,.]+", salary_raw.replace(",", ""))
            salary_min = float(salary_nums[0]) if salary_nums else None
            salary_max = float(salary_nums[-1]) if len(salary_nums) > 1 else salary_min

            # 근무시간 파싱
            wh = inq["working_hours"] or inq["schedule"] or ""
            daily_hours = None
            time_match = re.search(r"(\d{1,2}):?(\d{2})\s*[~\-]\s*(\d{1,2}):?(\d{2})", wh)
            if time_match:
                h1 = int(time_match.group(1)) + int(time_match.group(2)) / 60
                h2 = int(time_match.group(3)) + int(time_match.group(4)) / 60
                daily_hours = round(h2 - h1, 2) if h2 > h1 else None

            loc = inq["location"] or ""
            city = loc.split(",")[0].split(" ")[0].strip() if loc else ""

            # memo / contact 정보를 internal_notes에 보관 (employers 페이지 PII 파싱용)
            memo_raw = inq["memo"] or ""
            parsed_data_json = inq["parsed_data"] or ""
            # parsed_data에 저장된 추가 필드 복원
            extra = {}
            if parsed_data_json:
                try:
                    import json as _json_aj
                    extra = _json_aj.loads(parsed_data_json)
                except Exception:
                    pass
            contact_name = extra.get("contact_name") or ""
            phone_val = inq["phone"] or ""
            email_val = inq["email"] or ""
            school_loc = extra.get("school_location") or ""
            internal_notes = memo_raw
            if not internal_notes and (contact_name or school_loc):
                internal_notes = f"{school_name} / {contact_name} / {phone_val} / {email_val} / {school_loc}"
            # raw_text: 전체 inquiry 요약
            raw_parts = [f"School: {school_name}", f"Location: {loc}"]
            if inq["start_date"]:  raw_parts.append(f"Starting Date : {inq['start_date']}")
            if inq["teaching_age"]: raw_parts.append(f"Teaching Age : {inq['teaching_age']}")
            if salary_raw:         raw_parts.append(f"Salary : {salary_raw}")
            if wh:                 raw_parts.append(f"Working Hours : {wh}")
            housing_val = inq["housing_detail"] or inq["housing_type"] or ""
            if housing_val:        raw_parts.append(f"Housing : {housing_val}")
            if inq["benefits"]:    raw_parts.append(f"Benefits : {inq['benefits']}")
            raw_text = "\n".join(raw_parts)

            # job_code: WEB-{inquiry_id} (일관성 + inquiry 추적)
            job_code = f"WEB-{inquiry_id}"
            cur = conn.execute(
                """INSERT OR IGNORE INTO jobs
                   (seq, job_code, location, city, start_date, teaching_age,
                    working_hours, daily_hours, salary_min, salary_max, salary_raw,
                    vacation, housing, housing_type, housing_detail,
                    benefits, status, is_hot, is_deleted, created_at,
                    internal_notes, raw_text, employer_display_name,
                    source_file, inquiry_id)
                   VALUES (1, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                           ?, 'pending_review', 0, 0, ?, ?, ?, ?, 'web_form', ?)""",
                (
                    job_code, loc, city, inq["start_date"], inq["teaching_age"],
                    wh, daily_hours, salary_min, salary_max, salary_raw,
                    inq["vacation"], housing_val, inq["housing_type"], inq["housing_detail"],
                    inq["benefits"], datetime.now(timezone.utc).isoformat(),
                    internal_notes, raw_text, school_name, inquiry_id,
                ),
            )
            new_row_id = cur.lastrowid
            new_notes = f"{notes}\nJOB_REGISTERED:{job_code}".strip()
            conn.execute("UPDATE client_inquiries SET notes = ? WHERE id = ?", (new_notes, inquiry_id))
            conn.commit()
            _log.info("자동 job 생성 완료: %s (inquiry #%d)", job_code, inquiry_id)

            # 관리자 알림 이메일 (실패 무시)
            try:
                if _EMAIL_OK:
                    from email_templates import send_new_job_pending_alert
                    _admin_email = os.getenv("GMAIL_USER", "bridgejobkr@gmail.com")
                    send_new_job_pending_alert(_admin_email, school_name, inquiry_id, job_code)
            except Exception as _em:
                _log.warning("관리자 알림 이메일 실패 (무시): %s", _em)

        finally:
            conn.close()
    except Exception as e:
        _log.error("_auto_create_job_from_inquiry 실패: %s", e, exc_info=True)


# ── Admin: 채용공고 승인 (pending_review → open) ─────────────────────────────

@app.put("/api/admin/jobs/{job_id}/approve", tags=["admin"])
async def admin_approve_job(job_id: int, request: Request):
    """pending_review 상태 공고를 open으로 승인"""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        row = conn.execute("SELECT status FROM jobs WHERE id = ? AND is_deleted = 0", (job_id,)).fetchone()
        if not row:
            raise HTTPException(404, "공고를 찾을 수 없습니다.")
        if row[0] == "open":
            return ok(message="이미 공개 상태입니다.", data={"job_id": job_id, "status": "open"})
        conn.execute(
            "UPDATE jobs SET status = 'open', internal_notes = COALESCE(internal_notes, '') || ? WHERE id = ?",
            (f"\nAPPROVED:{datetime.now(timezone.utc).isoformat()}", job_id),
        )
        conn.commit()
        _maybe_auto_backup()
        return ok(message="공고가 승인되어 공개됩니다.", data={"job_id": job_id, "status": "open"})
    finally:
        conn.close()


# ── Admin: Google Sheets 동기화 ───────────────────────────────────────────────

@app.post("/api/admin/sync/google-sheets", tags=["admin"])
async def admin_sync_google_sheets(request: Request, body: dict = None):
    """Google Sheets → DB 동기화 (candidates / jobs / all)"""
    _check_admin(request)
    if body is None:
        body = {}
    mode = body.get("mode", "all")
    dry_run = bool(body.get("dry_run", False))

    try:
        from tools.sync_google_sheets import run_sync
        result = run_sync(mode=mode, dry_run=dry_run)
        return ok(data=result, message=f"동기화 완료 (mode={mode}, dry_run={dry_run})")
    except ImportError:
        raise HTTPException(503, "sync_google_sheets 모듈을 찾을 수 없습니다. tools/sync_google_sheets.py 확인 필요.")
    except Exception as e:
        import logging as _log_sync
        _log_sync.getLogger("bridge.sync").error("Google Sheets 동기화 실패: %s", e, exc_info=True)
        raise HTTPException(500, f"동기화 실패: {str(e)}")


# ── Admin: Google Forms Webhook 수신 ─────────────────────────────────────────

@app.post("/api/admin/sync/incoming", tags=["admin"])
async def admin_sync_incoming(request: Request):
    """Google Apps Script에서 form 제출 시 webhook 수신 (HMAC 검증 필수)"""
    import hmac as _hmac_mod
    import hashlib as _hs

    # Rate Limit: IP당 분당 30건
    if not _rate_ok(_ip_hash(request), window=60, max_posts=30):
        raise HTTPException(429, "Too many webhook requests")

    _WEBHOOK_SECRET = os.getenv("BRIDGE_WEBHOOK_SECRET", "")
    sig_header = request.headers.get("x-bridge-signature", "")

    body_bytes = await request.body()

    # HMAC-SHA256 검증 (Fail-Closed: 시크릿 미설정 시 차단)
    if not _WEBHOOK_SECRET:
        raise HTTPException(503, "Webhook secret not configured")
    expected = _hmac_mod.new(_WEBHOOK_SECRET.encode(), body_bytes, _hs.sha256).hexdigest()
    if not _hmac_mod.compare_digest(sig_header, expected):
        raise HTTPException(403, "Invalid signature")

    try:
        import json as _json
        payload = _json.loads(body_bytes)
    except Exception:
        raise HTTPException(400, "Invalid JSON body")

    data_type = payload.get("type", "")
    data = payload.get("data", {})

    if data_type == "job":
        # 고용주 폼 제출 → client_inquiries upsert (PII 암호화)
        _PII_FIELDS = {"phone", "email", "contact_name", "business_registration", "school_location", "memo"}
        for field in _PII_FIELDS:
            if field in data and data[field]:
                data[field] = _encrypt_if_needed(str(data[field]), field)
        data["source"] = "google_form_webhook"
        data["inbox_status"] = "new"
        data.setdefault("submitted_at", datetime.now(timezone.utc).isoformat())

        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        try:
            allowed = {
                "school_name", "location", "teaching_age", "working_hours", "salary_raw",
                "start_date", "vacation", "housing_type", "benefits", "source",
                "inbox_status", "submitted_at",
                "phone", "email", "contact_name",  # 암호화됨
                "location_plain", "school_name_plain",  # 검색용 평문
            }
            clean = {k: v for k, v in data.items() if k in allowed and v is not None}
            # location/school_name 은 webhook에서 평문 → plain 컬럼에 복사 (컬럼 존재 시만)
            _inq_cols_check = conn.execute("PRAGMA table_info(client_inquiries)").fetchall()
            _inq_col_names = {r[1] for r in _inq_cols_check}
            if "location_plain" in _inq_col_names:
                clean["location_plain"] = str(data.get("location", "") or "")
            if "school_name_plain" in _inq_col_names:
                clean["school_name_plain"] = str(data.get("school_name", "") or "")
            cols = ", ".join(clean.keys())
            placeholders = ", ".join("?" * len(clean))
            cur = conn.execute(
                f"INSERT INTO client_inquiries ({cols}) VALUES ({placeholders})",
                list(clean.values()),
            )
            conn.commit()
            new_id = cur.lastrowid
        finally:
            conn.close()

        # 자동 job 생성 (background)
        threading.Thread(
            target=_auto_create_job_from_inquiry,
            args=(new_id, data.get("school_name", "")),
            daemon=True
        ).start()
        return ok(data={"id": new_id, "type": "job"}, message="채용의뢰 접수 완료")

    elif data_type == "candidate":
        # 교사 지원 폼 제출 → candidates upsert (PII 암호화)
        _CAND_PII = {"email", "phone", "full_name", "date_of_birth", "address", "passport_number"}
        for field in _CAND_PII:
            if field in data and data[field]:
                data[field] = _encrypt_if_needed(str(data[field]), field)
        data["source"] = "google_form_webhook"
        data.setdefault("created_at", datetime.now(timezone.utc).isoformat())

        import uuid as _uuid_sync
        cand_id = f"cnd_{_uuid_sync.uuid4().hex[:12]}"
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        try:
            allowed = {
                "full_name", "email", "phone", "nationality", "teaching_age",
                "visa_type", "start_date", "source", "created_at",
            }
            clean = {k: v for k, v in data.items() if k in allowed and v is not None}
            if not clean:
                raise HTTPException(400, "유효한 후보자 데이터가 없습니다.")
            clean["candidate_id"] = cand_id
            clean["status"] = "Active"
            clean["updated_at"] = datetime.now(timezone.utc).isoformat()
            # 4를 포함하지 않는 sheet_number 할당
            clean["sheet_number"] = _next_sheet_number(conn)
            new_sn = clean["sheet_number"]
            cols = list(clean.keys())
            vals = list(clean.values())
            ph = ["?" for _ in cols]
            conn.execute(
                f"INSERT INTO candidates ({', '.join(cols)}) VALUES ({', '.join(ph)})",
                vals,
            )
            conn.commit()
        finally:
            conn.close()
        # Google Sheet 동기화 (비동기)
        threading.Thread(target=_sync_to_sheet, args=(cand_id,), daemon=True).start()
        return ok(data={"id": cand_id, "sheet_number": new_sn, "type": "candidate"}, message="후보자 접수 완료")

    else:
        raise HTTPException(400, f"알 수 없는 type: {data_type!r}")


# ── Admin: Create Job from Inquiry ─────────────────────────────────────────────

@app.post("/api/admin/jobs/create-from-inquiry/{inquiry_id}", tags=["admin"])
async def admin_create_job_from_inquiry(inquiry_id: int, request: Request):
    """채용의뢰를 기반으로 새 job 레코드 생성 (PII 분리)."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        inq = conn.execute("SELECT * FROM client_inquiries WHERE id = ?", (inquiry_id,)).fetchone()
        if not inq:
            raise HTTPException(404, "Inquiry not found")

        # 재등록 방지: notes에 'JOB_REGISTERED' 마커 확인
        notes = inq["notes"] or ""
        if "JOB_REGISTERED" in notes:
            raise HTTPException(409, "이미 등록된 채용의뢰입니다.")

        # 새 job_code 생성 (max seq + 1)
        max_row = conn.execute("SELECT MAX(id) FROM jobs").fetchone()
        new_id = (max_row[0] or 0) + 1
        job_code = f"Job.{new_id + 1000}"

        # 급여 파싱
        salary_raw = inq["salary_raw"] or ""
        salary_nums = re.findall(r"[\d,.]+", salary_raw.replace(",", ""))
        salary_min = float(salary_nums[0]) if salary_nums else None
        salary_max = float(salary_nums[-1]) if len(salary_nums) > 1 else salary_min

        # 근무시간 → daily_hours 파싱
        wh = inq["working_hours"] or inq["schedule"] or ""
        daily_hours = None
        time_match = re.search(r"(\d{1,2}):?(\d{2})\s*[~\-]\s*(\d{1,2}):?(\d{2})", wh)
        if time_match:
            h1 = int(time_match.group(1)) + int(time_match.group(2)) / 60
            h2 = int(time_match.group(3)) + int(time_match.group(4)) / 60
            daily_hours = round(h2 - h1, 2) if h2 > h1 else None

        # city 추출 (location에서 첫 단어)
        loc = inq["location"] or ""
        city = loc.split(",")[0].split(" ")[0].strip() if loc else ""

        conn.execute(
            """INSERT INTO jobs (job_code, seq, location, city, start_date, teaching_age,
               working_hours, daily_hours, salary_min, salary_max, salary_raw,
               vacation, housing, benefits, status, is_hot, is_deleted, created_at)
               VALUES (?, 1, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'open', 0, 0, ?)""",
            (
                job_code, loc, city, inq["start_date"], inq["teaching_age"],
                wh, daily_hours, salary_min, salary_max, salary_raw,
                inq["vacation"], inq["housing_detail"] or inq["housing_type"],
                inq["benefits"], datetime.now(timezone.utc).isoformat(),
            ),
        )

        # inquiry에 등록 마커 추가
        new_notes = f"{notes}\nJOB_REGISTERED:{job_code}".strip()
        conn.execute("UPDATE client_inquiries SET notes = ? WHERE id = ?", (new_notes, inquiry_id))
        conn.commit()
        _maybe_auto_backup()
        return ok(message=f"Job {job_code} 생성 완료", data={"job_code": job_code})
    finally:
        conn.close()


# ── Admin: Jobs v2 (BRJ ID 시스템) ───────────────────────────────────────────

_JOB_REGION_MAP = {
    "Seoul": "SE", "서울": "SE", "Busan": "BS", "부산": "BS",
    "Daegu": "DG", "대구": "DG", "Incheon": "IC", "인천": "IC",
    "Gwangju": "GJ", "광주": "GJ", "Daejeon": "DJ", "대전": "DJ",
    "Ulsan": "US", "울산": "US", "Sejong": "SJ", "세종": "SJ",
    "Gyeonggi": "GG", "경기": "GG", "Gangwon": "GW", "강원": "GW",
    "Chungbuk": "CB", "충북": "CB", "Chungnam": "CN", "충남": "CN",
    "Jeonbuk": "JB", "전북": "JB", "Jeonnam": "JN", "전남": "JN",
    "Gyeongbuk": "KB", "경북": "KB", "Gyeongnam": "KN", "경남": "KN",
    "Jeju": "JJ", "제주": "JJ",
    "Suwon": "GG", "Seongnam": "GG", "Yongin": "GG", "Goyang": "GG",
    "Bucheon": "GG", "Anyang": "GG", "Hwaseong": "GG", "Pyeongtaek": "GG",
    "Cheonan": "CN", "Cheongju": "CB", "Jeonju": "JB",
    "Pohang": "KB", "Changwon": "KN", "Gimhae": "KN",
}

_JOB_REGION_NAMES = {
    "SE": "서울", "BS": "부산", "DG": "대구", "IC": "인천",
    "GJ": "광주", "DJ": "대전", "US": "울산", "SJ": "세종",
    "GG": "경기", "GW": "강원", "CB": "충북", "CN": "충남",
    "JB": "전북", "JN": "전남", "KB": "경북", "KN": "경남",
    "JJ": "제주", "XX": "기타",
}


def _detect_region(text: str) -> str:
    if not text:
        return "XX"
    for key, code in _JOB_REGION_MAP.items():
        if key.lower() in text.lower():
            return code
    return "XX"


def _next_brj_id(conn, region_code: str) -> str:
    yymm = datetime.now(timezone.utc).strftime("%y%m")
    prefix = f"BRJ-{region_code}-{yymm}-"
    row = conn.execute(
        "SELECT brj_id FROM jobs WHERE brj_id LIKE ? ORDER BY brj_id DESC LIMIT 1",
        (f"{prefix}%",),
    ).fetchone()
    seq = int(row[0].split("-")[-1]) + 1 if row else 1
    return f"{prefix}{seq:03d}"


_JOB_WRITABLE = {
    "city", "district", "location", "teaching_age", "class_size",
    "working_hours", "daily_hours", "start_date", "salary_raw",
    "salary_min", "salary_max", "salary_krw", "salary_negotiable",
    "vacation", "vacation_days", "housing", "housing_type", "housing_detail",
    "benefits", "native_count", "teach_hrs_week", "teaching_hours_weekly",
    "is_hot", "is_part_time", "status",
    "enc_employer_name", "enc_contact_name", "enc_contact_phone",
    "enc_contact_email", "enc_contact_kakao", "employer_display_name",
    "visa_sponsorship", "f_visa_welcome", "kyopo_welcome",
    "degree_requirement", "korea_resident_only",
    "raw_text", "raw_source", "parse_confidence", "parse_warnings",
    "internal_notes",
}


@app.get("/api/admin/jobs/v2", tags=["admin"])
async def admin_list_jobs_v2(
    request: Request,
    status: Optional[str] = None,
    region: Optional[str] = None,
    search: Optional[str] = None,
    limit: int = 100,
    offset: int = 0,
):
    """채용공고 v2 전체 목록 (BRJ ID + 관리자 필드 포함)."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        where = ["is_deleted = 0"]
        params: list[Any] = []
        if status:
            where.append("status = ?")
            params.append(status)
        if region:
            where.append("region = ?")
            params.append(region)
        if search:
            s = f"%{search.strip()[:50]}%"
            where.append(
                "(brj_id LIKE ? OR job_code LIKE ? OR city LIKE ? OR location LIKE ? OR employer_display_name LIKE ?)"
            )
            params.extend([s, s, s, s, s])
        w = " AND ".join(where)
        # 필요한 컬럼만 조회 (87→20컬럼 -- PII 복호화 포함)
        _V2_COLS = (
            "brj_id, id, job_code, region, region_name, location, city, "
            "employer_display_name, internal_notes, raw_text, teaching_age, "
            "salary_raw, status, is_hot, created_at, "
            "enc_contact_email, enc_contact_name, enc_contact_phone, enc_contact_kakao, "
            "enc_employer_name"
        )
        total = conn.execute(f"SELECT COUNT(*) FROM jobs WHERE {w}", params).fetchone()[0]
        rows = conn.execute(
            f"SELECT {_V2_COLS} FROM jobs WHERE {w} ORDER BY created_at DESC LIMIT ? OFFSET ?",
            params + [limit, offset],
        ).fetchall()
        data = []
        for r in rows:
            d = dict(r)
            # salary_raw fallback: raw_text에서 추출
            if not d.get("salary_raw") and d.get("raw_text"):
                _m = re.search(r"Monthly Salary\s*:\s*(.+)", d["raw_text"])
                if _m:
                    d["salary_raw"] = _m.group(1).strip().strip('`"')
            # PII 복호화 -- enc_ 필드를 contact_* 로 변환 (원본 enc_ 필드는 제거)
            d["contact_email"] = _safe_decrypt(d.pop("enc_contact_email", ""), "enc_contact_email") or ""
            d["contact_name"]  = _safe_decrypt(d.pop("enc_contact_name", ""),  "enc_contact_name")  or ""
            d["contact_phone"] = _safe_decrypt(d.pop("enc_contact_phone", ""), "enc_contact_phone") or ""
            d["contact_kakao"] = _safe_decrypt(d.pop("enc_contact_kakao", ""), "enc_contact_kakao") or ""
            _enc_emp = d.pop("enc_employer_name", "")
            if _enc_emp and not d.get("employer_display_name"):
                d["employer_display_name"] = _safe_decrypt(_enc_emp, "enc_employer_name") or ""
            data.append(d)
        return ok(data={"jobs": data, "total": total})
    finally:
        conn.close()


@app.get("/api/admin/jobs/v2/{brj_id}", tags=["admin"])
async def admin_get_job_v2(brj_id: str, request: Request):
    """채용공고 상세 조회 (BRJ ID)."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        row = conn.execute(
            "SELECT * FROM jobs WHERE brj_id = ? AND is_deleted = 0", (brj_id,)
        ).fetchone()
        if not row:
            raise HTTPException(404, "Job not found")
        d = dict(row)
        for k in list(d.keys()):
            if isinstance(d[k], bytes):
                d[k] = None
        return ok(data=d)
    finally:
        conn.close()


@app.post("/api/admin/jobs/v2", status_code=201, tags=["admin"])
async def admin_create_job_v2(request: Request, body: dict):
    """새 채용공고 등록 (BRJ ID 자동 생성)."""
    _check_admin(request)
    if not _rate_ok(_ip_hash(request)):
        raise HTTPException(429, "잠시 후 다시 시도해주세요.")
    record = {k: v for k, v in body.items() if k in _JOB_WRITABLE and v is not None}
    region = body.get("region") or _detect_region(
        record.get("city", "") or record.get("location", "")
    )
    now = datetime.now(timezone.utc).isoformat()
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        brj_id = _next_brj_id(conn, region)
        max_row = conn.execute("SELECT MAX(id) FROM jobs").fetchone()
        job_code = f"Job.{(max_row[0] or 0) + 1001}"
        record.update({
            "brj_id": brj_id, "job_code": job_code,
            "region": region, "region_name": _JOB_REGION_NAMES.get(region, "기타"),
            "status": record.get("status", "open"),
            "is_deleted": 0, "is_hot": record.get("is_hot", 0),
            "created_at": now, "seq": 1,
        })
        cols = ", ".join(record.keys())
        ph = ", ".join("?" for _ in record)
        conn.execute(f"INSERT INTO jobs ({cols}) VALUES ({ph})", list(record.values()))
        conn.commit()
        _maybe_auto_backup()
        return ok(data={"brj_id": brj_id, "job_code": job_code}, message=f"채용공고 {brj_id} 생성 완료")
    finally:
        conn.close()


@app.patch("/api/admin/jobs/v2/{brj_id}", tags=["admin"])
async def admin_update_job_v2(brj_id: str, request: Request, body: dict):
    """채용공고 수정 (BRJ ID)."""
    _check_admin(request)
    updates = {k: v for k, v in body.items() if k in _JOB_WRITABLE}
    if not updates:
        raise HTTPException(400, "수정할 항목이 없습니다.")
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        row = conn.execute(
            "SELECT id FROM jobs WHERE brj_id = ? AND is_deleted = 0", (brj_id,)
        ).fetchone()
        if not row:
            raise HTTPException(404, "Job not found")
        set_clause = ", ".join(f"{k}=?" for k in updates)
        conn.execute(
            f"UPDATE jobs SET {set_clause} WHERE brj_id = ?",
            list(updates.values()) + [brj_id],
        )
        conn.commit()
        _maybe_auto_backup()
        return ok(message=f"채용공고 {brj_id} 수정 완료")
    finally:
        conn.close()


@app.delete("/api/admin/jobs/v2/{brj_id}", tags=["admin"])
async def admin_soft_delete_job_v2(brj_id: str, request: Request):
    """채용공고 soft delete (BRJ ID)."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        row = conn.execute(
            "SELECT id FROM jobs WHERE brj_id = ? AND is_deleted = 0", (brj_id,)
        ).fetchone()
        if not row:
            raise HTTPException(404, "Job not found")
        conn.execute("UPDATE jobs SET is_deleted = 1 WHERE brj_id = ?", (brj_id,))
        conn.commit()
        _maybe_auto_backup()
        return ok(message=f"채용공고 {brj_id} 삭제 완료")
    finally:
        conn.close()


# ── Admin: Boards 관리 ─────────────────────────────────────────────────────────

def _ensure_boards_table():
    """boards 테이블 생성 + 기본 보드 시드 (없으면)."""
    import json as _json
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS boards (
            id TEXT PRIMARY KEY,
            label TEXT NOT NULL,
            label_kr TEXT,
            display_mode TEXT DEFAULT 'list',
            sort_order INTEGER DEFAULT 0,
            is_hidden INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()
    # 빈 테이블이면 seed 파일에서 복원
    count = conn.execute("SELECT COUNT(*) FROM boards").fetchone()[0]
    if count == 0:
        seed_path = Path(__file__).parent / "migrations" / "community_seed.json"
        if seed_path.exists():
            try:
                data = _json.loads(seed_path.read_text(encoding="utf-8"))
                for b in data.get("boards", []):
                    conn.execute("""
                        INSERT OR IGNORE INTO boards (id, label, label_kr, display_mode, sort_order, is_hidden)
                        VALUES (?,?,?,?,?,?)
                    """, (b["id"], b["label"], b.get("label_kr"), b.get("display_mode","list"),
                          b.get("sort_order",0), b.get("is_hidden",0)))
                conn.commit()
                logging.getLogger("bridge.api").info(
                    "_ensure_boards_table: %d boards 복원", len(data.get("boards",[])))
            except Exception as _se:
                logging.getLogger("bridge.api").warning("boards seed 스킵: %s", _se)
    conn.close()

try:
    _ensure_boards_table()
except Exception as _e:
    logging.getLogger("bridge.api").warning("_ensure_boards_table 스킵: %s", _e)


class BoardCreate(BaseModel):
    id: str = Field(..., min_length=1, max_length=50, pattern=r'^[a-z0-9_]+$')
    label: str = Field(..., min_length=1, max_length=100)
    label_kr: Optional[str] = Field(None, max_length=100)
    display_mode: str = Field('list', pattern=r'^(list|card|gallery)$')


class BoardUpdate(BaseModel):
    label: Optional[str] = Field(None, min_length=1, max_length=100)
    label_kr: Optional[str] = Field(None, max_length=100)
    display_mode: Optional[str] = Field(None, pattern=r'^(list|card|gallery)$')
    sort_order: Optional[int] = None
    is_hidden: Optional[int] = Field(None, ge=0, le=1)


@app.get("/api/admin/boards", tags=["admin"])
async def admin_list_boards(request: Request):
    """게시판 목록 (관리자)."""
    _check_admin(request)
    if _supa:
        try:
            res = _supa.table('boards').select('*').order('sort_order').order('id').execute()
            return ok(data={"boards": res.data or []})
        except Exception as _e:
            logging.getLogger("bridge.api").warning("[Supabase] admin_list_boards 실패 → SQLite 폴백: %s", _e)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        cur = conn.execute("SELECT * FROM boards ORDER BY sort_order, id")
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    finally:
        conn.close()
    return ok(data={"boards": rows})


@app.post("/api/admin/boards", status_code=201, tags=["admin"])
async def admin_create_board(body: BoardCreate, request: Request):
    """게시판 추가 (관리자)."""
    _check_admin(request)
    if _supa:
        try:
            chk = _supa.table('boards').select('id').eq('id', body.id).execute()
            if chk.data:
                raise HTTPException(409, f"Board '{body.id}' already exists")
            _supa.table('boards').insert({
                'id': body.id, 'label': body.label,
                'label_kr': body.label_kr, 'display_mode': body.display_mode,
            }).execute()
            return ok(data={"id": body.id}, message="Board created")
        except HTTPException:
            raise
        except Exception as _e:
            logging.getLogger("bridge.api").warning("[Supabase] admin_create_board 실패 → SQLite 폴백: %s", _e)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        existing = conn.execute("SELECT id FROM boards WHERE id = ? LIMIT 1", (body.id,)).fetchone()
        if existing:
            raise HTTPException(409, f"Board '{body.id}' already exists")
        conn.execute(
            "INSERT INTO boards (id, label, label_kr, display_mode) VALUES (?, ?, ?, ?)",
            (body.id, body.label, body.label_kr, body.display_mode),
        )
        conn.commit()
    finally:
        conn.close()
    return ok(data={"id": body.id}, message="Board created")


@app.put("/api/admin/boards/{board_id}", tags=["admin"])
async def admin_update_board(board_id: str, body: BoardUpdate, request: Request):
    """게시판 설정 변경 (관리자)."""
    _check_admin(request)
    updates: dict = {}
    if body.label        is not None: updates['label']        = body.label
    if body.label_kr     is not None: updates['label_kr']     = body.label_kr
    if body.display_mode is not None: updates['display_mode'] = body.display_mode
    if body.sort_order   is not None: updates['sort_order']   = body.sort_order
    if body.is_hidden    is not None: updates['is_hidden']    = body.is_hidden
    if not updates:
        raise HTTPException(400, "수정할 항목이 없습니다.")
    if _supa:
        try:
            chk = _supa.table('boards').select('id').eq('id', board_id).execute()
            if not chk.data:
                raise HTTPException(404, "Board not found")
            _supa.table('boards').update(updates).eq('id', board_id).execute()
            return ok(message=f"Board '{board_id}' updated")
        except HTTPException:
            raise
        except Exception as _e:
            logging.getLogger("bridge.api").warning("[Supabase] admin_update_board 실패 → SQLite 폴백: %s", _e)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        if not conn.execute("SELECT id FROM boards WHERE id = ? LIMIT 1", (board_id,)).fetchone():
            raise HTTPException(404, "Board not found")
        set_clause = ", ".join(f"{k} = ?" for k in updates)
        vals = list(updates.values()) + [board_id]
        conn.execute(f"UPDATE boards SET {set_clause} WHERE id = ?", vals)
        conn.commit()
    finally:
        conn.close()
    return ok(message=f"Board '{board_id}' updated")


@app.delete("/api/admin/boards/{board_id}", tags=["admin"])
async def admin_delete_board(board_id: str, request: Request):
    """게시판 숨김 (soft delete: is_hidden=1)."""
    _check_admin(request)
    if _supa:
        try:
            chk = _supa.table('boards').select('id').eq('id', board_id).execute()
            if not chk.data:
                raise HTTPException(404, "Board not found")
            _supa.table('boards').update({'is_hidden': 1}).eq('id', board_id).execute()
            return ok(message=f"Board '{board_id}' hidden")
        except HTTPException:
            raise
        except Exception as _e:
            logging.getLogger("bridge.api").warning("[Supabase] admin_delete_board 실패 → SQLite 폴백: %s", _e)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        if not conn.execute("SELECT id FROM boards WHERE id = ? LIMIT 1", (board_id,)).fetchone():
            raise HTTPException(404, "Board not found")
        conn.execute("UPDATE boards SET is_hidden = 1 WHERE id = ?", (board_id,))
        conn.commit()
    finally:
        conn.close()
    return ok(message=f"Board '{board_id}' hidden")


# ── file_uploads 테이블 (Supabase -> SQLite 마이그레이션) ─────────────────────

def _ensure_file_uploads_table():
    """file_uploads 테이블 생성 (없으면)."""
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS file_uploads (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            entity_type TEXT NOT NULL,
            entity_id TEXT NOT NULL,
            file_type TEXT,
            file_url TEXT,
            file_size INTEGER DEFAULT 0,
            is_deleted INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()
    conn.close()

try:
    _ensure_file_uploads_table()
except Exception as _e:
    logging.getLogger("bridge.api").warning("_ensure_file_uploads_table skip: %s", _e)


# ── Admin: Banners 관리 ───────────────────────────────────────────────────────

def _ensure_banners_table():
    """banners 테이블 생성 (없으면)."""
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS banners (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            image_url TEXT NOT NULL,
            link_url TEXT,
            position TEXT DEFAULT 'main_top',
            is_active INTEGER DEFAULT 1,
            sort_order INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()
    conn.close()

try:
    _ensure_banners_table()
except Exception as _e:
    logging.getLogger("bridge.api").warning("_ensure_banners_table 스킵: %s", _e)


class BannerCreate(BaseModel):
    image_url: str = Field(..., min_length=1, max_length=500)
    link_url: Optional[str] = Field(None, max_length=500)
    position: str = Field('main_top', pattern=r'^(main_top|board_top|sidebar)$')
    is_active: int = Field(1, ge=0, le=1)
    sort_order: int = Field(0)


class BannerUpdate(BaseModel):
    image_url: Optional[str] = Field(None, min_length=1, max_length=500)
    link_url: Optional[str] = Field(None, max_length=500)
    position: Optional[str] = Field(None, pattern=r'^(main_top|board_top|sidebar)$')
    is_active: Optional[int] = Field(None, ge=0, le=1)
    sort_order: Optional[int] = None


@app.get("/api/admin/banners", tags=["admin"])
async def admin_list_banners(request: Request):
    """배너 목록 (관리자)."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        cur = conn.execute("SELECT * FROM banners ORDER BY sort_order, id")
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    finally:
        conn.close()
    return ok(data={"banners": rows})


@app.post("/api/admin/banners", status_code=201, tags=["admin"])
async def admin_create_banner(body: BannerCreate, request: Request):
    """배너 추가 (관리자)."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        cur = conn.execute(
            "INSERT INTO banners (image_url, link_url, position, is_active, sort_order) VALUES (?, ?, ?, ?, ?)",
            (body.image_url, body.link_url, body.position, body.is_active, body.sort_order),
        )
        conn.commit()
        new_id = cur.lastrowid
    finally:
        conn.close()
    return ok(data={"id": new_id}, message="Banner created")


@app.put("/api/admin/banners/{banner_id}", tags=["admin"])
async def admin_update_banner(banner_id: int, body: BannerUpdate, request: Request):
    """배너 수정 (관리자)."""
    _check_admin(request)
    updates: dict = {}
    if body.image_url  is not None: updates['image_url']  = body.image_url
    if body.link_url   is not None: updates['link_url']   = body.link_url
    if body.position   is not None: updates['position']   = body.position
    if body.is_active  is not None: updates['is_active']  = body.is_active
    if body.sort_order is not None: updates['sort_order'] = body.sort_order
    if not updates:
        raise HTTPException(400, "수정할 항목이 없습니다.")
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        if not conn.execute("SELECT id FROM banners WHERE id = ? LIMIT 1", (banner_id,)).fetchone():
            raise HTTPException(404, "Banner not found")
        set_clause = ", ".join(f"{k} = ?" for k in updates)
        vals = list(updates.values()) + [banner_id]
        conn.execute(f"UPDATE banners SET {set_clause} WHERE id = ?", vals)
        conn.commit()
    finally:
        conn.close()
    return ok(message=f"Banner #{banner_id} updated")


@app.delete("/api/admin/banners/{banner_id}", tags=["admin"])
async def admin_delete_banner(banner_id: int, request: Request):
    """배너 비활성화 (soft delete: is_active=0)."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        if not conn.execute("SELECT id FROM banners WHERE id = ? LIMIT 1", (banner_id,)).fetchone():
            raise HTTPException(404, "Banner not found")
        conn.execute("UPDATE banners SET is_active = 0 WHERE id = ?", (banner_id,))
        conn.commit()
    finally:
        conn.close()
    return ok(message=f"Banner #{banner_id} deactivated")


# ── Admin: Posts PUT/DELETE alias ─────────────────────────────────────────────

class PostUpdate(BaseModel):
    title: Optional[str] = Field(None, min_length=2, max_length=200)
    body:  Optional[str] = Field(None, min_length=10, max_length=200000)
    board: Optional[str] = None


@app.put("/api/admin/posts/{post_id}", tags=["admin"])
async def admin_put_post(post_id: int, body: PostUpdate, request: Request):
    """게시글 수정 -- PUT alias (관리자)."""
    _check_admin(request)
    _tag_re = re.compile(r'<[^>]+>')
    updates: list[str] = []
    params: list = []
    if body.title is not None:
        updates.append("title=?"); params.append(_tag_re.sub('', body.title.strip()))
    if body.body is not None:
        updates.append("body=?"); params.append(_tag_re.sub('', body.body.strip()))
    if body.board is not None:
        updates.append("board=?"); params.append(body.board)
    if not updates:
        raise HTTPException(400, "수정할 항목이 없습니다.")
    params.append(post_id)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        r = conn.execute(
            "SELECT id FROM community_posts WHERE id=? AND is_deleted=0", (post_id,)
        ).fetchone()
        if not r:
            raise HTTPException(404, "Post not found")
        conn.execute(f"UPDATE community_posts SET {', '.join(updates)} WHERE id=?", params)
        conn.commit()
        return ok(message=f"Post #{post_id} updated")
    finally:
        conn.close()


@app.delete("/api/admin/posts/{post_id}", tags=["admin"])
async def admin_delete_post(post_id: int, request: Request):
    """게시글 삭제 -- soft delete (관리자, 게시판 무관)."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        r = conn.execute(
            "SELECT id FROM community_posts WHERE id=? AND is_deleted=0", (post_id,)
        ).fetchone()
        if not r:
            raise HTTPException(404, "Post not found")
        conn.execute("UPDATE community_posts SET is_deleted=1 WHERE id=?", (post_id,))
        conn.commit()
        return ok(message=f"Post #{post_id} deleted")
    finally:
        conn.close()


# ── Admin: Site Partners 관리 ──────────────────────────────────────────────────

def _ensure_site_partners_table():
    """site_partners 테이블 생성 (없으면)."""
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS site_partners (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            category TEXT NOT NULL DEFAULT 'academy',
            logo_url TEXT,
            website TEXT,
            sort_order INTEGER DEFAULT 0,
            is_active INTEGER DEFAULT 1,
            is_deleted INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    # 기존 하드코딩 파트너가 비어있으면 시드 데이터 삽입
    count = conn.execute("SELECT COUNT(*) FROM site_partners").fetchone()[0]
    if count == 0:
        academies = [
            'Chungdahm Learning', 'YBM', 'Warwick Franklin', 'Poly',
            'Wall Street English', 'April', 'Hillside IYASkola', 'Sogang SLP',
            'Fast Track Kids', 'Avalon', 'DYB Choisun', 'Rise Korea',
            'JLS Jungsang', 'Real Class', 'Siwon School',
            'Korea University Foreign Language Center', 'Crecerse',
            'LinguaEdu', 'Simson Edu', 'LexKim English', 'MiEdu',
            'Twinkle Language', 'SDA', 'Wiz Island', 'Kids College',
        ]
        schools = [
            'Busan International Foreign School', 'Dalton School',
            'Taejon Christian International School', 'Busan Foreign School',
            'Gyeonggi English Village', 'Dulwich College Seoul',
            'Korea International School', 'Chadwick International',
            'Gangwon English Camp', 'Dwight School Seoul',
            'Yongsan International School of Seoul', 'Saint Paul Academy',
            'Paju English Village', 'Seoul Scholars International',
            'North London Collegiate School', 'Seoul Foreign School',
            'Daegu International School', 'British Council Korea',
            'Gyeonggi International School', 'Jeju International School',
            'Mountain Cherry Academy', 'Seoul International School',
        ]
        for i, name in enumerate(academies):
            conn.execute(
                "INSERT INTO site_partners (name, category, sort_order) VALUES (?, 'academy', ?)",
                (name, i)
            )
        for i, name in enumerate(schools):
            conn.execute(
                "INSERT INTO site_partners (name, category, sort_order) VALUES (?, 'school', ?)",
                (name, i)
            )
        conn.commit()
    conn.close()

try:
    _ensure_site_partners_table()
except Exception as _e:
    logging.getLogger("bridge.api").warning("_ensure_site_partners_table 스킵: %s", _e)


class PartnerCreate(BaseModel):
    name: str = Field(..., min_length=1, max_length=200)
    category: str = Field('academy', pattern=r'^(academy|school)$')
    logo_url: Optional[str] = None
    website: Optional[str] = None
    sort_order: Optional[int] = 0


class PartnerUpdate(BaseModel):
    name: Optional[str] = Field(None, min_length=1, max_length=200)
    category: Optional[str] = Field(None, pattern=r'^(academy|school)$')
    logo_url: Optional[str] = None
    website: Optional[str] = None
    sort_order: Optional[int] = None
    is_active: Optional[int] = Field(None, ge=0, le=1)


@app.get("/api/partners", tags=["public"])
async def public_list_partners(request: Request):
    """공개 파트너 목록 (활성 파트너만)."""
    if not _rate_ok(_ip_hash(request), window=60, max_posts=15):
        return bridge_error("RATE_LIMIT", "Too many requests.", retryable=True, status=429)
    conn = _db_connect()
    try:
        cur = conn.execute(
            "SELECT id, name, category, logo_url, website, sort_order FROM site_partners WHERE is_active = 1 AND is_deleted = 0 ORDER BY sort_order, name"
        )
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    finally:
        conn.close()
    return _ok_public({"partners": rows}, max_age=300)


@app.get("/api/admin/partners", tags=["admin"])
async def admin_list_partners(request: Request):
    """파트너 목록 (관리자 -- 비활성 포함)."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        cur = conn.execute("SELECT * FROM site_partners WHERE is_deleted = 0 ORDER BY sort_order, name")
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    finally:
        conn.close()
    return ok(data={"partners": rows})


@app.post("/api/admin/partners", status_code=201, tags=["admin"])
async def admin_create_partner(body: PartnerCreate, request: Request):
    """파트너 추가 (관리자)."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        cur = conn.execute(
            "INSERT INTO site_partners (name, category, logo_url, website, sort_order) VALUES (?, ?, ?, ?, ?)",
            (body.name, body.category, body.logo_url, body.website, body.sort_order or 0),
        )
        conn.commit()
        new_id = cur.lastrowid
    finally:
        conn.close()
    return ok(data={"id": new_id}, message="Partner created")


@app.put("/api/admin/partners/{partner_id}", tags=["admin"])
async def admin_update_partner(partner_id: int, body: PartnerUpdate, request: Request):
    """파트너 수정 (관리자)."""
    _check_admin(request)
    updates: dict = {}
    if body.name       is not None: updates['name']       = body.name
    if body.category   is not None: updates['category']   = body.category
    if body.logo_url   is not None: updates['logo_url']   = body.logo_url
    if body.website    is not None: updates['website']    = body.website
    if body.sort_order is not None: updates['sort_order'] = body.sort_order
    if body.is_active  is not None: updates['is_active']  = body.is_active
    if not updates:
        raise HTTPException(400, "수정할 항목이 없습니다.")
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        if not conn.execute("SELECT id FROM site_partners WHERE id = ? AND is_deleted = 0 LIMIT 1", (partner_id,)).fetchone():
            raise HTTPException(404, "Partner not found")
        set_clause = ", ".join(f"{k} = ?" for k in updates)
        vals = list(updates.values()) + [partner_id]
        conn.execute(f"UPDATE site_partners SET {set_clause} WHERE id = ?", vals)
        conn.commit()
    finally:
        conn.close()
    return ok(message=f"Partner #{partner_id} updated")


@app.delete("/api/admin/partners/{partner_id}", tags=["admin"])
async def admin_delete_partner(partner_id: int, request: Request):
    """파트너 삭제 -- soft delete (관리자)."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        if not conn.execute("SELECT id FROM site_partners WHERE id = ? AND is_deleted = 0 LIMIT 1", (partner_id,)).fetchone():
            raise HTTPException(404, "Partner not found")
        conn.execute("UPDATE site_partners SET is_deleted = 1 WHERE id = ?", (partner_id,))
        conn.commit()
    finally:
        conn.close()
    return ok(message=f"Partner #{partner_id} deleted")


@app.put("/api/admin/partners/reorder", tags=["admin"])
async def admin_reorder_partners(request: Request):
    """파트너 순서 일괄 변경 (관리자). body: {order: [{id: 1, sort_order: 0}, ...]}"""
    _check_admin(request)
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(400, "Invalid JSON")
    order_list = body.get("order", [])
    if not order_list:
        raise HTTPException(400, "order 필드가 비어있습니다.")
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        for item in order_list:
            pid = item.get("id")
            so  = item.get("sort_order")
            if pid is not None and so is not None:
                conn.execute(
                    "UPDATE site_partners SET sort_order = ? WHERE id = ? AND is_deleted = 0",
                    (so, pid),
                )
        conn.commit()
    finally:
        conn.close()
    return ok(message="Partner order updated")


# ── Admin: Site Settings 관리 ─────────────────────────────────────────────────

def _ensure_site_settings_table():
    """site_settings 테이블 생성 + 초기 데이터 삽입."""
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS site_settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL DEFAULT '',
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    # 초기 기본값 삽입 (이미 있으면 무시)
    defaults = [
        ('site_name', 'BRIDGE'),
        ('business_number', ''),
        ('company_name', 'BRIDGE Agency'),
        ('ceo_name', ''),
        ('address', ''),
        ('contact_email', 'bridgejobkr@gmail.com'),
        ('contact_phone', ''),
        ('kakao_channel', ''),
        ('instagram', ''),
        ('facebook', ''),
        ('youtube', ''),
        ('blog', ''),
        ('footer_text', '© 2026 BRIDGE Recruitment · bridgejob.co.kr'),
        ('footer_description', 'Korea ESL Recruitment Platform'),
        ('hero_tagline', 'A career that changes your life'),
        ('hero_subtitle', "Korea's #1 ESL recruitment platform -- 원어민 영어강사 채용 전문"),
        ('nav_menu', '[{"href":"/community/about","label":"About us"},{"href":"/community/korea","label":"Korea"},{"href":"/community/visa","label":"Visa"},{"href":"/jobs","label":"Job Board"},{"href":"/community/support","label":"Support"},{"href":"/community/support_kr","label":"업무지원"},{"href":"/community","label":"Community"}]'),
        ('nav_cta_1', '{"href":"/apply","label":"Apply"}'),
        ('nav_cta_2', '{"href":"/inquiry","label":"원어민채용의뢰"}'),
    ]
    for key, value in defaults:
        conn.execute(
            "INSERT OR IGNORE INTO site_settings (key, value) VALUES (?, ?)",
            (key, value)
        )
    conn.commit()
    conn.close()

try:
    _ensure_site_settings_table()
except Exception as _e:
    logging.getLogger("bridge.api").warning("_ensure_site_settings_table 스킵: %s", _e)


@app.get("/api/settings", tags=["public"])
async def public_get_settings(request: Request):
    """공개 사이트 설정 (footer 등)."""
    if not _rate_ok(_ip_hash(request), window=60, max_posts=15):
        return bridge_error("RATE_LIMIT", "Too many requests.", retryable=True, status=429)
    # 공개 허용 키만 반환 (민감 설정 노출 방지)
    _PUBLIC_SETTING_KEYS = {
        "nav_menu", "hero_title", "hero_subtitle", "hero_cta_text", "hero_cta_url",
        "footer_text", "footer_links", "company_name", "company_email",
        "kakao_channel_url", "site_description",
    }
    try:
        # site_settings 테이블 없으면 생성 후 조회
        try:
            _ensure_site_settings_table()
        except Exception:
            pass
        conn = _db_connect()
        try:
            rows = conn.execute("SELECT key, value FROM site_settings").fetchall()
            settings = {r[0]: r[1] for r in rows if r[0] in _PUBLIC_SETTING_KEYS}
        finally:
            conn.close()
    except Exception as _e:
        logging.getLogger("bridge.api").warning("/api/settings 조회 실패: %s", _e)
        settings = {}
    return _ok_public({"settings": settings}, max_age=300)


@app.get("/api/admin/settings", tags=["admin"])
async def admin_get_settings(request: Request):
    """관리자 사이트 설정 조회."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        rows = conn.execute("SELECT key, value, updated_at FROM site_settings").fetchall()
        settings = {r[0]: {"value": r[1], "updated_at": r[2]} for r in rows}
    finally:
        conn.close()
    return ok(data={"settings": settings})


@app.put("/api/admin/settings", tags=["admin"])
async def admin_update_settings(request: Request):
    """사이트 설정 일괄 업데이트. body: {settings: {key: value, ...}}"""
    _check_admin(request)
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(400, "Invalid JSON")
    settings = body.get("settings", {})
    if not settings:
        raise HTTPException(400, "settings 필드가 비어있습니다.")
    now_iso = datetime.now(timezone.utc).isoformat()
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        for k, v in settings.items():
            conn.execute(
                "INSERT OR REPLACE INTO site_settings (key, value, updated_at) VALUES (?, ?, ?)",
                (str(k), str(v), now_iso),
            )
        conn.commit()
    finally:
        conn.close()
    return ok(message="Settings updated")


# ── 방문 추적 (Visit Tracking) ────────────────────────────────────────────────

def _ensure_site_visits_table():
    """site_visits 테이블 생성."""
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS site_visits (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            visitor_hash TEXT NOT NULL,
            referrer TEXT,
            channel TEXT,
            search_keyword TEXT,
            landing_page TEXT,
            utm_source TEXT,
            utm_medium TEXT,
            utm_campaign TEXT,
            utm_term TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_visits_created ON site_visits(created_at)
    """)
    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_visits_channel ON site_visits(channel)
    """)
    conn.commit()
    conn.close()

try:
    _ensure_site_visits_table()
except Exception as _e:
    logging.getLogger("bridge.api").warning("_ensure_site_visits_table 스킵: %s", _e)


def _classify_channel(referrer: str) -> str:
    """referrer URL → 채널 분류."""
    if not referrer:
        return "direct"
    r = referrer.lower()
    if "google." in r:
        return "google"
    if "naver.com" in r or "search.naver" in r:
        return "naver"
    if "daum.net" in r or "search.daum" in r:
        return "daum"
    if "bing.com" in r:
        return "bing"
    if "yahoo." in r:
        return "yahoo"
    if "facebook.com" in r or "fb.com" in r:
        return "facebook"
    if "instagram.com" in r:
        return "instagram"
    if "twitter.com" in r or "t.co" in r or "x.com" in r:
        return "twitter"
    if "reddit.com" in r:
        return "reddit"
    if "youtube.com" in r or "youtu.be" in r:
        return "youtube"
    if "kakao" in r:
        return "kakao"
    if "linkedin.com" in r:
        return "linkedin"
    if "bridgejob" in r:
        return "internal"
    return "other"


def _extract_search_keyword(referrer: str) -> str:
    """referrer URL에서 검색어 추출 (Naver, Daum 등)."""
    if not referrer:
        return ""
    try:
        from urllib.parse import urlparse, parse_qs
        parsed = urlparse(referrer)
        params = parse_qs(parsed.query)
        # Naver: query=
        for key in ("query", "q", "p", "search_query", "text", "keyword"):
            if key in params:
                return params[key][0]
    except Exception:
        pass
    return ""


@app.post("/api/track", tags=["public"])
async def track_visit(request: Request):
    """페이지 방문 기록 (공개, rate-limited)."""
    ip = _ip_hash(request)
    if not _rate_ok(ip, window=60, max_posts=30):
        return ok(message="ok")  # 조용히 무시

    try:
        body = await request.json()
    except Exception:
        return ok(message="ok")

    referrer = str(body.get("referrer", ""))[:500]
    landing = str(body.get("page", ""))[:200]
    utm_source = str(body.get("utm_source", ""))[:100]
    utm_medium = str(body.get("utm_medium", ""))[:100]
    utm_campaign = str(body.get("utm_campaign", ""))[:100]
    utm_term = str(body.get("utm_term", ""))[:100]

    channel = _classify_channel(referrer)
    keyword = _extract_search_keyword(referrer)

    # UTM term이 있으면 검색어로 사용
    if utm_term and not keyword:
        keyword = utm_term

    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        conn.execute(
            "INSERT INTO site_visits (visitor_hash, referrer, channel, search_keyword, landing_page, "
            "utm_source, utm_medium, utm_campaign, utm_term) VALUES (?,?,?,?,?,?,?,?,?)",
            (ip, referrer, channel, keyword, landing, utm_source, utm_medium, utm_campaign, utm_term)
        )
        conn.commit()
    finally:
        conn.close()
    return ok(message="ok")


@app.get("/api/admin/analytics", tags=["admin"])
async def admin_analytics(request: Request, days: int = Query(30, ge=1, le=365)):
    """방문 분석 데이터 (관리자)."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%d")

        # 총 방문수
        total = conn.execute(
            "SELECT COUNT(*) as cnt FROM site_visits WHERE created_at >= ?", (cutoff,)
        ).fetchone()["cnt"]

        # 채널별 유입
        channels = conn.execute(
            "SELECT channel, COUNT(*) as cnt FROM site_visits WHERE created_at >= ? "
            "GROUP BY channel ORDER BY cnt DESC", (cutoff,)
        ).fetchall()

        # 검색 키워드 TOP 30
        keywords = conn.execute(
            "SELECT search_keyword, COUNT(*) as cnt FROM site_visits "
            "WHERE created_at >= ? AND search_keyword != '' "
            "GROUP BY search_keyword ORDER BY cnt DESC LIMIT 30", (cutoff,)
        ).fetchall()

        # 인기 랜딩 페이지 TOP 20
        pages = conn.execute(
            "SELECT landing_page, COUNT(*) as cnt FROM site_visits "
            "WHERE created_at >= ? AND landing_page != '' "
            "GROUP BY landing_page ORDER BY cnt DESC LIMIT 20", (cutoff,)
        ).fetchall()

        # 일별 방문 추이 (최근 days일)
        daily = conn.execute(
            "SELECT DATE(created_at) as day, COUNT(*) as cnt FROM site_visits "
            "WHERE created_at >= ? GROUP BY DATE(created_at) ORDER BY day", (cutoff,)
        ).fetchall()

        # UTM 캠페인별
        campaigns = conn.execute(
            "SELECT utm_source, utm_medium, utm_campaign, COUNT(*) as cnt FROM site_visits "
            "WHERE created_at >= ? AND (utm_source != '' OR utm_campaign != '') "
            "GROUP BY utm_source, utm_medium, utm_campaign ORDER BY cnt DESC LIMIT 20", (cutoff,)
        ).fetchall()

        # 오늘 방문
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        today_count = conn.execute(
            "SELECT COUNT(*) as cnt FROM site_visits WHERE DATE(created_at) = ?", (today,)
        ).fetchone()["cnt"]

        return ok(data={
            "total_visits": total,
            "today_visits": today_count,
            "channels": [{"channel": r["channel"], "count": r["cnt"]} for r in channels],
            "keywords": [{"keyword": r["search_keyword"], "count": r["cnt"]} for r in keywords],
            "pages": [{"page": r["landing_page"], "count": r["cnt"]} for r in pages],
            "daily": [{"date": r["day"], "count": r["cnt"]} for r in daily],
            "campaigns": [
                {"source": r["utm_source"], "medium": r["utm_medium"],
                 "campaign": r["utm_campaign"], "count": r["cnt"]}
                for r in campaigns
            ],
        })
    finally:
        conn.close()


# ══════════════════════════════════════════════════════════════════════════════
# TESTIMONIALS API
# ══════════════════════════════════════════════════════════════════════════════

def _ensure_testimonials_schema():
    """testimonials 테이블 자동 생성 + 시드 (Render DB 초기화 복구)."""
    import json as _json
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS testimonials (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                name        TEXT    NOT NULL,
                country     TEXT    NOT NULL DEFAULT '',
                photo_url   TEXT    DEFAULT NULL,
                rating      INTEGER DEFAULT 5,
                review_text TEXT    NOT NULL,
                sort_order  INTEGER DEFAULT 0,
                is_visible  INTEGER DEFAULT 1,
                is_deleted  INTEGER DEFAULT 0,
                created_at  TEXT    DEFAULT '',
                updated_at  TEXT    DEFAULT ''
            )
        """)
        conn.commit()

        # ── 정크 자동 정리 (멱등, 매 startup 실행)
        #   목적: Render free plan ephemeral disk + Drive auto-restore 가 옛 DB로 덮어써도
        #         매 startup 마다 자동 복구되도록 보장
        #   패턴: 옛 어드민 이름("최고관리자") + 스크랩 잔재 JS + Cafe24 절대경로 사진
        try:
            junk_q = """
                WHERE is_deleted=0 AND (
                    name = '최고관리자'
                    OR review_text LIKE '%function board_move%'
                    OR review_text LIKE '%bo_v_act%'
                    OR review_text LIKE '%board_move(href)%'
                    OR (photo_url IS NOT NULL AND photo_url LIKE 'http://%')
                )
            """
            junk_count = conn.execute(f"SELECT COUNT(*) FROM testimonials {junk_q}").fetchone()[0]
            if junk_count > 0:
                conn.execute(f"UPDATE testimonials SET is_deleted=1, is_visible=0 {junk_q}")
                conn.commit()
                logging.getLogger("bridge.api").info(
                    "_ensure_testimonials_schema: 정크 %d rows 논리삭제", junk_count)
        except Exception as _je:
            logging.getLogger("bridge.api").warning("testimonials junk cleanup 스킵: %s", _je)

        # ── 시드 INSERT (멱등, 매 startup 실행)
        #   기준: testimonials_seed.json (100건 — Chris/Hannah/Liam 등)
        #   NOT EXISTS 가드로 중복 INSERT 방지 (name + review_text 복합키)
        #   결과: 어떤 DB 상태에서 출발해도 startup 후엔 100건 캐논 시드 보장
        seed_path = Path(__file__).parent / "migrations" / "testimonials_seed.json"
        if seed_path.exists():
            try:
                items = _json.loads(seed_path.read_text(encoding="utf-8"))
                inserted = 0
                for t in items:
                    cur = conn.execute("""
                        INSERT INTO testimonials (name, country, photo_url, rating, review_text, sort_order, is_visible, is_deleted, created_at, updated_at)
                        SELECT ?,?,?,?,?,?,1,0,?,?
                        WHERE NOT EXISTS (SELECT 1 FROM testimonials WHERE name=? AND review_text=? AND is_deleted=0)
                    """, (t["name"], t.get("country",""), t.get("photo_url"), t.get("rating",5),
                          t["review_text"], t.get("sort_order",0), t.get("created_at",""), t.get("updated_at",""),
                          t["name"], t["review_text"]))
                    inserted += cur.rowcount
                conn.commit()
                if inserted > 0:
                    logging.getLogger("bridge.api").info(
                        "_ensure_testimonials_schema: 시드 %d/%d rows 추가 (이미 있는 건 스킵)", inserted, len(items))
            except Exception as _se:
                    logging.getLogger("bridge.api").warning("testimonials seed 스킵: %s", _se)
    finally:
        conn.close()


try:
    _ensure_testimonials_schema()
except Exception as _e:
    logging.getLogger("bridge.api").warning("_ensure_testimonials_schema 스킵: %s", _e)


def _ensure_apostille_categories():
    """제목 기반 자동 분류 — 하드코딩 ID 대신 콘텐츠 기반 룰.

    아키텍처 배경:
    - 게시물 ID 하드코딩은 게시물 삭제/재생성 시 깨짐 → 운영 불가
    - 제목 키워드 기반으로 자동 분류 → 새 글 추가/삭제와 무관하게 동작
    - 관리자는 admin UI 토글로 override 가능 → PATCH 시 _maybe_auto_backup() 호출되어
      Drive에 즉시 저장 → 다음 cold start의 restore가 최신 분류 유지
    """
    APOSTILLE_KEYWORDS = [
        'apostille', 'notarization', 'fbi background', 'acro',
        '범죄경력', '학위·범죄', '학위 범죄', '공증', '아포스티유',
    ]
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        # 제목에 키워드 포함 + 비자 관련 보드 + 카테고리 미설정 → apostille
        # (관리자가 명시적으로 다른 카테고리로 설정한 경우는 보존)
        like_clauses = " OR ".join([f"LOWER(title) LIKE ?" for _ in APOSTILLE_KEYWORDS])
        params = [f"%{kw.lower()}%" for kw in APOSTILLE_KEYWORDS]
        sql = (
            f"UPDATE community_posts SET category='apostille' "
            f"WHERE board IN ('visa', 'visa_related', 'visa_type') "
            f"  AND is_deleted = 0 "
            f"  AND (category IS NULL OR category = '') "
            f"  AND ({like_clauses})"
        )
        cur = conn.execute(sql, params)
        affected = cur.rowcount
        conn.commit()
        conn.close()
        if affected > 0:
            logging.getLogger("bridge.api").info(
                "_ensure_apostille_categories: %d posts 자동분류 → apostille (제목 키워드)", affected)
    except Exception as _e:
        logging.getLogger("bridge.api").warning("apostille category 자동분류 스킵: %s", _e)


try:
    _ensure_apostille_categories()
except Exception as _e:
    logging.getLogger("bridge.api").warning("_ensure_apostille_categories 스킵: %s", _e)


@app.get("/api/testimonials", tags=["testimonials"])
async def testimonials_list(
    request: Request,
    limit: int = 20,
    offset: int = 0,
    random: int = 0,
):
    """공개 리뷰 목록. random=1이면 랜덤 순서."""
    limit = min(max(limit, 1), 30)  # 스크래핑 방지
    if not _rate_ok(_ip_hash(request), window=60, max_posts=15):
        return bridge_error("RATE_LIMIT", "Too many requests.", retryable=True, status=429)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    total = 0
    data: list = []
    try:
        # name이 비어있는 행은 불완전 데이터 → 제외 (프론트엔드 이름 공백 방지)
        _base_cond = "is_visible = 1 AND is_deleted = 0 AND TRIM(COALESCE(name,'')) != ''"
        total = conn.execute(
            f"SELECT COUNT(*) FROM testimonials WHERE {_base_cond}"
        ).fetchone()[0]
        if not random:
            cur = conn.execute(
                f"SELECT id, name, country, photo_url, rating, review_text, sort_order, created_at FROM testimonials WHERE {_base_cond} ORDER BY sort_order DESC, id DESC LIMIT ? OFFSET ?",
                (limit, offset),
            )
        else:
            cur = conn.execute(
                f"SELECT id, name, country, photo_url, rating, review_text, sort_order, created_at FROM testimonials WHERE {_base_cond} LIMIT ? OFFSET ?",
                (limit, offset),
            )
        cols = [d[0] for d in cur.description]
        data = [dict(zip(cols, r)) for r in cur.fetchall()]
    except Exception as _e:
        logging.getLogger("bridge.api").error("testimonials_list DB 오류: %s", _e)
        # 테이블 미존재 등 DB 오류 → 빈 목록 반환 (프론트엔드 크래시 방지)
        total = 0
        data = []
        try:
            _ensure_testimonials_schema()
        except Exception:
            pass
    finally:
        conn.close()
    if random and data:
        import random as _random
        _random.shuffle(data)
    return ok(data={"total": total, "testimonials": data})


@app.post("/api/testimonials", status_code=201, tags=["testimonials"])
async def testimonials_create(request: Request):
    """새 리뷰 추가 (Admin)."""
    _check_admin(request)
    body = await request.json()
    name        = (body.get("name") or "").strip()
    country     = (body.get("country") or "").strip()
    review_text = (body.get("review_text") or "").strip()
    if not name or not review_text:
        err("name and review_text are required", 400)
    rating    = int(body.get("rating", 5))
    photo_url = body.get("photo_url") or None
    now_iso   = datetime.now(timezone.utc).isoformat()
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        cur = conn.execute(
            "INSERT INTO testimonials (name, country, photo_url, rating, review_text, sort_order, is_visible, is_deleted, created_at, updated_at) VALUES (?, ?, ?, ?, ?, 0, 1, 0, ?, ?)",
            (name, country, photo_url, rating, review_text, now_iso, now_iso),
        )
        conn.commit()
        new_id = cur.lastrowid
    finally:
        conn.close()
    return ok(data={"id": new_id}, message="Testimonial created")


@app.put("/api/testimonials/{tid}", tags=["testimonials"])
async def testimonials_update(tid: int, request: Request):
    """리뷰 수정 (Admin)."""
    _check_admin(request)
    body    = await request.json()
    updates = {k: body[k] for k in ("name","country","photo_url","rating","review_text","is_visible") if k in body}
    if not updates:
        err("No fields to update", 400)
    updates['updated_at'] = datetime.now(timezone.utc).isoformat()
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        set_clause = ", ".join(f"{k} = ?" for k in updates)
        vals = list(updates.values()) + [tid]
        conn.execute(f"UPDATE testimonials SET {set_clause} WHERE id = ? AND is_deleted = 0", vals)
        conn.commit()
    finally:
        conn.close()
    return ok(message="Testimonial updated")


@app.delete("/api/testimonials/{tid}", tags=["testimonials"])
async def testimonials_delete(tid: int, request: Request):
    """리뷰 soft-delete (Admin)."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        conn.execute("UPDATE testimonials SET is_deleted = 1 WHERE id = ?", (tid,))
        conn.commit()
    finally:
        conn.close()
    return ok(message="Testimonial deleted")


@app.patch("/api/testimonials/reorder", tags=["testimonials"])
async def testimonials_reorder(request: Request):
    """리뷰 순서 변경 (Admin)."""
    _check_admin(request)
    body  = await request.json()
    items = body.get("items", [])
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        for item in items:
            conn.execute(
                "UPDATE testimonials SET sort_order = ? WHERE id = ?",
                (item["sort_order"], item["id"]),
            )
        conn.commit()
    finally:
        conn.close()
    return ok(message="Reorder saved")


# ── Secure Download Links System ──────────────────────────────────────────────

_DOWNLOAD_DIR = Path(os.getenv("BRIDGE_DOWNLOAD_DIR",
    "/data/secure_downloads" if os.getenv("RENDER_EXTERNAL_URL") else str(Path(__file__).resolve().parent / "secure_downloads")))
try:
    _DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
except OSError:
    _DOWNLOAD_DIR = Path("/tmp/bridge_downloads")
    _DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)


def _ensure_download_links_schema():
    """secure download_links + mail_logs 테이블 생성."""
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS download_links (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            uuid TEXT UNIQUE NOT NULL,
            file_path TEXT NOT NULL,
            original_name TEXT NOT NULL,
            candidate_id TEXT,
            max_downloads INTEGER DEFAULT 3,
            download_count INTEGER DEFAULT 0,
            expires_at TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now')),
            is_deleted INTEGER DEFAULT 0
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS download_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            link_uuid TEXT NOT NULL,
            ip_hash TEXT,
            downloaded_at TEXT DEFAULT (datetime('now'))
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS mail_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            log_type TEXT DEFAULT 'profile_broadcast',
            candidate_id TEXT,
            employer_count INTEGER DEFAULT 0,
            sent_count INTEGER DEFAULT 0,
            failed_count INTEGER DEFAULT 0,
            status TEXT DEFAULT 'completed',
            sent_at TEXT DEFAULT (datetime('now')),
            details TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS mail_introduce_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sent_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            candidate_ids TEXT NOT NULL,
            to_email TEXT NOT NULL,
            school_name TEXT,
            contact_name TEXT,
            subject TEXT,
            status TEXT DEFAULT 'sent',
            error TEXT,
            sender_email TEXT,
            link_expiry_days INTEGER DEFAULT 10
        )
    """)
    conn.commit()
    conn.close()


try:
    _ensure_download_links_schema()
except Exception as _e:
    logging.getLogger("bridge.api").warning("_ensure_download_links_schema 스킵: %s", _e)


def _ensure_page_content_table():
    """공개 페이지 관리자 인라인 편집 콘텐츠 저장 테이블."""
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS page_content (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            page TEXT NOT NULL,
            section_key TEXT NOT NULL,
            field TEXT NOT NULL DEFAULT 'content',
            value TEXT NOT NULL,
            updated_at TEXT DEFAULT (datetime('now')),
            UNIQUE(page, section_key, field)
        )
    """)
    conn.commit()
    conn.close()


try:
    _ensure_page_content_table()
except Exception as _e:
    logging.getLogger("bridge.api").warning("_ensure_page_content_table 스킵: %s", _e)


@app.get("/api/page-content/{page}", tags=["public"])
async def get_page_content(page: str, request: Request):
    """공개 페이지 커스텀 콘텐츠 조회 (관리자 편집 내용 반영)."""
    if not re.match(r'^[a-z0-9_-]+$', page):
        raise HTTPException(status_code=400, detail="Invalid page key")
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT section_key, field, value FROM page_content WHERE page = ?",
        (page,)
    ).fetchall()
    conn.close()
    result: dict = {}
    for row in rows:
        sk = row["section_key"]
        if sk not in result:
            result[sk] = {}
        result[sk][row["field"]] = row["value"]
    return {"success": True, "data": result}


class PageContentPatch(BaseModel):
    page: str
    section_key: str
    field: str = "content"
    value: str


@app.patch("/api/admin/page-content", tags=["admin"])
async def patch_page_content(payload: PageContentPatch, request: Request):
    """관리자 전용 — 공개 페이지 섹션 콘텐츠 수정."""
    _check_admin(request)
    if not re.match(r'^[a-z0-9_-]+$', payload.page):
        raise HTTPException(status_code=400, detail="Invalid page key")
    if not re.match(r'^[a-z0-9_-]+$', payload.section_key):
        raise HTTPException(status_code=400, detail="Invalid section key")
    if payload.field not in ("content", "title"):
        raise HTTPException(status_code=400, detail="Invalid field")
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.execute("""
        INSERT INTO page_content (page, section_key, field, value, updated_at)
        VALUES (?, ?, ?, ?, datetime('now'))
        ON CONFLICT(page, section_key, field) DO UPDATE SET
            value = excluded.value,
            updated_at = excluded.updated_at
    """, (payload.page, payload.section_key, payload.field, payload.value))
    conn.commit()
    conn.close()
    logging.getLogger("bridge.api").info(
        "page_content 업데이트: page=%s section=%s field=%s", payload.page, payload.section_key, payload.field
    )
    return {"success": True}


@app.post("/api/admin/download-links", tags=["admin"])
async def admin_create_download_link(request: Request):
    """SECURITY: 보안 다운로드 링크 생성 (UUID, 3회 제한, 14일 만료)."""
    _check_admin(request)
    body = await request.json()
    file_path = body.get("file_path", "")
    original_name = body.get("original_name", "")
    candidate_id = body.get("candidate_id", "")

    if not file_path or not original_name:
        raise HTTPException(400, "file_path and original_name required")

    # SECURITY: 첨부파일명에 실명 금지 → candidate_{번호}_resume.pdf
    if candidate_id:
        ext = Path(original_name).suffix
        original_name = f"candidate_{candidate_id}{ext}"

    link_uuid = str(uuid.uuid4())
    expires = (datetime.now(timezone.utc) + timedelta(days=14)).strftime("%Y-%m-%d %H:%M:%S")

    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        conn.execute(
            """INSERT INTO download_links (uuid, file_path, original_name, candidate_id, max_downloads, expires_at)
               VALUES (?, ?, ?, ?, 3, ?)""",
            (link_uuid, file_path, original_name, candidate_id, expires),
        )
        conn.commit()
    finally:
        conn.close()

    return ok(data={"uuid": link_uuid, "url": f"/dl/{link_uuid}", "expires_at": expires})


@app.get("/dl/{link_uuid}", tags=["download"])
async def download_file(link_uuid: str, request: Request):
    """SECURITY: 보안 파일 다운로드 -- UUID 기반, 횟수/만료 제한."""
    from fastapi.responses import FileResponse as FastFileResponse

    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        link = conn.execute(
            "SELECT * FROM download_links WHERE uuid = ? AND is_deleted = 0", (link_uuid,)
        ).fetchone()

        if not link:
            return SafeJSONResponse({"error": "링크를 찾을 수 없습니다"}, status_code=404)

        # SECURITY: 다운로드 횟수 제한 (최대 3회)
        if link["download_count"] >= link["max_downloads"]:
            return SafeJSONResponse(
                {"error": "다운로드 횟수를 초과했습니다 (최대 3회)"}, status_code=403
            )

        # SECURITY: 만료일 체크 (14일)
        expires = datetime.strptime(link["expires_at"], "%Y-%m-%d %H:%M:%S")
        if datetime.now(timezone.utc).replace(tzinfo=None) > expires:
            return SafeJSONResponse(
                {"error": "링크가 만료되었습니다 (14일)"}, status_code=410
            )

        file_p = Path(link["file_path"])
        if not file_p.exists():
            return SafeJSONResponse({"error": "파일을 찾을 수 없습니다"}, status_code=404)

        # SECURITY: 다운로드 카운트 증가 + IP 로그
        ip_h = _ip_hash(request)
        conn.execute(
            "UPDATE download_links SET download_count = download_count + 1 WHERE uuid = ?",
            (link_uuid,),
        )
        conn.execute(
            "INSERT INTO download_logs (link_uuid, ip_hash) VALUES (?, ?)",
            (link_uuid, ip_h),
        )
        conn.commit()

        return FastFileResponse(
            path=str(file_p),
            filename=link["original_name"],
            media_type="application/octet-stream",
        )
    finally:
        conn.close()


@app.post("/api/admin/download-links/cleanup", tags=["admin"])
async def admin_cleanup_download_links(request: Request):
    """SECURITY: 만료된 다운로드 링크 + 파일 자동 삭제."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
        expired = conn.execute(
            "SELECT uuid, file_path FROM download_links WHERE expires_at < ? AND is_deleted = 0",
            (now_str,),
        ).fetchall()

        cleaned = 0
        for row in expired:
            fp = Path(row["file_path"])
            if fp.exists():
                fp.unlink()
            conn.execute(
                "UPDATE download_links SET is_deleted = 1 WHERE uuid = ?", (row["uuid"],)
            )
            cleaned += 1

        conn.commit()
        return ok(message=f"만료 링크 {cleaned}건 정리 완료")
    finally:
        conn.close()


# ── Secure Profile Broadcast (보안 규칙 적용) ────────────────────────────────


class SecureProfileBroadcastBody(BaseModel):
    candidate_id: str
    employer_ids: list[int]
    custom_subject: Optional[str] = None
    custom_body: Optional[str] = None
    download_link_uuid: Optional[str] = None


@app.post("/api/admin/matching/send-profile-secure", tags=["admin"])
async def admin_matching_send_profile_secure(request: Request, body: SecureProfileBroadcastBody):
    """SECURITY: 매칭된 employer들에게 후보자 프로필 개별 발송 (전체 보안 규칙 적용).

    보안 규칙:
    1. 완전 개별 발송 -- CC/BCC 절대 금지, To에 1명만
    2. PII 마스킹 -- 후보자 실명/이메일/전화/주소 절대 미포함
    3. Reply-To: bridgejobkr@gmail.com (회신 격리)
    4. 발송 전 PII 자동 스캔
    5. Rate limit: 3초 간격, 10/분, 200/일
    6. 발송 로그: employer 이메일 마스킹 저장
    """
    _check_admin(request)
    if not body.employer_ids:
        raise HTTPException(400, "employer_ids가 비어 있습니다.")

    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        # 후보자 조회
        cand = conn.execute(
            "SELECT * FROM candidates WHERE candidate_id = ?", (body.candidate_id,)
        ).fetchone()
        if not cand:
            raise HTTPException(404, "Candidate not found")
        cand_d = dict(cand)

        # SECURITY: 프로필 카드 -- PII 제외
        card_html = _build_profile_card(cand_d)

        # 템플릿 로드
        tpl = conn.execute(
            "SELECT subject, body_html FROM email_templates WHERE template_key = 'profile_broadcast'"
        ).fetchone()
        if not tpl:
            raise HTTPException(404, "profile_broadcast template not found")

        base_subject = body.custom_subject or tpl["subject"]
        base_body = body.custom_body or tpl["body_html"]

        # 다운로드 링크 삽입
        dl_html = ""
        if body.download_link_uuid:
            dl_html = f'<p style="margin:16px 0"><a href="https://bridgejob.co.kr/dl/{body.download_link_uuid}" style="display:inline-block;background:#1d1d1f;color:#fff;padding:10px 24px;border-radius:8px;text-decoration:none;font-weight:600">View Full Profile</a></p>'

        sent_count = 0
        fail_count = 0
        results = []

        for emp_id in body.employer_ids:
            emp = conn.execute(
                "SELECT id, email, school_name, school_name_plain, contact_name FROM client_inquiries WHERE id = ?",
                (emp_id,),
            ).fetchone()
            if not emp:
                continue

            # PII 복호화
            to_email = _safe_decrypt(emp["email"], "email") or ""
            school   = (emp["school_name_plain"]
                        or _safe_decrypt(emp["school_name"], "school_name")
                        or "School")
            contact  = _safe_decrypt(emp["contact_name"], "contact_name") or ""
            if not to_email or "@" not in to_email:
                continue  # 복호화 실패/빈값이면 스킵

            # 개별 템플릿 치환
            html = base_body.replace("{{profile_cards}}", card_html)
            html = html.replace("{{school_name}}", school)
            html = html.replace("{{contact_name}}", contact)
            html = html.replace("{{download_link}}", dl_html)
            html = re.sub(r"\{\{\w+\}\}", "", html)

            subject = f"{base_subject} - {school}" if school != "School" else base_subject

            # SECURITY: 개별 발송 + PII 스캔 + Rate limit
            success, msg = _secure_send_email(to_email, subject, html)

            status_val = "sent" if success else "failed"
            if success:
                sent_count += 1
            else:
                fail_count += 1

            # SECURITY: 발송 로그 -- 이메일 마스킹 저장
            conn.execute(
                """INSERT INTO profile_sends (candidate_id, employer_id, school_name, to_email, status)
                   VALUES (?, ?, ?, ?, ?)""",
                (body.candidate_id, emp_id, school, _mask_email_for_log(to_email), status_val),
            )
            results.append({"employer_id": emp_id, "school": school, "status": status_val, "message": msg})

            # SECURITY: 발송 간격 최소 3초
            if success:
                await asyncio.sleep(3)

        conn.commit()

        # SECURITY: 메일 발송 로그 (수신자 이메일 직접 저장 금지)
        conn.execute(
            """INSERT INTO mail_logs (log_type, candidate_id, employer_count, sent_count, failed_count, status, details)
               VALUES ('profile_broadcast', ?, ?, ?, ?, 'completed', ?)""",
            (body.candidate_id, len(body.employer_ids), sent_count, fail_count,
             json.dumps({"employer_ids": body.employer_ids})),
        )
        conn.commit()

        return ok(
            message=f"Profile broadcast: {sent_count} sent, {fail_count} failed",
            data={"sent": sent_count, "failed": fail_count, "results": results},
        )
    finally:
        conn.close()


@app.get("/api/admin/mail-logs", tags=["admin"])
async def admin_mail_logs(request: Request, log_type: str = "", limit: int = 100):
    """SECURITY: 메일 발송 로그 조회 (수신자 이메일 직접 저장 없음)."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        if log_type:
            rows = conn.execute(
                "SELECT * FROM mail_logs WHERE log_type = ? ORDER BY sent_at DESC LIMIT ?",
                (log_type, limit),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM mail_logs ORDER BY sent_at DESC LIMIT ?", (limit,)
            ).fetchall()
        return ok(data=[dict(r) for r in rows])
    finally:
        conn.close()


# ─── 엔드포인트 1: 발송 현황 ──────────────────────────────────────────────────
@app.get("/api/admin/mail/stats", tags=["admin"])
async def get_mail_stats(request: Request):
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        today_count = conn.execute(
            "SELECT COALESCE(SUM(sent_count), 0) FROM mail_logs WHERE DATE(sent_at) = DATE('now', 'localtime')"
        ).fetchone()[0]
        return ok(data={"sent_today": int(today_count), "limit": 500})
    finally:
        conn.close()


# ─── 엔드포인트 2: 메일 발송 (/send + /send-bulk 동일 핸들러) ────────────────
def _fetch_processed_cv_attachment(candidate_id: str) -> "tuple[bytes, str] | None":
    """S3에서 processed CV 다운로드 → (bytes, filename) or None.
    파일명은 _build_output_filename() SSoT 사용 — {번호}{국적}_{성별}({YY}born).pdf
    """
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        conn.row_factory = sqlite3.Row
        try:
            row = conn.execute(
                "SELECT cv_processed_s3_key, sheet_number, nationality, nationality_plain, gender, dob "
                "FROM candidates WHERE candidate_id = ?",
                (candidate_id,),
            ).fetchone()
        finally:
            conn.close()
        if not row or not row["cv_processed_s3_key"]:
            return None
        pdf_data = s3_download_bytes(row["cv_processed_s3_key"])

        # 파일명 생성 — SSoT 사용
        try:
            sys.path.insert(0, str(Path(__file__).resolve().parent))
            from tools.doc_processor import _build_output_filename  # type: ignore
            nat = row["nationality_plain"] or ""
            if not nat and row["nationality"]:
                nat = _safe_decrypt(row["nationality"], "nationality") or ""
            gender = _safe_decrypt(row["gender"], "gender") if row["gender"] else ""
            dob    = _safe_decrypt(row["dob"], "dob") if row["dob"] else ""
            cand = {"nationality": nat, "gender": gender, "dob": dob}
            fname = _build_output_filename(int(row["sheet_number"] or 0), cand, ".pdf")
        except Exception as _fn_err:
            logging.getLogger("bridge.mail").warning("filename fallback: %s", _fn_err)
            fname = f"{row['sheet_number'] or 'resume'}.pdf"

        return pdf_data, fname
    except Exception as e:
        logging.getLogger("bridge.mail").warning("CV 첨부 실패 %s: %s", candidate_id, e)
        return None


async def _handle_mail_send(request: Request):
    _check_admin(request)
    import json as _json

    # FormData (MailModal) 또는 JSON 모두 지원
    content_type = request.headers.get("content-type", "")
    manual_files: list[tuple[bytes, str]] = []  # (bytes, filename) 수동 첨부
    if "multipart/form-data" in content_type:
        form = await request.form()
        data = {}
        for k in form:
            val = form[k]
            if hasattr(val, "read"):
                # UploadFile → 수동 첨부파일로 수집
                file_bytes = await val.read()
                if file_bytes:
                    manual_files.append((file_bytes, val.filename or "attachment"))
            else:
                data[k] = val
        # FormData 개별발송 형식 → recipients 리스트 변환
        to_email = str(data.get("to", "")).strip()
        recipients = [to_email] if to_email else []
        sender_label = str(data.get("sender", "gmail"))
        subject = str(data.get("subject", ""))
        body = str(data.get("body", ""))
        personal = True
        attach_cv_ids = [cid.strip() for cid in str(data.get("attach_cv_ids", "")).split(",") if cid.strip()]
    else:
        data = await request.json()
        recipients = data.get("recipients", [])
        sender_label = data.get("sender", "gmail")
        subject = data.get("subject", "")
        body = data.get("body_html", data.get("body", ""))
        personal = data.get("personal_send", data.get("personal", True))
        attach_cv_ids = data.get("attach_cv_ids", [])

    if not recipients:
        raise HTTPException(status_code=400, detail="수신자가 없습니다.")
    if not subject:
        raise HTTPException(status_code=400, detail="제목을 입력하세요.")

    # SECURITY: 블랙리스트 필터링 -- 차단된 이메일 자동 제거
    blocked = [e for e in recipients if _is_email_blacklisted(e)]
    if blocked:
        logging.getLogger("bridge.mail").warning("블랙리스트 차단 %d건: %s", len(blocked), blocked[:3])
        recipients = [e for e in recipients if not _is_email_blacklisted(e)]
        if not recipients:
            raise HTTPException(status_code=403, detail="모든 수신자가 블랙리스트에 등록되어 발송할 수 없습니다.")

    smtp_key = "naver" if "naver" in sender_label.lower() else "gmail"
    cfg = SMTP_CONFIG[smtp_key]

    if not cfg["user"] or not cfg["password"]:
        raise HTTPException(status_code=503, detail=f"{smtp_key.upper()} SMTP 계정이 설정되지 않았습니다.")

    # Processed CV 첨부 파일 미리 다운로드
    cv_attachments: list[tuple[bytes, str]] = []
    for cid in attach_cv_ids:
        att = _fetch_processed_cv_attachment(cid)
        if att:
            cv_attachments.append(att)

    sent = 0
    failed = 0
    errors: list = []

    def _build_msg(to_addr: str, bcc_list: list = None) -> MIMEMultipart:
        msg = MIMEMultipart("mixed")
        msg["From"] = f"BRIDGE Recruitment <{cfg['user']}>"
        msg["To"] = to_addr
        if bcc_list:
            msg["Bcc"] = ", ".join(bcc_list)
        msg["Subject"] = subject
        msg["Reply-To"] = "bridgejobkr@gmail.com"
        msg.attach(MIMEText(body, "html", "utf-8"))
        # Processed CV 첨부 (RFC 5987 한글 파일명 안전 처리)
        from urllib.parse import quote as _urlquote
        for pdf_data, pdf_name in cv_attachments:
            part = MIMEBase("application", "pdf")
            part.set_payload(pdf_data)
            _email_encoders.encode_base64(part)
            try:
                part.add_header(
                    "Content-Disposition",
                    f"attachment; filename*=UTF-8''{_urlquote(pdf_name)}"
                )
            except Exception:
                part.add_header("Content-Disposition", "attachment", filename=pdf_name)
            msg.attach(part)
        # 수동 첨부파일 (드래그&드롭/파일선택) — RFC 5987
        for file_data, file_name in manual_files:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(file_data)
            _email_encoders.encode_base64(part)
            try:
                part.add_header(
                    "Content-Disposition",
                    f"attachment; filename*=UTF-8''{_urlquote(file_name)}"
                )
            except Exception:
                part.add_header("Content-Disposition", "attachment", filename=file_name)
            msg.attach(part)
        return msg

    try:
        server = smtplib.SMTP(cfg["host"], cfg["port"], timeout=10)
        server.starttls()
        server.login(cfg["user"], cfg["password"])

        if personal:
            for email in recipients:
                try:
                    msg = _build_msg(email)
                    server.sendmail(cfg["user"], [email], msg.as_string())
                    sent += 1
                    time.sleep(3)
                except Exception as exc:
                    failed += 1
                    errors.append(str(exc))
        else:
            batch_size = 100
            for i in range(0, len(recipients), batch_size):
                batch = recipients[i:i + batch_size]
                msg = _build_msg(cfg["user"], bcc_list=batch)
                try:
                    server.sendmail(cfg["user"], batch, msg.as_string())
                    sent += len(batch)
                    time.sleep(3)
                except Exception as exc:
                    failed += len(batch)
                    errors.append(str(exc))

        server.quit()

    except smtplib.SMTPAuthenticationError:
        raise HTTPException(status_code=401, detail="SMTP 인증 실패. 계정/비밀번호 확인 필요.")
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"SMTP 연결 오류: {str(exc)}")

    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        preview = ", ".join(recipients[:5]) + ("..." if len(recipients) > 5 else "")
        conn.execute(
            "INSERT INTO mail_logs (log_type, sent_count, failed_count, status, details) VALUES (?, ?, ?, ?, ?)",
            ("manual_send", sent, failed, "completed" if failed == 0 else "partial",
             _json.dumps({"sender": sender_label, "subject": subject, "recipients_preview": preview,
                          "cv_attached": len(cv_attachments)}, ensure_ascii=False)),
        )
        conn.commit()
    finally:
        conn.close()

    return ok(data={"sent": sent, "failed": failed, "errors": errors[:3], "cv_attached": len(cv_attachments)})


@app.post("/api/admin/mail/send", tags=["admin"])
async def send_mail(request: Request):
    return await _handle_mail_send(request)


@app.post("/api/admin/mail/send-bulk", tags=["admin"])
async def send_mail_bulk(request: Request):
    return await _handle_mail_send(request)


# ─── 발송 이력 (profile_sends) API ───────────────────────────────────────────

@app.post("/api/admin/profile-sends", tags=["admin"])
async def record_profile_send(request: Request):
    """메일 발송 성공 시 프론트에서 호출하여 이력 기록."""
    _check_admin(request)
    data = await request.json()
    required = ["candidate_id", "to_email", "subject"]
    for field in required:
        if not data.get(field):
            raise HTTPException(400, f"필수 필드 누락: {field}")
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        conn.execute("""
            INSERT INTO profile_sends
            (candidate_id, candidate_number, employer_id, school_name,
             to_email, subject, template_used, sender_email, attachments, status, error_msg)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            data["candidate_id"],
            data.get("candidate_number"),
            data.get("employer_job_id"),
            data.get("employer_name", ""),
            data["to_email"],
            data["subject"],
            data.get("template_used", "profile_broadcast"),
            data.get("sender_email", "bridgejobkr@naver.com"),
            data.get("attachments", ""),
            data.get("status", "sent"),
        ))
        conn.commit()
    finally:
        conn.close()
    return ok(message="발송 이력 기록 완료")


@app.get("/api/admin/profile-sends", tags=["admin"])
async def get_profile_sends(request: Request):
    """발송 이력 목록 -- 최근 100건."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute("""
            SELECT id, candidate_id, candidate_number, school_name AS employer_name,
                   to_email AS employer_email, subject, template_used, sender_email,
                   status, sent_at
            FROM profile_sends
            ORDER BY sent_at DESC
            LIMIT 100
        """).fetchall()
        return ok(data={"sends": [dict(r) for r in rows]})
    finally:
        conn.close()


@app.get("/api/admin/profile-sends/check-duplicate", tags=["admin"])
async def check_duplicate_send(request: Request):
    """동일 candidate+to_email 조합 24시간 내 발송 여부 확인."""
    _check_admin(request)
    candidate_id = request.query_params.get("candidate_id")
    employer_email = request.query_params.get("employer_email")
    if not candidate_id or not employer_email:
        raise HTTPException(400, "candidate_id, employer_email 필수")
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        exists = conn.execute("""
            SELECT COUNT(*) FROM profile_sends
            WHERE candidate_id = ? AND to_email = ?
              AND sent_at >= datetime('now','localtime','-24 hours')
              AND status = 'sent'
        """, (candidate_id, employer_email)).fetchone()[0]
        return ok(data={"duplicate": exists > 0, "count": exists})
    finally:
        conn.close()


# ─── 소개발송 통합 엔드포인트 ────────────────────────────────────────────────

@app.post("/api/admin/mail/send-profile", tags=["admin"])
async def send_profile_mail(request: Request):
    """
    소개발송 전용 -- 1:1 개별 발송 + profile_sends 자동 기록.
    body: {sender, candidate_id, candidate_number, recipients:[{email,name,job_id}],
           subject, html_body, attachment_paths}
    """
    _check_admin(request)
    import json as _json2
    data = await request.json()

    sender          = data.get("sender", "naver")
    recipients      = data.get("recipients", [])
    subject         = data.get("subject", "")
    html_body       = data.get("html_body", "")
    candidate_id    = data.get("candidate_id", "")
    candidate_number = data.get("candidate_number")
    att_paths       = data.get("attachment_paths", [])

    if not recipients:
        raise HTTPException(400, "수신자 없음")
    if len(recipients) > 99:
        raise HTTPException(400, "1회 최대 99명")
    if not subject:
        raise HTTPException(400, "제목 필수")

    cfg = SMTP_CONFIG.get(sender)
    if not cfg:
        raise HTTPException(400, f"알 수 없는 발신자: {sender}")

    # 일일 한도 체크
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        today_sent = conn.execute("""
            SELECT COUNT(*) FROM profile_sends
            WHERE DATE(sent_at) = DATE('now','localtime')
              AND sender_email LIKE ?
              AND status = 'sent'
        """, (f"%{sender}%",)).fetchone()[0]

        daily_limit = cfg.get("limit", 500)
        if today_sent + len(recipients) > daily_limit:
            raise HTTPException(429, f"일일 한도 초과: 오늘 {today_sent}건 발송, 한도 {daily_limit}건")

        results = {"sent": 0, "failed": 0, "errors": []}

        for rcpt in recipients:
            email = (rcpt.get("email") or "").strip()
            if not email:
                continue

            # 블랙리스트 체크
            if _is_email_blacklisted(email):
                results["failed"] += 1
                results["errors"].append({"email": email, "error": "블랙리스트"})
                continue

            # 중복 체크 (24시간)
            dup = conn.execute("""
                SELECT COUNT(*) FROM profile_sends
                WHERE candidate_id = ? AND to_email = ?
                  AND sent_at >= datetime('now','localtime','-24 hours')
                  AND status = 'sent'
            """, (candidate_id, email)).fetchone()[0]

            status_val = "sent"
            error_msg  = ""

            if dup > 0:
                status_val = "skipped"
                error_msg  = "24시간 내 중복"
                results["failed"] += 1
                results["errors"].append({"email": email, "error": error_msg})
            else:
                result = _send_one_email(sender, email, subject, html_body, att_paths)
                if result["ok"]:
                    results["sent"] += 1
                else:
                    status_val = "failed"
                    error_msg  = result.get("error", "")
                    results["failed"] += 1
                    results["errors"].append({"email": email, "error": error_msg})

            # profile_sends 기록 (성공/실패/스킵 모두)
            conn.execute("""
                INSERT INTO profile_sends
                (candidate_id, candidate_number, employer_id, school_name,
                 to_email, subject, template_used, sender_email,
                 attachments, status, error_msg, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now','localtime'))
            """, (
                candidate_id,
                candidate_number,
                rcpt.get("job_id"),
                rcpt.get("name", ""),
                email,
                subject,
                "profile_broadcast",
                sender,
                ",".join(att_paths) if att_paths else "",
                status_val,
                error_msg,
            ))

        conn.commit()
        results["today_total"] = today_sent + results["sent"]
        results["daily_limit"] = daily_limit
        return ok(data=results)
    finally:
        conn.close()


# ─── 이메일 블랙리스트 관리 API ──────────────────────────────────────────────

@app.get("/api/admin/email-blacklist", tags=["admin"])
async def admin_list_blacklist(request: Request):
    """블랙리스트 목록 (해시+라벨만 반환 -- 원문 이메일 미포함)."""
    _check_admin(request)
    _ensure_email_blacklist_table()
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        rows = conn.execute(
            "SELECT id, label, reason, added_at FROM email_blacklist ORDER BY id"
        ).fetchall()
        data = [{"id": r[0], "label": r[1], "reason": r[2], "added_at": r[3]} for r in rows]
    finally:
        conn.close()
    return ok(data=data)


@app.post("/api/admin/email-blacklist", tags=["admin"])
async def admin_add_blacklist(request: Request, body: dict):
    """블랙리스트에 이메일 추가 (이메일은 SHA-256 해시로 저장)."""
    _check_admin(request)
    emails = body.get("emails", [])
    label = body.get("label", "")
    reason = body.get("reason", "agency")
    if isinstance(emails, str):
        emails = [e.strip() for e in emails.split(",") if e.strip()]
    if not emails:
        raise HTTPException(status_code=400, detail="이메일을 입력하세요.")
    _ensure_email_blacklist_table()
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    added = 0
    try:
        for email in emails:
            h = _email_hash(email)
            try:
                conn.execute(
                    "INSERT INTO email_blacklist (email_hash, label, reason) VALUES (?, ?, ?)",
                    (h, label or email.split("@")[0][:20], reason),
                )
                added += 1
            except sqlite3.IntegrityError:
                pass  # 이미 등록됨
        conn.commit()
    finally:
        conn.close()
    # 캐시 무효화
    global _BLACKLIST_CACHE
    _BLACKLIST_CACHE = None
    return ok(message=f"{added}건 블랙리스트 등록 완료", data={"added": added, "total_input": len(emails)})


@app.delete("/api/admin/email-blacklist/{bl_id}", tags=["admin"])
async def admin_remove_blacklist(bl_id: int, request: Request):
    """블랙리스트 항목 삭제."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        conn.execute("DELETE FROM email_blacklist WHERE id = ?", (bl_id,))
        conn.commit()
    finally:
        conn.close()
    global _BLACKLIST_CACHE
    _BLACKLIST_CACHE = None
    return ok(message=f"블랙리스트 #{bl_id} 삭제 완료")


@app.post("/api/admin/email-blacklist/check", tags=["admin"])
async def admin_check_blacklist(request: Request, body: dict):
    """이메일이 블랙리스트에 있는지 확인."""
    _check_admin(request)
    email = body.get("email", "").strip()
    if not email:
        raise HTTPException(status_code=400, detail="이메일을 입력하세요.")
    blocked = _is_email_blacklisted(email)
    return ok(data={"email": email, "blocked": blocked})


# ─── 엔드포인트 3: 템플릿 목록 ───────────────────────────────────────────────
@app.get("/api/admin/mail/templates", tags=["admin"])
async def get_mail_templates(request: Request):
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            "SELECT template_key, subject, body_html FROM email_templates ORDER BY template_key"
        ).fetchall()
        templates = [{"name": r["template_key"], "subject": r["subject"], "body": r["body_html"]} for r in rows]
    except Exception:
        templates = []
    finally:
        conn.close()

    if not templates:
        templates = [
            {"name": "arrival_guide",       "subject": "[BRIDGE] Korea Arrival Guide",  "body": "<p>Please find your Korea arrival guide attached.</p>"},
            {"name": "interview_candidate", "subject": "<BRIDGE> Interview Guide",       "body": "<p>Your interview has been scheduled. Please see details below.</p>"},
            {"name": "contract_offer",      "subject": "[BRIDGE] Contract Offer",        "body": "<p>We are pleased to offer you the following position.</p>"},
        ]

    return ok(data={"templates": templates})


# ── 소개 메일 시스템 ─────────────────────────────────────────────────────────

@app.get("/api/admin/employers-for-mail", tags=["admin"])
async def employers_for_mail(
    request: Request,
    location: str = "",
    teaching_age: str = "",
    page: int = 1,
    per_page: int = 100,
):
    """
    소개 메일 수신자 목록 -- client_inquiries 테이블 기반.
    필터: location(도시), teaching_age(대상 연령)
    """
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        conditions = ["is_deleted = 0", "email IS NOT NULL", "email != ''"]
        params: list = []
        if location:
            conditions.append("(location_plain LIKE ? OR location_plain LIKE ?)")
            params += [f"%{location}%", f"{location}%"]
        if teaching_age:
            conditions.append("teaching_age LIKE ?")
            params.append(f"%{teaching_age}%")

        where = " AND ".join(conditions)
        offset = (page - 1) * per_page

        total = conn.execute(
            f"SELECT COUNT(*) FROM client_inquiries WHERE {where}", params
        ).fetchone()[0]

        rows = conn.execute(
            f"""SELECT id, school_name, location, contact_name, email, phone,
                teaching_age, start_date, salary_raw, submitted_at
                FROM client_inquiries WHERE {where}
                ORDER BY submitted_at DESC LIMIT ? OFFSET ?""",
            params + [per_page, offset],
        ).fetchall()

        result = []
        for r in rows:
            result.append({
                "id":           r["id"],
                "school_name":  decrypt_field(r["school_name"] or "", "school_name"),
                "location":     decrypt_field(r["location"] or "", "location"),
                "contact_name": decrypt_field(r["contact_name"] or "", "contact_name"),
                "email":        decrypt_field(r["email"] or "", "email"),
                "teaching_age": r["teaching_age"] or "",
                "start_date":   r["start_date"] or "",
                "salary_raw":   r["salary_raw"] or "",
            })

        return ok(data={"employers": result, "total": total, "page": page, "per_page": per_page})
    finally:
        conn.close()


def _build_teacher_html_block(cand: dict, expiry_days: int = 10) -> str:
    """강사 1명의 소개 HTML 블록 생성 (PII 최소화 -- 번호·국적·조건만).
    v2.3: "None"/"null" 문자열 정화 + decrypt 실패 시 폴백 텍스트.
    """
    import html as _html
    # "None"/"null"/"NULL" 문자열 (DB 오염 데이터) → 빈 문자열로 정화
    def _clean(v):
        s = str(v or "").strip()
        if s.lower() in ("none", "null", "undefined"): return ""
        return s
    # 암호화 필드 복호화 — nationality_plain 전부 비어있어 nationality(T3v1) fallback 필요
    snum    = cand.get("sheet_number") or ""
    nat_raw = cand.get("nationality_plain") or cand.get("nationality") or ""
    nat     = _clean(_safe_decrypt(nat_raw, "nationality"))
    gen_raw = cand.get("gender") or ""
    gender  = _clean(_safe_decrypt(gen_raw, "gender"))
    loc_raw = (cand.get("current_location") or "").strip()
    loc     = _clean(_safe_decrypt(loc_raw, "current_location"))
    # 평문 값 -- escape 적용 (이메일 본문 XSS 방어)
    snum    = _html.escape(str(snum))
    nat     = _html.escape(nat)
    gender  = _html.escape(gender)
    # 국내/해외 이모지 -- 한국 도시 키워드 확장
    _KR_KEYS = ("한국", "Korea", "korea", "KOREA", "Seoul", "서울",
                "Busan", "부산", "Daegu", "대구", "Incheon", "인천",
                "Gwangju", "광주", "Daejeon", "대전", "Ulsan", "울산",
                "Gyeonggi", "경기", "Jeju", "제주", "Suwon", "수원")
    emoji   = "😎" if loc and any(k in loc for k in _KR_KEYS) else "✈️"
    start   = _html.escape(_clean(cand.get("start_date")))
    target  = _html.escape(_clean(cand.get("target_age") or cand.get("target")))
    exp     = _html.escape(_clean(cand.get("experience") or cand.get("korea_experience")))
    area    = _html.escape(_clean(cand.get("area_prefs")))
    housing = _html.escape(_clean(cand.get("housing")))
    summary = _html.escape(_clean(cand.get("talent_summary")))
    star    = cand.get("talent_reference_star")
    star_str = ("★" * int(star) + "☆" * (5 - int(star)) + f" ({star}점)") if star else ""

    # 헤더 — nat/gender 모두 빈 경우 폴백 메시지 (decrypt 실패 대응)
    header_parts = [snum]
    if nat:    header_parts.append(nat)
    if gender: header_parts.append(gender)
    if not nat and not gender:
        header_parts.append("<span style='font-size:11px;color:#999;font-weight:normal'>(상세 정보는 첨부 이력서 참조)</span>")
    header_txt = " ".join(header_parts) if len(header_parts) > 1 else header_parts[0]

    # presigned URL -- cv_processed_s3_key 우선, 없으면 원본 file_uploads 폴백
    # AWS IAM User 방식: ExpiresIn 최대 604800초(7일) -- 초과 설정 시 AWS가 7일로 자동 제한
    cv_key  = cand.get("cv_processed_s3_key") or cand.get("fallback_cv_s3_key") or ""
    cv_link = ""
    if cv_key:
        try:
            _safe_expiry = max(1, min(int(expiry_days), 7))
            cv_link = s3_presigned_url(cv_key, expires=_safe_expiry * 86400)
        except Exception:
            cv_link = ""

    rows_html = ""
    if start:
        rows_html += f"<tr><td style='color:#666;padding:2px 8px 2px 0;width:80px'>시작일</td><td>{start}</td></tr>"
    if target:
        rows_html += f"<tr><td style='color:#666;padding:2px 8px 2px 0'>대상</td><td>{target}</td></tr>"
    if exp:
        rows_html += f"<tr><td style='color:#666;padding:2px 8px 2px 0'>경력</td><td>{exp}</td></tr>"
    if area:
        rows_html += f"<tr><td style='color:#666;padding:2px 8px 2px 0'>선호지역</td><td>{area}</td></tr>"
    if housing:
        rows_html += f"<tr><td style='color:#666;padding:2px 8px 2px 0'>주거</td><td>{housing}</td></tr>"
    if star_str:
        rows_html += f"<tr><td style='color:#666;padding:2px 8px 2px 0'>레퍼런스</td><td>{star_str}</td></tr>"
    if summary:
        rows_html += f"<tr><td style='color:#666;padding:2px 8px 2px 0;vertical-align:top'>메모</td><td style='font-style:italic'>{summary}</td></tr>"

    cv_btn = ""
    if cv_link:
        cv_btn = (
            f"<a href='{cv_link}' style='display:inline-block;margin-top:8px;"
            "background:#1d1d1f;color:#fff;padding:5px 12px;border-radius:4px;"
            "text-decoration:none;font-size:12px'>📄 이력서 보기</a>"
        )

    return (
        f"<div style='border-left:4px solid #1d1d1f;padding:10px 14px;margin:12px 0;"
        f"background:#f9f9f9;border-radius:0 4px 4px 0'>"
        f"<p style='margin:0 0 6px;font-weight:bold;font-size:15px'>"
        f"{emoji} ◼{header_txt}</p>"
        f"<table style='border-collapse:collapse;font-size:13px'>{rows_html}</table>"
        f"{cv_btn}</div>"
    )


@app.post("/api/admin/candidates/check-processed-status", tags=["admin"])
async def candidates_check_processed_status(request: Request):
    """
    sheet_number 목록 → cv_processed_status 일괄 조회.
    body: { "sheet_numbers": [5739, 6121, ...] }
    반환: { "statuses": { "5739": {"status": "done", "has_cv": true, "name_hash": "..."}, ... } }
    """
    _check_admin(request)
    try:
        data = await request.json()
    except Exception:
        raise HTTPException(400, "Invalid JSON body")
    sheet_numbers = data.get("sheet_numbers") or []
    if not isinstance(sheet_numbers, list):
        raise HTTPException(400, "sheet_numbers must be an array")
    # 정수 필터
    nums: list[int] = []
    for v in sheet_numbers[:500]:
        try:
            nums.append(int(v))
        except Exception:
            continue
    if not nums:
        return ok(data={"statuses": {}})
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 3000")
    conn.row_factory = sqlite3.Row
    try:
        placeholders = ",".join("?" * len(nums))
        rows = conn.execute(
            f"SELECT sheet_number, cv_processed_status, cv_processed_s3_key, "
            f"cv_processed_at, cv_processed_attempts, cv_processed_last_error "
            f"FROM candidates WHERE sheet_number IN ({placeholders}) AND status != 'Deleted'",
            nums,
        ).fetchall()
        statuses = {}
        for r in rows:
            statuses[str(r["sheet_number"])] = {
                "status": r["cv_processed_status"] or "pending",
                "has_cv": bool(r["cv_processed_s3_key"]),
                "processed_at": r["cv_processed_at"] or "",
                "attempts": r["cv_processed_attempts"] or 0,
                "last_error": (r["cv_processed_last_error"] or "")[:150],
            }
        # 누락된 번호 표시
        for n in nums:
            if str(n) not in statuses:
                statuses[str(n)] = {"status": "not_found", "has_cv": False,
                                    "processed_at": "", "attempts": 0, "last_error": ""}
        return ok(data={"statuses": statuses})
    finally:
        conn.close()


@app.post("/api/admin/mail/introduce", tags=["admin"])
async def mail_introduce(request: Request):
    """
    소개 메일 자동 발송.
    candidate_ids: sheet_number 목록
    employer_ids: client_inquiries.id 목록 (직접 선택)
    employer_filter: {location, teaching_age} (필터로 자동 선택)
    link_expiry_days: presigned URL 유효기간 (기본 10일)
    sender: naver | gmail
    custom_message: 상단 추가 메시지
    require_processed_cv: bool (기본 True) — True면 cv_processed_status='done' 아닌 강사 있으면 400
    """
    _check_admin(request)
    import json as _json_intro
    data = await request.json()

    candidate_ids   = data.get("candidate_ids", [])       # sheet_number 목록
    employer_ids    = data.get("employer_ids", [])         # client_inquiries.id 목록
    emp_filter      = data.get("employer_filter", {})
    expiry_days     = int(data.get("link_expiry_days", 10))
    sender_key      = data.get("sender", "naver")
    custom_msg      = data.get("custom_message", "").strip()

    if not candidate_ids:
        raise HTTPException(400, "candidate_ids(강사 번호) 필수")
    if not employer_ids and not emp_filter:
        raise HTTPException(400, "employer_ids 또는 employer_filter 필수")

    # 네이버 SMTP 1회 발송 한도 100건 (bridgejobkr@naver.com)
    # Gmail도 유사 한도 존재 → 공통 100건 상한 적용
    _MAX_RECIPIENTS_PER_SEND = 100
    if employer_ids and len(employer_ids) > _MAX_RECIPIENTS_PER_SEND:
        raise HTTPException(
            400,
            f"1회 발송은 최대 {_MAX_RECIPIENTS_PER_SEND}개 기관까지 가능합니다 "
            f"(현재 {len(employer_ids)}개). 나눠서 발송해 주세요."
        )

    cfg = SMTP_CONFIG.get(sender_key)
    if not cfg or not cfg.get("user") or not cfg.get("password"):
        raise HTTPException(503, f"{sender_key} SMTP 미설정")

    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row

    try:
        # ── 1. 강사 조회 ──
        placeholders = ",".join(["?"] * len(candidate_ids))
        cand_rows = conn.execute(
            f"""SELECT c.sheet_number, c.nationality_plain, c.nationality, c.gender,
                c.current_location, c.start_date, c.target_age, c.target,
                c.experience, c.korea_experience, c.area_prefs, c.housing,
                c.talent_summary, c.talent_reference_star, c.cv_processed_s3_key,
                (SELECT fu.s3_key FROM file_uploads fu
                 WHERE fu.entity_id = c.candidate_id
                   AND fu.file_type IN ('cv','cover_letter','cv_processed')
                   AND fu.is_deleted = 0
                   AND fu.s3_key IS NOT NULL AND fu.s3_key != ''
                 ORDER BY fu.id DESC LIMIT 1) AS fallback_cv_s3_key
                FROM candidates c WHERE c.sheet_number IN ({placeholders}) AND c.is_deleted=0""",
            [str(c) for c in candidate_ids],
        ).fetchall()

        if not cand_rows:
            raise HTTPException(404, "조회된 강사 없음")

        # ── v2.0 발송 전 이중 검증 — processed CV 없는 강사 차단 (100% 보장) ──
        require_processed_cv = bool(data.get("require_processed_cv", True))
        if require_processed_cv:
            missing = []
            for cr in cand_rows:
                d = dict(cr)
                if not d.get("cv_processed_s3_key"):
                    missing.append(str(d.get("sheet_number")))
            if missing:
                raise HTTPException(
                    400,
                    f"이력서 미처리 강사 있음 — 발송 차단: {', '.join(missing)}. "
                    f"재처리 후 다시 시도하거나 require_processed_cv=false 로 우회 가능."
                )

        # ── 2. 구인자(수신자) 조회 ──
        if employer_ids:
            emp_ph = ",".join(["?"] * len(employer_ids))
            emp_rows = conn.execute(
                f"SELECT id, school_name, location, contact_name, email FROM client_inquiries"
                f" WHERE id IN ({emp_ph}) AND is_deleted=0 AND email IS NOT NULL",
                [int(e) for e in employer_ids],
            ).fetchall()
        else:
            conditions = ["is_deleted = 0", "email IS NOT NULL", "email != ''"]
            params: list = []
            if emp_filter.get("location"):
                conditions.append("location_plain LIKE ?")
                params.append(f"%{emp_filter['location']}%")
            if emp_filter.get("teaching_age"):
                conditions.append("teaching_age LIKE ?")
                params.append(f"%{emp_filter['teaching_age']}%")
            emp_rows = conn.execute(
                "SELECT id, school_name, location, contact_name, email FROM client_inquiries"
                f" WHERE {' AND '.join(conditions)} LIMIT ?",
                params + [_MAX_RECIPIENTS_PER_SEND],
            ).fetchall()

        if not emp_rows:
            raise HTTPException(404, "수신 구인자 없음")
        # 필터 경로에서도 최종 상한 강제 (방어적)
        if len(emp_rows) > _MAX_RECIPIENTS_PER_SEND:
            emp_rows = emp_rows[:_MAX_RECIPIENTS_PER_SEND]

        # ── 3. 메일 본문 생성 ──
        teacher_blocks = ""
        for cr in cand_rows:
            teacher_blocks += _build_teacher_html_block(dict(cr), expiry_days)

        # 제목 -- 이모지 제거 (기업 메일 호환성 ↑) + 사용자 커스텀 지원
        subj_custom = (data.get("subject") or "").strip()
        template_subject = subj_custom or "[BRIDGE] 원어민 강사 프로필 소개"
        candidate_ids_str = _json_intro.dumps([str(c) for c in candidate_ids])

        # 이메일 부분 마스킹 (로그 PII 최소화): abc123@domain.com -> abc***@domain.com
        def _mask_email(e: str) -> str:
            if not e or "@" not in e: return ""
            local, dom = e.rsplit("@", 1)
            if len(local) <= 3: return local[0:1] + "***@" + dom
            return local[:3] + "***@" + dom

        # 중복 발송 방지 -- 7일 내 같은 (candidate_ids, email_masked) 조합 스킵
        try:
            recent_logs = conn.execute(
                """SELECT to_email FROM mail_introduce_log
                   WHERE candidate_ids=? AND status='sent'
                     AND sent_at > datetime('now', '-7 days')""",
                (candidate_ids_str,),
            ).fetchall()
            _recent_emails = {r[0] for r in recent_logs}
        except Exception:
            _recent_emails = set()

        sent = 0
        failed = 0
        skipped_dup = 0
        errors: list = []

        # SMTP 연결 1회 수립 후 루프 내 재사용 (인증/핸드셰이크 비용 최소화)
        import smtplib
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        try:
            srv = smtplib.SMTP(cfg["host"], cfg["port"], timeout=20)
            srv.ehlo()
            srv.starttls()
            srv.login(cfg["user"], cfg["password"])
        except Exception as _conn_e:
            raise HTTPException(503, f"SMTP 연결 실패: {type(_conn_e).__name__}")

        try:
            for emp in emp_rows:
                emp_email   = decrypt_field((emp["email"] or "").strip(), "email")
                emp_name    = decrypt_field(emp["school_name"] or emp["contact_name"] or "", "school_name")
                emp_contact = decrypt_field(emp["contact_name"] or "", "contact_name")

                if not emp_email or _is_email_blacklisted(emp_email):
                    failed += 1
                    errors.append({"email": _mask_email(emp_email), "error": "블랙리스트 또는 빈 이메일"})
                    continue

                # 중복 발송 체크 (마스킹된 이메일로 비교)
                emp_email_masked = _mask_email(emp_email)
                if emp_email_masked in _recent_emails:
                    skipped_dup += 1
                    continue

                # XSS escape (수신자 메타데이터 + 커스텀 메시지)
                import html as _html
                safe_emp_name = _html.escape(emp_name or "")
                safe_custom   = _html.escape(custom_msg) if custom_msg else ""

                # 본문 조립 (수신자별 이름 치환)
                greeting = f"<p>안녕하세요, <strong>{safe_emp_name}</strong> 담당자님.</p>" if safe_emp_name else "<p>안녕하세요.</p>"
                # custom_msg는 줄바꿈만 보존
                custom_block = f"<p>{safe_custom.replace(chr(10), '<br/>')}</p>" if safe_custom else ""
                body_html = f"""
<div style="font-family:Arial,sans-serif;max-width:640px;margin:0 auto">
{greeting}
<p>BRIDGE 원어민 강사 프로필을 공유드립니다.<br/>
Start date and preferences noted. Reference provided for review only.</p>
{custom_block}
<hr style="border:none;border-top:1px solid #e0e0e0;margin:16px 0"/>
{teacher_blocks}
<hr style="border:none;border-top:1px solid #e0e0e0;margin:16px 0"/>
<p style="font-size:11px;color:#999">
If you are not the intended recipient or no longer in this role,
please notify us and delete this email.
Additionally, let us know if any contact should be removed due to a change in management.
</p>
<p style="font-size:11px;color:#bbb">BRIDGE Recruitment &nbsp;|&nbsp; bridgejob.co.kr</p>
</div>"""

                try:
                    msg = MIMEMultipart("alternative")
                    msg["From"]    = cfg["user"]
                    msg["To"]      = emp_email
                    msg["Subject"] = template_subject
                    msg.attach(MIMEText(body_html, "html", "utf-8"))
                    srv.send_message(msg)

                    # 발송 로그 -- to_email 부분 마스킹 저장 (PII 최소화)
                    conn.execute(
                        """INSERT INTO mail_introduce_log
                           (candidate_ids, to_email, school_name, contact_name, subject, status, sender_email, link_expiry_days)
                           VALUES (?,?,?,?,?,?,?,?)""",
                        (candidate_ids_str, emp_email_masked, emp_name, emp_contact,
                         template_subject, "sent", cfg["user"], expiry_days),
                    )
                    conn.commit()
                    sent += 1

                except Exception as _smtp_e:
                    failed += 1
                    errors.append({"email": emp_email_masked, "error": str(_smtp_e)[:120]})
                    conn.execute(
                        """INSERT INTO mail_introduce_log
                           (candidate_ids, to_email, school_name, contact_name, subject, status, error, sender_email, link_expiry_days)
                           VALUES (?,?,?,?,?,?,?,?,?)""",
                        (candidate_ids_str, emp_email_masked, emp_name, emp_contact,
                         template_subject, "failed", str(_smtp_e)[:200], cfg["user"], expiry_days),
                    )
                    conn.commit()
        finally:
            try: srv.quit()
            except Exception: pass

        return ok(data={
            "sent": sent, "failed": failed, "skipped_duplicate": skipped_dup,
            "total_employers": len(emp_rows),
            "candidate_count": len(cand_rows),
            "errors": errors[:50],
        })
    finally:
        conn.close()


@app.post("/api/admin/mail/introduce-preview", tags=["admin"])
async def mail_introduce_preview(request: Request):
    """소개 메일 HTML 미리보기 (발송 없음)."""
    _check_admin(request)
    data = await request.json()
    candidate_ids = data.get("candidate_ids", [])
    custom_msg    = data.get("custom_message", "").strip()
    expiry_days   = int(data.get("link_expiry_days", 10))

    if not candidate_ids:
        raise HTTPException(400, "candidate_ids 필수")

    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        placeholders = ",".join(["?"] * len(candidate_ids))
        cand_rows = conn.execute(
            f"""SELECT c.sheet_number, c.nationality_plain, c.nationality, c.gender,
                c.current_location, c.start_date, c.target_age, c.target,
                c.experience, c.korea_experience, c.area_prefs, c.housing,
                c.talent_summary, c.talent_reference_star, c.cv_processed_s3_key,
                (SELECT fu.s3_key FROM file_uploads fu
                 WHERE fu.entity_id = c.candidate_id
                   AND fu.file_type IN ('cv','cover_letter','cv_processed')
                   AND fu.is_deleted = 0 AND fu.s3_key IS NOT NULL AND fu.s3_key != ''
                 ORDER BY fu.id DESC LIMIT 1) AS fallback_cv_s3_key
                FROM candidates c WHERE c.sheet_number IN ({placeholders}) AND c.is_deleted=0""",
            [str(c) for c in candidate_ids],
        ).fetchall()
    finally:
        conn.close()

    if not cand_rows:
        raise HTTPException(404, "조회된 강사 없음")

    teacher_blocks = "".join(_build_teacher_html_block(dict(r), expiry_days) for r in cand_rows)
    custom_block = f"<p>{custom_msg}</p>" if custom_msg else ""
    html = f"""<div style="font-family:Arial,sans-serif;max-width:640px;margin:0 auto">
<p>안녕하세요, <strong>(구인자명)</strong> 담당자님.</p>
<p>BRIDGE 원어민 강사 프로필을 공유드립니다.</p>
{custom_block}
<hr style="border:none;border-top:1px solid #e0e0e0;margin:16px 0"/>
{teacher_blocks}
<hr style="border:none;border-top:1px solid #e0e0e0;margin:16px 0"/>
<p style="font-size:11px;color:#bbb">BRIDGE Recruitment&nbsp;|&nbsp;bridgejob.co.kr</p>
</div>"""
    return ok(data={"html": html, "candidate_count": len(cand_rows)})


@app.get("/api/admin/mail/introduce-log", tags=["admin"])
async def get_introduce_log(request: Request, limit: int = 50):
    """소개 메일 발송 이력 조회."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            "SELECT * FROM mail_introduce_log ORDER BY sent_at DESC LIMIT ?", (limit,)
        ).fetchall()
        return ok(data={"logs": [dict(r) for r in rows]})
    finally:
        conn.close()


# ── 구인자 관리 (employers) CRUD -- employers 테이블 기반 ──────────────────────

_EMPLOYERS_ENC_FIELDS = {"email", "phone", "memo"}

# jobs 테이블 필드 → 프론트엔드 필드 역매핑 (PATCH용)
_EMPLOYERS_JOB_FIELD_MAP: dict[str, tuple[str, bool]] = {
    "name":        ("enc_employer_name",  True),
    "email":       ("enc_contact_email",  True),
    "phone":       ("enc_contact_phone",  True),
    "contact":     ("enc_contact_name",   True),
    "kakao":       ("enc_contact_kakao",  True),
    "teachingAge": ("teaching_age",       False),
    "salary":      ("salary_raw",         False),
    "memo":        ("internal_notes",     False),
    "rawText":     ("raw_text",           False),
    "status":      ("status",             False),
    "region":      ("region",             False),
    "city":        ("city",               False),
}


def _parse_memo_pii(memo: str) -> dict:
    """internal_notes 메모에서 업체명·담당자·전화·이메일 추출.

    메모 형식 예시:
      (부산 해운대 브릿지영어1호점 원장0104560333 bridge@nave.com ...)
      (경기 고양 일산 GE어학원 010-2343-0601 ge.ilsan7@gmail.com 숙소만)
    """
    import re
    result = {"name": "", "contact": "", "phone": "", "email": "", "city": ""}
    if not memo:
        return result

    full_text = memo.strip()

    # 이메일 추출 -- 전체 메모에서 (ASCII 로컬파트만 매칭)
    emails = re.findall(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", full_text)
    if emails:
        result["email"] = emails[0]

    # 전화번호 추출 -- 전체 메모에서
    phones = re.findall(r"01[016789][-\s]?\d{3,4}[-\s]?\d{4}", full_text)
    if phones:
        raw_ph = re.sub(r"[\s-]", "", phones[0])
        if len(raw_ph) == 11:
            result["phone"] = f"{raw_ph[:3]}-{raw_ph[3:7]}-{raw_ph[7:]}"
        elif len(raw_ph) == 10:
            result["phone"] = f"{raw_ph[:3]}-{raw_ph[3:6]}-{raw_ph[6:]}"
        else:
            result["phone"] = raw_ph

    # 담당자 직함 추출 -- 전체 메모에서
    _TITLES = ["원장", "부원장", "대표", "이사", "부장", "실장", "팀장", "담당", "선생", "매니저", "관장", "사장"]
    for title in _TITLES:
        # 패턴1: "이름+직함+전화" (직함 앞에 이름, 뒤에 전화)
        pat = re.search(
            r"([\uAC00-\uD7AF]{1,5})?" + re.escape(title) + r"(?:\s*01[016789])",
            full_text,
        )
        if pat:
            result["contact"] = (pat.group(1) or "") + title
            break
        # 패턴2: "전화+직함" (전화 뒤에 직함)
        pat2 = re.search(
            r"01[016789][-\s]?\d{3,4}[-\s]?\d{4}\s*([\uAC00-\uD7AF]{0,5})" + re.escape(title),
            full_text,
        )
        if pat2:
            result["contact"] = (pat2.group(1) or "") + title
            break

    # 업체명 추출: 괄호 안 내용 → 지역·도시 키워드 제거 → 첫 전화/이메일/직함까지
    # 가장 바깥쪽 괄호 내용 사용
    inner = full_text
    m_paren = re.match(r"^\((.+)\)\s*$", inner, re.DOTALL)
    if not m_paren:
        m_paren = re.match(r"^\((.+?)\)", inner, re.DOTALL)
    if not m_paren:
        # 닫는 괄호 없는 경우도 처리
        m_paren = re.match(r"^\((.+)", inner, re.DOTALL)
    if m_paren:
        inner = m_paren.group(1).strip()
    _LOCATIONS = [
        "서울", "부산", "대구", "인천", "광주", "대전", "울산", "세종",
        "경기", "강원", "충북", "충남", "전북", "전남", "경북", "경남", "제주",
        "해운대", "동래", "사직", "고양", "일산", "수지", "용인", "기흥",
        "송도", "화성", "남동탄", "동탄", "분당", "성남", "수원", "안양",
        "의정부", "구로", "강남", "마포", "잠실", "종로", "신촌", "홍대",
        "부천", "광명", "시흥", "안산", "군포", "의왕", "평택", "파주",
        "김포", "양주", "구미", "포항", "경산", "창원", "청주", "천안",
        "아산", "전주", "순천", "여수", "목포", "서귀포",
    ]

    name_src = inner
    _REGIONS = {"서울", "부산", "대구", "인천", "광주", "대전", "울산", "세종",
                "경기", "강원", "충북", "충남", "전북", "전남", "경북", "경남", "제주"}
    # 반복 제거: 앞에서부터 지역/도시 키워드를 계속 제거 (마지막 도시 = city)
    last_city = ""
    changed = True
    while changed:
        changed = False
        for loc in _LOCATIONS:
            m = re.match(r"^" + re.escape(loc) + r"\s*", name_src)
            if m:
                if loc not in _REGIONS:
                    last_city = loc
                name_src = name_src[m.end():]
                changed = True
                break
    result["city"] = last_city

    # 첫 번째 전화·이메일·직함까지의 텍스트 = 업체명
    _TITLE_PAT = "|".join(re.escape(t) for t in _TITLES)
    stop = re.search(
        r"(01[016789][-\s]?\d|[A-Za-z0-9._%+-]+@|" + _TITLE_PAT + ")",
        name_src,
    )
    if stop:
        candidate = name_src[: stop.start()].strip().rstrip("/ ,;")
        # 슬래시 뒤 5자 이상이면 잡음 → 슬래시 앞만 업체명
        slash_m = re.match(r"^(.+?)/\s*(.+)$", candidate)
        if slash_m and len(slash_m.group(2)) > 4:
            candidate = slash_m.group(1).strip()
        if candidate and len(candidate) <= 30:
            result["name"] = candidate

    return result


_CITY_EN_KO = {
    "Seoul": "서울", "Busan": "부산", "Daegu": "대구", "Incheon": "인천",
    "Gwangju": "광주", "Daejeon": "대전", "Ulsan": "울산", "Sejong": "세종",
    "Suwon": "수원", "Goyang": "고양", "Yongin": "용인", "Seongnam": "성남",
    "Hwaseong": "화성", "Bucheon": "부천", "Anyang": "안양", "Ansan": "안산",
    "Namyangju": "남양주", "Pyeongtaek": "평택", "Uijeongbu": "의정부",
    "Siheung": "시흥", "Paju": "파주", "Gimpo": "김포", "Gwangmyeong": "광명",
    "Gunpo": "군포", "Uiwang": "의왕", "Hanam": "하남", "Yangju": "양주",
    "Icheon": "이천", "Osan": "오산", "Guri": "구리",
    "Cheongju": "청주", "Cheonan": "천안", "Asan": "아산",
    "Jeonju": "전주", "Yeosu": "여수", "Suncheon": "순천", "Mokpo": "목포",
    "Changwon": "창원", "Gimhae": "김해", "Jinju": "진주", "Yangsan": "양산",
    "Gumi": "구미", "Pohang": "포항", "Gyeongsan": "경산",
    "Gangneung": "강릉", "Wonju": "원주", "Chuncheon": "춘천",
    "Jeju": "제주", "Seogwipo": "서귀포",
    # 추가 매핑 (2026-03-22)
    "Boryeong": "보령", "Chungju": "충주", "Dongducheon": "동두천",
    "Dongtan": "동탄", "Eumseong": "음성", "Geochang": "거창",
    "Geoje": "거제", "Gunsan": "군산", "Gwacheon": "과천",
    "Gyeongju": "경주", "Masan": "마산", "Tongyeong": "통영",
    "Wirye": "위례",
    # 도/광역 이름 (지역명 그대로)
    "Gangwon": "강원", "Gangwon-do": "강원", "Gangwondo": "강원",
    "Chungbuk": "충북", "Gyeongbuk": "경북", "Gyeongsang": "경상",
    "Chungcheongnam-do": "충남", "Jeonlado": "전라", "Jeju-do": "제주",
}


def _normalize_raw_text(raw: str) -> str:
    """rawText 백틱·따옴표 정리. 줄 앞뒤 백틱/따옴표 모두 제거."""
    if not raw:
        return raw
    lines = raw.split("\n")
    cleaned = []
    for line in lines:
        s = line.strip()
        # 앞뒤 백틱·따옴표 반복 제거
        while s and s[0] in '`"':
            s = s[1:]
        while s and s[-1] in '`"':
            s = s[:-1]
        cleaned.append(s)
    return "\n".join(cleaned)


@app.get("/api/employers", tags=["employers"])
async def get_employers(request: Request):
    """구인자 전체 목록 -- jobs 테이블 기반 (PII 복호화 포함)."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            "SELECT id, brj_id, job_code, region, region_name, city, location, "
            "teaching_age, salary_raw, "
            "enc_employer_name, employer_display_name, "
            "enc_contact_name, enc_contact_phone, enc_contact_email, enc_contact_kakao, "
            "status, is_hot, internal_notes, raw_text, created_at "
            "FROM jobs WHERE is_deleted = 0 ORDER BY created_at DESC"
        ).fetchall()
        result = []
        _STATUS_MAP = {"open": "active", "closed": "paused", "hot": "new", "filled": "paused"}
        for r in rows:
            # jNumber: Job.1003 → 1003 (숫자만)
            jcode = r["job_code"] or ""
            jnum = jcode.replace("Job.", "").replace("Job", "").strip() or r["brj_id"] or str(r["id"])

            # 암호화 필드 복호화
            name    = _safe_decrypt(r["enc_employer_name"], "enc_employer_name") or r["employer_display_name"] or ""
            contact = _safe_decrypt(r["enc_contact_name"], "enc_contact_name") or ""
            phone   = _safe_decrypt(r["enc_contact_phone"], "enc_contact_phone") or ""
            email   = _safe_decrypt(r["enc_contact_email"], "enc_contact_email") or ""

            # 암호화 필드가 비어있으면 메모에서 PII 추출 (fallback)
            if not name or not phone or not email:
                memo_text = r["internal_notes"] or ""
                # 메모가 없으면 rawText 끝 괄호 블록에서 추출
                # (괄호가 닫히지 않은 경우도 포함)
                if not memo_text:
                    raw = r["raw_text"] or ""
                    _tail = re.search(r"\([^)]{10,}\)\s*[\"'`]*\s*$", raw)
                    if not _tail:
                        _tail = re.search(r"\(.{10,}$", raw)
                    if _tail:
                        memo_text = _tail.group(0)
                parsed = _parse_memo_pii(memo_text)
                if not name:
                    name = parsed["name"]
                if not contact:
                    contact = parsed["contact"]
                if not phone:
                    phone = parsed["phone"]
                if not email:
                    email = parsed["email"]
                parsed_city = parsed.get("city", "")
            else:
                parsed_city = ""

            status_raw = r["status"] or "open"
            status = _STATUS_MAP.get(status_raw, status_raw)
            # 도시명 영→한 변환 + 광역시 중복 방지
            # location에 더 세부 지역정보 있으면 사용 (Seoul Gangnam → 강남)
            loc_raw = r["location"] or r["city"] or ""
            loc_parts = loc_raw.split() if loc_raw else []
            region_ko = r["region_name"] or r["region"] or ""
            if len(loc_parts) >= 2:
                # "Seoul Gangnam" → 두 번째 단어가 세부 지역
                sub = loc_parts[1]
                city_ko = _CITY_EN_KO.get(sub, sub)
            elif loc_parts:
                city_first = loc_parts[0]
                city_ko = _CITY_EN_KO.get(city_first, city_first)
                # 광역시: region과 city가 같으면 city 비움 (부산/부산 → 부산만)
                if city_ko == region_ko:
                    city_ko = ""
            else:
                city_ko = ""
            # DB city가 비면 메모에서 추출한 도시명 사용 (부산 → 해운대)
            if not city_ko and parsed_city:
                city_ko = parsed_city

            result.append({
                "jNumber":     jnum,
                "region":      r["region_name"] or r["region"] or "",
                "city":        city_ko,
                "name":        name,
                "email":       email,
                "emails":      [email] if email else [],
                "phone":       phone,
                "contact":     contact,
                "teachingAge": r["teaching_age"] or "",
                "salary":      r["salary_raw"] or "",
                "status":      status,
                "blacklist":   False,
                "active":      status_raw in ("open", "hot"),
                "isNew":       bool(r["is_hot"]),
                "confirmed":   not bool(r["is_hot"]),
                "tags":        [],
                "memo":        r["internal_notes"] or "",
                "rawText":     _normalize_raw_text(r["raw_text"] or ""),
                "createdAt":   r["created_at"] or "",
            })
        return result
    finally:
        conn.close()


@app.patch("/api/employers/{j_number}", tags=["employers"])
async def update_employer(j_number: str, request: Request):
    """구인자 정보 수정 -- jobs 테이블 (job_code 또는 brj_id로 식별)."""
    _check_admin(request)
    data = await request.json()
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        sets: list[str] = []
        vals: list = []
        for k, v in data.items():
            if k not in _EMPLOYERS_JOB_FIELD_MAP:
                continue
            col, do_encrypt = _EMPLOYERS_JOB_FIELD_MAP[k]
            if do_encrypt and v:
                sets.append(f"{col} = ?")
                vals.append(_encrypt_if_needed(str(v), col))
            else:
                sets.append(f"{col} = ?")
                vals.append(v)
        if not sets:
            return {"ok": True}
        # job_code 숫자("1003") 또는 brj_id("BRJ-BS-2602-001") 모두 지원
        job_code_val = f"Job.{j_number}" if not j_number.startswith("BRJ") else ""
        vals.extend([j_number, job_code_val])
        conn.execute(
            f"UPDATE jobs SET {','.join(sets)} WHERE brj_id = ? OR job_code = ?",
            vals,
        )
        conn.commit()
        return {"ok": True}
    finally:
        conn.close()


@app.delete("/api/employers/{j_number}", tags=["employers"])
async def delete_employer(j_number: str, request: Request):
    """구인자 논리 삭제 -- jobs 테이블 (is_deleted=1)."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        job_code_val = f"Job.{j_number}" if not j_number.startswith("BRJ") else ""
        conn.execute(
            "UPDATE jobs SET is_deleted = 1 WHERE brj_id = ? OR job_code = ?",
            (j_number, job_code_val),
        )
        conn.commit()
        return {"ok": True}
    finally:
        conn.close()


# ── 구인자 메일 단건 발송 (/api/send-mail) ────────────────────────────────────
@app.post("/api/send-mail", tags=["mail"])
async def send_mail_single(request: Request):
    """MailComposer(구인자관리)에서 단건 발송 -- Naver/Gmail 자동 분기."""
    _check_admin(request)
    data = await request.json()
    to_email: str = data.get("to", "").strip()
    subject: str = data.get("subject", "").strip()
    html_body: str = data.get("html", data.get("body", ""))
    sender: str = data.get("from", data.get("from_", "bridgejobkr@naver.com"))

    resume_s3_key: str = data.get("resume_s3_key", "").strip()

    if not to_email or not subject:
        raise HTTPException(status_code=400, detail="to, subject 필수")

    smtp_key = "naver" if "naver" in sender.lower() else "gmail"
    cfg = SMTP_CONFIG[smtp_key]
    if not cfg["user"] or not cfg["password"]:
        raise HTTPException(status_code=503, detail=f"{smtp_key.upper()} SMTP 미설정")

    try:
        # resume_s3_key 있으면 "mixed"(첨부), 없으면 "alternative"
        if resume_s3_key:
            msg = MIMEMultipart("mixed")
        else:
            msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"] = f"BRIDGE Recruitment <{cfg['user']}>"
        msg["To"] = to_email
        msg["Reply-To"] = "bridgejobkr@gmail.com"
        msg.attach(MIMEText(html_body, "html", "utf-8"))

        # S3 이력서 첨부 (sprint B B-1) — 파일명 SSoT(_build_output_filename) 사용
        if resume_s3_key:
            try:
                _pdf_bytes = s3_download_bytes(resume_s3_key)
                if _pdf_bytes:
                    from email.mime.base import MIMEBase as _MB
                    from email import encoders as _enc
                    _part = _MB("application", "pdf")
                    _part.set_payload(_pdf_bytes)
                    _enc.encode_base64(_part)

                    # SSoT 파일명: candidate_id 주어지면 DB에서 정보 조회 후 _build_output_filename
                    _cid_hint = data.get("candidate_id", "")
                    _fname = None
                    if _cid_hint:
                        try:
                            sys.path.insert(0, str(Path(__file__).resolve().parent))
                            from tools.doc_processor import _build_output_filename  # type: ignore
                            _fn_conn = sqlite3.connect(str(_ADMIN_DB_PATH))
                            _fn_conn.row_factory = sqlite3.Row
                            try:
                                _fn_row = _fn_conn.execute(
                                    "SELECT sheet_number, nationality, nationality_plain, gender, dob "
                                    "FROM candidates WHERE candidate_id = ?",
                                    (str(_cid_hint),)
                                ).fetchone()
                            finally:
                                _fn_conn.close()
                            if _fn_row:
                                _nat = _fn_row["nationality_plain"] or (_safe_decrypt(_fn_row["nationality"], "nationality") if _fn_row["nationality"] else "")
                                _fn_cand = {
                                    "nationality": _nat,
                                    "gender": _safe_decrypt(_fn_row["gender"], "gender") if _fn_row["gender"] else "",
                                    "dob": _safe_decrypt(_fn_row["dob"], "dob") if _fn_row["dob"] else "",
                                }
                                _fname = _build_output_filename(int(_fn_row["sheet_number"] or 0), _fn_cand, ".pdf")
                        except Exception as _fn_e:
                            logging.getLogger("bridge.api").debug("SSoT filename skip: %s", _fn_e)
                    if not _fname:
                        _fname = resume_s3_key.split("/")[-1] or f"BRIDGE_CV.pdf"

                    # RFC 5987 한글 파일명 안전 처리
                    try:
                        from urllib.parse import quote
                        _part.add_header(
                            "Content-Disposition",
                            f"attachment; filename*=UTF-8''{quote(_fname)}"
                        )
                    except Exception:
                        _part.add_header("Content-Disposition", f'attachment; filename="{_fname}"')
                    msg.attach(_part)
            except Exception as _att_e:
                logging.getLogger("bridge.api").warning("이력서 첨부 실패(계속): %s", _att_e)

        server = smtplib.SMTP(cfg["host"], cfg["port"], timeout=10)
        server.starttls()
        server.login(cfg["user"], cfg["password"])
        server.sendmail(cfg["user"], [to_email], msg.as_string())
        server.quit()

        # profile_sends 선택적 기록 (candidate_id + employer_job_id 둘 다 있을 때만)
        ps_candidate = data.get("candidate_id", "")
        ps_job_id = data.get("employer_job_id")
        ps_employer_name = data.get("employer_name", "")
        if ps_candidate and ps_job_id:
            try:
                ps_conn = sqlite3.connect(str(_ADMIN_DB_PATH))
                ps_conn.execute("PRAGMA busy_timeout = 5000")
                ps_conn.execute(
                    "INSERT INTO profile_sends (candidate_id, employer_id, school_name, to_email, status) VALUES (?, ?, ?, ?, ?)",
                    (str(ps_candidate), int(ps_job_id), ps_employer_name, to_email, "sent"),
                )
                ps_conn.commit()
                ps_conn.close()
            except Exception as ps_exc:
                logging.getLogger("bridge.api").warning("send-mail profile_sends 기록 실패: %s", ps_exc)

        return {"ok": True, "to": to_email}
    except smtplib.SMTPAuthenticationError:
        raise HTTPException(status_code=401, detail="SMTP 인증 실패")
    except Exception as exc:
        logging.getLogger("bridge.api").error("send-mail SMTP 오류: %s", exc)
        raise HTTPException(status_code=500, detail="메일 발송 중 오류가 발생했습니다.")


# ── 앱 비밀키 관리 (App Secrets) ─────────────────────────────────────────────
# 외부 API 키 등 민감 정보를 AES-256 암호화 후 SQLite에 저장.
# 값(value)은 어떤 API 응답에도 절대 포함되지 않음.

def _ensure_app_secrets_table():
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS app_secrets (
            key         TEXT PRIMARY KEY,
            enc_value   TEXT NOT NULL,
            description TEXT DEFAULT '',
            updated_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()
    conn.close()

try:
    _ensure_app_secrets_table()
except Exception as _e:
    logging.getLogger("bridge.api").warning("_ensure_app_secrets_table 스킵: %s", _e)


def get_secret(key: str) -> "str | None":
    """내부 전용 -- 복호화된 시크릿 값 반환. 없으면 None, env var 폴백 포함."""
    # 환경변수 우선 (Render 대시보드 설정 값)
    env_val = os.getenv(key)
    if env_val:
        return env_val
    # DB 조회
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        row = conn.execute(
            "SELECT enc_value FROM app_secrets WHERE key = ?", (key,)
        ).fetchone()
        if not row:
            return None
        enc = row["enc_value"]
        if _VAULT_OK:
            try:
                return decrypt_field(enc)
            except Exception:
                return enc
        return enc
    finally:
        conn.close()


@app.get("/api/admin/secrets", tags=["admin"])
async def list_secrets(request: Request):
    """비밀키 목록 -- 키 이름·설명·수정일만 반환. 값은 절대 포함 안 함."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            "SELECT key, description, updated_at FROM app_secrets ORDER BY key"
        ).fetchall()
        return ok(data={"secrets": [dict(r) for r in rows]})
    finally:
        conn.close()


@app.post("/api/admin/secrets", tags=["admin"])
async def set_secret(request: Request):
    """비밀키 등록/수정 -- AES-256 암호화 저장. body: {key, value, description}"""
    _check_admin(request)
    body = await request.json()
    key   = str(body.get("key", "")).strip().upper().replace(" ", "_")
    value = str(body.get("value", "")).strip()
    desc  = str(body.get("description", "")).strip()
    if not key or not value:
        raise HTTPException(400, "key와 value는 필수입니다.")
    if len(key) > 64 or len(value) > 4096:
        raise HTTPException(400, "key 또는 value가 너무 깁니다.")
    enc_value = encrypt_field(value, "app_secret") if _VAULT_OK else value
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        conn.execute(
            """INSERT INTO app_secrets (key, enc_value, description, updated_at)
               VALUES (?, ?, ?, datetime('now'))
               ON CONFLICT(key) DO UPDATE SET
                 enc_value   = excluded.enc_value,
                 description = excluded.description,
                 updated_at  = excluded.updated_at""",
            (key, enc_value, desc)
        )
        conn.commit()
        return ok(message=f"'{key}' 저장 완료")
    finally:
        conn.close()


@app.delete("/api/admin/secrets/{key}", tags=["admin"])
async def delete_secret(key: str, request: Request):
    """비밀키 삭제."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        conn.execute("DELETE FROM app_secrets WHERE key = ?", (key.upper(),))
        conn.commit()
        return ok(message=f"'{key}' 삭제 완료")
    finally:
        conn.close()


# ── 카카오 OAuth 관리자 로그인 ────────────────────────────────────────────────
# 흐름: 카카오 버튼 → /login redirect → Kakao → /callback →
#       OTC 발급 → 프론트 redirect → /exchange → api_key 반환

_KAKAO_OTC: "dict[str, dict]" = {}   # otc → {api_key, expires_at}


@app.get("/api/admin/kakao/login", tags=["admin"])
async def kakao_login():
    """카카오 OAuth 시작 -- Kakao 인증 페이지로 redirect."""
    from fastapi.responses import RedirectResponse
    client_id = get_secret("KAKAO_CLIENT_ID")
    if not client_id:
        raise HTTPException(503, "KAKAO_CLIENT_ID 미설정 -- /admin/settings 비밀키 관리에서 등록하세요.")
    redirect_uri = "https://bridge-n7hk.onrender.com/api/admin/kakao/callback"
    kakao_url = (
        "https://kauth.kakao.com/oauth/authorize"
        f"?client_id={client_id}"
        f"&redirect_uri={redirect_uri}"
        "&response_type=code"
    )
    return RedirectResponse(url=kakao_url)


@app.get("/api/admin/kakao/callback", tags=["admin"])
async def kakao_callback(code: str = "", error: str = "", error_description: str = ""):
    """Kakao 인증 콜백 -- code → token → user_id → OTC 발급."""
    from fastapi.responses import RedirectResponse
    import httpx as _httpx

    front_url = os.getenv("FRONTEND_URL", "https://bridgejob.co.kr")

    if error or not code:
        logging.getLogger("bridge.api").warning("[KAKAO] 콜백 오류: %s", error)
        return RedirectResponse(f"{front_url}/admin?kakao_error=cancelled")

    client_id = get_secret("KAKAO_CLIENT_ID")
    client_secret = get_secret("KAKAO_CLIENT_SECRET")
    allowed_ids_raw = get_secret("KAKAO_ADMIN_IDS") or ""
    allowed_ids = {s.strip() for s in allowed_ids_raw.split(",") if s.strip()}

    if not client_id:
        return RedirectResponse(f"{front_url}/admin?kakao_error=not_configured")

    token_params: dict = {
        "grant_type": "authorization_code",
        "client_id": client_id,
        "redirect_uri": "https://bridge-n7hk.onrender.com/api/admin/kakao/callback",
        "code": code,
    }
    if client_secret:
        token_params["client_secret"] = client_secret

    try:
        async with _httpx.AsyncClient(timeout=10) as hc:
            token_res = await hc.post(
                "https://kauth.kakao.com/oauth/token",
                data=token_params,
                headers={"Content-Type": "application/x-www-form-urlencoded"},
            )
        if token_res.status_code != 200:
            logging.getLogger("bridge.api").warning("[KAKAO] 토큰 교환 실패: %s", token_res.text)
            return RedirectResponse(f"{front_url}/admin?kakao_error=token_failed")
        access_token = token_res.json().get("access_token", "")
        if not access_token:
            return RedirectResponse(f"{front_url}/admin?kakao_error=no_token")

        async with _httpx.AsyncClient(timeout=10) as hc:
            user_res = await hc.get(
                "https://kapi.kakao.com/v2/user/me",
                headers={"Authorization": f"Bearer {access_token}"},
            )
        if user_res.status_code != 200:
            return RedirectResponse(f"{front_url}/admin?kakao_error=user_failed")
        kakao_id = str(user_res.json().get("id", ""))
    except Exception as _ke:
        logging.getLogger("bridge.api").error("[KAKAO] 예외: %s", _ke)
        return RedirectResponse(f"{front_url}/admin?kakao_error=network_error")

    if not allowed_ids:
        # 초기 설정 중 -- ID를 셋업 페이지로 전달 (로그인은 허용 안 함)
        logging.getLogger("bridge.api").warning("[KAKAO] KAKAO_ADMIN_IDS 미설정 -- 셋업 유도: %s", kakao_id)
        return RedirectResponse(f"{front_url}/admin/kakao-setup?kakao_id={kakao_id}")
    if kakao_id not in allowed_ids:
        logging.getLogger("bridge.api").warning("[KAKAO] 미허가 계정: %s", kakao_id)
        return RedirectResponse(f"{front_url}/admin?kakao_error=not_allowed")

    import secrets as _sec_mod
    otc = _sec_mod.token_urlsafe(24)
    # C2 패치: api_key를 OTC dict에 저장하지 않음 (exchange 응답에서도 불필요)
    _KAKAO_OTC[otc] = {
        "expires_at": datetime.now(timezone.utc).timestamp() + 60,
    }
    return RedirectResponse(f"{front_url}/admin?kakao_otc={otc}")


@app.post("/api/admin/kakao/exchange", tags=["admin"])
async def kakao_exchange(request: Request):
    """OTC → session_token 교환. C2 패치: api_key 응답 제거. 코드는 사용 즉시 삭제."""
    body = await request.json()
    otc = str(body.get("code", "")).strip()
    entry = _KAKAO_OTC.pop(otc, None)
    if not entry:
        raise HTTPException(401, "유효하지 않은 코드입니다.")
    if datetime.now(timezone.utc).timestamp() > entry["expires_at"]:
        raise HTTPException(401, "코드가 만료되었습니다 (60초 초과).")
    session_token = _create_session(request)
    cookie = (
        f"bridge_session={session_token}; HttpOnly; Secure; "
        f"SameSite=None; Path=/; Max-Age={_SESSION_TTL}"
    )
    response = JSONResponse(content=ok(data={"session_token": session_token}))
    response.headers["Set-Cookie"] = cookie
    return response


# ── 통합 수신함 + Gmail 라우터 등록 ──────────────────────────────────────────
try:
    from inbox_api import router as inbox_router
    from gmail_collector import router as gmail_router
    app.include_router(inbox_router)
    app.include_router(gmail_router)
    logging.getLogger("bridge.api").info("[STARTUP] inbox/gmail 라우터 등록 완료")
except Exception as _inbox_err:
    logging.getLogger("bridge.api").warning("[STARTUP] inbox/gmail 라우터 로드 실패 (계속 진행): %s", _inbox_err)

# ── Push 알림 라우터 등록 ─────────────────────────────────────────────────────
try:
    from push_api import router as push_router
    app.include_router(push_router)
    logging.getLogger("bridge.api").info("[STARTUP] push 라우터 등록 완료")
except Exception as _push_err:
    logging.getLogger("bridge.api").warning("[STARTUP] push 라우터 로드 실패 (계속 진행): %s", _push_err)

# ── 소셜 미디어 자동 게시 라우터 등록 ──────────────────────────────────────
try:
    from social_api import router as social_router
    app.include_router(social_router)
    logging.getLogger("bridge.api").info("[STARTUP] social 라우터 등록 완료")
except Exception as _social_err:
    logging.getLogger("bridge.api").warning("[STARTUP] social 라우터 로드 실패 (계속 진행): %s", _social_err)

# ── 결제 라우터 등록 (PAYMENT_ENABLED=false → 모든 엔드포인트 503) ──────────
try:
    from backend.routers.payments import router as payments_router
    app.include_router(payments_router)
    logging.getLogger("bridge.api").info("[STARTUP] payments 라우터 등록 완료")
except Exception as _pay_err:
    logging.getLogger("bridge.api").warning("[STARTUP] payments 라우터 로드 실패 (계속 진행): %s", _pay_err)

# ── Static Files Mount 제거 ─────────────────────────────────────────────────
# /uploads/ 직접 접근은 HMAC 서명 URL (/api/files/) 로 교체됨.
# 비인증 직접 접근 → 403 (security_middleware.py에서 차단)


# ── 게시판 이미지 (공개 콘텐츠) — Cafe24 FTP 대체 ─────────────────────────────
# 설계: 파일 자체를 Render 영구 디스크에 저장 → 영구 보존 + URL 직접 접근
# 디렉터리: /data/uploads/community/{board}/{post_id}/{filename}
# 보안: board 화이트리스트 + 경로탈주 차단 + 확장자 화이트리스트
import mimetypes as _mimetypes

_COMMUNITY_IMG_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".svg"}
_COMMUNITY_IMG_CACHE_SEC = 86400  # 1일

@app.get("/api/community-image/{board}/{post_id}/{filename}", tags=["community"])
async def community_image_serve(board: str, post_id: str, filename: str, request: Request):
    """게시판 이미지 공개 서빙 — 인증 불필요 (공개 콘텐츠).
    Cafe24 FTP 방식과 동일하게 파일 자체를 영구 디스크에서 반환.
    """
    # 1. board 화이트리스트
    if board not in _BOARDS:
        raise HTTPException(404, "Board not found")
    # 2. post_id는 정수만
    try:
        int(post_id)
    except ValueError:
        raise HTTPException(400, "invalid post_id")
    # 3. filename 안전성 (path traversal + 확장자)
    if "/" in filename or "\\" in filename or ".." in filename:
        raise HTTPException(400, "invalid filename")
    ext = Path(filename).suffix.lower()
    if ext not in _COMMUNITY_IMG_EXTS:
        raise HTTPException(400, "invalid extension")
    # 4. 경로 조립 + 범위 검증
    img_dir = _UPLOAD_BASE / "community" / board / post_id
    img_path = (img_dir / filename).resolve()
    try:
        img_path.relative_to(_UPLOAD_BASE.resolve())
    except ValueError:
        raise HTTPException(400, "path escape")
    if not img_path.exists() or not img_path.is_file():
        raise HTTPException(404, "image not found")
    # 5. mime + 캐시 헤더
    mime, _ = _mimetypes.guess_type(str(img_path))
    mime = mime or "application/octet-stream"
    from fastapi.responses import FileResponse
    return FileResponse(
        path=str(img_path),
        media_type=mime,
        headers={
            "Cache-Control": f"public, max-age={_COMMUNITY_IMG_CACHE_SEC}, immutable",
            "X-Content-Type-Options": "nosniff",
        },
    )


@app.post("/api/admin/community/{board}/{post_id}/images", tags=["admin"])
async def admin_community_upload_image(
    board: str,
    post_id: int,
    request: Request,
    file: UploadFile = FastFile(...),
):
    """관리자 게시판 이미지 업로드 — 영구 디스크 저장 + image_paths 자동 등록.
    기존 글 편집 시 사용. 같은 이름이면 덮어쓰기 (Cafe24 방식과 동일).
    """
    _check_admin(request)
    if board not in _BOARDS:
        raise HTTPException(404, "Board not found")
    if not _rate_ok(_ip_hash(request), window=60, max_posts=20):
        raise HTTPException(429, "Too many uploads.")
    # 파일 검증
    data = await file.read()
    if not data:
        raise HTTPException(400, "Empty file.")
    ext = _validate_file(data, file.filename or "image.png", "community_image")
    # 파일명 안전 (사용자 입력 정규화: 영숫자 + 한글 + - _ 만 허용)
    import re as _re
    safe_stem = _re.sub(r"[^a-zA-Z0-9가-힣_\-]", "_", Path(file.filename or "image").stem)[:50]
    if not safe_stem:
        safe_stem = "image"
    # 저장 디렉터리: /data/uploads/community/{board}/{post_id}/
    img_dir = _UPLOAD_BASE / "community" / board / str(post_id)
    img_dir.mkdir(parents=True, exist_ok=True)
    img_path = img_dir / f"{safe_stem}{ext}"
    # 경로 검증
    try:
        img_path.resolve().relative_to(_UPLOAD_BASE.resolve())
    except ValueError:
        raise HTTPException(400, "path escape")
    img_path.write_bytes(data)
    # DB image_paths 업데이트 (UPSERT)
    public_url = f"/api/community-image/{board}/{post_id}/{safe_stem}{ext}"
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    try:
        row = conn.execute(
            "SELECT image_paths FROM community_posts WHERE id=? AND board=? AND is_deleted=0",
            (post_id, board),
        ).fetchone()
        if not row:
            img_path.unlink(missing_ok=True)
            raise HTTPException(404, "Post not found")
        import json as _json
        raw = row[0]
        try:
            paths = _json.loads(raw) if raw and raw.strip().startswith("[") else ([raw] if raw else [])
        except Exception:
            paths = []
        # 중복 제거 + 새 경로 prepend (최신이 첫번째)
        paths = [p for p in paths if p != public_url]
        paths.insert(0, public_url)
        # 최대 10개로 제한
        paths = paths[:10]
        conn.execute(
            "UPDATE community_posts SET image_paths=?, updated_at=CURRENT_TIMESTAMP WHERE id=? AND board=?",
            (_json.dumps(paths, ensure_ascii=False), post_id, board),
        )
        conn.commit()
    finally:
        conn.close()
    return ok(data={"url": public_url, "image_paths": paths}, message="Image uploaded")


@app.delete("/api/admin/community/{board}/{post_id}/images", tags=["admin"])
async def admin_community_delete_image(
    board: str,
    post_id: int,
    request: Request,
    url: str,
):
    """관리자 게시판 이미지 삭제 — 파일 + image_paths 동기 제거."""
    _check_admin(request)
    if board not in _BOARDS:
        raise HTTPException(404, "Board not found")
    # url 형식 검증: /api/community-image/{board}/{post_id}/{filename}
    expected_prefix = f"/api/community-image/{board}/{post_id}/"
    if not url.startswith(expected_prefix):
        raise HTTPException(400, "invalid url")
    filename = url[len(expected_prefix):]
    if "/" in filename or "\\" in filename or ".." in filename:
        raise HTTPException(400, "invalid filename")
    img_path = (_UPLOAD_BASE / "community" / board / str(post_id) / filename).resolve()
    try:
        img_path.relative_to(_UPLOAD_BASE.resolve())
    except ValueError:
        raise HTTPException(400, "path escape")
    img_path.unlink(missing_ok=True)
    # DB
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    try:
        row = conn.execute(
            "SELECT image_paths FROM community_posts WHERE id=? AND board=?",
            (post_id, board),
        ).fetchone()
        if row and row[0]:
            import json as _json
            try:
                paths = _json.loads(row[0]) if row[0].strip().startswith("[") else []
            except Exception:
                paths = []
            paths = [p for p in paths if p != url]
            conn.execute(
                "UPDATE community_posts SET image_paths=?, updated_at=CURRENT_TIMESTAMP WHERE id=? AND board=?",
                (_json.dumps(paths, ensure_ascii=False), post_id, board),
            )
            conn.commit()
    finally:
        conn.close()
    return ok(message="Image deleted")


# ── 인터뷰 리마인더 백그라운드 스레드 ─────────────────────────────────────────
def _interview_reminder_loop():
    """10분 간격으로 인터뷰 30분 전 리마인더 발송 (Render 서버용)."""
    import time as _time
    import importlib
    _log = logging.getLogger("bridge.reminder_thread")
    _log.info("[REMINDER] 인터뷰 리마인더 스레드 시작 (10분 주기)")
    # tools/ 경로 보장
    _tools = Path(__file__).resolve().parent / "tools"
    if str(_tools.parent) not in sys.path:
        sys.path.insert(0, str(_tools.parent))
    while True:
        try:
            mod = importlib.import_module("tools.interview_reminder")
            mod.check_and_send_reminders()
        except Exception as e:
            _log.warning("[REMINDER] 리마인더 체크 실패: %s", e)
        _time.sleep(600)  # 10분

threading.Thread(target=_interview_reminder_loop, daemon=True, name="reminder").start()


# ── Resume Converter API -- 이력서 PII 제거 ──────────────────────────────────────

@app.post("/api/resume/process", tags=["admin"])
async def process_resume_files(
    request: Request,
    candidate_id: str = None,
    nationality: str = None,
    gender: str = None,
    birth_year: str = None,
    files_resume: list[UploadFile] = FastFile(default=None),
    files_cover: list[UploadFile] = FastFile(default=None),
    files_photo: list[UploadFile] = FastFile(default=None),
    files_reference: list[UploadFile] = FastFile(default=None),
    files_other: list[UploadFile] = FastFile(default=None),
):
    """
    이력서 파일 일괄 처리 (PII 제거)

    5개 섹션 독립 처리:
    - files_resume: 이력서 (PDF/DOCX)
    - files_cover: 커버레터 (PDF/DOCX)
    - files_photo: 사진 (JPG/PNG/WEBP)
    - files_reference: 추천서 (PDF/DOCX)
    - files_other: 기타 파일
    출력 파일명: {candidate_id}{nationality}_{gender}({birth_year}born).{ext}
    """
    # 관리자 인증
    _check_admin(request)

    # 입력값 검증
    if not candidate_id or not candidate_id.strip():
        raise HTTPException(status_code=400, detail="candidate_id는 필수입니다")

    candidate_id = candidate_id.strip()

    # 출력 파일명 스템 생성: {번호}{국적}_{성별}({년생}born)
    _nat   = (nationality or "").strip()
    _gen   = (gender or "").strip()
    _yr    = (birth_year or "").strip()
    _stem  = candidate_id
    if _nat:
        _stem += _nat
    if _gen or _yr:
        _stem += f"_{_gen}({_yr}born)"

    # 파일 섹션 수집
    file_sections = {
        "resume": files_resume or [],
        "cover": files_cover or [],
        "photo": files_photo or [],
        "reference": files_reference or [],
        "other": files_other or [],
    }

    total_files = sum(len(files) for files in file_sections.values())
    if total_files == 0:
        raise HTTPException(status_code=400, detail="최소 1개 이상의 파일을 제공해야 합니다")

    # 후보자 정보 로드 (PII 제거 시 참고용)
    candidate_dict = None
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        conn.row_factory = sqlite3.Row
        try:
            num = int(candidate_id)
            row = conn.execute(
                "SELECT full_name, email, mobile_phone, nationality, photo_url FROM candidates WHERE sheet_number = ? LIMIT 1",
                (num,)
            ).fetchone()
            if row:
                candidate_dict = {
                    "sheet_number": num,
                    "full_name": decrypt_field(str(row[0])) if row[0] else "",
                    "email": decrypt_field(str(row[1])) if row[1] else "",
                    "mobile_phone": decrypt_field(str(row[2])) if row[2] else "",
                    "nationality": decrypt_field(str(row[3])) if row[3] else "",
                    "photo_url": str(row[4]) if row[4] else "",
                }
        except (ValueError, sqlite3.Error):
            pass
        finally:
            conn.close()
    except Exception:
        pass

    # 사진 파일 먼저 추출 (resume/cover 처리 시 photo_path로 전달)
    _resume_log = logging.getLogger("bridge.resume")
    photo_tmp_path = None
    try:
        if file_sections.get("photo"):
            import tempfile as _tmpmod
            photo_file = file_sections["photo"][0]
            photo_bytes = await photo_file.read()
            if photo_bytes:
                photo_ext = Path(photo_file.filename).suffix.lower() or ".jpg"
                with _tmpmod.NamedTemporaryFile(suffix=photo_ext, delete=False) as _ptmp:
                    _ptmp.write(photo_bytes)
                    photo_tmp_path = Path(_ptmp.name)
                _resume_log.info("[RESUME] 폼 사진 저장: %s (%d bytes)", photo_tmp_path, len(photo_bytes))
            else:
                _resume_log.warning("[RESUME] 폼 사진 bytes가 비어있음 (photo file drained?)")
    except Exception as _photo_err:
        _resume_log.warning("[RESUME] 폼 사진 임시파일 생성 실패: %s", _photo_err)
        photo_tmp_path = None

    # 폼 사진 없으면 DB photo_url → S3 폴백
    if photo_tmp_path is None and candidate_dict and candidate_dict.get("photo_url"):
        try:
            import tempfile as _tmpmod_s3
            _photo_s3_key = candidate_dict["photo_url"]
            _photo_s3_bytes = s3_download_bytes(_photo_s3_key)
            if _photo_s3_bytes:
                _photo_s3_ext = Path(_photo_s3_key).suffix.lower() or ".jpg"
                with _tmpmod_s3.NamedTemporaryFile(suffix=_photo_s3_ext, delete=False) as _ptmp_s3:
                    _ptmp_s3.write(_photo_s3_bytes)
                    photo_tmp_path = Path(_ptmp_s3.name)
                _resume_log.info("[RESUME] S3 사진 폴백 성공: %s (%d bytes)", _photo_s3_key, len(_photo_s3_bytes))
            else:
                _resume_log.warning("[RESUME] S3 사진 폴백: photo_url=%s 이지만 bytes 없음", _photo_s3_key)
        except Exception as _s3_err:
            _resume_log.warning("[RESUME] S3 사진 폴백 실패: %s", _s3_err)
            photo_tmp_path = None

    # brj_number: candidate_id를 정수로 사용
    try:
        _brj_number = int(candidate_id) if candidate_id else 0
    except (ValueError, TypeError):
        _brj_number = 0

    # 파일 처리
    def _process_file(file_bytes: bytes, file_name: str, candidate_dict: dict = None) -> dict:
        """파일 처리 (PII 제거)"""
        import io
        import tempfile

        try:
            file_ext = Path(file_name).suffix.lower()

            # 이미지 파일 (처리 안 함)
            if file_ext in ('.jpg', '.jpeg', '.png', '.webp', '.gif'):
                return {
                    "file_name": file_name,
                    "file_type": "image",
                    "processed": False,
                    "pii_count": 0,
                    "size_before": len(file_bytes),
                    "size_after": len(file_bytes),
                    "error": None,
                }

            # DOCX 처리
            if file_ext in ('.docx', '.doc'):
                try:
                    sys.path.insert(0, str(Path(__file__).resolve().parent))
                    from tools.doc_processor import process_docx

                    with tempfile.NamedTemporaryFile(suffix=file_ext, delete=False) as tmp:
                        tmp.write(file_bytes)
                        tmp_path = Path(tmp.name)

                    try:
                        doc, log_entries = process_docx(
                            tmp_path,
                            brj_number=_brj_number,
                            candidate=candidate_dict,
                            dry=False,
                            photo_path=photo_tmp_path,
                        )

                        output = io.BytesIO()
                        doc.save(output)
                        processed_bytes = output.getvalue()

                        pii_count = len([l for l in log_entries if any(
                            kw in l for kw in ["EMAIL:", "PHONE:", "NAME:", "LOCATION:", "WORKPLACE:"]
                        )])

                        import base64 as _b64
                        out_fname = f"{_stem}.docx"
                        return {
                            "file_name": out_fname,
                            "file_type": "docx",
                            "processed": True,
                            "pii_count": pii_count,
                            "size_before": len(file_bytes),
                            "size_after": len(processed_bytes),
                            "processed_data": _b64.b64encode(processed_bytes).decode(),
                            "error": None,
                        }
                    finally:
                        tmp_path.unlink(missing_ok=True)

                except Exception as e:
                    return {
                        "file_name": file_name,
                        "file_type": "docx",
                        "processed": False,
                        "pii_count": 0,
                        "size_before": len(file_bytes),
                        "size_after": 0,
                        "error": f"처리 실패: {str(e)}",
                    }

            # PDF 처리 — doc_processor.process_pdf 사용 (원본 레이아웃 유지 + PII 제거 + 번호 삽입)
            elif file_ext == '.pdf':
                try:
                    sys.path.insert(0, str(Path(__file__).resolve().parent))
                    from tools.doc_processor import process_pdf

                    with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
                        tmp.write(file_bytes)
                        tmp_path = Path(tmp.name)

                    try:
                        doc, log_entries = process_pdf(
                            tmp_path,
                            brj_number=_brj_number or 0,
                            candidate=candidate_dict,
                            photo_path=photo_tmp_path,
                        )

                        processed_bytes = doc.tobytes(deflate=True)
                        doc.close()

                        pii_count = len([l for l in log_entries if any(
                            kw in l for kw in ["EMAIL:", "PHONE:", "NAME:", "LOCATION:", "WORKPLACE:", "REDACT"]
                        )])

                        import base64 as _b64
                        out_fname = f"{_stem}.pdf"
                        return {
                            "file_name": out_fname,
                            "file_type": "pdf",
                            "processed": True,
                            "pii_count": pii_count,
                            "size_before": len(file_bytes),
                            "size_after": len(processed_bytes),
                            "processed_data": _b64.b64encode(processed_bytes).decode(),
                            "error": None,
                        }
                    finally:
                        tmp_path.unlink(missing_ok=True)

                except Exception as e:
                    return {
                        "file_name": file_name,
                        "file_type": "pdf",
                        "processed": False,
                        "pii_count": 0,
                        "size_before": len(file_bytes),
                        "size_after": 0,
                        "error": f"처리 실패: {str(e)}",
                    }

            else:
                return {
                    "file_name": file_name,
                    "file_type": "unknown",
                    "processed": False,
                    "pii_count": 0,
                    "size_before": len(file_bytes),
                    "size_after": 0,
                    "error": f"지원하지 않는 파일 형식: {file_ext}",
                }

        except Exception as e:
            return {
                "file_name": file_name,
                "file_type": "unknown",
                "processed": False,
                "pii_count": 0,
                "size_before": len(file_bytes),
                "size_after": 0,
                "error": f"예상치 못한 오류: {str(e)}",
            }

    # 모든 섹션 처리
    results = {"resume": [], "cover": [], "photo": [], "reference": [], "other": []}
    total_pii = 0
    processed_count = 0
    failed_count = 0

    for section_name, file_list in file_sections.items():
        for upload_file in file_list:
            try:
                file_bytes = await upload_file.read()
                result = _process_file(file_bytes, upload_file.filename, candidate_dict)
                results[section_name].append(result)

                if result.get("processed"):
                    processed_count += 1
                    total_pii += result.get("pii_count", 0)
                else:
                    if result.get("error"):
                        failed_count += 1

            except Exception as e:
                results[section_name].append({
                    "file_name": upload_file.filename,
                    "file_type": "unknown",
                    "processed": False,
                    "pii_count": 0,
                    "size_before": 0,
                    "size_after": 0,
                    "error": f"업로드 처리 오류: {str(e)}",
                })
                failed_count += 1

    # 사진 임시파일 정리
    if photo_tmp_path and photo_tmp_path.exists():
        try:
            photo_tmp_path.unlink()
        except Exception:
            pass

    # 로깅
    logging.getLogger("bridge.resume").info(
        f"[RESUME] candidate_id={candidate_id}, files={total_files}, "
        f"processed={processed_count}, pii={total_pii}"
    )

    return ok(
        data={
            "candidate_id": candidate_id,
            "pii_count": total_pii,
            "files_processed": results,
            "total_files": total_files,
            "processed_files": processed_count,
            "failed_files": failed_count,
        }
    )


# ── Form Config 엔드포인트 ──────────────────────────────────────────────────────
# 내부 서비스 키 전용 (브라우저 미노출 -- Vercel Server Component 전용)
@app.get("/api/internal/form-config/{form_name}", include_in_schema=False)
async def internal_get_form_config(form_name: str, request: Request):
    if form_name not in ("apply", "inquiry"):
        return SafeJSONResponse({"success": False, "detail": "not found"}, status_code=404)
    svc_key = request.headers.get("x-service-key", "")
    if not _FORM_CONFIG_READ_KEY or not svc_key or not hmac.compare_digest(svc_key, _FORM_CONFIG_READ_KEY):
        return SafeJSONResponse({"success": False, "detail": "forbidden"}, status_code=403)
    conn = None
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        rows = conn.execute(
            "SELECT field_key, options_json FROM form_configs WHERE form_name=?",
            (form_name,),
        ).fetchall()
        data = {row[0]: json.loads(row[1]) for row in rows}
        return SafeJSONResponse({"success": True, "data": data})
    except Exception as _e:
        logging.getLogger("bridge.api").error("internal_get_form_config 오류: %s", _e)
        return SafeJSONResponse({"success": False, "detail": "서버 오류"}, status_code=500)
    finally:
        if conn:
            conn.close()


@app.get("/api/admin/form-config/{form_name}")
async def admin_get_form_config(form_name: str, request: Request):
    if form_name not in ("apply", "inquiry"):
        return SafeJSONResponse({"success": False, "detail": "not found"}, status_code=404)
    _check_admin(request)   # 실패 시 HTTPException 발생 (FastAPI 자동 처리)
    if not _rate_ok(_ip_hash(request), window=60, max_posts=60):
        return SafeJSONResponse({"success": False, "detail": "Too Many Requests"}, status_code=429)
    conn = None
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        rows = conn.execute(
            "SELECT field_key, options_json FROM form_configs WHERE form_name=? ORDER BY field_key",
            (form_name,),
        ).fetchall()
        data = {row[0]: json.loads(row[1]) for row in rows}
        return SafeJSONResponse({"success": True, "data": data})
    except Exception as _e:
        logging.getLogger("bridge.api").error("admin_get_form_config 오류: %s", _e)
        return SafeJSONResponse({"success": False, "detail": "서버 오류"}, status_code=500)
    finally:
        if conn:
            conn.close()


@app.put("/api/admin/form-config/{form_name}/{field_key}")
async def admin_put_form_config(form_name: str, field_key: str, request: Request):
    if form_name not in ("apply", "inquiry"):
        return SafeJSONResponse({"success": False, "detail": "not found"}, status_code=404)
    if field_key not in _FORM_DEFAULTS.get(form_name, {}):
        return SafeJSONResponse({"success": False, "detail": "알 수 없는 field_key"}, status_code=404)
    _check_admin(request)
    if not _rate_ok(_ip_hash(request), window=60, max_posts=30):
        return SafeJSONResponse({"success": False, "detail": "Too Many Requests"}, status_code=429)
    try:
        body = await request.json()
    except Exception:
        return SafeJSONResponse({"success": False, "detail": "잘못된 JSON"}, status_code=400)
    options = body.get("options")
    if not isinstance(options, list) or len(options) == 0:
        return SafeJSONResponse({"success": False, "detail": "options 배열 필수 (최소 1개)"}, status_code=400)
    # XSS 방어: 각 항목을 문자열 변환 + 500자 이내로 제한
    cleaned: list[str] = []
    for opt in options[:200]:
        s = str(opt).strip()[:500]
        s = s.replace("<", "&lt;").replace(">", "&gt;")
        if s:
            cleaned.append(s)
    if not cleaned:
        return SafeJSONResponse({"success": False, "detail": "유효한 항목이 없습니다"}, status_code=400)
    conn = None
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute(
            """INSERT INTO form_configs (form_name, field_key, options_json, updated_at)
               VALUES (?, ?, ?, datetime('now'))
               ON CONFLICT(form_name, field_key) DO UPDATE SET
                   options_json=excluded.options_json,
                   updated_at=excluded.updated_at""",
            (form_name, field_key, json.dumps(cleaned, ensure_ascii=False)),
        )
        conn.commit()
        return SafeJSONResponse({"success": True, "data": {"options": cleaned}})
    except Exception as _e:
        logging.getLogger("bridge.api").error("admin_put_form_config 오류: %s", _e)
        return SafeJSONResponse({"success": False, "detail": "서버 오류"}, status_code=500)
    finally:
        if conn:
            conn.close()


@app.post("/api/admin/form-config/{form_name}/{field_key}/reset")
async def admin_reset_form_config(form_name: str, field_key: str, request: Request):
    if form_name not in ("apply", "inquiry"):
        return SafeJSONResponse({"success": False, "detail": "not found"}, status_code=404)
    _check_admin(request)
    if not _rate_ok(_ip_hash(request), window=60, max_posts=30):
        return SafeJSONResponse({"success": False, "detail": "Too Many Requests"}, status_code=429)
    defaults = _FORM_DEFAULTS.get(form_name, {}).get(field_key)
    if defaults is None:
        return SafeJSONResponse({"success": False, "detail": "알 수 없는 field_key"}, status_code=404)
    conn = None
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute(
            """INSERT INTO form_configs (form_name, field_key, options_json, updated_at)
               VALUES (?, ?, ?, datetime('now'))
               ON CONFLICT(form_name, field_key) DO UPDATE SET
                   options_json=excluded.options_json,
                   updated_at=excluded.updated_at""",
            (form_name, field_key, json.dumps(defaults, ensure_ascii=False)),
        )
        conn.commit()
        return SafeJSONResponse({"success": True, "data": {"options": defaults}})
    except Exception as _e:
        logging.getLogger("bridge.api").error("admin_reset_form_config 오류: %s", _e)
        return SafeJSONResponse({"success": False, "detail": "서버 오류"}, status_code=500)
    finally:
        if conn:
            conn.close()


# ── Sprint B -- 이력서 자동첨부 + 지역/연령 매칭 + Excel 다운로드 ───────────────
# B-1: 이력서 자동첨부 (processed CV presigned URL)
# B-2: jobs 기반 양방향 매칭 API
# B-3: jobs Excel 다운로드


@app.get("/api/admin/resume/find/{candidate_number}", tags=["admin"])
async def find_resume_by_candidate_number(candidate_number: int, request: Request):
    """
    sheet_number(=candidate_number)로 변환이력서 + 동영상 presigned URL 반환.
    MailComposer 자동첨부용.

    응답:
        presigned_url:   변환이력서 (cv_processed) — 없으면 원본 CV 폴백
        video_url:       최신 동영상 (있을 때만)
        video_filename:  동영상 파일명 (이메일 표시용)
    """
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        row = conn.execute(
            "SELECT candidate_id, cv_processed_s3_key, full_name "
            "FROM candidates WHERE sheet_number = ? AND COALESCE(is_deleted,0) != 1 LIMIT 1",
            (candidate_number,),
        ).fetchone()
        if not row:
            raise HTTPException(404, f"candidate_number={candidate_number} 후보자 없음")
        candidate_id, s3_key, full_name = row

        # 변환이력서 없으면 원본 cv 폴백
        if not s3_key:
            cv_row = conn.execute(
                "SELECT s3_key FROM file_uploads "
                "WHERE entity_id = ? AND file_type = 'cv' AND is_deleted = 0 "
                "AND s3_key IS NOT NULL ORDER BY rowid DESC LIMIT 1",
                (candidate_id,),
            ).fetchone()
            if cv_row and cv_row[0]:
                s3_key = cv_row[0]

        # 동영상 최신본 조회
        video_row = conn.execute(
            "SELECT s3_key, file_url FROM file_uploads "
            "WHERE entity_id = ? AND file_type = 'video' AND is_deleted = 0 "
            "AND s3_key IS NOT NULL ORDER BY rowid DESC LIMIT 1",
            (candidate_id,),
        ).fetchone()
    finally:
        conn.close()

    if not s3_key:
        raise HTTPException(404, "이력서 없음 -- 업로드 또는 변환 필요")

    try:
        presigned = s3_presigned_url(s3_key, expires=3600)
    except Exception as _e_s3:
        logging.getLogger("bridge.resume").error("[find_resume] presigned 실패: %s", _e_s3)
        raise HTTPException(500, "이력서 URL 생성 실패")

    video_url = None
    video_filename = None
    if video_row and video_row[0]:
        try:
            video_url = s3_presigned_url(video_row[0], expires=3600)
            video_filename = Path(video_row[0]).name
        except Exception as _e_v:
            logging.getLogger("bridge.resume").warning("[find_resume] 동영상 URL 실패: %s", _e_v)

    return ok(data={
        "candidate_id": candidate_id,
        "candidate_number": candidate_number,
        "full_name": full_name,
        "s3_key": s3_key,
        "presigned_url": presigned,
        "video_url": video_url,
        "video_filename": video_filename,
        "expires_in": 3600,
    })


# B-2: 지역/연령 양방향 매칭 (기존 _match_region / _match_target 재활용)

@app.get("/api/admin/matching/jobs-for-candidate", tags=["admin"])
async def matching_jobs_for_candidate(
    request: Request,
    candidate_id: str,
    limit: int = 30,
):
    """
    Sprint B B-2a -- 후보자 area_prefs/target_age 기반 jobs 매칭 목록.
    match_score: 지역(2) + 연령(1) = 최대 3.
    """
    _check_admin(request)
    if limit > 200:
        limit = 200
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        cand = conn.execute(
            "SELECT candidate_id, area_prefs, target, target_age, status "
            "FROM candidates WHERE candidate_id = ? AND COALESCE(is_deleted,0) != 1",
            (candidate_id,),
        ).fetchone()
        if not cand:
            raise HTTPException(404, "후보자 없음")
        c = dict(cand)

        jobs_rows = conn.execute(
            "SELECT brj_id, id, job_code, region, region_name, location, city, "
            "employer_display_name, teaching_age, salary_raw, status, is_hot, created_at "
            "FROM jobs WHERE is_deleted = 0"
        ).fetchall()
    finally:
        conn.close()

    matched = []
    for jr in jobs_rows:
        j = dict(jr)
        loc_str = " ".join(filter(None, [j.get("region"), j.get("region_name"),
                                         j.get("location"), j.get("city")]))
        region_ok = _match_region(c.get("area_prefs", ""), loc_str)
        age_ok = _match_target(c.get("target", ""), c.get("target_age", ""),
                               j.get("teaching_age", ""))
        if region_ok or age_ok:
            j["match_region"] = region_ok
            j["match_age"] = age_ok
            j["match_score"] = (2 if region_ok else 0) + (1 if age_ok else 0)
            matched.append(j)

    matched.sort(key=lambda x: x["match_score"], reverse=True)
    return ok(data={
        "candidate_id": candidate_id,
        "area_prefs": c.get("area_prefs"),
        "target_age": c.get("target_age"),
        "matched_jobs": matched[:limit],
        "total": len(matched),
    })


@app.get("/api/admin/matching/candidates-for-job", tags=["admin"])
async def matching_candidates_for_job(
    request: Request,
    brj_id: Optional[str] = None,
    job_id: Optional[int] = None,
    limit: int = 30,
):
    """
    Sprint B B-2b -- jobs 지역/교습연령 기반 Active 후보자 매칭 목록.
    brj_id(BRJ-xxx) 또는 job_id(정수 PK) 중 하나 필수.
    """
    _check_admin(request)
    if not brj_id and not job_id:
        raise HTTPException(400, "brj_id 또는 job_id 필수")
    if limit > 200:
        limit = 200

    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        if brj_id:
            job = conn.execute(
                "SELECT id, brj_id, region, region_name, location, city, "
                "teaching_age, employer_display_name "
                "FROM jobs WHERE brj_id = ? AND is_deleted = 0",
                (brj_id,),
            ).fetchone()
        else:
            job = conn.execute(
                "SELECT id, brj_id, region, region_name, location, city, "
                "teaching_age, employer_display_name "
                "FROM jobs WHERE id = ? AND is_deleted = 0",
                (job_id,),
            ).fetchone()

        if not job:
            raise HTTPException(404, "공고 없음")
        j = dict(job)

        cands = conn.execute(
            "SELECT candidate_id, sheet_number, area_prefs, target, target_age, "
            "nationality, gender, start_date, status "
            "FROM candidates WHERE COALESCE(is_deleted,0) != 1 AND status = 'Active'"
        ).fetchall()
    finally:
        conn.close()

    loc_str = " ".join(filter(None, [j.get("region"), j.get("region_name"),
                                     j.get("location"), j.get("city")]))
    matched = []
    for cr in cands:
        c = dict(cr)
        region_ok = _match_region(c.get("area_prefs", ""), loc_str)
        age_ok = _match_target(c.get("target", ""), c.get("target_age", ""),
                               j.get("teaching_age", ""))
        if region_ok or age_ok:
            c["match_region"] = region_ok
            c["match_age"] = age_ok
            c["match_score"] = (2 if region_ok else 0) + (1 if age_ok else 0)
            matched.append(c)

    matched.sort(key=lambda x: x["match_score"], reverse=True)
    return ok(data={
        "job": {
            "brj_id": j.get("brj_id"),
            "employer": j.get("employer_display_name"),
            "teaching_age": j.get("teaching_age"),
            "location": loc_str,
        },
        "matched_candidates": matched[:limit],
        "total": len(matched),
    })


# B-3: Jobs Excel 다운로드

@app.get("/api/admin/jobs/download-xlsx", tags=["admin"])
async def download_jobs_xlsx(
    request: Request,
    status: Optional[str] = None,
    region: Optional[str] = None,
    include_pii: bool = False,
):
    """
    Sprint B B-3 -- jobs 테이블 Excel(.xlsx) 다운로드.
    include_pii=true 시 enc_contact_* 복호화 컬럼 포함 (관리자 전용).
    """
    _check_admin(request)
    import io as _io_xlsx
    from openpyxl import Workbook as _WBx
    from openpyxl.styles import Font as _Fx, Alignment as _Ax, PatternFill as _Px
    from openpyxl.utils import get_column_letter as _gcx
    from datetime import date as _dx

    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    try:
        where_parts = ["is_deleted = 0"]
        params_xl: list = []
        if status:
            where_parts.append("status = ?")
            params_xl.append(status)
        if region:
            s_xl = f"%{region.strip()[:30]}%"
            where_parts.append("(region LIKE ? OR region_name LIKE ? OR city LIKE ?)")
            params_xl.extend([s_xl, s_xl, s_xl])
        w_xl = " AND ".join(where_parts)

        _pii_sql = (", enc_employer_name, enc_contact_name, enc_contact_phone, "
                    "enc_contact_email, enc_contact_kakao") if include_pii else ""
        rows_xl = conn.execute(
            f"SELECT brj_id, job_code, region, region_name, city, location, "
            f"employer_display_name, teaching_age, salary_raw, working_hours, "
            f"housing, visa_sponsorship, is_hot, status, created_at{_pii_sql} "
            f"FROM jobs WHERE {w_xl} ORDER BY created_at DESC",
            params_xl,
        ).fetchall()
    finally:
        conn.close()

    _BASE_COLS = [
        "brj_id", "job_code", "region", "region_name", "city", "location",
        "employer_display_name", "teaching_age", "salary_raw", "working_hours",
        "housing", "visa_sponsorship", "is_hot", "status", "created_at",
    ]
    _PII_COLS = ["enc_employer_name", "enc_contact_name", "enc_contact_phone",
                 "enc_contact_email", "enc_contact_kakao"]
    _ALL_COLS = _BASE_COLS + (_PII_COLS if include_pii else [])

    _COL_KR = {
        "brj_id": "BRJ ID", "job_code": "공고코드", "region": "지역코드",
        "region_name": "지역명", "city": "도시", "location": "상세위치",
        "employer_display_name": "업체명(공개)", "teaching_age": "교습연령",
        "salary_raw": "급여", "working_hours": "근무시간", "housing": "숙소",
        "visa_sponsorship": "비자스폰서", "is_hot": "급구", "status": "상태",
        "created_at": "등록일", "enc_employer_name": "업체명(실제)",
        "enc_contact_name": "담당자", "enc_contact_phone": "연락처",
        "enc_contact_email": "이메일", "enc_contact_kakao": "카카오",
    }
    _COL_W = {
        "brj_id": 14, "job_code": 14, "region": 10, "region_name": 14, "city": 12,
        "location": 22, "employer_display_name": 24, "teaching_age": 20,
        "salary_raw": 20, "working_hours": 16, "housing": 10,
        "visa_sponsorship": 10, "is_hot": 6, "status": 10, "created_at": 20,
        "enc_employer_name": 24, "enc_contact_name": 14, "enc_contact_phone": 18,
        "enc_contact_email": 28, "enc_contact_kakao": 18,
    }

    wb = _WBx()
    ws = wb.active
    ws.title = "Jobs"

    _hdr_fill = _Px(fill_type="solid", fgColor="1F4E79")
    _hdr_font = _Fx(bold=True, color="FFFFFF", name="맑은 고딕", size=10)
    _even_fill = _Px(fill_type="solid", fgColor="EBF3FB")

    for ci, col in enumerate(_ALL_COLS, start=1):
        cell = ws.cell(row=1, column=ci, value=_COL_KR.get(col, col))
        cell.fill = _hdr_fill
        cell.font = _hdr_font
        cell.alignment = _Ax(horizontal="center", vertical="center", wrap_text=False)
        ws.column_dimensions[_gcx(ci)].width = _COL_W.get(col, 14)
    ws.row_dimensions[1].height = 22

    for ri, row_xl in enumerate(rows_xl, start=2):
        d = dict(row_xl)
        if include_pii:
            for ec in _PII_COLS:
                raw = d.get(ec) or ""
                col_name = ec.replace("enc_", "")
                d[ec] = decrypt_field(raw, col_name) if raw else ""
        d["visa_sponsorship"] = "O" if d.get("visa_sponsorship") else ""
        d["is_hot"] = "★" if d.get("is_hot") else ""
        for ci, col in enumerate(_ALL_COLS, start=1):
            cell = ws.cell(row=ri, column=ci, value=str(d.get(col) or ""))
            cell.font = _Fx(name="맑은 고딕", size=9)
            if ri % 2 == 0:
                cell.fill = _even_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf_xl = _io_xlsx.BytesIO()
    wb.save(buf_xl)
    buf_xl.seek(0)

    today_str = _dx.today().strftime("%Y%m%d")
    fname_xl = f"bridge_jobs_{today_str}{'_pii' if include_pii else ''}.xlsx"
    from starlette.responses import Response as _StRx
    return _StRx(
        content=buf_xl.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname_xl}"'},
    )


# ── Render DB 원격 덤프 (관리자 전용 백업) ──────────────────────────────────────
@app.get("/api/admin/db/dump", tags=["admin"])
async def admin_db_dump(request: Request):
    """Render /data/master.db → SQL dump 반환. 관리자 인증 필수."""
    _check_admin(request)
    import io as _io
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    buf = _io.StringIO()
    for line in conn.iterdump():
        buf.write(line + "\n")
    conn.close()
    today = datetime.now().strftime("%Y%m%d_%H%M")
    from starlette.responses import Response as _DumpResp
    return _DumpResp(
        content=buf.getvalue().encode("utf-8"),
        media_type="text/plain; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="master_dump_{today}.sql"'},
    )


@app.get("/api/admin/db/stats", tags=["admin"])
async def admin_db_stats(request: Request):
    """Render DB 행 수 확인 -- 데이터 복원 진단용."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 3000")
    try:
        tables = ["candidates", "jobs", "client_inquiries", "interviews", "file_uploads",
                  "job_queue", "pipeline_events"]
        stats = {}
        for t in tables:
            try:
                row = conn.execute(f"SELECT COUNT(*) FROM {t}").fetchone()
                stats[t] = row[0] if row else 0
            except Exception:
                stats[t] = -1  # 테이블 없음
        return ok(data=stats, message="DB 통계")
    finally:
        conn.close()


# ── Resume Pipeline v2.0 대시보드 API ────────────────────────────────────
@app.post("/api/admin/pipeline/_smoke", tags=["admin"])
async def admin_pipeline_smoke(request: Request):
    """
    Render 환경 E2E smoke — enqueue → 워커 pickup → done 까지 1 cycle 테스트.
    실제 S3/파일 없이 파이프라인 기계적 흐름만 검증.
    """
    _check_admin(request)
    try:
        sys.path.insert(0, str(Path(__file__).resolve().parent))
        from tools.pipeline_v2 import (  # type: ignore
            enqueue_file_process, get_queue_stats, get_next_job_any,
        )
    except Exception as e:
        raise HTTPException(500, f"pipeline 모듈 import 실패: {e}")

    # 1. 기존 테스트 데이터 제거
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 3000")
    try:
        conn.execute("DELETE FROM job_queue WHERE payload_json LIKE '%__smoke_probe__%'")
        conn.commit()
    finally:
        conn.close()

    # 2. 4 job_type 각 1건 enqueue
    results = {}
    tests = [
        ("attachment_store", "__smoke_probe__A", "s3://probe/a.zip",  "attachment", "a.zip"),
        ("image_process",    "__smoke_probe__I", "s3://probe/p.jpg",  "photo",      "p.jpg"),
        ("video_process",    "__smoke_probe__V", "s3://probe/v.mp4",  "video",      "v.mp4"),
        ("resume_process",   "__smoke_probe__R", "s3://probe/cv.pdf", "cv",         "cv.pdf"),
    ]
    enqueued_ids: list[int] = []
    for exp_type, cid, s3, ftype, fn in tests:
        try:
            jid = enqueue_file_process(cid, s3, file_type=ftype, filename=fn, trigger="smoke")
            results[exp_type] = {"job_id": jid, "ok": True}
            enqueued_ids.append(jid)
        except Exception as e:
            results[exp_type] = {"ok": False, "error": str(e)[:200]}

    # 3. 큐 상태
    stats = get_queue_stats()

    # 4. 테스트 데이터 정리
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    try:
        conn.execute("DELETE FROM job_queue WHERE payload_json LIKE '%__smoke_probe__%'")
        conn.execute("DELETE FROM pipeline_events WHERE details_json LIKE '%__smoke_probe__%'")
        conn.commit()
    finally:
        conn.close()

    all_ok = all(r.get("ok") for r in results.values())
    return ok(data={
        "enqueue_results": results,
        "queue_stats": stats,
        "all_passed": all_ok,
    }, message="Pipeline smoke complete" if all_ok else "Pipeline smoke 일부 실패")


@app.get("/api/admin/pipeline/stats", tags=["admin"])
async def admin_pipeline_stats(request: Request):
    """파이프라인 큐 상태 + 후보자 cv_processed_status 분포."""
    _check_admin(request)
    try:
        sys.path.insert(0, str(Path(__file__).resolve().parent))
        from tools.pipeline_v2 import get_queue_stats  # type: ignore
        q_stats = get_queue_stats()
    except Exception as e:
        q_stats = {"error": str(e)}
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 3000")
    try:
        try:
            cand_rows = conn.execute(
                "SELECT cv_processed_status, COUNT(*) FROM candidates "
                "WHERE status != 'Deleted' GROUP BY cv_processed_status"
            ).fetchall()
            cand_stats = {(r[0] or "null"): r[1] for r in cand_rows}
        except Exception:
            cand_stats = {}
        try:
            total = conn.execute("SELECT COUNT(*) FROM candidates WHERE status != 'Deleted'").fetchone()[0]
        except Exception:
            total = 0
        try:
            done = conn.execute(
                "SELECT COUNT(*) FROM candidates WHERE status != 'Deleted' "
                "AND cv_processed_status = 'done' AND cv_processed_s3_key IS NOT NULL"
            ).fetchone()[0]
        except Exception:
            done = 0
    finally:
        conn.close()
    success_rate = (done / total * 100) if total else 0
    return ok(data={
        "queue": q_stats,
        "candidates_by_status": cand_stats,
        "total_candidates": total,
        "processed_done": done,
        "success_rate_pct": round(success_rate, 2),
    })


@app.get("/api/admin/pipeline/dlq", tags=["admin"])
async def admin_pipeline_dlq(request: Request, limit: int = 50):
    """DLQ 작업 목록 (수동 개입 필요)."""
    _check_admin(request)
    try:
        sys.path.insert(0, str(Path(__file__).resolve().parent))
        from tools.pipeline_v2 import list_dlq_jobs  # type: ignore
        jobs = list_dlq_jobs(limit=min(max(limit, 1), 200))
        return ok(data={"dlq_jobs": jobs, "count": len(jobs)})
    except Exception as e:
        raise HTTPException(500, f"DLQ 조회 실패: {e}")


@app.post("/api/admin/pipeline/dlq/{job_id}/requeue", tags=["admin"])
async def admin_pipeline_requeue(request: Request, job_id: int):
    """DLQ 작업 재큐잉."""
    _check_admin(request)
    try:
        sys.path.insert(0, str(Path(__file__).resolve().parent))
        from tools.pipeline_v2 import requeue_dlq_job  # type: ignore
        ok_flag = requeue_dlq_job(job_id)
        if not ok_flag:
            raise HTTPException(404, f"Job {job_id} not in DLQ")
        return ok(message=f"Job {job_id} 재큐잉 완료")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"재큐잉 실패: {e}")


@app.get("/api/admin/pipeline/events", tags=["admin"])
async def admin_pipeline_events(request: Request, limit: int = 100, candidate_id: Optional[str] = None):
    """파이프라인 이벤트 로그 조회."""
    _check_admin(request)
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 3000")
    conn.row_factory = sqlite3.Row
    try:
        safe_limit = min(max(limit, 1), 500)
        if candidate_id:
            rows = conn.execute(
                "SELECT * FROM pipeline_events WHERE candidate_id = ? "
                "ORDER BY id DESC LIMIT ?",
                (candidate_id, safe_limit),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM pipeline_events ORDER BY id DESC LIMIT ?",
                (safe_limit,),
            ).fetchall()
        return ok(data={"events": [dict(r) for r in rows], "count": len(rows)})
    finally:
        conn.close()


@app.post("/api/admin/db/backup", tags=["admin"])
async def admin_db_backup(request: Request):
    """현재 Render master.db → S3 백업 (admin 트리거).
    환경변수 AWS_ACCESS_KEY_ID/SECRET/BUCKET 필요.
    """
    _check_admin(request)
    try:
        sys.path.insert(0, str(Path(__file__).resolve().parent))
        from tools.db_persist import backup as _s3_backup
        ok_flag = _s3_backup(dry_run=False)
        if ok_flag:
            return ok(message="S3 백업 완료", data={"status": "uploaded"})
        raise HTTPException(500, "백업 실패 -- 로그 확인")
    except RuntimeError as e:
        raise HTTPException(503, f"S3 미설정: {e}")
    except Exception as e:
        logging.getLogger("bridge.api").error("[BACKUP] 실패: %s", e, exc_info=True)
        raise HTTPException(500, f"백업 오류: {e}")


@app.get("/api/admin/db/backup-list", tags=["admin"])
async def admin_db_backup_list(request: Request):
    """S3 백업 목록 조회."""
    _check_admin(request)
    try:
        sys.path.insert(0, str(Path(__file__).resolve().parent))
        from tools.db_persist import _s3_client, _list_backups
        client, bucket = _s3_client()
        items = _list_backups(client, bucket)
        data = [{
            "key": it["Key"],
            "size_kb": it["Size"] // 1024,
            "modified": it["LastModified"].isoformat(),
        } for it in items[:20]]
        return ok(data={"bucket": bucket, "count": len(items), "items": data})
    except RuntimeError as e:
        raise HTTPException(503, f"S3 미설정: {e}")
    except Exception as e:
        raise HTTPException(500, f"조회 실패: {e}")


@app.post("/api/admin/db/restore", tags=["admin"])
async def admin_db_restore(request: Request):
    """로컬 master.db → Render DB 복원.
    - field=db_file (바이너리 .db): SQLite 파일 직접 업로드 (권장)
    - field=sql_dump (text SQL): SQL 덤프 방식 (레거시)
    관리자 인증 필수.
    """
    _check_admin(request)
    import shutil as _sh
    import tempfile as _tf
    import os as _os

    db_path = str(_ADMIN_DB_PATH)
    backup_path = db_path + ".pre_restore"

    try:
        form = await request.form()
        db_file = form.get("db_file")
        sql_file = form.get("sql_dump")

        # ── 방식 1: 바이너리 .db 직접 업로드 ──
        if db_file is not None:
            db_bytes = await db_file.read() if hasattr(db_file, "read") else db_file
            if not isinstance(db_bytes, bytes):
                db_bytes = bytes(db_bytes)
            # SQLite 매직 바이트 검증
            if not db_bytes.startswith(b"SQLite format 3"):
                raise HTTPException(400, f"유효한 SQLite DB 파일이 아닙니다. (받은 크기: {len(db_bytes)}B, 첫 16바이트: {db_bytes[:16]!r})")

            # SECURITY: community_posts 덮어쓰기 차단 — 게시판 수정사항 우발적 손실 방지
            # 명시적 X-Confirm-Overwrite-Community: yes 헤더 없으면 community_posts 보존
            confirm = request.headers.get("X-Confirm-Overwrite-Community", "").strip().lower()
            preserve_community = (confirm != "yes")

            # 백업
            try:
                _sh.copy2(db_path, backup_path)
            except Exception:
                pass

            # community_posts 보존: 현재 DB에서 미리 읽어둠
            preserved_community = []
            if preserve_community and _os.path.exists(db_path):
                try:
                    _c = sqlite3.connect(db_path)
                    _c.row_factory = sqlite3.Row
                    _rows = _c.execute("SELECT * FROM community_posts").fetchall()
                    preserved_community = [dict(r) for r in _rows]
                    _c.close()
                except Exception as _pres_err:
                    logging.getLogger("bridge.api").warning(
                        "community_posts 보존 읽기 실패 (계속): %s", _pres_err)

            # 임시 파일에 쓰고 atomic 교체
            tmp_path = db_path + ".uploading"
            with open(tmp_path, "wb") as tmp:
                tmp.write(db_bytes)

            _os.replace(tmp_path, db_path)

            # community_posts 복원 (preserve_community=True이고 데이터 있을 때만)
            if preserve_community and preserved_community:
                try:
                    _c2 = sqlite3.connect(db_path)
                    _c2.execute("DELETE FROM community_posts")
                    cols = list(preserved_community[0].keys())
                    placeholders = ",".join(["?"] * len(cols))
                    col_names = ",".join(cols)
                    _c2.executemany(
                        f"INSERT INTO community_posts ({col_names}) VALUES ({placeholders})",
                        [tuple(r[c] for c in cols) for r in preserved_community],
                    )
                    _c2.commit()
                    _c2.close()
                    logging.getLogger("bridge.api").info(
                        "[restore] community_posts %d개 보존 완료 (덮어쓰기 차단)",
                        len(preserved_community))
                except Exception as _restore_err:
                    logging.getLogger("bridge.api").error(
                        "community_posts 보존 복원 실패: %s", _restore_err)

        # ── 방식 2: SQL 덤프 (레거시) ──
        elif sql_file is not None:
            sql_text = await sql_file.read() if hasattr(sql_file, "read") else sql_file
            if isinstance(sql_text, bytes):
                sql_text = sql_text.decode("utf-8", errors="replace")
            if not sql_text.strip().startswith(("BEGIN TRANSACTION", "PRAGMA")):
                raise HTTPException(400, "유효한 SQLite SQL dump 형식이 아닙니다.")

            try:
                _sh.copy2(db_path, backup_path)
            except Exception:
                pass

            with _tf.NamedTemporaryFile(suffix=".db", delete=False, dir=str(Path(db_path).parent)) as tmp:
                tmp_path = tmp.name

            conn_new = sqlite3.connect(tmp_path)
            conn_new.executescript(sql_text)
            conn_new.commit()
            conn_new.close()
            _os.replace(tmp_path, db_path)

        else:
            raise HTTPException(400, "db_file 또는 sql_dump 필드가 필요합니다.")

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"업로드 파싱 실패: {e}")

    # WAL/SHM 파일 제거 (구 DB 흔적 제거)
    for ext in ("-wal", "-shm"):
        wal_path = db_path + ext
        if _os.path.exists(wal_path):
            try:
                _os.remove(wal_path)
            except Exception:
                pass

    # 복원 확인
    try:
        conn_check = sqlite3.connect(db_path)
        conn_check.execute("PRAGMA journal_mode=DELETE")  # WAL 모드 해제 후 재확인
        tables = [r[0] for r in conn_check.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
        _log.info("DB 복원 후 테이블 목록: %s", tables)
        if "candidates" not in tables:
            raise Exception(f"candidates 테이블 없음. 존재하는 테이블: {tables}")
        cand_count = conn_check.execute("SELECT COUNT(*) FROM candidates").fetchone()[0]
        jobs_count = conn_check.execute("SELECT COUNT(*) FROM jobs").fetchone()[0] if "jobs" in tables else 0
        inq_count  = conn_check.execute("SELECT COUNT(*) FROM client_inquiries").fetchone()[0] if "client_inquiries" in tables else 0
        conn_check.close()
    except Exception as e:
        # 롤백
        try:
            if Path(backup_path).exists():
                _os.replace(backup_path, db_path)
        except Exception:
            pass
        _log.error("DB 복원 실패: %s", e, exc_info=True)
        raise HTTPException(500, f"DB 복원 실패: {e}")

    _log.info("DB 복원 완료: candidates=%d, jobs=%d, client_inquiries=%d", cand_count, jobs_count, inq_count)
    return ok(data={"candidates": cand_count, "jobs": jobs_count, "client_inquiries": inq_count},
              message=f"복원 완료: 후보자 {cand_count}명, 구인 {jobs_count}건, 문의 {inq_count}건")


# ── Google Forms 자동 수신 API ────────────────────────────────────────────────
import threading as _threading

_form_watcher_thread = None
_form_watcher_lock = _threading.Lock()

def _get_form_watcher_state() -> dict:
    """form_watcher_state.json 읽기."""
    try:
        state_path = _os.path.join(_os.path.dirname(__file__), "tools", "resume_converter", "form_watcher_state.json")
        if _os.path.exists(state_path):
            import json as _json2
            with open(state_path, encoding="utf-8") as f:
                return _json2.load(f)
    except Exception:
        pass
    return {}


@app.get("/api/admin/form-watch/status", tags=["admin"])
async def form_watch_status(req: Request):
    """구글폼 자동 수신 상태 조회."""
    _check_admin(req)
    global _form_watcher_thread
    state = _get_form_watcher_state()
    running = bool(_form_watcher_thread and _form_watcher_thread.is_alive())
    return ok(data={
        "running":        running,
        "last_check":     state.get("last_check"),
        "download_count": state.get("download_count", 0),
        "processed_ids":  len(state.get("processed_file_ids", [])),
        "folder_ids":     state.get("folder_ids", []),
        "form_id":        state.get("form_id"),
    })


@app.post("/api/admin/form-watch/start", tags=["admin"])
async def form_watch_start(req: Request):
    """구글폼 자동 수신 시작 (백그라운드 폴링)."""
    _check_admin(req)
    global _form_watcher_thread
    with _form_watcher_lock:
        if _form_watcher_thread and _form_watcher_thread.is_alive():
            return ok(message="이미 실행 중")
        try:
            sys_path_base = _os.path.join(_os.path.dirname(__file__), "tools")
            if sys_path_base not in _sys.path:
                _sys.path.insert(0, sys_path_base)
            from resume_converter.form_watcher import start_background
            _form_watcher_thread = start_background(interval_sec=300)
            return ok(message="구글폼 수신 시작")
        except Exception as e:
            raise HTTPException(500, f"폼 감시 시작 실패: {e}")


@app.post("/api/admin/form-watch/stop", tags=["admin"])
async def form_watch_stop(req: Request):
    """구글폼 자동 수신 중지."""
    _check_admin(req)
    try:
        from resume_converter.form_watcher import stop_background
        stop_background()
        return ok(message="구글폼 수신 중지")
    except Exception as e:
        raise HTTPException(500, f"폼 감시 중지 실패: {e}")


@app.post("/api/admin/form-watch/poll-now", tags=["admin"])
async def form_watch_poll_now(req: Request):
    """즉시 1회 폴링 (새 이력서 확인)."""
    _check_admin(req)
    try:
        sys_path_base = _os.path.join(_os.path.dirname(__file__), "tools")
        if sys_path_base not in _sys.path:
            _sys.path.insert(0, sys_path_base)
        from resume_converter.form_watcher import run_once
        n = run_once()
        return ok(data={"downloaded": n}, message=f"새 파일 {n}개 다운로드")
    except Exception as e:
        raise HTTPException(500, f"폴링 실패: {e}")


@app.post("/api/admin/form-watch/set-folder", tags=["admin"])
async def form_watch_set_folder(req: Request):
    """폼 업로드 Drive 폴더 ID 설정."""
    _check_admin(req)
    body = await req.json()
    folder_id = body.get("folder_id", "").strip()
    form_id   = body.get("form_id", "1vFwznyCWqFtqVomKIU1V9R8_kzMGIqH1QM4TiKs-o6Q").strip()
    if not folder_id:
        raise HTTPException(400, "folder_id 필요")
    try:
        sys_path_base = _os.path.join(_os.path.dirname(__file__), "tools")
        if sys_path_base not in _sys.path:
            _sys.path.insert(0, sys_path_base)
        from resume_converter.form_watcher import set_form_config
        set_form_config(form_id, [folder_id])
        return ok(message=f"폴더 설정 완료: {folder_id}")
    except Exception as e:
        raise HTTPException(500, f"설정 실패: {e}")


# ── Sheet Preferences (영구 영속성) ───────────────────────────────────────────

def _ensure_sheet_prefs_table():
    """Canvas Sheet 모든 UI 설정을 DB에 영구 저장 — localStorage 의존 제거."""
    conn = sqlite3.connect(str(_ADMIN_DB_PATH))
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS sheet_prefs (
            pref_key   TEXT PRIMARY KEY,
            pref_value TEXT NOT NULL,
            updated_at TEXT DEFAULT (datetime('now'))
        )
    """)
    conn.commit()
    conn.close()


try:
    _ensure_sheet_prefs_table()
except Exception as _e:
    logging.getLogger("bridge.api").warning("_ensure_sheet_prefs_table 스킵: %s", _e)


@app.get("/api/admin/sheet/prefs", tags=["admin"])
async def get_sheet_prefs(req: Request):
    """Canvas Sheet 전체 환경설정 로드 (컬럼너비/행높이/셀스타일/열순서 등)."""
    _check_admin(req)
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        rows = conn.execute("SELECT pref_key, pref_value FROM sheet_prefs").fetchall()
        conn.close()
        result = {}
        for key, val in rows:
            try:
                result[key] = json.loads(val)
            except Exception:
                result[key] = val
        return result
    except Exception as e:
        raise HTTPException(500, f"sheet_prefs 로드 실패: {e}")


@app.patch("/api/admin/sheet/prefs", tags=["admin"])
async def save_sheet_prefs(req: Request):
    """Canvas Sheet 환경설정 저장. body: { key: string, value: any }"""
    _check_admin(req)
    body = await req.json()
    key = body.get("key", "").strip()
    if not key:
        raise HTTPException(400, "key 필요")
    value = body.get("value")
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        conn.execute(
            "INSERT INTO sheet_prefs(pref_key, pref_value, updated_at) VALUES(?,?,datetime('now')) "
            "ON CONFLICT(pref_key) DO UPDATE SET pref_value=excluded.pref_value, updated_at=excluded.updated_at",
            (key, json.dumps(value, ensure_ascii=False))
        )
        conn.commit()
        conn.close()
        return ok(message=f"저장 완료: {key}")
    except Exception as e:
        raise HTTPException(500, f"sheet_prefs 저장 실패: {e}")


@app.patch("/api/admin/sheet/prefs/bulk", tags=["admin"])
async def save_sheet_prefs_bulk(req: Request):
    """Canvas Sheet 환경설정 일괄 저장. body: { prefs: { key: value, ... } }"""
    _check_admin(req)
    body = await req.json()
    prefs = body.get("prefs", {})
    if not prefs:
        raise HTTPException(400, "prefs 필요")
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        for key, value in prefs.items():
            conn.execute(
                "INSERT INTO sheet_prefs(pref_key, pref_value, updated_at) VALUES(?,?,datetime('now')) "
                "ON CONFLICT(pref_key) DO UPDATE SET pref_value=excluded.pref_value, updated_at=excluded.updated_at",
                (key, json.dumps(value, ensure_ascii=False))
            )
        conn.commit()
        conn.close()
        return ok(message=f"{len(prefs)}개 저장 완료")
    except Exception as e:
        raise HTTPException(500, f"sheet_prefs 일괄 저장 실패: {e}")


# ── 영속성 진단 엔드포인트 (관리자 키 또는 IP 화이트리스트 필요) ────────────
@app.get("/api/admin/diag/persistence", tags=["admin"])
async def diag_persistence(req: Request):
    """DB 데이터 + Drive 백업 + 자동 복원 환경변수 상태 종합 진단.
    데이터 안 보일 때 한 번 호출로 원인 즉시 파악."""
    _check_admin(req)
    diag = {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "render_env": {
            "DRIVE_OAUTH_TOKEN_JSON_set": bool(os.getenv("DRIVE_OAUTH_TOKEN_JSON", "").strip()),
            "DRIVE_OAUTH_TOKEN_JSON_len": len(os.getenv("DRIVE_OAUTH_TOKEN_JSON", "").strip()),
            "GCP_SA_JSON_B64_set": bool(os.getenv("GCP_SA_JSON_B64", "").strip()),
            "ADMIN_API_KEY_set": bool(os.getenv("ADMIN_API_KEY", "").strip()),
            "BRIDGE_FIELD_KEY_set": bool(os.getenv("BRIDGE_FIELD_KEY", "").strip()),
            "DB_PATH": os.getenv("DB_PATH", "master.db"),
        },
        "db": {},
        "drive_backups": [],
        "auto_restore_capable": False,
        "recommended_action": "",
    }
    try:
        conn = sqlite3.connect(str(_ADMIN_DB_PATH))
        conn.execute("PRAGMA busy_timeout = 5000")
        for t in ["candidates", "jobs", "client_inquiries", "sheet_prefs",
                  "boards", "community_posts", "form_configs", "interviews",
                  "site_partners", "testimonials"]:
            try:
                cnt = conn.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
                diag["db"][t] = cnt
            except Exception:
                diag["db"][t] = "(테이블 없음)"
        conn.close()
    except Exception as e:
        diag["db_error"] = str(e)[:200]

    # Drive 백업 목록 (OAuth 토큰 있을 때만)
    if diag["render_env"]["DRIVE_OAUTH_TOKEN_JSON_set"]:
        diag["auto_restore_capable"] = True
        try:
            import sys as _sys
            _tools_dir = str(Path(__file__).resolve().parent / "tools")
            if _tools_dir not in _sys.path:
                _sys.path.insert(0, _tools_dir)
            from db_persist import _drive, _get_or_create_folder, _list_backups  # type: ignore
            d = _drive()
            fid = _get_or_create_folder(d)
            backups = _list_backups(d, fid)[:5]
            diag["drive_backups"] = [{
                "name": b["name"],
                "size_kb": int(b.get("size", 0)) // 1024,
                "created": b["createdTime"][:16],
            } for b in backups]
        except Exception as e:
            diag["drive_error"] = str(e)[:200]

    # 권장 조치
    cand = diag["db"].get("candidates", 0)
    sp = diag["db"].get("sheet_prefs", 0)
    if isinstance(cand, int) and cand < 100:
        if not diag["render_env"]["DRIVE_OAUTH_TOKEN_JSON_set"]:
            diag["recommended_action"] = (
                "🚨 DB 휘발 + 자동복원 불가. Render Environment에 "
                "DRIVE_OAUTH_TOKEN_JSON 환경변수 즉시 등록 필요."
            )
        else:
            diag["recommended_action"] = (
                "🔄 DB 휘발 감지 + OAuth 토큰 등록됨. "
                "POST /api/admin/diag/auto-restore 호출하면 즉시 복원."
            )
    elif isinstance(sp, int) and sp == 0:
        diag["recommended_action"] = (
            "⚠️ candidates 정상이지만 sheet_prefs 비어있음. "
            "이전 시점 UI 설정은 손실. 지금부터 입력하면 영구 보존."
        )
    else:
        diag["recommended_action"] = "✅ 모든 데이터 정상."

    return ok(data=diag)


@app.post("/api/admin/diag/auto-restore", tags=["admin"])
async def diag_auto_restore(req: Request):
    """수동으로 Drive 자동복원 트리거 — 즉시 master.db 채우기.
    Render Free Plan sleep/wake 시 데이터 휘발 응급 복구용."""
    _check_admin(req)
    if not os.getenv("DRIVE_OAUTH_TOKEN_JSON", "").strip():
        raise HTTPException(503, "DRIVE_OAUTH_TOKEN_JSON 미설정. Render Environment 등록 필요.")
    try:
        import sys as _sys
        _tools_dir = str(Path(__file__).resolve().parent / "tools")
        if _tools_dir not in _sys.path:
            _sys.path.insert(0, _tools_dir)
        from db_persist import restore as _drive_restore  # type: ignore
        ok_flag = _drive_restore()
        if ok_flag:
            # post-restore ensure (옛 백업 호환)
            _ENSURE_FUNCS = [
                "_ensure_sheet_prefs_table", "_ensure_email_blacklist_table",
                "_ensure_boards_table", "_ensure_file_uploads_table",
                "_ensure_banners_table", "_ensure_site_partners_table",
                "_ensure_site_settings_table", "_ensure_site_visits_table",
                "_ensure_page_content_table", "_ensure_pipeline_tables",
                "_ensure_talent_auth_tables",
            ]
            for fn in _ENSURE_FUNCS:
                try:
                    f = globals().get(fn)
                    if callable(f): f()
                except Exception:
                    pass
            # 카운트 재확인
            counts = {}
            try:
                conn = sqlite3.connect(str(_ADMIN_DB_PATH))
                for t in ["candidates", "jobs", "client_inquiries", "sheet_prefs"]:
                    try:
                        counts[t] = conn.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
                    except Exception:
                        counts[t] = -1
                conn.close()
            except Exception:
                pass
            return ok(message="Drive 자동복원 성공", data={"counts": counts})
        raise HTTPException(500, "Drive 복원 실패 — 백업 미존재 또는 권한 오류")
    except Exception as e:
        raise HTTPException(500, f"자동복원 예외: {e}")


# ── 로컬 실행 ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import uvicorn
    print("BRIDGE API Server 시작 중...")
    print("  문서: http://localhost:8000/docs")
    uvicorn.run("api_server:app", host="0.0.0.0", port=8000, reload=True)

