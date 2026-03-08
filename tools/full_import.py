import openpyxl, sqlite3, hashlib, datetime, sys, shutil
sys.stdout.reconfigure(encoding='utf-8')

BASE      = "Q:/Claudework/bridge base"
CAND_XLSX = f"{BASE}/original_candidates/CANDIDATES_MASTER_COMBINED.xlsx"
JOBS_XLSX = f"{BASE}/original_jobs/JOBS_MASTER_COMBINED.xlsx"
UNI_XLSX  = f"{BASE}/archive/old_data/unified_bridge_data.xlsx"
DB_PATH   = f"{BASE}/master.db"

lines = []
def p(msg=""):
    print(msg)
    lines.append(str(msg))

def cv(v):
    if v is None: return None
    s = str(v).strip()
    return s if s and s not in ('None','nan','') else None

# ─── 1. PRE 체크섬 + 백업 ────────────────────────────────────────
conn = sqlite3.connect(DB_PATH)
h = hashlib.sha256('\n'.join(conn.iterdump()).encode()).hexdigest()
conn.close()
ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
with open(f"{BASE}/tasks/db_checksum.log", 'a') as f:
    f.write(f"{datetime.datetime.now().isoformat()} PRE-FULL-IMPORT {h}\n")
shutil.copy(DB_PATH, f"{DB_PATH}.backup_{ts}")
p(f"PRE checksum: {h[:16]}...  백업: master.db.backup_{ts}")

# ─── 2. CANDIDATES 임포트 ────────────────────────────────────────
p()
p("=== CANDIDATES_MASTER_COMBINED.xlsx → candidates ===")

wb = openpyxl.load_workbook(CAND_XLSX, read_only=True, data_only=True)
sh = wb['Sheet1']

# col_index → db_field
COL_MAP = {
    16: 'preferences',
    18: 'notice',
    19: 'applied',
    22: 'interview_time',
    24: 'major',
    33: 'religion',
    39: 'consent',
    43: 'placed_company',
    44: 'placed_salary',
    45: 'start_month',
    46: 'housing_detail',
    47: 'referral_fee',
    51: 'visa_type',
    53: 'education_level',
    54: 'health_info',
    57: 'tattoo',
    58: 'piercings',
    59: 'married',
    60: 'korean_criminal_record',
    61: 'documents',
    64: 'notes',
}

# 이메일/이름 → {col_i: value} 매핑 구축 (row3부터, row2=한글헤더 skip)
email_map = {}  # email.lower() → {col_i: val}
name_map  = {}  # name.lower() → {col_i: val}

for row in sh.iter_rows(min_row=3, values_only=True):
    raw_email = row[1] if len(row) > 1 else None
    raw_name  = row[2] if len(row) > 2 else None

    entry = {}
    for col_i in COL_MAP:
        v = cv(row[col_i] if len(row) > col_i else None)
        if v:
            entry[col_i] = v
    if not entry:
        continue

    if raw_email and '@' in str(raw_email):
        key = str(raw_email).strip().lower()
        if key not in email_map:
            email_map[key] = entry
        else:
            for k, v in entry.items():
                if k not in email_map[key]:
                    email_map[key][k] = v

    if raw_name:
        key = str(raw_name).strip().lower()
        if key not in name_map:
            name_map[key] = entry
        else:
            for k, v in entry.items():
                if k not in name_map[key]:
                    name_map[key][k] = v

wb.close()
p(f"Excel 매핑: email={len(email_map)}개, name={len(name_map)}개")

conn = sqlite3.connect(DB_PATH)
conn.execute("PRAGMA busy_timeout = 10000")
cur = conn.cursor()

# DB 컬럼 목록
cur.execute("PRAGMA table_info(candidates)")
db_cols = {r[1] for r in cur.fetchall()}

# 실제 존재하는 컬럼만 사용
EFFECTIVE = {ci: f for ci, f in COL_MAP.items() if f in db_cols}
MISSING_COLS = {ci: f for ci, f in COL_MAP.items() if f not in db_cols}
if MISSING_COLS:
    p(f"DB에 없는 컬럼 (skip): {list(MISSING_COLS.values())}")

# BEFORE 채움률
p()
p("BEFORE 채움률:")
before = {}
for ci, f in EFFECTIVE.items():
    cur.execute(f"SELECT COUNT(*) FROM candidates WHERE {f} IS NOT NULL AND TRIM(CAST({f} AS TEXT)) != ''")
    cnt = cur.fetchone()[0]
    before[f] = cnt
    p(f"  {f:30s}: {cnt:5d}건  ({cnt/3059*100:.1f}%)")

# 업데이트 실행
OVERWRITE = {'notes', 'preferences', 'applied', 'notice'}
stats = {f: 0 for f in EFFECTIVE.values()}
match_email = 0
match_name  = 0
no_match    = 0

cur.execute("SELECT candidate_id, email, full_name FROM candidates")
all_cands = cur.fetchall()

for (cid, raw_email, raw_name) in all_cands:
    entry = None
    if raw_email and '@' in str(raw_email):
        ekey = str(raw_email).strip().lower()
        if ekey in email_map:
            entry = email_map[ekey]
            match_email += 1

    if entry is None and raw_name:
        nkey = str(raw_name).strip().lower()
        if nkey in name_map:
            entry = name_map[nkey]
            match_name += 1

    if entry is None:
        no_match += 1
        continue

    updates = []
    params  = []
    for col_i, db_field in EFFECTIVE.items():
        new_val = entry.get(col_i)
        if not new_val:
            continue
        cur.execute(f"SELECT {db_field} FROM candidates WHERE candidate_id = ?", (cid,))
        row_r = cur.fetchone()
        cur_val = str(row_r[0]).strip() if row_r and row_r[0] else ''
        if not cur_val or db_field in OVERWRITE:
            updates.append(f"{db_field} = ?")
            params.append(new_val)
            stats[db_field] += 1

    if updates:
        params.append(cid)
        cur.execute(f"UPDATE candidates SET {', '.join(updates)} WHERE candidate_id = ?", params)

conn.commit()

# AFTER 채움률
p()
p("AFTER 채움률:")
for ci, f in EFFECTIVE.items():
    cur.execute(f"SELECT COUNT(*) FROM candidates WHERE {f} IS NOT NULL AND TRIM(CAST({f} AS TEXT)) != ''")
    cnt = cur.fetchone()[0]
    diff = cnt - before.get(f, 0)
    p(f"  {f:30s}: {cnt:5d}건  ({cnt/3059*100:.1f}%)  [+{diff}]")

p()
p("업데이트 통계 (field별):")
for f, cnt in stats.items():
    p(f"  {f}: {cnt}건")
p(f"  email_match: {match_email}")
p(f"  name_match:  {match_name}")
p(f"  no_match:    {no_match}")

# ─── 3. CLIENT_INQUIRIES 임포트 ────────────────────────────────────
p()
p("=== JOBS_MASTER_COMBINED.xlsx → client_inquiries ===")

wb2 = openpyxl.load_workbook(JOBS_XLSX, read_only=True, data_only=True)
sh2 = wb2['Sheet1']

cur.execute("PRAGMA table_info(client_inquiries)")
ci_cols = {r[1] for r in cur.fetchall()}

JOBS_COL_MAP = {
    1:  'submitted_at',
    2:  'email',
    4:  'contact_name',
    5:  'phone',
    7:  'school_name',
    9:  'location',
    11: 'start_date',
    12: 'teaching_age',
    14: 'schedule',
    15: 'working_hours',
    16: 'vacancies',
    22: 'salary_raw',
    23: 'meal',
    24: 'housing_type',
    25: 'travel_support',
    26: 'benefits',
    27: 'vacation',
    28: 'sick_leave',
    29: 'memo',
}

# DB 기존 레코드 인덱스
cur.execute("SELECT id, email, school_name FROM client_inquiries WHERE is_deleted = 0")
ci_existing = {}
for (rid, e, s) in cur.fetchall():
    if e and '@' in str(e):
        k = str(e).strip().lower()
        ci_existing[k] = rid
    if s:
        k2 = f"school:{str(s).strip().lower()}"
        if k2 not in ci_existing:
            ci_existing[k2] = rid

insert_count = 0
update_count = 0
skip_count   = 0

for row in sh2.iter_rows(min_row=2, values_only=True):
    raw_email  = row[2] if len(row) > 2 else None
    raw_school = row[7] if len(row) > 7 else None

    has_email  = raw_email and '@' in str(raw_email)
    has_school = raw_school and str(raw_school).strip()

    if not has_email and not has_school:
        skip_count += 1
        continue

    # 매칭
    matched_id = None
    if has_email:
        ekey = str(raw_email).strip().lower()
        matched_id = ci_existing.get(ekey)
    if matched_id is None and has_school:
        skey = f"school:{str(raw_school).strip().lower()}"
        matched_id = ci_existing.get(skey)

    if matched_id:
        # 기존 레코드: 빈 필드 채우기
        upd = []
        prm = []
        for col_i, db_field in JOBS_COL_MAP.items():
            if db_field in ('email','school_name') or db_field not in ci_cols:
                continue
            v = cv(row[col_i] if len(row) > col_i else None)
            if not v:
                continue
            cur.execute(f"SELECT {db_field} FROM client_inquiries WHERE id = ?", (matched_id,))
            cv_row = cur.fetchone()
            cur_v = str(cv_row[0]).strip() if cv_row and cv_row[0] else ''
            if not cur_v:
                upd.append(f"{db_field} = ?")
                prm.append(v)
        if upd:
            prm.append(matched_id)
            cur.execute(f"UPDATE client_inquiries SET {', '.join(upd)} WHERE id = ?", prm)
            update_count += 1
    else:
        # 신규 INSERT
        fields = []
        vals   = []
        for col_i, db_field in JOBS_COL_MAP.items():
            if db_field not in ci_cols:
                continue
            v = cv(row[col_i] if len(row) > col_i else None)
            if v:
                fields.append(db_field)
                vals.append(v)
        if not fields:
            skip_count += 1
            continue
        fields += ['is_deleted', 'source']
        vals   += [0, 'jobs_excel']
        ph = ','.join(['?']*len(vals))
        cur.execute(f"INSERT INTO client_inquiries ({','.join(fields)}) VALUES ({ph})", vals)
        insert_count += 1

conn.commit()
wb2.close()
p(f"신규 INSERT: {insert_count}건 / 업데이트: {update_count}건 / 스킵: {skip_count}건")

# ─── 4. unified_bridge_data Candidates_Master (21행) ──────────────
p()
p("=== unified_bridge_data.xlsx Candidates_Master (21행) ===")
wb3 = openpyxl.load_workbook(UNI_XLSX, read_only=True, data_only=True)
sh3 = wb3['Candidates_Master']

UNI_MAP = {
    1:  'email',
    10: 'education_level',
    11: 'major',
    12: 'certification',
    21: 'job_prefs',
    22: 'experience',
    34: 'interview_time',
    35: 'kakaotalk',
    36: 'mobile_phone',
}
uni_upd = 0
uni_skip = 0

for row in sh3.iter_rows(min_row=2, values_only=True):
    raw_email = row[1] if len(row) > 1 else None
    if not raw_email or '@' not in str(raw_email):
        uni_skip += 1
        continue
    ekey = str(raw_email).strip().lower()
    cur.execute("SELECT candidate_id FROM candidates WHERE LOWER(TRIM(email)) = ?", (ekey,))
    m = cur.fetchone()
    if not m:
        uni_skip += 1
        continue
    cid = m[0]
    upd = []
    prm = []
    for col_i, db_field in UNI_MAP.items():
        if db_field == 'email' or db_field not in db_cols:
            continue
        v = cv(row[col_i] if len(row) > col_i else None)
        if not v:
            continue
        cur.execute(f"SELECT {db_field} FROM candidates WHERE candidate_id = ?", (cid,))
        cv_r = cur.fetchone()
        cur_v = str(cv_r[0]).strip() if cv_r and cv_r[0] else ''
        if not cur_v:
            upd.append(f"{db_field} = ?")
            prm.append(v)
    if upd:
        prm.append(cid)
        cur.execute(f"UPDATE candidates SET {', '.join(upd)} WHERE candidate_id = ?", prm)
        uni_upd += 1

conn.commit()
wb3.close()
p(f"unified 업데이트: {uni_upd}건 / 스킵: {uni_skip}건")

# ─── 5. POST 검증 ─────────────────────────────────────────────────
p()
p("=== POST 검증 ===")
cur.execute("PRAGMA integrity_check")
p(f"integrity: {cur.fetchone()[0]}")
for t in ['candidates', 'client_inquiries', 'jobs']:
    cur.execute(f"SELECT COUNT(*) FROM {t}")
    p(f"{t}: {cur.fetchone()[0]}")

p()
p("전체 채움률 최종 (candidates):")
check_fields = [
    'email','full_name','nationality','arc_holders','ancestry','dob','gender',
    'current_location','start_date','target','area_prefs','reference','experience',
    'preferences','applied','job_prefs','current_salary','desired_salary',
    'interview_time','education_level','major','certification','documents','doc_status',
    'health_info','tattoo','piercings','dependents','married','housing','housing_type',
    'religion','e_visa','visa_type','kakaotalk','mobile_phone','criminal_record',
    'criminal_record_check','korean_criminal_record','consent','fact_check',
    'how_to','source','placed_company','placed_salary','start_month','housing_detail',
    'referral_fee','process_date','past_placement','photo_url','status','notes',
]
for f in check_fields:
    if f not in db_cols:
        p(f"  {f:30s}: 컬럼없음")
        continue
    cur.execute(f"SELECT COUNT(*) FROM candidates WHERE {f} IS NOT NULL AND TRIM(CAST({f} AS TEXT)) != ''")
    cnt = cur.fetchone()[0]
    p(f"  {f:30s}: {cnt:5d}건  ({cnt/3059*100:.1f}%)")

# POST 체크섬
h2 = hashlib.sha256('\n'.join(conn.iterdump()).encode()).hexdigest()
conn.close()
with open(f"{BASE}/tasks/db_checksum.log", 'a') as f:
    f.write(f"{datetime.datetime.now().isoformat()} POST-FULL-IMPORT {h2}\n")
p()
p(f"POST checksum: {h2[:16]}...")
p("=== 완료 ===")

with open(f"{BASE}/tasks/import_report_full.txt", 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))
print(f"\n리포트 저장: tasks/import_report_full.txt")
