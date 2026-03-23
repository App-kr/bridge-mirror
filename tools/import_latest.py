"""
import_latest.py — 최신 Excel 원본 → master.db 동기화
=======================================================
대상 파일: original_candidates/BRIDGE_Candidates_Source.xlsx
  - Source 시트 (3122행) → candidates 테이블 (UPSERT)
  - New 시트 (66행) → candidates 테이블 (status='Active' 마킹)
  - Client 시트 (2689행) → client_inquiries 테이블 (UPSERT)

사용법:
  python tools/import_latest.py --dry       # 미리보기 (DB 변경 없음)
  python tools/import_latest.py             # 실제 임포트
  python tools/import_latest.py --sheet source   # Source만
  python tools/import_latest.py --sheet client   # Client만
"""
import openpyxl
import sqlite3
import hashlib
import datetime
import sys
import shutil
import os
import re

sys.stdout.reconfigure(encoding='utf-8')

BASE = "Q:/Claudework/bridge base"
XLSX_PATH = f"{BASE}/original_candidates/BRIDGE_Candidates_Source.xlsx"
DB_PATH = f"{BASE}/master.db"

DRY_RUN = "--dry" in sys.argv
SHEET_FILTER = None
if "--sheet" in sys.argv:
    idx = sys.argv.index("--sheet")
    if idx + 1 < len(sys.argv):
        SHEET_FILTER = sys.argv[idx + 1].lower()

lines = []
def p(msg=""):
    print(msg)
    lines.append(str(msg))


def cv(v):
    """셀 값 정규화."""
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s not in ('None', 'nan', '') else None


# ── Source 시트 컬럼 매핑 (54열) ──────────────────────────────────────
SOURCE_COL_MAP = {
    0:  'email',
    1:  'full_name',
    # 2: Photo (skip)
    # 3: No (skip)
    5:  'nationality',
    6:  'ancestry',
    7:  'dob',
    8:  'gender',
    9:  'current_location',
    10: 'start_date',
    11: 'target',
    12: 'area_prefs',
    13: 'reference',
    14: 'experience',
    15: 'employment',
    16: 'job_prefs',
    17: 'contract_offered',   # Apply
    19: 'current_salary',
    20: 'desired_salary',
    21: 'interview_time',
    23: 'major',
    24: 'certification',
    25: 'documents',
    26: 'health_info',
    27: 'personal_consideration',
    28: 'pets',
    29: 'tattoo',
    30: 'piercings',
    31: 'married',
    32: 'housing',
    33: 'religion',
    34: 'e_visa',
    35: 'kakaotalk',
    36: 'mobile_phone',
    37: 'criminal_record',
    39: 'passport',
    40: 'consent',
    41: 'fact_check',
    42: 'how_to',
    45: 'placed_company',
    46: 'placed_salary',
    47: 'start_month',
    48: 'housing_detail',
    49: 'referral_fee',
    50: 'process_date',
}

# ── New 시트 컬럼 매핑 (50열) ─────────────────────────────────────────
NEW_COL_MAP = {
    0:  'email',
    1:  'full_name',
    4:  'arc_holders',
    5:  'nationality',
    6:  'ancestry',
    7:  'dob',
    8:  'gender',
    9:  'current_location',
    10: 'start_date',
    11: 'target',
    12: 'area_prefs',
    13: 'reference',
    14: 'experience',
    15: 'employment',
    16: 'job_prefs',
    17: 'interview_time',
    18: 'contract_offered',
    19: 'current_salary',
    20: 'desired_salary',
    21: 'interview_time',
    22: 'education_level',
    23: 'major',
    24: 'certification',
    25: 'documents',
    26: 'health_info',
    27: 'personal_consideration',
    28: 'piercings',
    29: 'dependents',
    30: 'married',
    31: 'housing',
    32: 'religion',
    33: 'e_visa',
    34: 'kakaotalk',
    35: 'mobile_phone',
    36: 'criminal_record',
    37: 'korean_criminal_record',
    38: 'consent',
    39: 'fact_check',
    40: 'how_to',
    42: 'placed_company',
    43: 'placed_salary',
    44: 'start_month',
    45: 'housing_detail',
    46: 'referral_fee',
    47: 'process_date',
    48: 'past_placement',
}

# ── Client 시트 컬럼 매핑 (12열) ──────────────────────────────────────
CLIENT_COL_MAP = {
    # 0: 지역 (special handling)
    # 1: 도시 (special handling → location)
    2: 'school_name',
    3: 'email',
    4: 'phone',
    5: 'teaching_age',
    # 6: 분류 (skip)
    7: 'travel_support',
    8: 'memo',
    9: 'vacation',
    10: 'notes',
    # 11: 기타/메일 (special handling)
}


def _stable_id(email: str, name: str) -> str:
    """email+name 기반 안정적 candidate_id 생성."""
    raw = f"{email or ''}|{name or ''}".lower().strip()
    return hashlib.sha1(raw.encode()).hexdigest()[:12]


def _parse_location(region_val, city_val) -> str:
    """지역+도시 → location 문자열."""
    parts = []
    if region_val:
        parts.append(str(region_val).strip())
    if city_val:
        parts.append(str(city_val).strip())
    return " ".join(parts) if parts else ""


def import_source_sheet(wb, conn, db_cols):
    """Source 시트 → candidates 테이블 (UPSERT)."""
    p()
    p("=" * 60)
    p("=== Source 시트 → candidates (UPSERT) ===")

    ws = wb['Source']
    cur = conn.cursor()

    # 유효 컬럼만
    effective = {ci: f for ci, f in SOURCE_COL_MAP.items() if f in db_cols}
    skipped_cols = {ci: f for ci, f in SOURCE_COL_MAP.items() if f not in db_cols}
    if skipped_cols:
        p(f"  DB에 없는 컬럼 (skip): {list(skipped_cols.values())}")

    # 기존 DB 인덱스 (email → candidate_id)
    cur.execute("SELECT candidate_id, LOWER(TRIM(email)), LOWER(TRIM(full_name)) FROM candidates")
    email_to_id = {}
    name_to_id = {}
    for cid, e, n in cur.fetchall():
        if e and '@' in e:
            email_to_id[e] = cid
        if n:
            name_to_id[n] = cid

    update_count = 0
    insert_count = 0
    skip_count = 0
    field_updates = {f: 0 for f in effective.values()}

    # OVERWRITE 필드: 항상 최신 값으로 덮어씀
    OVERWRITE = {'notes', 'placed_company', 'placed_salary', 'start_month',
                 'housing_detail', 'referral_fee', 'process_date', 'contract_offered'}

    row_num = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1
        raw_email = cv(row[0] if len(row) > 0 else None)
        raw_name = cv(row[1] if len(row) > 1 else None)

        if not raw_email and not raw_name:
            skip_count += 1
            continue

        # 매칭
        matched_id = None
        if raw_email and '@' in raw_email:
            matched_id = email_to_id.get(raw_email.lower().strip())
        if matched_id is None and raw_name:
            matched_id = name_to_id.get(raw_name.lower().strip())

        if matched_id:
            # UPDATE: 빈 필드 채우기 + OVERWRITE 필드 갱신
            updates = []
            params = []
            for col_i, db_field in effective.items():
                if db_field in ('email', 'full_name'):
                    continue
                v = cv(row[col_i] if len(row) > col_i else None)
                if not v:
                    continue
                cur.execute(f"SELECT {db_field} FROM candidates WHERE candidate_id = ?", (matched_id,))
                cur_row = cur.fetchone()
                cur_val = str(cur_row[0]).strip() if cur_row and cur_row[0] else ''
                if not cur_val or db_field in OVERWRITE:
                    updates.append(f"{db_field} = ?")
                    params.append(v)
                    field_updates[db_field] = field_updates.get(db_field, 0) + 1
            if updates:
                updates.append("updated_at = ?")
                params.append(datetime.datetime.now().isoformat())
                params.append(matched_id)
                if not DRY_RUN:
                    cur.execute(
                        f"UPDATE candidates SET {', '.join(updates)} WHERE candidate_id = ?",
                        params,
                    )
                update_count += 1
        else:
            # INSERT: 신규 후보자
            fields_vals = {}
            for col_i, db_field in effective.items():
                v = cv(row[col_i] if len(row) > col_i else None)
                if v:
                    fields_vals[db_field] = v

            if not fields_vals.get('email') and not fields_vals.get('full_name'):
                skip_count += 1
                continue

            # candidate_id 생성
            cid = _stable_id(fields_vals.get('email', ''), fields_vals.get('full_name', ''))
            fields_vals['candidate_id'] = cid
            fields_vals['status'] = 'Inactive'
            fields_vals['source'] = 'excel_source'
            fields_vals['inbox_status'] = 'new'
            fields_vals['is_deleted'] = 0
            fields_vals['created_at'] = datetime.datetime.now().isoformat()

            # DB에 있는 컬럼만
            final_fields = {k: v for k, v in fields_vals.items() if k in db_cols or k == 'candidate_id'}
            cols = list(final_fields.keys())
            vals = list(final_fields.values())

            if not DRY_RUN:
                try:
                    ph = ','.join(['?'] * len(vals))
                    cur.execute(
                        f"INSERT INTO candidates ({','.join(cols)}) VALUES ({ph})",
                        vals,
                    )
                except sqlite3.IntegrityError:
                    skip_count += 1
                    continue

            insert_count += 1
            # 인덱스 갱신
            if fields_vals.get('email') and '@' in fields_vals['email']:
                email_to_id[fields_vals['email'].lower()] = cid

    if not DRY_RUN:
        conn.commit()

    p(f"  총 행: {row_num}")
    p(f"  UPDATE: {update_count}건 / INSERT: {insert_count}건 / SKIP: {skip_count}건")
    p(f"  필드별 업데이트:")
    for f, cnt in sorted(field_updates.items(), key=lambda x: -x[1]):
        if cnt > 0:
            p(f"    {f:30s}: {cnt}건")

    return update_count, insert_count


def _clean_val(v):
    """Excel 숫자 소수점 제거 (93.0 → 93) + 정리."""
    if v is None:
        return None
    s = str(v).strip()
    if not s or s in ('None', 'nan'):
        return None
    # 숫자.0 패턴 정리 (93.0 → 93, 14253284554.0 → 14253284554)
    if re.match(r'^\d+\.0$', s):
        s = s[:-2]
    return s


def import_new_sheet(wb, conn, db_cols):
    """New 시트 → candidates 전체 덮어쓰기 (원어민관리 시트 = 최신 진실)."""
    p()
    p("=" * 60)
    p("=== New 시트 → candidates (전체 덮어쓰기) ===")

    ws = wb['New']
    cur = conn.cursor()

    effective = {ci: f for ci, f in NEW_COL_MAP.items() if f in db_cols}
    activated = 0
    overwritten = 0
    inserted = 0
    skip = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_email = cv(row[0] if len(row) > 0 else None)
        raw_name = cv(row[1] if len(row) > 1 else None)

        # 이메일이 없거나 @ 없으면 카테고리/메모 행 → 스킵
        if not raw_email or '@' not in raw_email:
            skip += 1
            continue

        # 매칭
        matched_id = None
        cur.execute(
            "SELECT candidate_id FROM candidates WHERE LOWER(TRIM(email)) = ?",
            (raw_email.lower().strip(),),
        )
        m = cur.fetchone()
        if m:
            matched_id = m[0]

        if matched_id is None and raw_name:
            cur.execute(
                "SELECT candidate_id FROM candidates WHERE LOWER(TRIM(full_name)) = ?",
                (raw_name.lower().strip(),),
            )
            m = cur.fetchone()
            if m:
                matched_id = m[0]

        if matched_id:
            # ★ 전체 덮어쓰기: Excel 값이 있으면 DB값 무조건 교체
            updates = ["status = 'Active'"]
            params = []
            changed_fields = []
            for col_i, db_field in effective.items():
                if db_field == 'email':
                    continue  # email은 매칭키 — full_name은 덮어쓰기 허용
                v = _clean_val(row[col_i] if len(row) > col_i else None)
                if not v:
                    continue
                # DB 현재값과 비교
                cur.execute(f"SELECT {db_field} FROM candidates WHERE candidate_id = ?", (matched_id,))
                cur_row = cur.fetchone()
                cur_val = str(cur_row[0]).strip() if cur_row and cur_row[0] else ''
                if v != cur_val:
                    updates.append(f"{db_field} = ?")
                    params.append(v)
                    changed_fields.append(db_field)

            updates.append("updated_at = ?")
            params.append(datetime.datetime.now().isoformat())
            params.append(matched_id)
            if not DRY_RUN:
                cur.execute(
                    f"UPDATE candidates SET {', '.join(updates)} WHERE candidate_id = ?",
                    params,
                )
            activated += 1
            if changed_fields:
                overwritten += 1
                if len(changed_fields) <= 5:
                    p(f"    {raw_name or raw_email}: {', '.join(changed_fields)}")
        else:
            # INSERT: 신규 후보자
            fields_vals = {}
            for col_i, db_field in effective.items():
                v = _clean_val(row[col_i] if len(row) > col_i else None)
                if v:
                    fields_vals[db_field] = v

            cid = _stable_id(fields_vals.get('email', ''), fields_vals.get('full_name', ''))
            fields_vals['candidate_id'] = cid
            fields_vals['status'] = 'Active'
            fields_vals['source'] = 'excel_new'
            fields_vals['inbox_status'] = 'new'
            fields_vals['is_deleted'] = 0
            fields_vals['created_at'] = datetime.datetime.now().isoformat()

            final_fields = {k: v for k, v in fields_vals.items() if k in db_cols or k == 'candidate_id'}
            cols = list(final_fields.keys())
            vals = list(final_fields.values())

            if not DRY_RUN:
                try:
                    ph = ','.join(['?'] * len(vals))
                    cur.execute(f"INSERT INTO candidates ({','.join(cols)}) VALUES ({ph})", vals)
                except sqlite3.IntegrityError:
                    skip += 1
                    continue
            inserted += 1

    if not DRY_RUN:
        conn.commit()

    p(f"  Active 마킹: {activated}건 (값 변경: {overwritten}건) / INSERT: {inserted}건 / SKIP: {skip}건")
    return activated, inserted


def import_client_sheet(wb, conn, ci_cols):
    """Client 시트 → client_inquiries 테이블 (UPSERT)."""
    p()
    p("=" * 60)
    p("=== Client 시트 → client_inquiries (UPSERT) ===")

    ws = wb['Client']
    cur = conn.cursor()

    # 기존 DB 인덱스
    cur.execute("SELECT id, LOWER(TRIM(email)), LOWER(TRIM(school_name)) FROM client_inquiries WHERE (is_deleted IS NULL OR is_deleted=0)")
    email_to_id = {}
    school_to_id = {}
    for rid, e, s in cur.fetchall():
        if e and '@' in e:
            # 쉼표로 여러 이메일 포함 가능 → 첫 번째만
            first_email = e.split(',')[0].strip()
            email_to_id[first_email] = rid
        if s:
            school_to_id[s] = rid

    update_count = 0
    insert_count = 0
    skip_count = 0
    next_id = cur.execute("SELECT COALESCE(MAX(CAST(id AS INTEGER)),0)+1 FROM client_inquiries").fetchone()[0]

    row_num = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1
        raw_region = cv(row[0] if len(row) > 0 else None)
        raw_city = cv(row[1] if len(row) > 1 else None)
        raw_school = cv(row[2] if len(row) > 2 else None)
        raw_email = cv(row[3] if len(row) > 3 else None)
        raw_phone = cv(row[4] if len(row) > 4 else None)
        raw_target = cv(row[5] if len(row) > 5 else None)
        raw_support = cv(row[7] if len(row) > 7 else None)
        raw_info = cv(row[8] if len(row) > 8 else None)
        raw_vacation = cv(row[9] if len(row) > 9 else None)
        raw_memo = cv(row[10] if len(row) > 10 else None)
        raw_other = cv(row[11] if len(row) > 11 else None)

        if not raw_school and not raw_email:
            skip_count += 1
            continue

        location = _parse_location(raw_region, raw_city)

        # 연락처 파싱: "010-xxxx-xxxx원장\n043-..." → phone + contact_name
        contact_name = ""
        phone_clean = ""
        if raw_phone:
            # 전화번호에서 직책(원장, 실장 등) 추출
            phone_str = str(raw_phone)
            name_match = re.search(r'(원장|실장|부장|대표|선생님|담당)', phone_str)
            if name_match:
                contact_name = name_match.group(0)
            # 전화번호만 추출 (첫 번째)
            phone_nums = re.findall(r'[\d\-]+', phone_str)
            phone_clean = phone_nums[0] if phone_nums else phone_str.split('\n')[0].strip()

        # 이메일 정규화 (쉼표 구분 → 첫 번째)
        first_email = ""
        if raw_email:
            first_email = raw_email.split(',')[0].strip().lower()

        # 기타/메일에서 추가 이메일 추출
        extra_email = ""
        if raw_other and '@' in str(raw_other):
            extra_email = str(raw_other).strip()

        # 매칭
        matched_id = None
        if first_email and '@' in first_email:
            matched_id = email_to_id.get(first_email)
        if matched_id is None and raw_school:
            matched_id = school_to_id.get(raw_school.lower().strip())

        if matched_id:
            # UPDATE: 빈 필드만 채우기
            updates = []
            params = []

            def _fill(field, val):
                if not val or field not in ci_cols:
                    return
                cur.execute(f"SELECT {field} FROM client_inquiries WHERE id = ?", (matched_id,))
                r = cur.fetchone()
                cur_val = str(r[0]).strip() if r and r[0] else ''
                if not cur_val:
                    updates.append(f"{field} = ?")
                    params.append(val)

            _fill('location', location)
            _fill('phone', phone_clean)
            _fill('teaching_age', raw_target)
            _fill('travel_support', raw_support)
            _fill('vacation', raw_vacation)
            if raw_info:
                _fill('memo', raw_info)
            if raw_memo:
                _fill('notes', raw_memo)

            if updates:
                params.append(matched_id)
                if not DRY_RUN:
                    cur.execute(
                        f"UPDATE client_inquiries SET {', '.join(updates)} WHERE id = ?",
                        params,
                    )
                update_count += 1
        else:
            # INSERT
            fields = {}
            fields['school_name'] = raw_school or ""
            fields['email'] = raw_email or extra_email or ""
            fields['location'] = location
            if phone_clean:
                fields['phone'] = phone_clean
            if contact_name:
                fields['contact_name'] = contact_name
            if raw_target:
                fields['teaching_age'] = raw_target
            if raw_support:
                fields['travel_support'] = raw_support
            if raw_info:
                fields['memo'] = raw_info
            if raw_vacation:
                fields['vacation'] = raw_vacation
            if raw_memo:
                fields['notes'] = raw_memo
            fields['source'] = 'excel_client'
            fields['is_deleted'] = 0
            fields['inbox_status'] = 'new'
            fields['submitted_at'] = datetime.datetime.now().isoformat()

            # DB에 있는 컬럼만
            final = {k: v for k, v in fields.items() if k in ci_cols}
            if not final.get('school_name') and not final.get('email'):
                skip_count += 1
                continue

            cols = list(final.keys())
            vals = list(final.values())

            if not DRY_RUN:
                ph = ','.join(['?'] * len(vals))
                cur.execute(f"INSERT INTO client_inquiries ({','.join(cols)}) VALUES ({ph})", vals)
            insert_count += 1

            # 인덱스 갱신
            if first_email and '@' in first_email:
                email_to_id[first_email] = next_id
            if raw_school:
                school_to_id[raw_school.lower().strip()] = next_id
            next_id += 1

    if not DRY_RUN:
        conn.commit()

    p(f"  총 행: {row_num}")
    p(f"  UPDATE: {update_count}건 / INSERT: {insert_count}건 / SKIP: {skip_count}건")
    return update_count, insert_count


# ══════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════
def main():
    mode = "DRY RUN (DB 변경 없음)" if DRY_RUN else "LIVE (DB 변경)"
    p(f"import_latest.py — {mode}")
    p(f"Excel: {XLSX_PATH}")
    p(f"DB: {DB_PATH}")
    p(f"시트 필터: {SHEET_FILTER or '전체'}")

    if not os.path.exists(XLSX_PATH):
        p(f"ERROR: Excel 파일 없음: {XLSX_PATH}")
        sys.exit(1)

    # 백업
    if not DRY_RUN:
        ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_path = f"{DB_PATH}.backup_{ts}"
        shutil.copy(DB_PATH, backup_path)
        p(f"백업: {backup_path}")

    # PRE 통계
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout = 10000")
    conn.row_factory = sqlite3.Row

    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM candidates")
    pre_cand = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM client_inquiries")
    pre_inq = cur.fetchone()[0]
    p(f"\nPRE: candidates={pre_cand} / client_inquiries={pre_inq}")

    # DB 컬럼 목록
    cur.execute("PRAGMA table_info(candidates)")
    cand_cols = {r[1] for r in cur.fetchall()}
    cur.execute("PRAGMA table_info(client_inquiries)")
    ci_cols = {r[1] for r in cur.fetchall()}

    # Excel 열기 (read_only + data_only → 대용량 지원)
    p(f"\nExcel 열기 중 (258MB, 잠시 대기)...")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    p(f"시트: {wb.sheetnames}")

    # 임포트 실행
    if SHEET_FILTER is None or SHEET_FILTER == 'source':
        import_source_sheet(wb, conn, cand_cols)

    if SHEET_FILTER is None or SHEET_FILTER == 'new':
        import_new_sheet(wb, conn, cand_cols)

    if SHEET_FILTER is None or SHEET_FILTER == 'client':
        import_client_sheet(wb, conn, ci_cols)

    wb.close()

    # POST 통계
    cur.execute("SELECT COUNT(*) FROM candidates")
    post_cand = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM client_inquiries")
    post_inq = cur.fetchone()[0]

    p()
    p("=" * 60)
    p("=== 최종 결과 ===")
    p(f"  candidates:       {pre_cand} → {post_cand}  (+{post_cand - pre_cand})")
    p(f"  client_inquiries: {pre_inq} → {post_inq}  (+{post_inq - pre_inq})")

    if not DRY_RUN:
        cur.execute("PRAGMA integrity_check")
        p(f"  integrity: {cur.fetchone()[0]}")

    conn.close()

    # 리포트 저장
    report_path = f"{BASE}/tasks/import_report_latest.txt"
    os.makedirs(os.path.dirname(report_path), exist_ok=True)
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    p(f"\n리포트: {report_path}")


if __name__ == '__main__':
    main()
